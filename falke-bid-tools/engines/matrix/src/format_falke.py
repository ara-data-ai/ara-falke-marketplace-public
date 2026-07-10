"""
FALKE Matrix Pipeline — Falke House Style (visual constants)
=============================================================
The exact colors, fonts, borders, and number formats of Falke's manual matrix
house style, extracted programmatically from Falke's own reference file
(SAMPLE TOWER - Bid Comparison Matrix.xlsx, sheet Bid_Form) —
see FALKE/03_Matrix/FALKE-HOUSE-FORMAT-SPEC.md for the full extraction.

Consumed by write_matrix._populate_leveled_sheet, which writes the
Leveled_Normalized sheet in the house geometry with these styles applied
inline at write time (v0.3.0 — the earlier post-pass formatter is gone; the
leveled sheet is now BORN in the house format).

GRAND TOTAL font note: the FEB 26 reference renders the GRAND TOTAL amount in
Calibri 14. DECIDED (Derick, 2026-07-02): that is drift, not design — GT
amounts are NORMALIZED to Avenir Book 12 bold white, single underline kept.
"""

from __future__ import annotations

from openpyxl.styles import Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Falke house palette (exact hex, extracted from the FEB 26 reference)
# ---------------------------------------------------------------------------

FALKE_FONT_NAME = "Avenir Book"   # exact name stored in the reference xlsx
FALKE_AQUA = "A3EAF3"             # light aqua — header band + subtotal rows
FALKE_TEAL = "00A9CA"             # strong teal — construction subtotal + grand total
FALKE_GRAY = "A6A6A6"             # gray — contractor header block (white text)
FALKE_WHITE = "FFFFFF"

# Accounting number format used on every amount / $/SF cell in the reference.
FALKE_AMOUNT_FORMAT = (
    '_([$$-409]* #,##0.00_);_([$$-409]* \\(#,##0.00\\);'
    '_([$$-409]* "-"??_);_(@_)'
)

# Percentage format for the VAR % and % SPREAD columns.
FALKE_PCT_FORMAT = "0.0%"

AQUA_FILL = PatternFill("solid", fgColor=FALKE_AQUA)
TEAL_FILL = PatternFill("solid", fgColor=FALKE_TEAL)
GRAY_FILL = PatternFill("solid", fgColor=FALKE_GRAY)

_THIN = Side(style="thin")
_DOUBLE = Side(style="double")

# v0.3.1 (v2-reference addendum): the reference's vertical rails are thin
# GRAY #7F7F7F (theme lt1 tint -0.5 / literal FF7F7F7F — same color), run
# through BLANK cells, and cover every table column. Left/right edges below
# are gray for full fidelity.
FALKE_RAIL_GRAY = "FF7F7F7F"
RAIL_SIDE = Side(style="thin", color=FALKE_RAIL_GRAY)

# Subtotal accents (APPROVED 2026-07-05): the reference's
# accounting top-thin/bottom-double on subtotal rows is DARK RED, not black.
# The reference stores two hues one RGB unit apart (#D74648 on every division
# subtotal row; #D74547 on the CCS row alone) — visually identical, same
# design intent stored twice; NORMALIZED to the dominant #D74648 per the
# drift-not-design precedent (see the spec addendum). Header/GT box
# horizontals stay black — the all-gray-box observation was NOT approved.
FALKE_ACCENT_RED = "FFD74648"
_RED_THIN = Side(style="thin", color=FALKE_ACCENT_RED)
_RED_DOUBLE = Side(style="double", color=FALKE_ACCENT_RED)

RAIL_BORDER = Border(left=RAIL_SIDE, right=RAIL_SIDE)    # column rails
BOX_BORDER = Border(left=RAIL_SIDE, right=RAIL_SIDE, top=_THIN, bottom=_THIN)
SUBTOTAL_BORDER = Border(left=RAIL_SIDE, right=RAIL_SIDE, top=_RED_THIN,
                         bottom=_RED_DOUBLE)

HEADER_ROW_HEIGHT = 46.5  # reference row 12

# ---------------------------------------------------------------------------
# House fonts (Avenir Book throughout; reference sizes)
# ---------------------------------------------------------------------------

BODY_FONT = Font(name=FALKE_FONT_NAME, size=12)
BOLD_FONT = Font(name=FALKE_FONT_NAME, size=12, bold=True)
ITALIC_FONT = Font(name=FALKE_FONT_NAME, size=12, italic=True)
WHITE_BOLD_FONT = Font(name=FALKE_FONT_NAME, size=12, bold=True, color=FALKE_WHITE)
BIDDER_BLOCK_FONT = Font(name=FALKE_FONT_NAME, size=11, bold=True, color=FALKE_WHITE)
BANNER_BOLD_FONT = Font(name=FALKE_FONT_NAME, size=12, bold=True, color=FALKE_WHITE)
BANNER_BODY_FONT = Font(name=FALKE_FONT_NAME, size=12, color=FALKE_WHITE)
DISCLAIMER_FONT = Font(name=FALKE_FONT_NAME, size=10, italic=True)

# GRAND TOTAL amount cells: NORMALIZED to the house font per Derick
# (2026-07-02) — the FEB 26 reference's literal Calibri-14 there is treated
# as drift, not design. Avenir Book 12 bold white, single underline kept,
# merge/center/height kept. $/SF span: Avenir Book 12 bold white,
# right-aligned, no underline (as reference G164).
GT_AMOUNT_FONT = Font(name=FALKE_FONT_NAME, size=12, bold=True,
                      underline="single", color=FALKE_WHITE)
GT_SF_FONT = Font(name=FALKE_FONT_NAME, size=12, bold=True, color=FALKE_WHITE)

# Reference row heights for the two teal total rows (FEB 26 rows 154 / 164).
CCS_ROW_HEIGHT = 16.5
GT_ROW_HEIGHT = 19.0
