"""Scorecard run-pack EMITTER (P1-4, verdict c; Marvin's ratification §3, §8.1).

The matrix run ends by writing ONE ``<Project> - Scorecard Inputs.xlsx`` beside
the matrix: four tabs (Settings / Baseline / Framework / Scores), with the
firms, the project identity, and the SF suggestion pre-filled from the run that
just produced the matrix, stamped with the matrix run_id + producer format
version.

WHY THIS LIVES IN THE MATRIX ENGINE
-----------------------------------
Because the disclosure moment is the end of the matrix run, days before scoring
(Boris §B.3, ratified). The operator learns about the two files with the longest
lead time when they can still act on it, and the firm names originate from the
pipeline instead of a human re-typing eight names into a grid that hard-stops on
a typo. That failure class is 100% manufactured; this file is where it dies.

WHAT THE PACK DOES *NOT* DO
---------------------------
It does not confirm anything. The SF gate and the baseline gate still run, per
run, answered by a human in the conversation. Pre-filled is not pre-confirmed
(Marvin §5): project name and address are IDENTITY (facts about the building,
free to pre-fill); SF is a JUDGMENT (which denominator is the right one for this
comparison), and a suggestion does not become a confirmation by traveling
through a spreadsheet. The schema contains NO confirmation field of any kind —
see pack_schema.R1 before you consider adding one.

The friction we remove is data entry. The friction we keep is judgment.
"""

from __future__ import annotations

import datetime as _dt
import importlib.util
import sys
from pathlib import Path
from typing import Optional, Sequence

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.protection import SheetProtection

from src.normalized_models import NormalizedBid
from src.run_config import RunInputs
from src.write_matrix import (PRODUCER_FORMAT_VERSION, PRODUCER_NAME)

# ---------------------------------------------------------------------------
# The producer -> consumer schema import
# ---------------------------------------------------------------------------
# Floyd's verdict (c): "The producer->consumer schema dependency is acceptable
# because both engines ship in one plugin at one version." This is that
# dependency, made explicit and single-sourced rather than duplicated: every
# label, sheet name, and enum comes from the scorecard engine's stdlib-only
# pack_schema module. Copying the labels into this file instead would recreate
# the two-copy drift that P0-1 existed to kill.
#
# It is loaded by path (not by `import`) because the two engines are separate
# Python packages run with different PYTHONPATHs; they are always siblings under
# engines/ in the same plugin, so the path is deterministic. If it is ever not
# there, we fail LOUDLY and immediately — a silently skipped pack is exactly the
# kind of quiet degradation this program exists to end.
_SCHEMA_PATH = (Path(__file__).resolve().parents[2]
                / "scorecard" / "scorecard" / "pack_schema.py")


def _load_schema():
    if not _SCHEMA_PATH.exists():
        raise RuntimeError(
            f"Run-pack schema not found at {_SCHEMA_PATH}. The matrix and "
            f"scorecard engines ship together in one plugin at one version; a "
            f"missing sibling means a broken install, not an optional feature. "
            f"Reinstall the plugin.")
    spec = importlib.util.spec_from_file_location(
        "falke_pack_schema", str(_SCHEMA_PATH))
    module = importlib.util.module_from_spec(spec)
    sys.modules.setdefault("falke_pack_schema", module)
    spec.loader.exec_module(module)
    return module


ps = _load_schema()

# ---------------------------------------------------------------------------
# Visual convention (Marvin R4) — carried from the existing templates
# ---------------------------------------------------------------------------
# | Cell class                    | Fill   | Locked | Meaning                  |
# | Producer-filled, verified     | F2F2F2 | yes    | Re-derived at parse.     |
# | Producer-filled, advisory echo| F2F2F2 | yes    | A suggestion (SF).       |
# | Operator-entry, required      | white  | no     | Blank = exit 2.          |
# | Operator-entry, optional      | white  | no     | Fill if applicable.      |
# | Header                        | D9E1F2 | yes    | Reshaping it = exit 2.   |
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                          fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                          fill_type="solid")
TITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
LABEL_FONT = Font(bold=True)
BLOCK_FONT = Font(bold=True, size=11)


def _protect(ws) -> None:
    """R3: "Locked" in xlsx is ADVISORY UI, never integrity.

    Sheet protection is trivially removed and Excel honors it only as friction.
    So: no password — a password implies a boundary that does not exist and
    would lock Falke out of their own file. The integrity is the parser's
    re-derivation of every producer-filled field (run_pack.py), which never
    trusts the lock.
    """
    ws.protection = SheetProtection(sheet=True, password=None)
    ws.protection.sheet = True


def _locked(cell, *, value=None, header: bool = False):
    if value is not None:
        cell.value = value
    cell.fill = HEADER_FILL if header else LOCKED_FILL
    if header:
        cell.font = HEADER_FONT
    cell.protection = openpyxl.styles.Protection(locked=True)
    return cell


def _operator(cell, *, value=None):
    if value is not None:
        cell.value = value
    cell.protection = openpyxl.styles.Protection(locked=False)
    return cell


# ---------------------------------------------------------------------------
# Falke's standing framework (Marvin §4.3) — the reference, when it exists
# ---------------------------------------------------------------------------

class StandingFramework:
    """Falke's standing evaluation framework: version, effective date, rows,
    and the semantic hash of its (short_label, weight) pairs.

    ``available`` is False when Falke supplied no standing-framework file. That
    is TODAY'S REALITY (§10.2) and it is not an error: the pack is still emitted,
    the Framework tab is still pre-filled from ARA's shipped default CONTENT,
    Settings records `none (shipped default)`, and the drift check degrades to
    W8 (WARN, always) claiming nothing about policy drift. It degrades honestly.
    """

    def __init__(self, *, available: bool, version: str, effective_date: str,
                 rows: Sequence, semantic_hash: str):
        self.available = available
        self.version = version
        self.effective_date = effective_date
        self.rows = list(rows)
        self.semantic_hash = semantic_hash

    @classmethod
    def shipped_default(cls) -> "StandingFramework":
        """No standing reference on file -> W8 bootstrap.

        The Framework tab still needs usable starting CONTENT, so it is
        pre-filled from ARA's shipped default rows. Settings then says
        `none (shipped default)` and the hash is EMPTY — deliberately. An empty
        hash is what makes W8 unfakeable downstream: there is nothing to compare
        against, so the check cannot claim what it does not know, and the card
        must not claim a standing framework that does not exist (§4.5).
        """
        return cls(available=False, version=ps.STANDING_NONE,
                   effective_date="", rows=ps.DEFAULT_FRAMEWORK_ROWS,
                   semantic_hash="")


def read_standing_framework(path: Optional[str]) -> StandingFramework:
    """Read Falke's ``standing-framework.xlsx`` — THE SEAM (§10.2).

    No such artifact exists today, so this returns ``shipped_default()`` when
    ``path`` is None. The schema it expects when Falke DOES adopt one is defined
    in pack_schema (STANDING_SHEET / SF_VERSION / SF_EFFECTIVE_DATE + a
    framework table identical in shape to the Framework tab). Nothing about the
    downstream design changes when it lands: the hash is already computed and
    stamped, the W1/W2 BLOCKER tiers already exist, and W8 simply stops firing.

    Deliberately NOT built here: any search for a standing framework by
    convention (beside the matrix, beside the project config, in the plugin).
    Guessing which file is Falke's evaluation policy is precisely the assertion
    the tool must not make. The path is explicit (--standing-framework) or there
    is no reference.
    """
    if not path:
        return StandingFramework.shipped_default()

    wb = openpyxl.load_workbook(path, data_only=True)
    if ps.STANDING_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Expected a sheet named '{ps.STANDING_SHEET}' in the standing "
            f"framework '{path}'; found: {', '.join(wb.sheetnames) or '(none)'}.")
    ws = wb[ps.STANDING_SHEET]

    scalars, table_header_row = _read_label_block(ws, ps.STANDING_SCALAR_FIELDS,
                                                  ps.FRAMEWORK_HEADERS, path)
    version = str(scalars.get(ps.SF_VERSION) or "").strip()
    effective = _fmt_date(scalars.get(ps.SF_EFFECTIVE_DATE))
    if not version or not effective:
        raise ValueError(
            f"The standing framework '{path}' must carry both a "
            f"'{ps.SF_VERSION}' and an '{ps.SF_EFFECTIVE_DATE}'. A framework "
            f"with no version and no effective date cannot be a reference "
            f"point for a drift check — it is the one artifact in this design "
            f"that predates bid opening, and that is the whole of its value.")

    rows = _read_framework_table(ws, table_header_row, path)
    return StandingFramework(
        available=True, version=version, effective_date=effective, rows=rows,
        semantic_hash=ps.framework_semantic_hash(
            [(r[1], r[2]) for r in rows]))


def _read_label_block(ws, scalar_fields, table_headers, path):
    """Label-addressed scalar read + a header-row scan for the table (R2)."""
    wanted = {ps.norm_label(f): f for f in scalar_fields}
    header_keys = [ps.norm_label(h) for h in table_headers]
    scalars = {}
    table_header_row = None
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        key = ps.norm_label(a)
        if not key:
            continue
        if key == header_keys[0]:
            got = [ps.norm_label(ws.cell(row=row, column=c).value)
                   for c in range(1, len(header_keys) + 1)]
            if got == header_keys:
                table_header_row = row
                break
        if key in wanted:
            scalars[wanted[key]] = ws.cell(row=row, column=2).value
    if table_header_row is None:
        raise ValueError(
            f"'{path}': could not find the framework table header row "
            f"({' | '.join(table_headers)}). Do not reshape the template.")
    return scalars, table_header_row


def _read_framework_table(ws, header_row, path):
    """Rows below ``header_row`` until the first fully-blank row (R2)."""
    rows = []
    r = header_row + 1
    while True:
        vals = [ws.cell(row=r, column=c).value for c in (1, 2, 3, 4)]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        category, short_label, weight, captures = vals
        if not str(category or "").strip() or not str(short_label or "").strip():
            raise ValueError(
                f"'{path}' row {r}: every framework row needs a Category and a "
                f"Short Label.")
        try:
            weight_f = float(weight)
        except (TypeError, ValueError):
            raise ValueError(
                f"'{path}' row {r}: 'Weight (%)' must be numeric; got "
                f"{weight!r}.")
        rows.append((str(category).strip(), str(short_label).strip(), weight_f,
                     str(captures).strip() if captures is not None else ""))
        r += 1
    if not rows:
        raise ValueError(f"'{path}': no framework rows found below the header.")
    total = sum(r[2] for r in rows)
    if abs(total - 100.0) > 0.01:
        raise ValueError(
            f"'{path}': framework weights must sum to 100; they sum to "
            f"{total:g}.")
    return rows


def _fmt_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


# ---------------------------------------------------------------------------
# The emitter
# ---------------------------------------------------------------------------

def emit_scorecard_pack(
    *,
    out_dir: Path,
    matrix_path: Path,
    run: RunInputs,
    run_id: str,
    bids: Sequence[NormalizedBid],
    matrix_exclusions: Sequence[tuple],
    standing: StandingFramework,
) -> Path:
    """Write ``<Project> - Scorecard Inputs.xlsx`` beside the matrix.

    ``bids`` is the matrix's SCORED-BIDDER ROSTER in matrix order — exactly the
    bidders that have a column in the workbook. ``matrix_exclusions`` is the run's
    own INPUT_EXCLUDED rulings as (identifier, reason) pairs: bids that never
    made it into the matrix, echoed here with their logged reason so the
    scorecard does not re-ask (§3.1). An operator may ADD exclusions with a
    reason; an operator may NOT un-exclude a matrix-excluded bidder through the
    pack — reversing a matrix ruling means re-running the matrix, where the
    ruling lives.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    pack_path = out_dir / ps.pack_filename(run.project_name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    framework_rows = list(standing.rows)
    _write_settings(wb, run=run, run_id=run_id, matrix_path=matrix_path,
                    bids=bids, matrix_exclusions=matrix_exclusions,
                    standing=standing)
    _write_baseline(wb)
    _write_framework(wb, framework_rows, standing)
    _write_scores(wb, bids, framework_rows)

    _stamp_pack(wb, run_id=run_id, standing=standing)
    wb.save(pack_path)
    return pack_path


def _stamp_pack(wb, *, run_id: str, standing: StandingFramework) -> None:
    """The pack's own stamp (§8.1): its schema version PLUS the matrix's
    producer / format version / run id, PLUS the standing-framework reference.

    The standing triple is stamped, not merely printed, because it is the ONE
    part of the pack that is not re-derivable from the matrix. If it lived only
    in a visible cell, an operator could mask a policy drift by editing the hash
    they are being measured against. The property is authoritative; the Settings
    row is courtesy (R3/§8.1).
    """
    from openpyxl.packaging.custom import StringProperty
    props = [
        (ps.PACK_STAMP_FORMAT_PROP, ps.PACK_FORMAT_VERSION),
        (ps.PACK_STAMP_PRODUCER_PROP, PRODUCER_NAME),
        (ps.PACK_STAMP_MATRIX_FORMAT_PROP, PRODUCER_FORMAT_VERSION),
        (ps.PACK_STAMP_MATRIX_RUN_ID_PROP, run_id),
        (ps.PACK_STAMP_STANDING_VERSION_PROP, standing.version),
        (ps.PACK_STAMP_STANDING_DATE_PROP, standing.effective_date),
        (ps.PACK_STAMP_STANDING_HASH_PROP, standing.semantic_hash),
    ]
    for name, value in props:
        wb.custom_doc_props.append(
            StringProperty(name=name, value=str(value if value is not None else "")))


def _title(ws, text: str, span: str) -> None:
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT
    ws.merge_cells(span)


def _write_settings(wb, *, run, run_id, matrix_path, bids, matrix_exclusions,
                    standing) -> None:
    """The identity-and-binding tab (§3.1).

    Its scope boundary (§7) is what keeps it from becoming the killed
    run-manifest YAML by another name: Settings consolidates INPUTS. It never
    consolidates DECISIONS. Not present, by ruling: any gate answer, --sheet
    selection (a pack could otherwise circulate with Bid_Form baked in and
    produce an un-leveled board card without anyone typing a flag — a gate
    bypass by data), out-dir, the band (R6 — it lives on Baseline), any tuning
    knob, and any send target.
    """
    ws = wb.create_sheet(title=ps.SHEET_SETTINGS)
    _title(ws, "SCORECARD RUN PACK — SETTINGS (identity, lineage, and "
               "evaluation conventions)", "A1:C1")

    values = {
        ps.S_PACK_FORMAT_VERSION: ps.PACK_FORMAT_VERSION,
        ps.S_PRODUCER: PRODUCER_NAME,
        ps.S_MATRIX_FORMAT_VERSION: PRODUCER_FORMAT_VERSION,
        ps.S_MATRIX_RUN_ID: run_id,
        # Human-legible lineage; a NAME, never a path (R7 — a machine path is a
        # fact about the session, and a travelling pack carries a wrong one).
        ps.S_MATRIX_FILE_NAME: Path(matrix_path).name,
        ps.S_EMITTED_AT: _dt.datetime.now().replace(microsecond=0).isoformat(),
        ps.S_PROJECT_NAME: run.project_name,
        ps.S_PROJECT_ADDRESS: run.project_address,
        # ADVISORY ECHO — not identity, not a confirmation (§5.1). The matrix's
        # GSF is what the bid documents happened to say; which denominator is
        # right for this comparison is a fiduciary judgment that changes every
        # ratio on the card, and it is the whole reason the SF gate exists.
        # Editing this cell is an override PROPOSAL, surfaced at preview,
        # semantically identical to --sf-basis. It never silently becomes the
        # basis.
        ps.S_SF_BASIS_VALUE: float(run.gross_sf),
        ps.S_SF_BASIS_LABEL: run.sf_basis_label or "GSF",
        ps.S_STANDING_VERSION: standing.version,
        ps.S_STANDING_EFFECTIVE: standing.effective_date,
        ps.S_STANDING_HASH: standing.semantic_hash,
    }

    row = 3
    for field in ps.SETTINGS_PRODUCER_FIELDS:
        _locked(ws.cell(row=row, column=1, value=field)).font = LABEL_FONT
        _locked(ws.cell(row=row, column=2), value=values[field])
        row += 1

    # ---- operator-entry scalars ----
    _operator(ws.cell(row=row, column=1))
    ws.cell(row=row, column=1, value=ps.S_BID_OPENING_DATE).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3,
            value="REQUIRED (YYYY-MM-DD). The date bids were opened — the clock "
                  "this evaluation is measured against. The matrix cannot know "
                  "it.")
    row += 1
    ws.cell(row=row, column=1, value=ps.S_ADDENDA_THROUGH).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3,
            value="Optional. The last addendum this evaluation reflects.")
    row += 2

    # ---- Display Aliases (kills the operator-facing JSON — Boris §B.4/P1-5) --
    row = _write_table_block(
        ws, row, ps.T_DISPLAY_ALIASES, ps.ALIAS_HEADERS, [],
        note="Optional. Short board-facing names. Left column must match a firm "
             "in the matrix exactly as spelled on the Scores tab.",
        locked_rows=False)

    # ---- Matrix Exclusions: the run's OWN rulings, echoed and locked ---------
    # The matrix run already made these with logged reasons. Echoed so the
    # scorecard does not re-ask; locked because the pack must never become a
    # quiet route to re-admit a bidder the matrix ruled out. Editing this block
    # fails the parser's re-derivation against the matrix AUDIT sheet (I8).
    row = _write_table_block(
        ws, row, ps.T_MATRIX_EXCLUSIONS, ps.EXCLUSION_HEADERS,
        [(str(name), str(reason)) for name, reason in matrix_exclusions],
        note="Set by the matrix run — do not edit. To reverse one of these, "
             "re-run the matrix, where the ruling lives.",
        locked_rows=True)

    # ---- Additional Exclusions: operator rulings, reason MANDATORY ----------
    # Blank reason = exit 2. No silent drops, ever (Marvin §1.5).
    _write_table_block(
        ws, row, ps.T_ADDITIONAL_EXCLUSIONS, ps.EXCLUSION_HEADERS, [],
        note="Optional. A firm to set aside from the scored field. A REASON IS "
             "MANDATORY — a blank reason stops the run. Firm must match a name "
             "on the Scores tab.",
        locked_rows=False)

    for col, width in (("A", 42), ("B", 44), ("C", 80)):
        ws.column_dimensions[col].width = width
    _protect(ws)


def _write_table_block(ws, row, label, headers, rows, *, note, locked_rows):
    """One Settings table block: a label row, a header row, then data until a
    fully-blank row (R2). Returns the next free row (one blank row after)."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BLOCK_FONT
    ws.cell(row=row, column=3, value=note)
    row += 1
    for idx, header in enumerate(headers, start=1):
        _locked(ws.cell(row=row, column=idx, value=header), header=True)
    row += 1
    for values in rows:
        for idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=idx, value=value)
            if locked_rows:
                _locked(cell)
            else:
                _operator(cell)
        row += 1
    if not rows and not locked_rows:
        # leave a few unlocked rows so the operator has somewhere to type
        for _ in range(4):
            for idx in range(1, len(headers) + 1):
                _operator(ws.cell(row=row, column=idx))
            row += 1
    return row + 1


def _write_baseline(wb) -> None:
    """The cost yardstick and its provenance (§3.2).

    ENTIRELY operator-entered: the producer knows nothing about Falke's estimate
    and must never appear to. Note the geometry — the provenance block sits in
    the label-addressed header block ABOVE the trade lines, so P1-6 adds rows
    without shifting an index. That is the entire reason R2 is a hard rule
    rather than a style preference: the current baseline-template.xlsx puts the
    band at rows 3-5 and the table header at row 7, and inserting a six-field
    provenance block there under a row-index parser breaks every downstream
    read. Under R2 it is additive and free.
    """
    ws = wb.create_sheet(title=ps.SHEET_BASELINE)
    _title(ws, "COST BASELINE — the yardstick this scorecard measures against",
           "A1:E1")

    helptext = {
        ps.B_BAND_LOW: "REQUIRED. Low end of the modeled baseline band, $M.",
        ps.B_BAND_HIGH: "REQUIRED. High end of the modeled baseline band, $M.",
        ps.B_BAND_MID: "REQUIRED. Modeled mid (takeoff), $M.",
        ps.B_PROVENANCE: ("REQUIRED. One of: independent | bid-informed | "
                          "bid-derived. How this baseline was arrived at. The "
                          "tool does not know and will not guess."),
        ps.B_ESTIMATOR: "REQUIRED. Name + firm of the estimator of record.",
        ps.B_BASIS_DATE: "Optional. Date of the drawing/spec set priced.",
        ps.B_BASIS_DOCUMENTS: "Optional. Drawing/spec set identifier.",
        ps.B_MI_SIRS_DERIVED: "Optional. Y/N.",
        ps.B_MI_SIRS_CONFLICT: (
            "Optional. Y/N/Unknown. Facts only — the tool never opines; a Y "
            "routes to the association's attorney."),
    }

    row = 3
    for field in ps.BASELINE_SCALAR_FIELDS:
        ws.cell(row=row, column=1, value=field).font = LABEL_FONT
        _operator(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=helptext[field])
        row += 1
    row += 1

    for idx, header in enumerate(ps.BASELINE_TRADE_HEADERS, start=1):
        cell = _locked(ws.cell(row=row, column=idx, value=header), header=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1
    for _ in range(12):
        for idx in range(1, len(ps.BASELINE_TRADE_HEADERS) + 1):
            _operator(ws.cell(row=row, column=idx))
        row += 1

    for col, width in (("A", 46), ("B", 30), ("C", 76), ("D", 14), ("E", 10)):
        ws.column_dimensions[col].width = width
    _protect(ws)


def _write_framework(wb, framework_rows, standing) -> None:
    """The evaluation PLAN (§3.3) — pre-filled; the operator's job here is
    normally to do NOTHING.

    The framework table is deliberately UNLOCKED. Locking it would make the pack
    pretend the plan is immutable, which is false — a roofing-only package
    legitimately carries different weights than a full interior renovation. The
    control is not prevention; it is DECLARATION + DETECTION + DISCLOSURE (§4).
    Severity keys on the declaration, not on the drift: the harm was never the
    deviation, the harm is the silence.
    """
    ws = wb.create_sheet(title=ps.SHEET_FRAMEWORK)
    _title(ws, "SCORING FRAMEWORK — the evaluation plan (weights must sum to "
               "100)", "A1:D1")

    basis_note = (
        f"REQUIRED. One of: standing | project-specific | revised-post-opening. "
        f"Pre-filled '{ps.BASIS_STANDING}'. Leave it alone if the weights below "
        f"are unchanged.")
    if not standing.available:
        basis_note = (
            "REQUIRED. One of: standing | project-specific | "
            "revised-post-opening. No standing framework file was supplied for "
            "this run, so the weights below are ARA's shipped default starting "
            "point — declare what they actually are for this project.")

    row = 3
    ws.cell(row=row, column=1, value=ps.F_BASIS).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2), value=ps.BASIS_STANDING)
    ws.cell(row=row, column=3, value=basis_note)
    row += 1
    ws.cell(row=row, column=1, value=ps.F_LOCK_DATE).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2),
              value=standing.effective_date or None)
    ws.cell(row=row, column=3,
            value="Pre-filled with the standing framework's effective date. Set "
                  "it yourself ONLY when declaring 'project-specific' — the "
                  "date the weights were locked, before bids were opened.")
    row += 1
    ws.cell(row=row, column=1, value=ps.F_RULING_NOTE).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3,
            value="REQUIRED for any basis other than 'standing'. e.g. 'board "
                  "approved revised weights on <date>'. A blank note stops the "
                  "run.")
    row += 2

    for idx, header in enumerate(ps.FRAMEWORK_HEADERS, start=1):
        cell = _locked(ws.cell(row=row, column=idx, value=header), header=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1
    for category, short_label, weight, captures in framework_rows:
        _operator(ws.cell(row=row, column=1), value=category)
        _operator(ws.cell(row=row, column=2), value=short_label)
        cell = _operator(ws.cell(row=row, column=3), value=float(weight))
        cell.number_format = "0.##"
        cell = _operator(ws.cell(row=row, column=4), value=captures)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    for _ in range(4):
        for idx in range(1, len(ps.FRAMEWORK_HEADERS) + 1):
            _operator(ws.cell(row=row, column=idx))
        row += 1

    for col, width in (("A", 38), ("B", 16), ("C", 12), ("D", 72)):
        ws.column_dimensions[col].width = width
    _protect(ws)


def _write_scores(wb, bids, framework_rows) -> None:
    """The evaluation RECORD (§3.4), with one decisive improvement: the Firm
    column is producer-filled and LOCKED.

    THIS IS THE RE-KEYING KILLER — exactly the matrix's scored-bidder roster, in
    matrix order. Excluded bidders never appear. Deleting a row does not exclude
    a bidder; it breaks the roster reconcile and exits 2 naming the missing firm
    (exclusion is the Settings mechanism, with a reason).

    Blank-cell semantics are P1-2's, and P1-2 has landed: a blank means NOT YET
    SCORED. The pack CARRIES that semantic and does not redefine it — the pack
    parser and the individual-flag parser are the SAME code reading the SAME
    semantics (scoring_inputs.parse_scores_table), so the pack inherited it for
    free, exactly as designed. No divergence, ever.

    So a freshly emitted pack — a fully blank grid — is the maximally provisional
    case, and it is the product's DEFAULT starting state. The operator fills what
    they know and re-runs; the card renders PRELIMINARY, unranked, with a
    worklist. The one thing the scorecard refuses is a grid with NO scored cell
    anywhere: that is this template, unfilled, and there is nothing to combine.
    """
    ws = wb.create_sheet(title=ps.SHEET_SCORES)
    _title(ws, "CATEGORY SCORES — the evaluation record (1-10 per category; "
               "Overall is computed, never supplied)", "A1:D1")

    row = 3
    ws.cell(row=row, column=1, value=ps.SC_SCORING_COMPLETED_DATE).font = LABEL_FONT
    _operator(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3,
            value="Required once every score is entered — the date the "
                  "evaluation record was closed. Leave blank while scoring is "
                  "in progress; the card renders PRELIMINARY.")
    row += 2

    _locked(ws.cell(row=row, column=1, value=ps.SCORES_FIRM_HEADER), header=True)
    for idx, (_cat, short_label, _w, _c) in enumerate(framework_rows, start=2):
        cell = _locked(ws.cell(row=row, column=idx, value=short_label),
                       header=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    for bid in bids:
        _locked(ws.cell(row=row, column=1, value=bid.contractor_name))
        for idx in range(2, len(framework_rows) + 2):
            _operator(ws.cell(row=row, column=idx))
        row += 1

    ws.column_dimensions["A"].width = 34
    for idx in range(2, len(framework_rows) + 2):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(idx)].width = 13
    ws.freeze_panes = ws.cell(row=row - len(bids), column=2)
    _protect(ws)
