"""
FALKE Matrix Pipeline — Excel Writer (Fresh Workbook)
======================================================
Writes normalized bid data into a brand-new openpyxl Workbook().

No template is loaded or copied — the FEB 26 file's merged cells, conditional
formatting, and drawing objects caused corruption on save.  This module
creates a clean xlsx that mirrors the FEB 26 row/column structure closely
enough for side-by-side value comparison.

Pipeline position:
    list[NormalizedBid]  →  write_matrix()  →  bid-comparison .xlsx

Layout (Bid_Form mirror):
  Col A  : CSI code / label
  Col B  : Row description
  Col C  : Normalization Note
  Col D+ : Contractor groups, each 3 columns wide:
             +0  COST SUBTOTALS (main comparison number)
             +1  $/SF           (subtotal ÷ gsf)
             +2  blank separator

Layout (Leveled_Normalized, v0.3.0 — FEB 26 geometry): contractor groups are
5 columns wide (COST | COST SUBTOTALS | $/SF | $/SXFX SUBTOTALS | VAR %) with
a BENCHMARK block appended right of all groups; written by
_populate_leveled_sheet with the Falke house format + leveling rules.

  Row 1  : Project title
  Row 2  : Project details
  Row 3  : blank
  Row 4  : Column headers (CSI / Building System / COST SUBTOTALS / $/SF / …)
  Row 5  : Contractor names
  Row 6  : Project label per contractor
  Row 7  : GSF per contractor
  Row 8  : blank
  Rows 9+: CSI division data
  …      : blank separator, footer section, blank separator, qualifications
"""

from __future__ import annotations

import uuid
from decimal import Decimal
from math import ceil
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import falke_rules
from src import format_falke as ff
from src.audit import AuditItem, AuditStatus
from src.normalize import _div_short
from src.normalized_models import (
    GRAND_TOTAL_COMPONENT_KEYS,
    CellState,
    NormalizedBid,
    NormalizedDivision,
    ReclassRecommendation,
    grand_total_component_amounts,
)
from src.run_config import RunInputs

# Col C — the Normalization Note column on the Bid_Form mirror (Option C §2.1).
NOTE_COL = 3

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Project identity is a PER-RUN input (RunInputs), never hardcoded. These
# fallbacks exist only so a programmatic caller that passes a gsf but no
# RunInputs still produces a generic, non-client-specific title.
DEFAULT_SF_BASIS_LABEL = "GSF"

# ---------------------------------------------------------------------------
# Producer stamp (Floyd consolidated ruling 2026-07-15, verdict f — P0 wave 2)
# ---------------------------------------------------------------------------
# Written as workbook CUSTOM DOCUMENT PROPERTIES (invisible — never visible
# geometry) so the consuming scorecard can check the workbook against its
# SUPPORTED_PRODUCER range (engines/scorecard/scorecard/matrix.py).
# RELEASE.md step-2 TRIPWIRE: any minor+ (format-changing) release must bump
# PRODUCER_FORMAT_VERSION here AND revisit the scorecard's SUPPORTED_PRODUCER
# range in the same commit. The stamp also carries the run identity + project
# identity properties below; adding a property is additive (an older consumer
# ignores it) and does NOT itself require a format bump — changing the SHEET
# GEOMETRY does.
PRODUCER_NAME = "falke-bid-tools/matrix"
PRODUCER_FORMAT_VERSION = "0.4.0"
STAMP_PRODUCER_PROP = "falke_bid_tools.producer"
STAMP_FORMAT_PROP = "falke_bid_tools.format_version"

# --- Run identity (P1-4 dependency; Marvin §10.1 — belongs to P1-3's contract)
# The producer stamp carried producer + format version and NO run identity, so
# the run pack's pack->matrix binding (§8.3) rested on a field that did not
# exist. It does now: one more StringProperty, minted per run, opaque and
# collision-free. It is EVIDENCE, NOT A GATE (§8.4) — the roster is the gate.
STAMP_RUN_ID_PROP = "falke_bid_tools.run_id"

# --- Project identity. The pack's I3 rule ("pack project identity != matrix
# project identity -> exit 2, always") needs a matrix-side identity the consumer
# can re-derive INDEPENDENTLY of the pack. The sheets cannot supply it: the
# board-facing default sheet is Leveled_Normalized, whose geometry carries the
# project NAME only (row 6, per bidder) and never the ADDRESS, and row 1 is
# overwritten by the Stage-6b quarantine banner. So identity is stamped, using
# the same invisible-doc-property pattern Marvin ratified in §8.1 ("the
# properties are authoritative; the visible rows are courtesy").
STAMP_PROJECT_NAME_PROP = "falke_bid_tools.project_name"
STAMP_PROJECT_ADDRESS_PROP = "falke_bid_tools.project_address"
STAMP_SF_BASIS_LABEL_PROP = "falke_bid_tools.sf_basis_label"


def mint_run_id() -> str:
    """Mint an opaque, collision-free run identity for one matrix run.

    Called once at run start (src/pipeline.py) and carried into both the
    workbook stamp and the run pack's Settings tab, so the two are bound by
    construction rather than by the operator remembering which was which.
    """
    return uuid.uuid4().hex[:12]


def _stamp_workbook(wb: openpyxl.Workbook, run: RunInputs,
                    run_id: str) -> None:
    """Stamp producer, format version, run identity, and project identity as
    custom doc properties (additive; no cell, row, or sheet is touched)."""
    from openpyxl.packaging.custom import StringProperty

    props = [
        (STAMP_PRODUCER_PROP, PRODUCER_NAME),
        (STAMP_FORMAT_PROP, PRODUCER_FORMAT_VERSION),
        (STAMP_RUN_ID_PROP, run_id),
        (STAMP_PROJECT_NAME_PROP, run.project_name),
        (STAMP_PROJECT_ADDRESS_PROP, run.project_address),
        (STAMP_SF_BASIS_LABEL_PROP,
         run.sf_basis_label or DEFAULT_SF_BASIS_LABEL),
    ]
    for name, value in props:
        wb.custom_doc_props.append(
            StringProperty(name=name, value=str(value if value is not None else "")))

# Column widths
COL_A_WIDTH: float = 15.0
COL_B_WIDTH: float = 45.0
CONTRACTOR_COST_WIDTH: float = 18.0
CONTRACTOR_SF_WIDTH: float = 10.0
CONTRACTOR_SEP_WIDTH: float = 4.0

# Number format for dollar amounts
AMOUNT_FORMAT = "#,##0.00"
# Accounting-negative variant for net-credit cells on the mirror (W-D ruling
# 5.2: negatives render in parentheses, never clamped). The leveled sheet's
# FALKE_AMOUNT_FORMAT already carries the accounting parentheses.
AMOUNT_FORMAT_NEG = "#,##0.00;(#,##0.00)"

# ---------------------------------------------------------------------------
# Division and footer row definitions (FEB 26 CSI sequence)
# ---------------------------------------------------------------------------

DIVISION_ROWS: list[tuple[str, str]] = [
    ("DIV 01 00 00", "General Requirements"),
    ("DIV 02 00 00", "Existing Conditions"),
    ("DIV 03 00 00", "Concrete"),
    ("DIV 04 00 00", "Masonry"),
    ("DIV 05 00 00", "Metals"),
    ("DIV 06 00 00", "Wood, Plastics & Composites"),
    ("DIV 07 00 00", "Thermal & Moisture Protection"),
    ("DIV 08 00 00", "Openings"),
    ("DIV 09 00 00", "Finishes"),
    ("DIV 10 00 00", "Specialties"),
    ("DIV 11 00 00", "Equipment"),
    ("DIV 12 00 00", "Furnishings"),
    ("DIV 13 00 00", "Special Construction"),
    ("DIV 21 00 00", "Fire Suppression"),
    ("DIV 22 00 00", "Plumbing"),
    ("DIV 23 00 00", "HVAC"),
    ("DIV 25 00 00", "Integrated Automation"),
    ("DIV 26 00 00", "Electrical"),
    ("DIV 27 00 00", "Communications"),
    ("DIV 28 00 00", "Electronic Safety & Security"),
]

FOOTER_ROWS: list[tuple[str, str]] = [
    ("CONSTRUCTION_SUBTOTAL", "Construction Cost Subtotal"),
    ("GL_INSURANCE",          "General Liability Insurance"),
    ("BUILDERS_RISK",         "Builders Risk Insurance"),
    ("GC_FEE",                "GC Fee"),
    ("OVERHEAD_PROFIT",       "Overhead & Profit"),
    ("OTHER_FEES",            "Other Fees / Insurance"),
    ("BOND",                  "Bond"),
    ("FEES_SUBTOTAL",         "Fees Subtotal"),
    ("GRAND_TOTAL",           "GRAND TOTAL"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float(amount: Optional[Decimal]) -> float:
    """Convert a Decimal (or None) to float. Returns 0.0 for None."""
    if amount is None:
        return 0.0
    return float(amount)


def _cell_amount(state: CellState, amount: Optional[Decimal]) -> float:
    """
    Return the numeric value to write for a CellValue, given its state.

    AMOUNT / EXPLICIT_ZERO / ALLOWANCE / NOT_COMPARABLE → float (or 0.0)
    NULL_BLANK / EXCLUDED / BY_OWNER_OTHERS → 0.0

    NOT_COMPARABLE keeps its as-submitted amount in the bidder's own numbers
    (R33/ENC-2); only cross-bid benchmarks fence it out (falke_rules).
    """
    if state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                 CellState.ALLOWANCE, CellState.NOT_COMPARABLE):
        return _to_float(amount)
    return 0.0


def _sort_bids(bids: list[NormalizedBid]) -> list[NormalizedBid]:
    """
    Order columns by leveled_total ASCENDING — lowest leveled bid first, the
    natural board reading order, fully firm-agnostic (Marvin §2.5, Floyd C7).
    A bid with leveled_total None (e.g. no grand total extracted) sorts LAST so
    it never masquerades as the low bid. Ties break on contractor_name for
    determinism.
    """
    def _key(b: NormalizedBid) -> tuple[int, float, str]:
        lt = b.footer.leveled_total
        if lt is None:
            return (1, 0.0, b.contractor_name)
        return (0, float(lt), b.contractor_name)

    return sorted(bids, key=_key)


# Column A(1)=CSI, B(2)=Building System, C(3)=Normalization Note (Option C §2.1),
# then contractor groups start at D(4). Changing this ONE base offset shifts every
# contractor column right by one; reconcile.py reads _col_start back independently
# and inherits the shift (spec §7.2.5).
_CONTRACTOR_COL_BASE = 4  # D=4


def _col_start(contractor_index: int) -> int:
    """
    Return the 1-based openpyxl column index for the COST SUBTOTALS column
    of the given contractor (0-based index) — Bid_Form MIRROR geometry.

    Layout: A(1)=CSI, B(2)=description, C(3)=Normalization Note, D(4)=contractor 0.
    Each contractor group is 3 columns wide.
    """
    return _CONTRACTOR_COL_BASE + contractor_index * 3  # D=4, G=7, J=10 …


# ---------------------------------------------------------------------------
# Leveled_Normalized geometry (v0.3.0 — FEB 26 four-column bidder groups)
# ---------------------------------------------------------------------------
#
# Per Derick's approved sample (falke_rules_sample_v2.py): each bidder gets
# FOUR columns exactly as in Falke's FEB 26 matrix, plus the separator column
# repurposed as VAR % —
#
#     COST | COST SUBTOTALS | $/SF | $/SXFX SUBTOTALS | VAR %
#
# Derick's rules: (R1) per-row bid amounts in COST, never subtotals; (R2)
# division subtotals in COST SUBTOTALS; (R3) per-row $/SF in $/SF, never
# subtotals; (R4) subtotal $/SF in $/SXFX SUBTOTALS. A(1)=CSI, B(2)=Building
# System, C(3)=separator (as FEB 26 col D); groups start at D(4), stride 5.
# The BENCHMARK block (median / % spread / valid bids / confidence) is
# appended after the last group. reconcile.py reads these constants back so
# the writer and the Stage-6b checker can never drift.

LEVELED_GROUP_STRIDE = 5
LEVELED_COST_OFFSET = 0    # COST — per-row line amounts (R1)
LEVELED_CSUB_OFFSET = 1    # COST SUBTOTALS — division/fees subtotals (R2)
LEVELED_SF_OFFSET = 2      # $/SF — per-row (R3)
LEVELED_SXFX_OFFSET = 3    # $/SXFX SUBTOTALS — subtotal $/SF (R4)
LEVELED_VAR_OFFSET = 4     # separator column, repurposed as VAR % (R11)
LEVELED_BENCH_COLS = 4     # BENCHMARK | % SPREAD | VALID BIDS | CONFIDENCE


def _lev_col_start(contractor_index: int) -> int:
    """COST column of contractor i on Leveled_Normalized (stride 5)."""
    return _CONTRACTOR_COL_BASE + contractor_index * LEVELED_GROUP_STRIDE


def _lev_bench_col(num_contractors: int) -> int:
    """First column of the BENCHMARK block (right of the last VAR % col)."""
    return _lev_col_start(num_contractors - 1) + LEVELED_VAR_OFFSET + 1


def _lev_last_col(num_contractors: int) -> int:
    """Last used column on Leveled_Normalized (the CONFIDENCE column)."""
    return _lev_bench_col(num_contractors) + LEVELED_BENCH_COLS - 1


# Col-B display labels for the leveled footer block (FEB 26 verbatim for the
# two teal total rows; writer labels for the rest). Column A is BLANK for the
# entire leveled footer/alternates/qualifications block (Derick review,
# 2026-07-03) — these col-B labels are the ONLY anchors, so the writer, the
# Stage-6b reconciler, and the quarantine cell-marker ALL consume this ONE
# constant (same single-source-of-truth pattern as GRAND_TOTAL_COMPONENT_KEYS).
# Bid_Form keeps its col-A machine keys unchanged.
_LEVELED_FOOTER_DISPLAY_OVERRIDES = {
    "CONSTRUCTION_SUBTOTAL": "CONSTRUCTION COST SUBTOTAL",
    "GRAND_TOTAL": "GRAND TOTAL CONSTRUCTION COST",
}
LEVELED_FOOTER_LABELS: dict[str, str] = {
    key: _LEVELED_FOOTER_DISPLAY_OVERRIDES.get(key, label)
    for key, label in FOOTER_ROWS
}


# ---------------------------------------------------------------------------
# Worksheet construction
# ---------------------------------------------------------------------------

def _write_header_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    gsf: int,
    run: RunInputs,
) -> None:
    """Write rows 1–8: title, details, blank, column headers, contractor info.

    Project identity (title, details, per-contractor label) comes from the
    per-run RunInputs — never a hardcoded project (M1). The $/SF header carries
    the confirmed SF-basis label so the board knows what the denominator means
    (M2 / scoping §1.4).
    """
    bold = Font(bold=True)
    sf_label = run.sf_basis_label or DEFAULT_SF_BASIS_LABEL

    # Row 1 — project title
    ws.cell(row=1, column=1).value = f"{run.project_name} — Bid Comparison Matrix"
    ws.cell(row=1, column=1).font = bold

    # Row 2 — project details
    details = f"Project: {run.project_name} | {run.project_address} | {gsf:,.0f} {sf_label}"
    if run.rfp_label:
        details += f" | {run.rfp_label}"
    ws.cell(row=2, column=1).value = details

    # Row 3 — blank (intentional)

    # Row 4 — column headers
    ws.cell(row=4, column=1).value = "CSI"
    ws.cell(row=4, column=1).font = bold
    ws.cell(row=4, column=2).value = "Building System"
    ws.cell(row=4, column=2).font = bold
    # Col C — Normalization Note (Option C §2.1). Present on the Bid_Form mirror
    # so a known-firm reclass recommendation is on the FACE of the document.
    ws.cell(row=4, column=3).value = "Normalization Note"
    ws.cell(row=4, column=3).font = bold

    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        sf_col = cost_col + 1
        ws.cell(row=4, column=cost_col).value = "COST SUBTOTALS"
        ws.cell(row=4, column=cost_col).font = bold
        ws.cell(row=4, column=sf_col).value = f"$/{sf_label}"
        ws.cell(row=4, column=sf_col).font = bold

    # Row 5 — contractor names
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        ws.cell(row=5, column=cost_col).value = bid.contractor_name
        ws.cell(row=5, column=cost_col).font = bold

    # Row 6 — project label per contractor
    for i in range(len(bids)):
        cost_col = _col_start(i)
        ws.cell(row=6, column=cost_col).value = run.project_name

    # Row 7 — SF basis per contractor
    for i in range(len(bids)):
        cost_col = _col_start(i)
        ws.cell(row=7, column=cost_col).value = gsf

    # Row 8 — blank (intentional)


def _descriptions_match(a: str, b: str) -> bool:
    """
    Return True when two line-item description strings are semantically the same.
    Uses word overlap ≥ 60% (no external libraries needed).
    """
    a_words = set(a.lower().split())
    b_words = set(b.lower().split())
    if not a_words or not b_words:
        return False
    overlap = len(a_words & b_words) / max(len(a_words), len(b_words))
    return overlap >= 0.6


# M-3 (W-D): the mirror's state → rendering table. Numeric states write the
# number (EXPLICIT_ZERO writes a real 0.00 — the bidder DID write $0;
# NOT_COMPARABLE writes the number as submitted). Token states write the
# bidder's classification verbatim (italic, NO fill — the mirror asserts
# nothing). NULL_BLANK writes NOTHING (truly blank — never 0.00, never "-").
# Reconcile lockstep: `_as_decimal` coerces blank/token cells to 0 and
# `_expected_subtotal` maps non-amount-bearing states to 0, so written-blank
# == expected-0 holds (the false-quarantine guard).
_MIRROR_LINE_NUMERIC_STATES = (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                               CellState.ALLOWANCE, CellState.NOT_COMPARABLE)
# Subtotal numeric states mirror reconcile._AMOUNT_BEARING_STATES exactly
# (NC never occurs at subtotal level — normalize resolves NC-composed
# subtotals to state AMOUNT).
_MIRROR_SUBTOTAL_NUMERIC_STATES = (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                                   CellState.ALLOWANCE)
_MIRROR_TOKEN_FONT = Font(italic=True)


def _lookup_item_cell(
    div: "NormalizedDivision | None",
    target_desc: str,
) -> "Optional[object]":
    """Return the CellValue for target_desc from a NormalizedDivision's
    line_item_cells (exact key first, then _descriptions_match), or None."""
    if div is None:
        return None
    if target_desc in div.line_item_cells:
        return div.line_item_cells[target_desc]
    for key, cell in div.line_item_cells.items():
        if _descriptions_match(target_desc, key):
            return cell
    return None


def _build_unified_descriptions(
    bids: list[NormalizedBid],
    csi_code: str,
) -> list[str]:
    """
    For a given CSI division, collect all unique line-item descriptions across
    all contractors in display order.

    Algorithm: iterate contractors in order; for each contractor's line_item_cells
    append a description only if no existing entry in the running list matches it
    (case-insensitive substring OR ≥60% word overlap).

    LUMP_SUM contractors with no real items (all NULL_BLANK) are skipped —
    their placeholder descriptions must not pollute the unified list when they
    carry no pricing signal.
    """
    all_descs: list[str] = []

    for bid in bids:
        div = _find_div(bid, csi_code)
        if div is None:
            continue

        # Skip divisions where all line items are NULL_BLANK (lump-sum placeholder rows)
        has_priced_item = any(
            cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                           CellState.ALLOWANCE, CellState.NOT_COMPARABLE)
            for cell in div.line_item_cells.values()
        )
        if not has_priced_item and div.cost_structure.value == "LUMP_SUM":
            continue

        for desc in div.line_item_cells:
            already_present = any(
                _descriptions_match(desc, existing) or desc.lower() in existing.lower()
                or existing.lower() in desc.lower()
                for existing in all_descs
            )
            if not already_present:
                all_descs.append(desc)

    return all_descs


def _find_div(bid: NormalizedBid, csi_code: str) -> "NormalizedDivision | None":
    """Return the first NormalizedDivision matching csi_code for a bid, or None."""
    for div in bid.divisions:
        if div.csi_code == csi_code:
            return div
    return None


def _marker_text(rec: ReclassRecommendation) -> str:
    """Build the in-place Normalization Note marker for a recommendation (§2.2).

    With a priced amount:
      ``Dumpsters $54,959 — normalize → DIV 01 (General Requirements).
        As-submitted here; applied in Leveled_Normalized.``
    Without an amount, the leading ``{desc} {amount} — `` is dropped.
    """
    to_short = _div_short(rec.to_division)
    tail = (
        f"normalize → {to_short} ({rec.to_division_name}). "
        f"As-submitted here; applied in Leveled_Normalized."
    )
    if rec.amount is not None:
        return f"{rec.line_item_desc} {_fmt_money(rec.amount)} — {tail}"
    return f"Normalize → {to_short} ({rec.to_division_name}). As-submitted here; applied in Leveled_Normalized."


def _div_short(csi_code: str) -> str:
    """Render the bare `DIV NN` form (drop ` 00 00`) for marker readability."""
    parts = csi_code.split()
    if len(parts) >= 2 and parts[0] == "DIV":
        return f"DIV {parts[1]}"
    return csi_code


def _fmt_money(amount: Decimal) -> str:
    """Board-display dollar string '$54,959' (matches normalize._fmt)."""
    return f"${int(amount.quantize(Decimal('1'))):,}"


def _note_by_desc(bids: list[NormalizedBid], csi_code: str) -> dict[str, ReclassRecommendation]:
    """Map a line-item description → its reclass recommendation for one division.

    A recommendation belongs to a description row on the mirror when its
    ``from_division`` equals the row's division (the dollars sit there as
    submitted). Keyed by the matched line_item_desc so the writer can stamp the
    Normalization Note on the right row. First match wins if two bidders share a
    description+target (the recommended target is identical by rule).
    """
    out: dict[str, ReclassRecommendation] = {}
    for bid in bids:
        for rec in bid.reclass_recommendations:
            if rec.from_division == csi_code and rec.line_item_desc not in out:
                out[rec.line_item_desc] = rec
    return out


def _write_division_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
    gsf: int,
    show_notes: bool = True,
) -> tuple[int, dict[str, int]]:
    """
    Write dynamic division rows starting at start_row.

    ``show_notes`` controls the Col C Normalization Note (Option C §2): True on
    the Bid_Form mirror (the recommendation is shown in place), False on the
    Leveled_Normalized sheet (the move is already applied there).

    For each CSI division:
      1. Division header row (bold): CSI code | Division name
      2. One row per unified line-item description (across all contractors)
      3. Bold subtotal row: blank | DIVISION NAME SUBTOTAL | sub amounts
      4. Blank spacer row

    Returns (next_row_after_divisions, subtotal_row_by_csi_code).
    subtotal_row_by_csi_code maps each CSI code to the Excel row number of its
    SUBTOTAL row — used by the caller to apply audit-driven color fills.
    """
    bold = Font(bold=True)

    row = start_row
    subtotal_row_by_csi: dict[str, int] = {}

    for csi_code, div_name in DIVISION_ROWS:
        # --- Division header row (bold) ---
        c_csi = ws.cell(row=row, column=1)
        c_csi.value = csi_code
        c_csi.font = bold

        c_name = ws.cell(row=row, column=2)
        c_name.value = div_name
        c_name.font = bold

        row += 1

        # --- Gather per-bid NormalizedDivision objects for this CSI code ---
        bid_divs: list["NormalizedDivision | None"] = [
            _find_div(bid, csi_code) for bid in bids
        ]

        # --- Unified description list across all contractors ---
        all_descs = _build_unified_descriptions(bids, csi_code)

        # --- Normalization-Note recommendations for this division (Option C §2) ---
        notes_by_desc = _note_by_desc(bids, csi_code) if show_notes else {}

        # --- One row per line-item description ---
        for desc in all_descs:
            ws.cell(row=row, column=2).value = desc

            # Stamp the in-place Normalization Note (Col C) — text only, never
            # alters any subtotal; YELLOW recommend-review band on the note cell.
            rec = notes_by_desc.get(desc)
            if rec is not None:
                note_cell = ws.cell(row=row, column=NOTE_COL)
                note_cell.value = _marker_text(rec)
                note_cell.fill = YELLOW_FILL
                note_cell.alignment = Alignment(wrap_text=True)

            for i, div in enumerate(bid_divs):
                cost_col = _col_start(i)
                cellv = _lookup_item_cell(div, desc)
                if cellv is None:
                    continue
                c = ws.cell(row=row, column=cost_col)
                if cellv.state in _MIRROR_LINE_NUMERIC_STATES:
                    c.value = _to_float(cellv.amount)
                    c.number_format = (AMOUNT_FORMAT_NEG if c.value < 0
                                       else AMOUNT_FORMAT)
                elif cellv.state == CellState.EXCLUDED:
                    # M-3: "Excluded" italic (verbatim bidder token when
                    # extraction later carries one — v0.3.1-Q1).
                    c.value = "Excluded"
                    c.font = _MIRROR_TOKEN_FONT
                elif cellv.state == CellState.BY_OWNER_OTHERS:
                    c.value = (cellv.display or "").strip() or "BY OTHERS"
                    c.font = _MIRROR_TOKEN_FONT
                # NULL_BLANK: truly blank — write nothing.

            row += 1

        # --- Subtotal row (bold) ---
        subtotal_label = div_name.upper() + " SUBTOTAL"
        c_sub_label = ws.cell(row=row, column=2)
        c_sub_label.value = subtotal_label
        c_sub_label.font = bold

        subtotal_row_by_csi[csi_code] = row  # record for audit-fill pass

        for i, bid in enumerate(bids):
            cost_col = _col_start(i)
            sf_col = cost_col + 1
            c_cost = ws.cell(row=row, column=cost_col)

            # M-3 (W-D): render the subtotal per CellState. Numeric states
            # (AMOUNT / EXPLICIT_ZERO / ALLOWANCE — reconcile's exact
            # _AMOUNT_BEARING_STATES) write the aggregated number; EXCLUDED /
            # BY_OWNER write the bidder's token (italic, no fill); absent or
            # all-NULL_BLANK writes NOTHING (a blank cell means the bidder
            # left it blank).
            divs = [d for d in bid.divisions if d.csi_code == csi_code]
            states = [d.subtotal_cell.state for d in divs]

            if any(s in _MIRROR_SUBTOTAL_NUMERIC_STATES for s in states):
                amount = sum(
                    _to_float(d.subtotal_cell.amount) for d in divs
                    if d.subtotal_cell.state in _MIRROR_SUBTOTAL_NUMERIC_STATES
                    and d.subtotal_cell.amount is not None
                )
                sf_val = round(amount / gsf, 2) if gsf > 0 else 0.0
                c_cost.value = amount
                c_cost.number_format = (AMOUNT_FORMAT_NEG if amount < 0
                                        else AMOUNT_FORMAT)
                c_cost.font = bold
                c_sf = ws.cell(row=row, column=sf_col)
                c_sf.value = sf_val
                c_sf.number_format = AMOUNT_FORMAT
                c_sf.font = bold
            elif CellState.EXCLUDED in states:
                c_cost.value = "Excluded"
                c_cost.font = _MIRROR_TOKEN_FONT
            elif CellState.BY_OWNER_OTHERS in states:
                token = next(
                    ((d.subtotal_cell.display or "").strip() for d in divs
                     if d.subtotal_cell.state == CellState.BY_OWNER_OTHERS),
                    "",
                )
                c_cost.value = token or "BY OTHERS"
                c_cost.font = _MIRROR_TOKEN_FONT
            # else: no division / all NULL_BLANK → truly blank cell.

        row += 1

        # --- Blank spacer ---
        row += 1

    return row, subtotal_row_by_csi  # next row after all divisions


def _write_footer_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
    gsf: int,
) -> tuple[int, list[dict]]:
    """
    Write a blank separator then the footer rows (construction subtotal,
    insurance, GC fee, overhead & profit, other fees, fees subtotal, grand
    total, bond).

    Every component row that COMPOSES the grand total — GL, Builders Risk, GC
    fee, Overhead & Profit, and Other Fees/Insurance — is rendered as a labeled
    row, so the footer visibly ties to the grand total for a board. The amounts
    come from grand_total_component_amounts() (the single source of truth shared
    with audit.py and reconcile.py): a contractor that folds insurance into
    `other_fees_subtotal` (e.g. a recurring firm) shows it on the Other Fees row, while a memo
    `other_fees` line that merely duplicates fees already counted is rendered as
    0 so it is not double-counted. FEES_SUBTOTAL is the sum of every additive
    fee component, so CONSTRUCTION_SUBTOTAL + FEES_SUBTOTAL == GRAND_TOTAL.

    Returns (next_row_after_footer, list[per_bid_footer_summary]).
    """
    bold = Font(bold=True)
    row = start_row + 1  # +1 blank separator

    # Pre-resolve the additive grand-total composition per bid (single source of
    # truth) so the rendered component rows and FEES_SUBTOTAL agree with audit /
    # reconcile, and a memo other_fees is suppressed to 0.
    components = [grand_total_component_amounts(bid.footer) for bid in bids]

    # The additive fee components that roll up into FEES_SUBTOTAL (everything
    # composing the grand total EXCEPT construction). Derived from the single
    # source of truth (GRAND_TOTAL_COMPONENT_KEYS) minus construction so it can
    # never drift from grand_total_component_amounts(). Bond is an additive
    # component of the grand total (Marvin's ruling), so it rolls up here.
    _FEE_COMPONENT_KEYS = tuple(
        k for k in GRAND_TOTAL_COMPONENT_KEYS if k != "CONSTRUCTION_SUBTOTAL"
    )

    summaries: list[dict] = []
    for bid in bids:
        summaries.append({})

    for key, label in FOOTER_ROWS:
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = label
        if key == "GRAND_TOTAL":
            ws.cell(row=row, column=2).font = bold

        for i, bid in enumerate(bids):
            cost_col = _col_start(i)
            footer = bid.footer
            comp = components[i]

            if key == "CONSTRUCTION_SUBTOTAL":
                val = float(comp.get("CONSTRUCTION_SUBTOTAL", 0))
            elif key in _FEE_COMPONENT_KEYS:
                # Additive contribution per the shared composition (memo
                # other_fees → absent from comp → 0).
                val = float(comp.get(key, 0))
            elif key == "FEES_SUBTOTAL":
                val = sum(
                    summaries[i].get(k, 0.0) for k in _FEE_COMPONENT_KEYS
                )
            elif key == "GRAND_TOTAL":
                val = _cell_amount(footer.grand_total.state, footer.grand_total.amount)
            else:
                val = 0.0

            summaries[i][key] = val

            c = ws.cell(row=row, column=cost_col)
            c.value = val
            c.number_format = AMOUNT_FORMAT
            if key == "GRAND_TOTAL":
                c.font = bold

        row += 1

    return row, summaries


def _write_alternates(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
) -> int:
    """Write bid alternates in their OWN clearly-labeled section (M7).

    Alternates (add/deduct options) are NEVER folded into the base/leveled
    total — the base comparison stays apples-to-apples. Each contractor's
    alternates are listed under their column. If no bidder submitted any
    alternate, the section is omitted entirely. Returns the next free row.
    """
    if not any(bid.footer.alternates for bid in bids):
        return start_row

    bold = Font(bold=True)
    row = start_row + 1  # blank separator

    ws.cell(row=row, column=1).value = "ALTERNATES"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = (
        "Bid Alternates (add/deduct — NOT included in base comparison)"
    )
    ws.cell(row=row, column=2).font = bold
    row += 1

    # One row per (contractor, alternate). Description in col B, amount under the
    # contractor's cost column — kept visually separate from the base divisions.
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        for alt in bid.footer.alternates:
            ws.cell(row=row, column=1).value = bid.contractor_name
            ws.cell(row=row, column=2).value = alt.description
            c = ws.cell(row=row, column=cost_col)
            if alt.amount is not None:
                c.value = float(alt.amount)
                c.number_format = AMOUNT_FORMAT
            else:
                c.value = alt.display
            row += 1

    return row


def _write_qualifications(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    bids: list[NormalizedBid],
    start_row: int,
) -> int:
    """Write one qualifications row per contractor, separated by a blank row.

    Returns the next free row (the workbook LEGEND lands below — W-D B4/§2:
    on Bid_Form the legend sits below Qualifications)."""
    row = start_row + 1  # +1 blank separator

    ws.cell(row=row, column=1).value = "QUALIFICATIONS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "Contractor Qualifications"
    ws.cell(row=row, column=2).font = Font(bold=True)

    row += 1
    for i, bid in enumerate(bids):
        cost_col = _col_start(i)
        ws.cell(row=row, column=cost_col - 1).value = bid.contractor_name
        qual_cell = ws.cell(row=row, column=cost_col)
        qual_cell.value = bid.qualifications_text or ""
        qual_cell.alignment = Alignment(wrap_text=True)
    return row + 1


def _set_column_widths(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_contractors: int,
) -> None:
    """Set column widths for readability."""
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = COL_B_WIDTH

    for i in range(num_contractors):
        cost_col = _col_start(i)
        sf_col   = cost_col + 1
        sep_col  = cost_col + 2

        ws.column_dimensions[get_column_letter(cost_col)].width = CONTRACTOR_COST_WIDTH
        ws.column_dimensions[get_column_letter(sf_col)].width   = CONTRACTOR_SF_WIDTH
        ws.column_dimensions[get_column_letter(sep_col)].width  = CONTRACTOR_SEP_WIDTH


# ---------------------------------------------------------------------------
# Audit sheet fills and AUDIT worksheet
# ---------------------------------------------------------------------------

# Cell fill constants — PatternFill is safe on fresh workbooks (no existing styles to corrupt)
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")   # soft red
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")   # soft yellow
GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")   # soft green

_STATUS_FILL = {
    AuditStatus.RED:    RED_FILL,
    AuditStatus.YELLOW: YELLOW_FILL,
    AuditStatus.GREEN:  GREEN_FILL,
}

_STATUS_SORT_KEY = {
    AuditStatus.RED:    0,
    AuditStatus.YELLOW: 1,
    AuditStatus.GREEN:  2,
}


# M-3 / B4 §2 item 2 (W-D): the ARA audit-fill pass is REMOVED from the
# mirror — the mirror asserts nothing (values/tokens, the col-C Normalization
# Note, and quarantine marks only). Every ARA diagnostic already lives on the
# AUDIT sheet; the leveled sheet never used this pass (rules spec §4.4/A1).


# AUDIT View column labels — the actual SHEET NAMES they point at (W-D B4/§2:
# kills the "As-Submitted" third dialect). ONE constant for the initial write
# AND the appended quarantine rows so the two can never drift.
AUDIT_VIEW_LABELS = {
    "leveled": "Leveled_Normalized",
    "mirror": "Bid_Form",
    "both": "Both",
}

# Default AUDIT geometry (W-D S2-3: the key moved to the TOP of the tab).
# Rows 1-2 title/subtitle, 3 blank, 4-7 the RGY key, 8 the board pointer,
# 9 blank, 10 column headers, 11+ data. Readers must NOT hardcode these:
# find_audit_header_row() locates the header by its labels, so the key block
# and an inserted QUARANTINE line never break the read-back.
AUDIT_HEADER_ROW = 10
AUDIT_DATA_START_ROW = 11

_AUDIT_BOARD_POINTER = (
    "Board members: the decision view is Leveled_Normalized. This tab is the "
    "estimator's log of every check performed."
)


def find_audit_header_row(ws) -> Optional[int]:
    """Locate the AUDIT column-header row by its own labels (col A 'Status',
    col C 'Code') — label-anchored, same idiom as the data-sheet locators."""
    for row in range(1, ws.max_row + 1):
        if (ws.cell(row=row, column=1).value == "Status"
                and ws.cell(row=row, column=3).value == "Code"):
            return row
    return None


def _write_audit_sheet(
    wb: openpyxl.Workbook,
    audit_items: list[AuditItem],
) -> None:
    """
    Create and populate the AUDIT worksheet in wb.

    Layout (W-D S2-3 — key at the TOP):
      Row 1: Title
      Row 2: Subtitle
      Row 3: blank
      Rows 4-7: KEY — totals by status (the tally a reader needs FIRST)
      Row 8: board pointer (the decision view is Leveled_Normalized)
      Row 9: blank
      Row 10: Column headers
      Row 11+: One row per AuditItem sorted RED→YELLOW→GREEN, then
               contractor, then division
    """
    ws = wb.create_sheet(title="AUDIT")
    bold = Font(bold=True)

    # --- Column widths ---
    from openpyxl.utils import get_column_letter
    col_widths = [10, 20, 28, 30, 16, 35, 18, 60]  # A..H (View col carries sheet names)
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Row 1: Title ---
    ws.cell(row=1, column=1).value = (
        "FALKE Matrix — Extraction & Normalization Audit Report"
    )
    ws.cell(row=1, column=1).font = bold

    # --- Row 2: Subtitle ---
    ws.cell(row=2, column=1).value = (
        "Generated by ARA Pipeline | Items requiring action before bid award"
    )

    # --- Row 3: blank ---

    # --- Rows 4-8: the KEY, at the TOP (W-D S2-3) ---
    red_count    = sum(1 for a in audit_items if a.status == AuditStatus.RED)
    yellow_count = sum(1 for a in audit_items if a.status == AuditStatus.YELLOW)
    green_count  = sum(1 for a in audit_items if a.status == AuditStatus.GREEN)
    total_count  = len(audit_items)

    key_lines = [
        (f"Total items audited:   {total_count}", None),
        (f"RED Critical:          {red_count}  — must resolve before award", RED_FILL),
        (f"YELLOW Review:         {yellow_count}  — verify before finalizing", YELLOW_FILL),
        (f"GREEN Verified:        {green_count}  — clean", GREEN_FILL),
        (_AUDIT_BOARD_POINTER, None),
    ]
    for i, (text, fill) in enumerate(key_lines):
        c = ws.cell(row=4 + i, column=1)
        c.value = text
        c.font = bold
        if fill:
            c.fill = fill

    # --- Row 9: blank ---

    # --- Row 10: Column headers (View column carries the sheet names, §4/W-D) ---
    headers = ["Status", "View", "Code", "Contractor", "Division", "Line Item", "Value", "Message"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=AUDIT_HEADER_ROW, column=col_idx)
        c.value = header
        c.font = bold

    # --- Sort items: RED first, then YELLOW, then GREEN; then by contractor; then division ---
    sorted_items = sorted(
        audit_items,
        key=lambda a: (
            _STATUS_SORT_KEY[a.status],
            a.contractor_name,
            a.division_csi or "",
        ),
    )

    # --- Rows 11+: One row per AuditItem ---
    for row_offset, item in enumerate(sorted_items):
        row = AUDIT_DATA_START_ROW + row_offset
        fill = _STATUS_FILL[item.status]

        values = [
            item.status.value,
            AUDIT_VIEW_LABELS.get(item.view, item.view),
            item.code.value,
            item.contractor_name,
            item.division_csi or "",
            item.line_item_desc or "",
            item.value or "",
            item.message,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill
            # Bold the Status cell text
            if col_idx == 1:
                c.font = Font(bold=True)


# ---------------------------------------------------------------------------
# Leveled-view banner (Option C §3.3)
# ---------------------------------------------------------------------------

# Two banner lines, written into rows 1–2 of Leveled_Normalized (replacing the
# normal title/details). Rows 4–8 (headers, names, GSF) stay at the SAME rows as
# the mirror so reconcile.py's row-5 name read and label-anchored reads work on
# both sheets unchanged.
_LEVELED_BANNER_LINE_1 = (
    "ESTIMATOR-NORMALIZED VIEW — does NOT match the submitted bids. Dollars have "
    "been moved between divisions for apples-to-apples comparison. See the "
    "Bid_Form sheet for each bid exactly as submitted."
)
_LEVELED_BANNER_LINE_2 = (
    "Normalization applied: known-firm division reclassifications (see "
    "Normalization Note column on Bid_Form for each move and its rationale)."
)


def _populate_data_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    ordered_bids: list[NormalizedBid],
    gsf: int,
    run: RunInputs,
    show_notes: bool,
) -> list[dict]:
    """Fill the Bid_Form mirror worksheet end-to-end.

    Returns the per-bid footer summaries (used only for the mirror's report).
    ``show_notes`` writes the Col C Normalization Note. NO ARA audit fills are
    applied here (M-3/B4 — the mirror asserts nothing).
    """
    _write_header_rows(ws, ordered_bids, gsf, run)

    if len(ordered_bids) == 1:
        notice = ws.cell(row=3, column=1)
        notice.value = "Single bid — no competitive comparison available."
        notice.font = Font(bold=True, italic=True)

    DIVISION_START_ROW = 9
    next_row, _subtotal_row_by_csi = _write_division_rows(
        ws, ordered_bids, DIVISION_START_ROW, gsf, show_notes=show_notes
    )

    # NO ARA audit fills on the mirror (M-3/B4 — the mirror asserts nothing);
    # every ARA diagnostic lives on the AUDIT sheet.

    next_row, footer_summaries = _write_footer_rows(ws, ordered_bids, next_row, gsf)
    next_row = _write_alternates(ws, ordered_bids, next_row)
    next_row = _write_qualifications(ws, ordered_bids, next_row)
    # The ONE workbook LEGEND, identical to the leveled sheet's (W-D B4/§2 —
    # placed below Qualifications on Bid_Form).
    _write_workbook_legend(ws, next_row + 1, bold_font=Font(bold=True))
    _set_column_widths(ws, len(ordered_bids))
    return footer_summaries


def _used_width_chars(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    used_width: int,
) -> float:
    """Approximate the character capacity of one wrapped line across the merged
    banner (sum of the set column widths from A through ``used_width``).

    openpyxl column width is roughly in characters of the default font, so the
    sum is a usable estimate of how many characters fit on one line of the
    full-width merged banner — used to size the row height so the text shows
    horizontally instead of stacking in column A.
    """
    from openpyxl.utils import get_column_letter

    total = 0.0
    for col in range(1, used_width + 1):
        dim = ws.column_dimensions.get(get_column_letter(col))
        total += (dim.width if dim is not None and dim.width else 8.43)
    return total


def _render_full_width_banner(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    text: str,
    fill: PatternFill,
    used_width: int,
    bold: bool,
) -> None:
    """Render ONE banner row as a merged, full-width, readable block.

    Shared by both the YELLOW leveled-view banner and the RED quarantine banner:
    fill col A → ``used_width`` with ``fill``, MERGE the row across that width so
    the long text flows HORIZONTALLY (instead of stacking into a tall, narrow
    column-A block), set ``wrap_text``, and give the row an explicit height sized
    to the wrapped-line count so the full text is visible. Overrides whatever was
    in those cells; does not shift data rows.
    """
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = Font(bold=bold)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    # Fill across the used width FIRST so every underlying cell is filled even
    # before the merge collapses them visually.
    for col in range(1, used_width + 1):
        ws.cell(row=row, column=col).fill = fill
    # MERGE the row across the full used width so the text flows horizontally.
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=used_width)
    # Explicit row height: estimate wrapped-line count from text length vs the
    # merged width's character capacity, ~15 pts per line (default font).
    line_chars = _used_width_chars(ws, used_width)
    est_lines = max(1, ceil(len(text) / max(line_chars, 1.0)))
    ws.row_dimensions[row].height = est_lines * 15.0


# ---------------------------------------------------------------------------
# Leveled_Normalized writer (v0.3.0 — FEB 26 geometry + house format + rules)
# ---------------------------------------------------------------------------
#
# Ported from the Derick-approved sample builder (falke_rules_sample_v2.py).
# The leveled sheet is BORN in the Falke house format (format_falke constants)
# and speaks the pure Falke highlight vocabulary (falke_rules — Cyan/Yellow/
# Red/Neutral, red-first). NO legacy ARA audit fills appear here: every ARA
# diagnostic lives on the AUDIT sheet only (rules spec §4.4/A1). Quarantine
# overlays (apply_quarantine) still win — they run in a later re-open pass.

def _populate_leveled_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    ordered_leveled: list[NormalizedBid],
    gsf: int,
    run: RunInputs,
) -> dict:
    """Write the Leveled_Normalized sheet end-to-end; returns rule-fire counts.

    Geometry (per bidder, stride 5): COST | COST SUBTOTALS | $/SF | $/SXFX
    SUBTOTALS | VAR %; BENCHMARK block (median / % spread / valid bids /
    confidence) right of all groups. Rows: 1–2 gray banner, 3 disclaimer (R34),
    4 aqua header, 5–7 gray bidder block, 9+ divisions, then footer (col A
    BLANK — col-B display labels per LEVELED_FOOTER_LABELS), alternates,
    qualifications, summary (R32), legend, and the decided-assumptions block.
    """
    names = [b.contractor_name for b in ordered_leveled]
    n = len(names)
    col_of = {nm: i for i, nm in enumerate(names)}
    lev_by_name = {b.contractor_name: b for b in ordered_leveled}
    gsf_f = float(gsf)

    def gcost(i): return _lev_col_start(i)
    def gcsub(i): return _lev_col_start(i) + LEVELED_CSUB_OFFSET
    def gsfc(i): return _lev_col_start(i) + LEVELED_SF_OFFSET
    def gsxfx(i): return _lev_col_start(i) + LEVELED_SXFX_OFFSET
    def gvar(i): return _lev_col_start(i) + LEVELED_VAR_OFFSET

    BENCH = _lev_bench_col(n)
    SPREAD, NVALID, CONF = BENCH + 1, BENCH + 2, BENCH + 3
    LAST = _lev_last_col(n)

    tracker = falke_rules.PaintTracker(red_font=ff.WHITE_BOLD_FONT)

    def _sf(v: float) -> float:
        return v / gsf_f if gsf_f > 0 else 0.0

    # --- Column widths (FEB 26 reference widths) ---
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 3.8
    for i in range(n):
        for c, w in ((gcost(i), 15.8), (gcsub(i), 16.8), (gsfc(i), 9.8),
                     (gsxfx(i), 12.8), (gvar(i), 9)):
            ws.column_dimensions[get_column_letter(c)].width = w
    for c, w in ((BENCH, 19), (SPREAD, 10), (NVALID, 10), (CONF, 12)):
        ws.column_dimensions[get_column_letter(c)].width = w

    # --- Rows 1–2: leveled banner (neutral house GRAY per rules-spec A1 — the
    # legacy yellow banner fill is superseded so yellow means only "overpriced"
    # on this sheet; banner TEXT unchanged) + row 3 disclaimer (R34) ---
    for r, (text, font) in enumerate(
        [(_LEVELED_BANNER_LINE_1, ff.BANNER_BOLD_FONT),
         (_LEVELED_BANNER_LINE_2, ff.BANNER_BODY_FONT)], start=1
    ):
        for c in range(1, LAST + 1):
            ws.cell(row=r, column=c).fill = ff.GRAY_FILL
        lead = ws.cell(row=r, column=1)
        lead.value = text
        lead.font = font
        lead.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST)
        ws.row_dimensions[r].height = 30
    d = ws.cell(row=3, column=1)
    disclaimer = "DISCLAIMER: " + falke_rules.DISCLAIMER
    if n == 1:
        disclaimer = ("Single bid — no competitive comparison available. "
                      + disclaimer)
    d.value = disclaimer
    d.font = ff.DISCLAIMER_FONT
    d.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=LAST)
    ws.row_dimensions[3].height = 26

    # --- Row 4: header band (aqua, FEB 26 titles verbatim) ---
    HDR = 4
    titles = [(1, "CSI"), (2, "Building System")]
    for i in range(n):
        titles += [(gcost(i), "COST"), (gcsub(i), "COST \nSUBTOTALS"),
                   (gsfc(i), "$/SF"), (gsxfx(i), "$/SXFX \nSUBTOTALS"),
                   (gvar(i), "VAR %")]
    titles += [(BENCH, "BENCHMARK (median)"), (SPREAD, "% SPREAD"),
               (NVALID, "VALID BIDS"), (CONF, "CONFIDENCE")]
    for c in range(1, LAST + 1):
        cell = ws.cell(row=HDR, column=c)
        cell.fill = ff.AQUA_FILL
        cell.border = ff.BOX_BORDER
    for c, t in titles:
        cell = ws.cell(row=HDR, column=c)
        cell.value = t
        cell.font = ff.BOLD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[HDR].height = ff.HEADER_ROW_HEIGHT

    # --- Rows 5–7: bidder block (gray, white bold 11, merged across 4 cols) ---
    for r, val_of in ((5, lambda nm: nm), (6, lambda nm: run.project_name),
                      (7, lambda nm: int(gsf))):
        for i, nm in enumerate(names):
            for c in range(gcost(i), gsxfx(i) + 1):
                ws.cell(row=r, column=c).fill = ff.GRAY_FILL
            lead = ws.cell(row=r, column=gcost(i))
            lead.value = val_of(nm)
            lead.font = ff.BIDDER_BLOCK_FONT
            lead.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=r, start_column=gcost(i),
                           end_row=r, end_column=gsxfx(i))

    # --- Helpers ---
    def money(row, col, v, font=None, border=ff.RAIL_BORDER):
        c = ws.cell(row=row, column=col)
        c.value = v
        c.number_format = ff.FALKE_AMOUNT_FORMAT
        c.font = font or ff.BODY_FONT
        if border:
            c.border = border
        return c

    def write_bench(row, prices, on_subtotal):
        """R9/R10/R29: benchmark block cells for one row; returns the median."""
        n_valid = len(prices)
        bench, spread = falke_rules.benchmark_stats(prices)
        fill = ff.AQUA_FILL if on_subtotal else None
        for col, v, nf in ((BENCH, bench, ff.FALKE_AMOUNT_FORMAT),
                           (SPREAD, spread, ff.FALKE_PCT_FORMAT),
                           (NVALID, n_valid, "0"),
                           (CONF, falke_rules.confidence(n_valid, spread), None)):
            c = ws.cell(row=row, column=col)
            if v is not None:
                c.value = v
            if nf:
                c.number_format = nf
            c.font = ff.BODY_FONT
            if fill is not None:
                c.fill = fill
        return bench

    var_by_bidder: dict[str, list[float]] = {nm: [] for nm in names}

    # REM-2 (Marvin): divisions whose subtotal the ENGINE derived (no stated
    # subtotal on the form — normalize flags SUBTOTAL_DERIVED) get an on-cell
    # disclosure, with the composition numbers when the bidder's stated
    # Construction Cost Subtotal does not reconcile to the displayed division
    # subtotals (a recurring scissor-lift pattern).
    derived_csis: dict[str, set[str]] = {}
    stated_cs: dict[str, Optional[float]] = {}
    div_sum_by_name: dict[str, float] = {}
    for b in ordered_leveled:
        nm = b.contractor_name
        flagged = {
            d.csi_code for d in b.divisions
            if "SUBTOTAL_DERIVED" in d.subtotal_cell.flags
        }
        if flagged:
            derived_csis[nm] = flagged
        cs = b.footer.construction_subtotal
        stated_cs[nm] = (
            float(cs.amount)
            if (cs.state == CellState.AMOUNT and cs.amount is not None)
            else None
        )
        # Same shared sum the ENC-3 composition check consumes — the REM-2
        # disclosure and the register row can never cite different arithmetic.
        div_sum_by_name[nm] = falke_rules.displayed_priced_sum(
            b, [c for c, _n in DIVISION_ROWS]
        )

    def _derived_subtotal_note(nm: str) -> str:
        text = ("[ARA REM-2] Subtotal DERIVED from priced line items — "
                "the bidder's form "
                "showed no stated subtotal for this division (REM-2).")
        # Composition sentence still discloses any real delta (> $1). Its
        # AUDIT pointer is honest either way (the W-D fix for the dangling
        # "see AUDIT" — S2-1): above the Falke tolerance it names the
        # SUBTOTAL_COMPOSITION_DISCREPANCY row ENC-3 just emitted; within
        # tolerance it says so and notes the RISK-1 cap conversation instead
        # of pointing at a row that doesn't exist (the known within-tolerance
        # $18,500-on-$3.9M shape sits INSIDE max($5, 0.5%)).
        cs_val = stated_cs.get(nm)
        if cs_val is not None and abs(div_sum_by_name[nm] - cs_val) > 1.0:
            delta = abs(div_sum_by_name[nm] - cs_val)
            text += (
                f" The bidder's stated Construction Cost Subtotal "
                f"(${cs_val:,.2f}) does not reconcile to the displayed "
                f"division subtotals (sum ${div_sum_by_name[nm]:,.2f}, delta "
                f"${delta:,.2f}) — this derived "
                f"amount may not be carried in the bidder's stated totals; "
            )
            if delta > falke_rules.tol(cs_val):
                text += ("see the SUBTOTAL_COMPOSITION_DISCREPANCY row on "
                         "the AUDIT tab.")
            else:
                text += ("within Falke's max($5, 0.5%) math tolerance, so it "
                         "is not separately flagged on AUDIT (RISK-1 — "
                         "tolerance cap under discussion with Falke).")
        return text

    def variance_pass(row, statuses, target_col_of, on_subtotal,
                      member_prices=None):
        """R9/R11/R12/R13/R15/R16 on one row; paints target_col_of(name).

        Benchmark computes from classified STATE only (A6) — the ``statuses``
        map carries (kind, amount) per bidder and only kind=="priced" enters
        the median, so a blank/zero/excluded cell can never poison it. On
        SUBTOTAL rows the caller passes ``member_prices`` — the ONE M-2
        median-membership set (falke_rules.median_membership: priced AND
        amount > 0 AND not R20-failed) — so an R20-failed subtotal keeps its
        red cell and its VAR% but leaves the median (R7, GOLD-DEV-1). Paint
        requires ≥MIN_BIDS_FOR_PAINT valid bids (Q5/RISK-2); the benchmark
        still displays below that.
        """
        prices = (member_prices if member_prices is not None
                  else [amt for (k, amt) in statuses.values() if k == "priced"])
        if not prices:
            return
        if on_subtotal or len(prices) >= 2:
            write_bench(row, prices, on_subtotal)
        if len(prices) < 2:
            return
        bench, _spread = falke_rules.benchmark_stats(prices)
        if bench == 0:
            return
        for nm, (kind, amt) in statuses.items():
            if kind != "priced" or amt is None or amt <= 0:
                # A net-negative subtotal (unclassified credit, W-D ruling 5)
                # is not a price: no VAR%, no variance paint — its R22 red and
                # AUDIT row carry the story.
                continue
            var = (amt - bench) / bench
            vc = ws.cell(row=row, column=gvar(col_of[nm]))
            vc.value = var
            vc.number_format = ff.FALKE_PCT_FORMAT
            vc.font = ff.BODY_FONT
            var_by_bidder[nm].append(var)
            cell = ws.cell(row=row, column=target_col_of(nm))
            if tracker.kind_at(cell) == "red":
                continue  # R16: red never downgraded
            fired = falke_rules.variance_color(amt, bench)
            if fired and len(prices) < falke_rules.MIN_BIDS_FOR_PAINT:
                tracker.count_gate_suppressed()
                fired = None
            if fired:
                tracker.paint(cell, fired, nm)
                falke_rules.attach_comment(
                    cell,
                    "[FALKE R12] " + falke_rules.MSG_CYAN if fired == "cyan"
                    else "[FALKE R13] " + falke_rules.MSG_YELLOW,
                )
            else:
                tracker.count_neutral()

    # --- Division blocks ---
    row = 9
    for csi, div_name in DIVISION_ROWS:
        # Col-A DIV keys stay, REGULAR weight (Derick review, 2026-07-03).
        ws.cell(row=row, column=1).value = csi
        ws.cell(row=row, column=1).font = ff.BODY_FONT
        ws.cell(row=row, column=2).value = div_name
        ws.cell(row=row, column=2).font = ff.BOLD_FONT
        row += 1

        for desc in _build_unified_descriptions(ordered_leveled, csi):
            ws.cell(row=row, column=2).value = desc
            ws.cell(row=row, column=2).font = ff.BODY_FONT
            statuses = {nm: falke_rules.line_status(b, csi, desc)
                        for nm, b in zip(names, ordered_leveled)}
            for nm, (kind, amt) in statuses.items():
                i = col_of[nm]
                if kind in ("priced", "zero"):
                    money(row, gcost(i), amt)                    # R1: COST col
                    money(row, gsfc(i), _sf(amt))                # R3: $/SF col
                    if kind == "zero":
                        cell = ws.cell(row=row, column=gcost(i))
                        tracker.paint(cell, "red", nm)
                        falke_rules.attach_comment(
                            cell,
                            "[FALKE R6] " + falke_rules.MSG_RED
                            + " (Zero value without "
                            "approved classification — R6.)",
                        )
                elif kind == "not_comparable" and amt is not None:
                    # ENC-2: amount displayed as submitted, no paint, and it
                    # never enters the benchmark (kind != "priced").
                    money(row, gcost(i), amt)
                    money(row, gsfc(i), _sf(amt))
                    falke_rules.attach_comment(
                        ws.cell(row=row, column=gcost(i)),
                        "[ARA NC/R7] Not Comparable — excluded from "
                        "benchmark (R7/R8).",
                    )
                elif kind == "excluded":
                    # ENC-5 (W-D): a line-level exclusion inside a division is
                    # VISIBLE on the decision view — the bidder's
                    # classification token, italic, no paint (the legend's
                    # italic vocabulary; severity lives on the per-line
                    # EXPLICIT_EXCLUSION RED row of the AUDIT register).
                    c = ws.cell(row=row, column=gcost(i))
                    c.value = "Excluded"
                    c.font = ff.ITALIC_FONT
                elif kind == "by_owner":
                    # ENC-5: verbatim by-owner token (ENC-1 doctrine — never
                    # tell the board a wrong story about who carries scope).
                    cellv = _lookup_item_cell(
                        _find_div(lev_by_name[nm], csi), desc)
                    token = ((cellv.display or "").strip()
                             if cellv is not None else "")
                    c = ws.cell(row=row, column=gcost(i))
                    c.value = token or "By Owner"
                    c.font = ff.ITALIC_FONT
            n_priced = sum(1 for k, _ in statuses.values() if k == "priced")
            if n_priced >= 2:
                variance_pass(row, statuses,
                              lambda nm: gcost(col_of[nm]), on_subtotal=False)
            row += 1

        # Subtotal row — R2/R4: amounts in COST SUBTOTALS + $/SXFX only.
        label = div_name.upper() + " SUBTOTAL"
        lab = ws.cell(row=row, column=2)
        lab.value = label
        lab.font = ff.BOLD_FONT
        lab.fill = ff.AQUA_FILL
        statuses = {nm: falke_rules.div_status(b, csi)
                    for nm, b in zip(names, ordered_leveled)}
        any_priced = any(k == "priced" for k, _ in statuses.values())
        for nm, (kind, amt) in statuses.items():
            i = col_of[nm]
            csub = ws.cell(row=row, column=gcsub(i))
            sxfx = ws.cell(row=row, column=gsxfx(i))
            for c in (csub, sxfx):
                c.fill = ff.AQUA_FILL
                c.border = ff.SUBTOTAL_BORDER
                c.font = ff.BODY_FONT
            if kind == "priced":
                money(row, gcsub(i), amt, border=ff.SUBTOTAL_BORDER)
                money(row, gsxfx(i), _sf(amt), border=ff.SUBTOTAL_BORDER)
                csub.fill = ff.AQUA_FILL
                sxfx.fill = ff.AQUA_FILL
                if amt is not None and amt < 0:
                    # W-D ruling 5.3 (R22 adopted): a net-negative division
                    # subtotal is LEGAL and rendered (accounting-negative via
                    # the house format), but with no classification pathway
                    # yet (Q10) it is an error until classified a deductive
                    # alternate / approved credit. Fenced from benchmarks by
                    # M-2 membership (amount > 0); the bidder's own R20/CCS
                    # arithmetic keeps it.
                    tracker.paint(csub, "red", nm)
                    falke_rules.attach_comment(
                        csub,
                        "[FALKE R22] " + falke_rules.MSG_RED
                        + " (Negative value without deductive-alternate or "
                          "approved-credit classification — R22.)",
                    )
                else:
                    fail = falke_rules.r20_math_fail(lev_by_name[nm], csi)
                    if fail:
                        tracker.paint(csub, "red", nm)
                        falke_rules.attach_comment(
                            csub,
                            "[FALKE R20] " + falke_rules.MSG_RED
                            + f" (Submitted subtotal ${fail[0]:,.2f} vs "
                              f"line-item sum ${fail[1]:,.2f}; delta "
                              f"${fail[2]:,.2f} > max($5, 0.5%) — R20.)",
                        )
            elif kind == "not_comparable":
                # ENC-2 (division level): a subtotal composed entirely of
                # Not-Comparable lines is displayed but never benchmarked.
                money(row, gcsub(i), amt, border=ff.SUBTOTAL_BORDER)
                money(row, gsxfx(i), _sf(amt), border=ff.SUBTOTAL_BORDER)
                csub.fill = ff.AQUA_FILL
                sxfx.fill = ff.AQUA_FILL
                falke_rules.attach_comment(
                    csub,
                    "[ARA NC/R7] Not Comparable — excluded from benchmark "
                    "(R7/R8).",
                )
            elif kind == "by_owner":
                # ENC-1: render the bidder's VERBATIM classification token
                # ("Not Applicable", "By Owner", …) — never tell the board
                # the owner carries scope that was merely not applicable.
                token = falke_rules.by_owner_token(lev_by_name[nm], csi)
                csub.value = token
                csub.font = ff.ITALIC_FONT
                falke_rules.attach_comment(
                    csub,
                    f"[FALKE R3/R8] Approved classification: {token} "
                    "(R3/R6/R8). Excluded from benchmark calculations.",
                )
            elif kind == "excluded":
                csub.value = "Excluded"
                tracker.paint(csub, "red", nm)
                falke_rules.attach_comment(
                    csub,
                    "[FALKE R28] " + falke_rules.MSG_RED
                    + " (Exclusion without user approval — R28.)",
                )
            elif kind == "zero":
                money(row, gcsub(i), 0.0, border=ff.SUBTOTAL_BORDER)
                csub.fill = ff.AQUA_FILL
                tracker.paint(csub, "red", nm)
                falke_rules.attach_comment(
                    csub,
                    "[FALKE R6] " + falke_rules.MSG_RED
                    + " (Zero value without approved "
                    "classification — R6.)",
                )
            elif kind == "missing":
                # GOLD-DEV-8 (Marvin GOLD-DEV-10 ruling (3)): a division THIS
                # bidder's reclass fully vacated is not a missing scope — it
                # renders blank with the reclass story on-cell, NEVER an R5
                # red, even when a peer priced the FROM division. Consumes the
                # vacated set carried on the normalized view (Floyd W2-4) —
                # the same per-bidder suppression audit.py already applies.
                vacated_to = lev_by_name[nm].vacated_by_reclass.get(csi)
                if vacated_to is not None:
                    falke_rules.attach_comment(
                        csub,
                        f"[ARA RECLASS] Reclassified — this bidder's "
                        f"{_div_short(csi)} "
                        f"scope is carried in {_div_short(vacated_to)} on "
                        f"this view (KNOWN_FIRM_RECLASSIFIED; see Bid_Form "
                        f"Normalization Note and AUDIT).",
                    )
                elif any_priced:
                    tracker.paint(csub, "red", nm)
                    falke_rules.attach_comment(
                        csub,
                        "[FALKE R5] " + falke_rules.MSG_RED
                        + " (No pricing submitted for a "
                        "division priced by other bidders — R5.)",
                    )
            # missing & nobody priced & not vacated: blank aqua band, no paint
        member_prices = [
            p for p in (
                falke_rules.median_membership(lev_by_name[nm], csi)
                for nm in names
            ) if p is not None
        ]
        variance_pass(row, statuses,
                      lambda nm: gcsub(col_of[nm]), on_subtotal=True,
                      member_prices=member_prices)
        # REM-2: disclose engine-DERIVED subtotals on-cell. Runs AFTER the
        # variance pass so an existing cyan/yellow comment is appended to,
        # never overwritten.
        for nm, (kind, _amt) in statuses.items():
            if kind != "priced" or csi not in derived_csis.get(nm, ()):
                continue
            cell = ws.cell(row=row, column=gcsub(col_of[nm]))
            note = _derived_subtotal_note(nm)
            if cell.comment is not None:
                note = cell.comment.text + "\n\n" + note
            falke_rules.attach_comment(cell, note)
        row += 2  # subtotal + spacer

    # --- Footer (FEB 26 column conventions; col A BLANK per Derick review) ---
    components = [grand_total_component_amounts(b.footer) for b in ordered_leveled]
    _FEE_KEYS = tuple(k for k in GRAND_TOTAL_COMPONENT_KEYS
                      if k != "CONSTRUCTION_SUBTOTAL")
    row += 1
    fee_sub: dict[str, float] = {}
    for key, _writer_label in FOOTER_ROWS:
        label = LEVELED_FOOTER_LABELS[key]
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = ff.BOLD_FONT if key in (
            "CONSTRUCTION_SUBTOTAL", "FEES_SUBTOTAL", "GRAND_TOTAL"
        ) else ff.BODY_FONT
        for i, (nm, bid) in enumerate(zip(names, ordered_leveled)):
            comp = components[i]
            if key == "CONSTRUCTION_SUBTOTAL":
                val = float(comp.get("CONSTRUCTION_SUBTOTAL", 0))
            elif key in _FEE_KEYS:
                val = float(comp.get(key, 0))
            elif key == "FEES_SUBTOTAL":
                val = fee_sub.get(nm, 0.0)
            elif key == "GRAND_TOTAL":
                val = _cell_amount(bid.footer.grand_total.state,
                                   bid.footer.grand_total.amount)
            else:
                val = 0.0
            if key in _FEE_KEYS:
                fee_sub[nm] = fee_sub.get(nm, 0.0) + val

            if key in ("CONSTRUCTION_SUBTOTAL", "GRAND_TOTAL"):
                # FEB 26 rows 154/164 exact mechanism: merged COST:COST-
                # SUBTOTALS and $/SF:$/SXFX pairs, teal band, white text.
                # GT amount font NORMALIZED per Derick (format_falke note).
                is_gt = key == "GRAND_TOTAL"
                brd = ff.BOX_BORDER if is_gt else ff.SUBTOTAL_BORDER
                amt_font = ff.GT_AMOUNT_FONT if is_gt else ff.WHITE_BOLD_FONT
                sf_font = ff.GT_SF_FONT if is_gt else ff.WHITE_BOLD_FONT
                c1 = money(row, gcost(i), val, font=amt_font, border=brd)
                c1.alignment = Alignment(horizontal="center" if is_gt else "left")
                c2 = money(row, gsfc(i), _sf(val), font=sf_font, border=brd)
                c2.alignment = Alignment(horizontal="right" if is_gt else "left")
                for c in range(gcost(i), gsxfx(i) + 1):
                    ws.cell(row=row, column=c).fill = ff.TEAL_FILL
                    ws.cell(row=row, column=c).border = brd
                ws.merge_cells(start_row=row, start_column=gcost(i),
                               end_row=row, end_column=gcsub(i))
                ws.merge_cells(start_row=row, start_column=gsfc(i),
                               end_row=row, end_column=gsxfx(i))
                lab = ws.cell(row=row, column=2)
                lab.fill = ff.TEAL_FILL
                lab.font = ff.WHITE_BOLD_FONT
                lab.border = brd
                ws.row_dimensions[row].height = (
                    ff.GT_ROW_HEIGHT if is_gt else ff.CCS_ROW_HEIGHT
                )
                if is_gt:
                    # R21: grand total vs sum of approved components.
                    expected = float(sum(components[i].values()))
                    delta = abs(val - expected)
                    if delta > falke_rules.tol(expected):
                        tracker.paint(c1, "red", nm)
                        falke_rules.attach_comment(
                            c1,
                            "[FALKE R21] " + falke_rules.MSG_RED
                            + f" (Grand total ${val:,.2f} vs sum of components "
                              f"${expected:,.2f}; delta ${delta:,.2f} > "
                              f"max($5, 0.5%) — R21.)",
                        )
                else:
                    # ENC-3 (S2-1, W-D): the stated CCS must compose from the
                    # displayed division subtotals at max($5, 0.5%). Same
                    # shared check as the AUDIT register row
                    # (SUBTOTAL_COMPOSITION_DISCREPANCY).
                    comp_fail = falke_rules.composition_check(
                        lev_by_name[nm],
                        [c_code for c_code, _n in DIVISION_ROWS],
                        stated_cs.get(nm),
                    )
                    if comp_fail is not None:
                        tracker.paint(c1, "red", nm)
                        falke_rules.attach_comment(
                            c1,
                            "[FALKE R21] " + falke_rules.MSG_RED
                            + f" (Construction Cost Subtotal does not compose "
                              f"from the displayed division subtotals — delta "
                              f"${comp_fail[1]:,.2f}.)",
                        )
            elif key == "FEES_SUBTOTAL":
                csub = money(row, gcsub(i), val, border=ff.SUBTOTAL_BORDER)
                sxfx = money(row, gsxfx(i), _sf(val), border=ff.SUBTOTAL_BORDER)
                csub.fill = ff.AQUA_FILL
                sxfx.fill = ff.AQUA_FILL
                ws.cell(row=row, column=2).fill = ff.AQUA_FILL
            else:
                money(row, gcost(i), val)
                money(row, gsfc(i), _sf(val))
        row += 1

    # --- Rail pass (v0.3.1, house-format addendum §B): thin gray #7F7F7F
    # left/right borders on EVERY cell from the header row through the
    # GRAND TOTAL row (the last footer row, so row-1 here), all columns —
    # INCLUDING blanks ("each cell has left & right borders even if it's
    # blank"). Existing top/bottom edges (header box, subtotal thin/double,
    # teal rows) are preserved; only left/right are set. Runs before the
    # alternates/legend region, which stays rail-free like the reference. ---
    gt_row = row - 1
    for r in range(HDR, gt_row + 1):
        for c in range(1, LAST + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(left=ff.RAIL_SIDE, right=ff.RAIL_SIDE,
                                 top=b.top, bottom=b.bottom)

    # --- Alternates (kept out of base; only if any; col A blank) ---
    if any(b.footer.alternates for b in ordered_leveled):
        row += 1
        ws.cell(row=row, column=2).value = (
            "Bid Alternates (add/deduct — NOT included in base comparison)")
        ws.cell(row=row, column=2).font = ff.BOLD_FONT
        row += 1
        for i, bid in enumerate(ordered_leveled):
            for alt in bid.footer.alternates:
                # Bidder attribution in col B (col A stays blank).
                ws.cell(row=row, column=2).value = (
                    f"{bid.contractor_name} — {alt.description}")
                ws.cell(row=row, column=2).font = ff.BODY_FONT
                if alt.amount is not None:
                    money(row, gcost(i), float(alt.amount))
                else:
                    ws.cell(row=row, column=gcost(i)).value = alt.display
                    ws.cell(row=row, column=gcost(i)).font = ff.BODY_FONT
                row += 1

    # --- Qualifications (col A blank) ---
    row += 1
    ws.cell(row=row, column=2).value = "Contractor Qualifications"
    ws.cell(row=row, column=2).font = ff.BOLD_FONT
    row += 1
    for i, bid in enumerate(ordered_leveled):
        q = ws.cell(row=row, column=gcost(i))
        q.value = bid.qualifications_text or ""
        q.font = ff.BODY_FONT
        q.alignment = Alignment(wrap_text=True)
    row += 2

    # --- Summary block (R32; Risk Profile omitted per Q6 — no formula given) ---
    ws.cell(row=row, column=2).value = (
        "LEVELING SUMMARY (Falke §15; Risk Profile omitted — Q6)")
    ws.cell(row=row, column=2).font = ff.BOLD_FONT
    row += 1
    for label, fn, nf in (
        ("Total Bid Amount (submitted)",
         lambda nm: float(lev_by_name[nm].footer.grand_total.amount or 0),
         ff.FALKE_AMOUNT_FORMAT),
        ("Adjusted Total (= leveled total, Q10)",
         lambda nm: float(lev_by_name[nm].footer.leveled_total or 0),
         ff.FALKE_AMOUNT_FORMAT),
        ("Red Flags",
         lambda nm: tracker.by_bidder.get(nm, {}).get("red", 0), None),
        ("Cyan Flags",
         lambda nm: tracker.by_bidder.get(nm, {}).get("cyan", 0), None),
        ("Yellow Flags",
         lambda nm: tracker.by_bidder.get(nm, {}).get("yellow", 0), None),
        ("Average Variance",
         lambda nm: (sum(var_by_bidder[nm]) / len(var_by_bidder[nm]))
         if var_by_bidder[nm] else None, ff.FALKE_PCT_FORMAT),
    ):
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = ff.BODY_FONT
        for i, nm in enumerate(names):
            c = ws.cell(row=row, column=gcsub(i))
            v = fn(nm)
            if v is not None:
                c.value = v
            c.font = ff.BODY_FONT
            if nf:
                c.number_format = nf
        row += 1
    row += 1

    # --- Decided rules & assumptions, then the ONE workbook LEGEND below it
    # (W-D B4/§2: the old Falke-only legend is superseded by the workbook
    # legend, placed below the assumptions block on this sheet) ---
    ws.cell(row=row, column=2).value = (
        "LEVELING RULE DECISIONS & ASSUMPTIONS IN FORCE "
        "(see FALKE-LEVELING-RULES-SPEC.md)")
    ws.cell(row=row, column=2).font = ff.BOLD_FONT
    row += 1
    for a in _LEVELED_ASSUMPTION_LINES:
        ws.cell(row=row, column=2).value = a
        ws.cell(row=row, column=2).font = ff.BODY_FONT
        row += 1
    row += 1
    _write_workbook_legend(ws, row, bold_font=ff.BOLD_FONT,
                           body_font=ff.BODY_FONT)

    return tracker.counts


# On-sheet record of the decided rules and open-question defaults this view is
# built on (carried from the Derick-approved sample; Derick to confirm the
# remaining Q-items with Falke).
_LEVELED_ASSUMPTION_LINES: tuple[str, ...] = (
    "A#1 (layout): 4-column bidder groups per FEB 26 — COST | COST SUBTOTALS | "
    "$/SF | $/SXFX SUBTOTALS; line amounts only in COST/$-SF, subtotals only in "
    "COST SUBTOTALS/$-SXFX (Derick R1–R4).",
    "A#2 (layout): footer follows FEB 26 exactly — fee rows in COST+$/SF; FEES "
    "SUBTOTAL in COST SUBTOTALS+$/SXFX; CONSTRUCTION SUBTOTAL & GRAND TOTAL "
    "merged across COST:COST SUBTOTALS (reference rows 154/164).",
    "A#3 (layout): VAR % occupies each group's separator column; benchmark "
    "block (median/spread/valid/confidence) appended right of all groups.",
    "A#4 (paint): division-level rules paint the COST SUBTOTALS cell; "
    "line-level rules paint the COST cell; R21 paints the merged GRAND TOTAL "
    "cell. Red-first precedence (R16).",
    "A#5 (Q9): hues assumed 00FFFF / FFFF00 / FF0000 — the Falke docx names "
    "colors without hex.",
    "A#6 (Q1): ±20% boundary INCLUSIVE; (Q5/RISK-2): paint requires ≥3 valid "
    "bids; (Q2): confidence ladder; (Q4): priced allowances count in the "
    "median.",
    "A#7 (R5/Q8): blank division = RED missing-pricing only where ≥1 other "
    "bidder priced it; blanks NEVER enter the median (A6). Zero displays as "
    "accounting '$ -' per FEB 26.",
    "A#8 (Q10): Adjusted Total = leveled total; (Q6): Risk Profile omitted — "
    "no formula given.",
    "A#9 (Q7/R30): six-output-tab structure not yet in scope; R19 unit math "
    "skipped (no qty/unit carried by extraction).",
    "A#10: Normalization Note column not on this sheet (recommendations live "
    "on Bid_Form; the moves are applied here).",
    "A#11 (teal rows): CONSTRUCTION COST SUBTOTAL / GRAND TOTAL CONSTRUCTION "
    "COST replicate FEB 26 rows 154/164. DECIDED (Derick, 2026-07-02): GT "
    "amounts NORMALIZED to Avenir Book 12 bold white underlined — the "
    "reference's Calibri-14 there is drift, not design.",
    "A#12 (col A — Derick review 2026-07-03): DIV row keys kept at regular "
    "weight; column A is BLANK for the entire footer/alternates/qualifications "
    "block (col-B display labels carry the meaning and anchor the tie-out).",
    "A#13 (Marvin gold-standard diff, 2026-07-03): a STATED $0 division "
    "subtotal is never a valid benchmark price (REM-1 — red per R6, or R28 "
    "'Excluded' / by-owner when the lines are so classified); engine-DERIVED "
    "subtotals carry an on-cell disclosure (REM-2); by-owner divisions "
    "display the bidder's verbatim classification token (ENC-1); "
    "'Not Comparable' amounts are displayed but excluded from every "
    "benchmark median (ENC-2, R7/R8).",
    "A#14 (M-1, W-D 2026-07-15): Scope-gap register rows (AUDIT) are "
    "thresholded at field median > $20,000; the R5 red on the leveled sheet "
    "is threshold-free per the Falke program.",
)


# ---------------------------------------------------------------------------
# WORKBOOK LEGEND (W-D B4/§2) — ONE identical block on BOTH data sheets
# ---------------------------------------------------------------------------
# Marvin's exact content: header + precedence line + 4 sections. Rendered
# with the swatch in column B and the text in column C (the established
# legend idiom). The AUDIT-fills bullet's "[soft red/yellow/green swatches]"
# is rendered as a SINGLE soft-red swatch cell (one swatch column exists;
# the text itself carries all three severities) — flagged for Marvin's
# ratification in the W-D report.

LEGEND_HEADER = "LEGEND — READING THIS WORKBOOK (every signal, all three tabs)"
LEGEND_PRECEDENCE = (
    "When signals collide on one cell: Quarantine ⚠ overrides all; "
    "then Red > Cyan > Yellow > Neutral."
)

# (section title, [(fill or None, text), ...]) — fills resolved at write time.
_LEGEND_SECTIONS: tuple = (
    ("Section 1 — Leveling colors (Leveled_Normalized only — the decision view):", (
        ("falke_red",
         "Red — Error / Requires Correction: missing pricing, unapproved zero "
         "or exclusion, or a math inconsistency (R5 / R6 / R20 / R21 / R28)."),
        ("falke_cyan",
         "Cyan — Potentially Underpriced (at or below benchmark × 0.80) — "
         "confirm full scope is included (R12)."),
        ("falke_yellow",
         "Yellow — Potentially Overpriced (at or above benchmark × 1.20) — "
         "confirm pricing basis (R13)."),
        (None,
         "Neutral — within ±20% of benchmark: no fill (R15). The aqua band on "
         "subtotal rows is house formatting, not a signal."),
        (None,
         'Italic text ("Excluded", "Not Applicable", …) — the bidder\'s own '
         "classification, shown verbatim."),
        (None,
         "Benchmark = median of valid bids only (R9); coloring requires "
         "≥3 valid bids (Q5)."),
    )),
    ("Section 2 — Bid_Form (the verification view):", (
        (None,
         "No leveling colors appear on Bid_Form. Every value and token is "
         "transcribed as the bidder submitted it; a blank cell means the "
         "bidder left it blank."),
        ("soft_yellow",
         "Normalization Note (column C only) — an estimator recommendation. "
         "Dollars are NOT moved on this sheet; the move is applied on "
         "Leveled_Normalized."),
    )),
    ("Section 3 — AUDIT (the estimator's diagnostic log):", (
        # Marvin amendment 2026-07-15 (W-D build item 1): ONE swatch column,
        # so the ARA severities render THREE rows — one per severity, each
        # with its own swatch ("a single red swatch captioned red/yellow/
        # green is exactly the ambiguity the legend exists to kill").
        ("soft_red",
         "AUDIT rows filled soft red — ARA severity RED: resolve before "
         "award. Appears only on the AUDIT tab."),
        ("soft_yellow_audit",
         "AUDIT rows filled soft yellow — ARA severity YELLOW: review. "
         "Appears only on the AUDIT tab."),
        ("soft_green",
         "AUDIT rows filled soft green — ARA severity GREEN: verified. "
         "Appears only on the AUDIT tab."),
        (None,
         "Board members: the decision view is Leveled_Normalized. This tab "
         "is the estimator's log of every check performed."),
    )),
    ("Section 4 — Quarantine (may appear on any sheet):", (
        ("soft_red",
         '⚠ "does not reconcile to source — verify": the tool\'s own '
         "post-write self-check failed on this figure. Verify it against the "
         "submitted bid before any award. Overrides every other signal on "
         "the cell."),
    )),
)

_LEGEND_FILLS = {
    "falke_red": falke_rules.RED_FILL,
    "falke_cyan": falke_rules.CYAN_FILL,
    "falke_yellow": falke_rules.YELLOW_FILL,
    "soft_yellow": YELLOW_FILL,        # FFF2CC — the Normalization Note hue
    "soft_yellow_audit": YELLOW_FILL,  # FFF2CC — ARA severity YELLOW swatch
    "soft_red": RED_FILL,              # FFCCCC — ARA/quarantine soft red
    "soft_green": GREEN_FILL,          # CCFFCC — ARA severity GREEN swatch
}


def _write_workbook_legend(ws, start_row: int, bold_font=None,
                           body_font=None) -> int:
    """Write the ONE workbook LEGEND block (identical content on both data
    sheets) starting at ``start_row``. Returns the next free row.

    Layout: header (col B, bold) → precedence line (col B) → per section:
    title (col B, bold) then bullets (swatch col B, text col C).
    """
    header_cell = ws.cell(row=start_row, column=2)
    header_cell.value = LEGEND_HEADER
    if bold_font is not None:
        header_cell.font = bold_font
    row = start_row + 1
    prec = ws.cell(row=row, column=2)
    prec.value = LEGEND_PRECEDENCE
    if body_font is not None:
        prec.font = body_font
    row += 1
    for title, bullets in _LEGEND_SECTIONS:
        tc = ws.cell(row=row, column=2)
        tc.value = title
        if bold_font is not None:
            tc.font = bold_font
        row += 1
        for fill_key, text in bullets:
            if fill_key is not None:
                ws.cell(row=row, column=2).fill = _LEGEND_FILLS[fill_key]
            txt = ws.cell(row=row, column=3)
            txt.value = text
            if body_font is not None:
                txt.font = body_font
            row += 1
    return row


# ---------------------------------------------------------------------------
# Stage 6b LOUD QUARANTINE — post-reconcile annotation pass
# ---------------------------------------------------------------------------
#
# When Stage 6b (reconcile_written_matrix) returns ≥1 POST_WRITE_TIEOUT_FAILURE,
# the matrix is STILL delivered (Derick's decision) but every affected figure is
# loud-quarantined so a non-technical board cannot mistake it for a verified
# matrix (Marvin's STAGE6B-QUARANTINE-DISCLOSURE-SPEC.md). The disclosure is three
# stacked, board-facing RED signals:
#   1. a RED banner at the top of Bid_Form AND Leveled_Normalized (this module),
#   2. a RED fill + verify-against-source comment on each failing cell, and
#   3. a RED AUDIT row + a QUARANTINE summary line (rows added by the pipeline via
#      the audit_items list; the summary line is written here).
#
# This runs AFTER reconcile so the banner reflects the tie-out result. The banner
# is written into rows 1–3 by OVERRIDING the existing title/details cells (same
# override idiom as _write_leveled_banner) — it does NOT shift the data rows, so
# the contractor-name row (5), footer label rows, and SUBTOTAL label rows that
# reconcile.py read stay at their original positions and the cell marks below
# re-locate them by the same labels (Marvin §2: "rows reconcile.py reads stay
# anchored by label, not by absolute number"). On Leveled_Normalized the RED
# quarantine banner sits ABOVE the existing yellow normalization banner.

# Banner text — Marvin §2 / §6 (exact strings). {N} filled per workbook.
_QUARANTINE_BANNER_LINE_1 = (
    "⚠ AUTOMATED CHECK FAILED — DO NOT RELY ON THE FLAGGED FIGURES FOR AN AWARD "
    "DECISION."
)


def _quarantine_banner_line_2(n: Optional[int]) -> str:
    """Marvin §2 Line 2 with singular/plural and the structural fallback.

    ``n`` is the count of distinct flagged figures; ``None`` ⇒ the count cannot be
    cleanly enumerated (a structural failure), so the banner uses "one or more
    figures" and never under-counts itself into looking minor.
    """
    if n is None:
        figures = "one or more figures"
        verb = "do"
    else:
        figures = f"{n} figure{'' if n == 1 else 's'}"
        verb = "does" if n == 1 else "do"
    return (
        f"This bid-comparison matrix did not pass the tool's final self-check. "
        f"{figures} on this sheet {verb} not reconcile to the source bids — the "
        f"tool's written value does not match its own verified calculation. This "
        f"is a tool/formatting problem, not a finding about any contractor's bid."
    )


_QUARANTINE_BANNER_LINE_3 = (
    "Before this matrix is used to award, a person must verify each flagged figure "
    "directly against that contractor's submitted bid. Flagged figures are marked "
    'in their cells with "⚠ does not reconcile to source — verify". The full list '
    "is on the AUDIT tab (filter the Code column for POST_WRITE_TIEOUT_FAILURE). "
    "Do not award off this matrix until every flagged figure has been checked by "
    "hand."
)

# Cell-comment text — Marvin §2 / §6 (exact string).
_QUARANTINE_CELL_COMMENT = (
    "[QUARANTINE] ⚠ does not reconcile to source — verify. The tool wrote {written} here; its "
    "own verified calculation was {expected} (difference {delta}). Check this "
    "figure against {contractor}'s submitted bid before relying on it."
)

# AUDIT summary quarantine line — Marvin §4 / §6 (exact string).
_QUARANTINE_AUDIT_SUMMARY = (
    "QUARANTINE: {n} figure(s) failed the tool's self-check and are FLAGGED on the "
    "Bid_Form sheet. Verify each against the contractor's submitted bid before any "
    "award."
)

# Message substrings that map a POST_WRITE_TIEOUT_FAILURE to a markable cell.
_GT_CELL_MARKERS = (
    "Grand-total tie-out FAILED",
    "Footer arithmetic FAILED",
    "Mirror/leveled grand-total mismatch",
)
_DIV_CELL_MARKER = "Division subtotal tie-out FAILED"


def _write_quarantine_banner(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    num_contractors: int,
    n: Optional[int],
    start_row: int = 1,
    used_width: Optional[int] = None,
) -> None:
    """Write the RED 3-row quarantine banner into ``start_row .. start_row+2``.

    Each banner row is MERGED across the full used width (col A → the sheet's
    last used column) so the long L2/L3 paragraphs flow HORIZONTALLY as a proper
    full-width banner instead of stacking into a tall, narrow column-A block
    (shared ``_render_full_width_banner``). Solid RED across the used width,
    Line 1 bold. ``used_width`` defaults to the mirror's width; the caller
    passes the leveled width (``_lev_last_col``) for Leveled_Normalized.
    """
    if used_width is None:
        used_width = _col_start(max(num_contractors - 1, 0)) + 2
    lines = [
        (_QUARANTINE_BANNER_LINE_1, True),
        (_quarantine_banner_line_2(n), False),
        (_QUARANTINE_BANNER_LINE_3, False),
    ]
    for offset, (text, weight_bold) in enumerate(lines):
        _render_full_width_banner(
            ws, start_row + offset, text, RED_FILL, used_width, bold=weight_bold
        )


def _shift_merges_and_heights(ws, n: int) -> None:
    """Shift merged ranges and explicit row heights down by ``n`` rows.

    ``ws.insert_rows`` moves cell VALUES and STYLES down but leaves merged
    ranges and row heights at their old coordinates (openpyxl limitation) —
    without this, every merged block on the leveled sheet (banner, bidder
    block, teal total rows) would detach from its content after the quarantine
    banner insertion. Call immediately after ``insert_rows(1, n)``.
    """
    for rng in ws.merged_cells.ranges:
        rng.shift(0, n)
    heights = {
        r: dim.height for r, dim in list(ws.row_dimensions.items())
        if dim.height is not None
    }
    for r in heights:
        ws.row_dimensions[r].height = None
    for r, h in heights.items():
        ws.row_dimensions[r + n].height = h


def _mark_cell(ws, row: int, col: int, written: str, expected: str,
               delta: str, contractor: str) -> None:
    """RED fill + verify-against-source comment on a single failing cell (§2).

    COMPOSES, never overwrites (Marvin GOLD-DEV-6 ruling (3)): the quarantine
    fill wins VISUALLY (established precedence: Quarantine > Falke Red), but
    the quarantine text PREPENDS above any existing comment so a tool defect
    landing on an already-flagged cell (e.g. an R21 red) never erases the
    bid-level story.
    """
    cell = ws.cell(row=row, column=col)
    text = _QUARANTINE_CELL_COMMENT.format(
        written=written, expected=expected, delta=delta, contractor=contractor,
    )
    if cell.comment is not None and cell.comment.text:
        text = text + "\n\n--- prior flag on this cell ---\n" + cell.comment.text
    cell.fill = RED_FILL
    cell.comment = Comment(text, "FALKE Stage 6b")


def _fmt_q(amount: Decimal) -> str:
    """Board-display dollar string for the cell comment (matches reconcile._fmt)."""
    return f"${int(amount.quantize(Decimal('1'))):,}"


def _mark_failing_cells(
    ws,
    sheet_name: str,
    ordered_bids: list[NormalizedBid],
    failures: list[AuditItem],
) -> None:
    """Mark each failing GRAND_TOTAL / SUBTOTAL cell on ``ws`` (Marvin §2/§3).

    Cells are re-located by label — the same anchors reconcile.py uses — and the
    written/expected/delta are recomputed here from the workbook + the blessed
    bids (writer-independent), so a parse of reconcile's message string is never
    required. Per-sheet geometry: on Bid_Form the footer anchors are col-A
    machine keys and the marked cell is the contractor's cost column; on
    Leveled_Normalized (v0.3.0) the footer anchors are the col-B display labels
    (LEVELED_FOOTER_LABELS — col A is blank there), the GRAND TOTAL mark lands
    on the merged COST anchor, and a division-subtotal mark lands on the COST
    SUBTOTALS column (+1). Structural failures (no single wrong cell) are
    skipped here; they are disclosed by the banner count and the AUDIT row only.
    """
    leveled = sheet_name == "Leveled_Normalized"
    col_start_fn = _lev_col_start if leveled else _col_start
    sub_col_offset = LEVELED_CSUB_OFFSET if leveled else 0

    def _footer_row(key: str) -> Optional[int]:
        if leveled:
            return _find_label_row_col_b(ws, LEVELED_FOOTER_LABELS[key])
        return _find_label_row_col_a(ws, key)

    name_to_col = _find_contractor_cols(ws, ordered_bids, col_start_fn)
    grand_total_row = _footer_row("GRAND_TOTAL")
    subtotal_rows = _find_subtotal_label_rows(ws)
    component_rows = {
        key: _footer_row(key)
        for key in GRAND_TOTAL_COMPONENT_KEYS
    }
    blessed_by_name = {b.contractor_name: b for b in ordered_bids}

    # A GRAND_TOTAL cell can be implicated by more than one failure (e.g. both a
    # grand-total tie-out and a mirror mismatch). Mark each GT cell ONCE, choosing
    # the "expected" that best describes the defect: grand-total tie-out (blessed
    # GT) > mirror mismatch (blessed GT) > footer arithmetic (the summed
    # components). Collect per-contractor, then emit.
    gt_priority = {"GRAND": 0, "MISMATCH": 1, "FOOTER": 2}
    gt_marks: dict[str, tuple[int, Decimal]] = {}  # contractor → (priority, expected)

    for f in failures:
        # Only failures stamped for THIS sheet (or the cross-sheet GT mismatch,
        # which has no [sheet] prefix and applies to GRAND_TOTAL on both sheets).
        is_mismatch = "Mirror/leveled grand-total mismatch" in f.message
        if not is_mismatch and f"[{sheet_name}]" not in f.message:
            continue

        col = name_to_col.get(f.contractor_name)
        if col is None:
            continue
        bid = blessed_by_name.get(f.contractor_name)
        if bid is None:
            continue

        # Division-subtotal failure → mark the (contractor, division) SUBTOTAL cell.
        if f.division_csi and _DIV_CELL_MARKER in f.message:
            sub_row = subtotal_rows.get(f.division_csi)
            if sub_row is None:
                continue
            sub_col = col + sub_col_offset
            written = _as_dec(ws.cell(row=sub_row, column=sub_col).value)
            expected = _blessed_div_subtotal(bid, f.division_csi)
            _mark_cell(
                ws, sub_row, sub_col,
                _fmt_q(written), _fmt_q(expected), _fmt_q(abs(written - expected)),
                f.contractor_name,
            )
            continue

        # Grand-total / footer-arithmetic / mirror-mismatch → GRAND_TOTAL cell.
        if grand_total_row is None:
            continue
        if "Grand-total tie-out FAILED" in f.message:
            kind, expected = "GRAND", _blessed_grand_total(bid)
        elif is_mismatch:
            kind, expected = "MISMATCH", _blessed_grand_total(bid)
        elif "Footer arithmetic FAILED" in f.message:
            # "Verified calculation" here is the sum of the written components —
            # that is what the GRAND TOTAL should equal but doesn't.
            kind = "FOOTER"
            expected = sum(
                (
                    _as_dec(ws.cell(row=component_rows[key], column=col).value)
                    for key in GRAND_TOTAL_COMPONENT_KEYS
                    if component_rows.get(key) is not None
                ),
                Decimal("0"),
            )
        else:
            continue  # structural — no single cell (banner + AUDIT cover it)

        prev = gt_marks.get(f.contractor_name)
        if prev is None or gt_priority[kind] < prev[0]:
            gt_marks[f.contractor_name] = (gt_priority[kind], expected)

    # Emit one mark per implicated GRAND_TOTAL cell.
    if grand_total_row is not None:
        for name, (_prio, expected) in gt_marks.items():
            col = name_to_col[name]
            written = _as_dec(ws.cell(row=grand_total_row, column=col).value)
            _mark_cell(
                ws, grand_total_row, col,
                _fmt_q(written), _fmt_q(expected), _fmt_q(abs(written - expected)),
                name,
            )


def _find_name_row(ws) -> int:
    """Return the contractor-name row: the row just below the "CSI" header row.

    Anchored to the "CSI" label in col A (header row) so it survives a banner
    row-shift on Leveled_Normalized; falls back to row 5 (the writer's default).
    """
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "CSI":
            return row + 1
    return 5


def _find_contractor_cols(
    ws, bids: list[NormalizedBid], col_start_fn=_col_start
) -> dict[str, int]:
    """Map contractor name → its COST column (read back from the name row).

    ``col_start_fn`` selects the sheet geometry: mirror stride 3 (default) or
    leveled stride 5 (``_lev_col_start``).
    """
    name_row = _find_name_row(ws)
    out: dict[str, int] = {}
    for i in range(len(bids)):
        col = col_start_fn(i)
        name = ws.cell(row=name_row, column=col).value
        if isinstance(name, str) and name:
            out[name] = col
    return out


def _find_label_row_col_a(ws, key: str) -> Optional[int]:
    """Row whose col-A value equals a footer key (e.g. GRAND_TOTAL)."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == key:
            return row
    return None


def _find_label_row_col_b(ws, label: str) -> Optional[int]:
    """Row whose col-B value equals a display label (leveled footer anchors)."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == label:
            return row
    return None


def _find_subtotal_label_rows(ws) -> dict[str, int]:
    """Map each division CSI code → its SUBTOTAL row (read back from col B)."""
    label_to_code = {
        f"{name.upper()} SUBTOTAL": code for code, name in DIVISION_ROWS
    }
    out: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=2).value
        if isinstance(label, str) and label in label_to_code:
            out[label_to_code[label]] = row
    return out


def _as_dec(v: object) -> Decimal:
    """Coerce a cell value to Decimal; non-numeric → 0 (matches reconcile._as_decimal)."""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _blessed_grand_total(bid: NormalizedBid) -> Decimal:
    gt = bid.footer.grand_total
    if gt.amount is not None and gt.state == CellState.AMOUNT:
        return Decimal(str(gt.amount))
    return Decimal("0")


def _blessed_div_subtotal(bid: NormalizedBid, csi_code: str) -> Decimal:
    total = Decimal("0")
    for div in bid.divisions:
        if div.csi_code != csi_code:
            continue
        cell = div.subtotal_cell
        if cell.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                          CellState.ALLOWANCE) and cell.amount is not None:
            total += Decimal(str(cell.amount))
    return total


def _quarantine_figure_count(failures: list[AuditItem]) -> Optional[int]:
    """Distinct flagged-figure count for the banner ``{N}`` (Marvin §2).

    Counts one per failing GRAND_TOTAL (grand-total / footer-arithmetic /
    mirror-mismatch) and one per failing (contractor, division) SUBTOTAL. Returns
    ``None`` when a STRUCTURAL failure is present (missing row/sheet/column or
    audit-row parity) — the count can't be cleanly enumerated, so the banner falls
    back to "one or more figures" and never under-counts.
    """
    figures: set[tuple[str, str]] = set()
    structural = False
    for f in failures:
        if f.division_csi and _DIV_CELL_MARKER in f.message:
            figures.add((f.contractor_name, f.division_csi))
        elif any(m in f.message for m in _GT_CELL_MARKERS):
            figures.add((f.contractor_name, "GRAND_TOTAL"))
        else:
            structural = True
    if structural:
        return None
    return len(figures)


def _append_audit_failure_rows(
    wb: openpyxl.Workbook,
    failures: list[AuditItem],
) -> None:
    """Append the POST_WRITE_TIEOUT_FAILURE rows to the AUDIT sheet (Marvin §4).

    The banner Line 3 tells the board to "filter the Code column for
    POST_WRITE_TIEOUT_FAILURE" — so each tie-out failure must appear as a RED row
    on the AUDIT tab. Rows are inserted at the END of the existing data region,
    matching the 8-column layout written by ``_write_audit_sheet`` (Status,
    View, Code, Contractor, Division, Line Item, Value, Message). The data
    region is located from the header row (label-anchored,
    ``find_audit_header_row``) so the top-of-tab key never breaks it. All
    tie-out failures are RED.

    NOTE: this runs AFTER reconcile's check-4 (audit-row parity), which counted
    only the Stage-5b rows — so appending here does not retroactively trip parity.
    """
    if "AUDIT" not in wb.sheetnames or not failures:
        return
    ws = wb["AUDIT"]

    header_row = find_audit_header_row(ws)
    if header_row is None:
        return
    # Locate the end of the contiguous data region below the header.
    row = header_row + 1
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    insert_at = row  # first blank row after the data region

    ws.insert_rows(insert_at, len(failures))
    for offset, f in enumerate(failures):
        r = insert_at + offset
        values = [
            f.status.value,
            AUDIT_VIEW_LABELS.get(f.view, f.view),
            f.code.value,
            f.contractor_name,
            f.division_csi or "",
            f.line_item_desc or "",
            f.value or "",
            f.message,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.fill = RED_FILL
            if col_idx == 1:
                c.font = Font(bold=True)


def _append_audit_quarantine_line(wb: openpyxl.Workbook, n: Optional[int]) -> None:
    """Add the QUARANTINE summary line ABOVE the RED/YELLOW/GREEN tally (§4).

    Inserts a RED-filled line at the top of the existing summary block on the
    AUDIT sheet. ``n`` ⇒ the distinct flagged-figure count; ``None`` ⇒ "one or
    more".
    """
    if "AUDIT" not in wb.sheetnames:
        return
    ws = wb["AUDIT"]
    # Find the summary block: the first "Total items audited:" row (col A).
    summary_row = None
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.startswith("Total items audited:"):
            summary_row = row
            break
    count_str = "one or more" if n is None else str(n)
    text = _QUARANTINE_AUDIT_SUMMARY.format(n=count_str)
    if summary_row is None:
        # No summary block (no audit_items) — append at the next free row.
        summary_row = ws.max_row + 2
        c = ws.cell(row=summary_row, column=1)
    else:
        ws.insert_rows(summary_row, 1)
        c = ws.cell(row=summary_row, column=1)
    c.value = text
    c.font = Font(bold=True)
    c.fill = RED_FILL


def apply_quarantine(
    output_path: str | Path,
    failures: list[AuditItem],
    bids: list[NormalizedBid],
    leveled_bids: Optional[list[NormalizedBid]] = None,
) -> int:
    """Loud-quarantine the just-written workbook IN PLACE (Marvin's spec).

    Re-opens ``output_path``, writes the RED banner on Bid_Form + Leveled_Normalized,
    marks each failing cell (RED fill + verify comment), appends the RED
    POST_WRITE_TIEOUT_FAILURE rows + the QUARANTINE summary line to the AUDIT
    sheet, and re-saves. Call ONLY when ``failures`` is non-empty. Returns
    the distinct flagged-figure count actually rendered in the banner ``{N}``
    (``-1`` when the structural fallback "one or more figures" was used) for the
    pipeline's console summary.

    Cell-mark re-location is label-anchored (writer-independent), matching
    reconcile.py — so it survives the banner write (which does not shift rows).
    """
    output_path = Path(output_path)
    wb = openpyxl.load_workbook(output_path)

    ordered_mirror = _sort_bids(bids)
    lev_by_name = {b.contractor_name: b for b in (leveled_bids or bids)}
    ordered_leveled = [
        lev_by_name.get(b.contractor_name, b) for b in ordered_mirror
    ]

    n = _quarantine_figure_count(failures)
    num_contractors = len(ordered_mirror)

    # --- Banner + cell marks on Bid_Form ---
    if "Bid_Form" in wb.sheetnames:
        ws = wb["Bid_Form"]
        _write_quarantine_banner(ws, num_contractors, n, start_row=1)
        _mark_failing_cells(ws, "Bid_Form", ordered_mirror, failures)

    # --- Banner + cell marks on Leveled_Normalized (RED banner ABOVE gray) ---
    # The leveled sheet carries the gray normalization banner in rows 1–2 and
    # the disclaimer in row 3. Insert 3 rows at the top so the RED quarantine
    # banner sits ABOVE them (Marvin §2). insert_rows moves values + styles;
    # _shift_merges_and_heights moves the merged ranges (banner, bidder block,
    # teal total rows) and row heights with them. The shift is harmless:
    # reconcile.py already ran on the pre-quarantine file, and the cell marks
    # below re-locate cells by label (name row anchored to the "CSI" header).
    # Bid_Form is NOT shifted: its rows 1–3 are title/details/blank, which the
    # RED banner simply overrides.
    if "Leveled_Normalized" in wb.sheetnames:
        ws_lev = wb["Leveled_Normalized"]
        ws_lev.insert_rows(1, 3)
        _shift_merges_and_heights(ws_lev, 3)
        _write_quarantine_banner(
            ws_lev, num_contractors, n, start_row=1,
            used_width=_lev_last_col(num_contractors),
        )
        _mark_failing_cells(ws_lev, "Leveled_Normalized", ordered_leveled, failures)

    # --- AUDIT: append the RED tie-out failure rows + the QUARANTINE summary line ---
    _append_audit_failure_rows(wb, failures)
    _append_audit_quarantine_line(wb, n)

    wb.save(output_path)
    return -1 if n is None else n


# ---------------------------------------------------------------------------
# Primary entry point
# ---------------------------------------------------------------------------

def write_matrix(
    bids: list[NormalizedBid],
    output_path: str | Path,
    run: RunInputs,
    audit_items: Optional[list[AuditItem]] = None,
    leveled_bids: Optional[list[NormalizedBid]] = None,
    run_id: Optional[str] = None,
) -> list[dict]:
    """
    Write normalized bid data into a fresh openpyxl Workbook.

    Project identity (title/address/SF basis) is supplied per-run via ``run``
    (RunInputs) — never hardcoded (M1/M2). ``run.gross_sf`` is the confirmed
    $/SF denominator; ``run.sf_basis_label`` labels the $/SF header.

    ``run_id`` is the run identity stamped into the workbook and carried into
    the run pack so the two are bound (Marvin §10.1). The pipeline mints it at
    run start and passes it here; a direct/programmatic caller that omits it
    gets a freshly minted one, so every workbook this function writes is
    stamped — there is no unstamped path.

    Option C writes TWO data sheets:
      * ``Bid_Form`` — the faithful mirror (as-submitted ``bids``), with the
        Col C Normalization Note for each known-firm reclass recommendation.
      * ``Leveled_Normalized`` — the moved-dollar view (``leveled_bids``), with
        the estimator-normalized banner; cross-bid audit signals apply here only.
    Plus a single ``AUDIT`` sheet whose ``View`` column segments the slices.
    When ``leveled_bids`` is None the leveled sheet mirrors ``bids`` (no reclass).

    Returns a list of per-bid summary dicts for reporting (mirror values).
    """
    output_path = Path(output_path)
    gsf = int(run.gross_sf)
    run_id = run_id or mint_run_id()

    # Step 1: Sort bids (leveled-total ascending). The leveled bids are sorted by
    # the SAME order as the mirror so the two sheets line up cell-for-cell.
    ordered_bids = _sort_bids(bids)
    print(f"  [write_matrix] Contractor order: "
          f"{[b.contractor_name for b in ordered_bids]}")

    leveled_by_name: dict[str, NormalizedBid] = {}
    if leveled_bids is not None:
        leveled_by_name = {b.contractor_name: b for b in leveled_bids}
    ordered_leveled = [
        leveled_by_name.get(b.contractor_name, b) for b in ordered_bids
    ]

    # Step 2: Create fresh workbook + the Bid_Form mirror sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Form"

    # Mirror: as-submitted, Normalization Notes shown, NO ARA fills (M-3/B4).
    footer_summaries = _populate_data_sheet(
        ws, ordered_bids, gsf, run, show_notes=True
    )

    # Leveled_Normalized: moved dollars applied, FEB 26 geometry + house format
    # + Falke leveling rules (v0.3.0). NO legacy ARA audit fills here — every
    # ARA diagnostic lives on the AUDIT sheet only (rules spec §4.4/A1).
    ws_lev = wb.create_sheet(title="Leveled_Normalized")
    rule_counts = _populate_leveled_sheet(ws_lev, ordered_leveled, gsf, run)
    print(f"  [write_matrix] Falke rules on Leveled_Normalized: "
          f"red={rule_counts['red']} cyan={rule_counts['cyan']} "
          f"yellow={rule_counts['yellow']} neutral={rule_counts['neutral']} "
          f"paint_suppressed_lt3_bids={rule_counts['gate_suppressed']}")

    # Step 8: Write AUDIT sheet (single sheet, View column segments the slices)
    if audit_items:
        _write_audit_sheet(wb, audit_items)

    # Step 9: Stamp producer/format-version/run-id/project-identity (invisible
    # custom doc properties — the scorecard's SUPPORTED_PRODUCER check reads
    # this (verdict f) and the run pack binds against it (P1-4 §8.3)) + save
    _stamp_workbook(wb, run, run_id)
    wb.save(output_path)
    print(f"  [write_matrix] Saved fresh workbook → {output_path}")
    print(f"  [write_matrix] Sheet dimensions: "
          f"{ws.max_row} rows × {ws.max_column} cols")

    # Build per-bid summary dicts (same shape as original for pipeline.py).
    # Row numbers are no longer sequential (dynamic per-item layout), so we
    # report 0 as the row sentinel and let pipeline.py print amounts only.
    summaries: list[dict] = []
    for i, bid in enumerate(ordered_bids):
        # Aggregate subtotal per CSI code (handles duplicate codes)
        seen: dict[str, float] = {}
        for div in bid.divisions:
            amount = _cell_amount(div.subtotal_cell.state, div.subtotal_cell.amount)
            seen[div.csi_code] = seen.get(div.csi_code, 0.0) + amount

        divisions_written = [
            {
                "csi_code": csi_code,
                "row": 0,  # dynamic layout — row number not fixed
                "amount": seen.get(csi_code, 0.0),
                "state": "AMOUNT",
            }
            for csi_code, _ in DIVISION_ROWS
        ]

        summaries.append({
            "contractor": bid.contractor_name,
            "matched": True,
            "name_col": _col_start(i),
            "divisions_written": divisions_written,
            "footer_written": footer_summaries[i],
            "warnings": [],
        })

    return summaries
