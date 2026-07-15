"""
FALKE Matrix Pipeline — Stage 6b: Post-Write Reconciliation (closed-loop tie-out)
=================================================================================
The Stage-5b audit (`audit_bids`) runs on the in-memory NormalizedBids BEFORE
`write_matrix`. Nothing else reads the FINISHED .xlsx back, so any write/transfer
error introduced DOWNSTREAM of the audit — a grand total written from the wrong
field, a division amount in the wrong row/column, a footer that no longer sums,
a dropped audit row — is unguarded and could put a wrong total in front of a
condo board.

This module closes that loop. `reconcile_written_matrix()` runs AFTER the save,
re-opens the written workbook, and asserts four invariants by reading the cells
back and comparing them to the audit-blessed values and the normalized objects:

  1. Per-contractor grand-total tie-out (core): the GRAND TOTAL cell == that
     bid's normalized footer.grand_total (the value Stage 5b blessed), within $1.
  2. Footer arithmetic re-checked FROM the written cells: construction subtotal +
     GL + Builders Risk + GC fee (read OUT of the cells) sum to the GRAND TOTAL
     CELL within $1. (Catches a write error the pre-write audit cannot see.)
  3. Division subtotals: each (bidder, division) SUBTOTAL cell == the normalized
     division subtotal, in the correct bidder COLUMN and correct division ROW.
  4. Audit-row count parity: the number of AuditItems the pipeline generated ==
     the number written to the AUDIT sheet (guards against truncation/overflow).

Behavior on mismatch (LOUD QUARANTINE — Derick's decision, Marvin's disclosure
spec STAGE6B-QUARANTINE-DISCLOSURE-SPEC.md): any failure on checks 1–4 yields a
first-class RED ``POST_WRITE_TIEOUT_FAILURE`` AuditItem.
``reconcile_written_matrix`` returns the list of failures; the pipeline DELIVERS
the file (it no longer refuses) but loud-quarantines every affected figure — a
RED top-of-sheet banner on Bid_Form + Leveled_Normalized, a RED in-place cell
mark with a verify-against-source comment, and a RED AUDIT row — then exits with
a DISTINCT code 3 ("delivered with verification failures"), never a silent clean
exit. A tie-out failure means the ENGINE's rendering is defective (a validated
number landed in the wrong cell); it is NOT a finding about a contractor's bid,
so the disclosure says "verify this figure against the submitted bid," not
"refusing to deliver." Tolerance is $1, matching ARITHMETIC_DISCREPANCY /
FOOTER_DISCREPANCY.

The cells are located by reading the written sheet's OWN labels (contractor
name in row 5, ``… SUBTOTAL`` rows in col B, footer keys in col A) — an
independent read-back, not a coordinate map handed over from the writer. That is
deliberate: a row-misplacement bug in the writer would also corrupt a handed-over
map, defeating check 3.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from src.audit import AuditCode, AuditItem, AuditStatus
from src.normalized_models import (
    GRAND_TOTAL_COMPONENT_KEYS,
    CellState,
    NormalizedBid,
    grand_total_component_sum,
)
from src.write_matrix import (
    DIVISION_ROWS,
    LEVELED_CSUB_OFFSET,
    LEVELED_FOOTER_LABELS,
    _col_start,
    _lev_col_start,
    _sort_bids,
    find_audit_header_row,
)

# CellStates that contribute a real numeric amount to a subtotal. Held HERE so
# check 3 computes its expected division subtotal INDEPENDENTLY of the writer's
# `write_matrix._cell_amount` — a future regression in that shared helper can't
# blind both the writer and this checker at once (Floyd C-2, reviewer
# independence). All other states (NULL_BLANK / EXCLUDED / BY_OWNER_OTHERS)
# contribute 0, matching the writer's current semantics.
_AMOUNT_BEARING_STATES = (
    CellState.AMOUNT,
    CellState.EXPLICIT_ZERO,
    CellState.ALLOWANCE,
)


def _expected_subtotal(state: CellState, amount: Optional[Decimal]) -> Decimal:
    """Independent expected division-subtotal scalar (no writer-helper dependency)."""
    if state in _AMOUNT_BEARING_STATES and amount is not None:
        return Decimal(str(amount))
    return Decimal("0")

# $1 tolerance — matches ARITHMETIC_DISCREPANCY / FOOTER_DISCREPANCY (Marvin).
TIEOUT_TOLERANCE = Decimal("1")

# Footer label that anchors the GRAND TOTAL row (col A key written by the writer).
_GRAND_TOTAL_KEY = "GRAND_TOTAL"
_CONSTRUCTION_KEY = "CONSTRUCTION_SUBTOTAL"
_GL_KEY = "GL_INSURANCE"
_BR_KEY = "BUILDERS_RISK"
_GC_FEE_KEY = "GC_FEE"
_OHP_KEY = "OVERHEAD_PROFIT"
_OTHER_FEES_KEY = "OTHER_FEES"
_BOND_KEY = "BOND"

# Every component row that COMPOSES the grand total (construction + the additive
# fee/insurance rows, including Bond). Stage 6b check 2 re-sums these read-back
# cells and asserts they equal the written GRAND TOTAL. Derived from the SINGLE
# SOURCE OF TRUTH (normalized_models.GRAND_TOTAL_COMPONENT_KEYS) so this checker,
# write_matrix's rendered footer, and audit.py's grand_total composition can
# never drift — Bond is an additive component of the grand total (Marvin's
# ruling), so it is included.
_GRAND_TOTAL_COMPONENT_KEYS = GRAND_TOTAL_COMPONENT_KEYS

# Human-readable component names for the missing-row diagnostic.
_COMPONENT_LABELS = {
    _CONSTRUCTION_KEY: "construction subtotal",
    _GL_KEY: "GL insurance",
    _BR_KEY: "Builders Risk",
    _GC_FEE_KEY: "GC fee",
    _OHP_KEY: "Overhead & Profit",
    _OTHER_FEES_KEY: "Other Fees",
    _BOND_KEY: "Bond",
    _GRAND_TOTAL_KEY: "GRAND TOTAL",
}


def _as_decimal(v: object) -> Decimal:
    """Coerce a cell value (float/int/None/str) to Decimal; non-numeric → 0."""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _fmt(amount: Decimal) -> str:
    """Board-display dollar string, matching audit._fmt."""
    return f"${int(amount.quantize(Decimal('1'))):,}"


def _fail(
    contractor: str,
    message: str,
    value: Optional[str] = None,
    division_csi: Optional[str] = None,
) -> AuditItem:
    """Build a first-class RED POST_WRITE_TIEOUT_FAILURE AuditItem."""
    return AuditItem(
        contractor_name=contractor,
        division_csi=division_csi,
        status=AuditStatus.RED,
        code=AuditCode.POST_WRITE_TIEOUT_FAILURE,
        message=message,
        value=value,
    )


def _find_footer_row(ws, key: str) -> Optional[int]:
    """Return the row whose col-A value equals a footer key (e.g. GRAND_TOTAL)."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == key:
            return row
    return None


def _find_footer_row_leveled(ws, key: str) -> Optional[int]:
    """Return the leveled-sheet row for a footer key, anchored on its col-B
    DISPLAY label (LEVELED_FOOTER_LABELS — the single source of truth shared
    with the writer). Column A is blank in the leveled footer block (v0.3.0),
    so col-B labels are the only anchors there; Bid_Form keeps col-A keys.
    """
    label = LEVELED_FOOTER_LABELS[key]
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == label:
            return row
    return None


def _find_subtotal_rows(ws) -> dict[str, int]:
    """Map each division's SUBTOTAL label (col B) to its written row.

    The writer emits the subtotal label as ``{division_name.upper()} SUBTOTAL``;
    we key the result by canonical CSI code so the caller can compare to the
    normalized division by code, independent of the writer's row math.
    """
    label_to_code = {
        f"{name.upper()} SUBTOTAL": code for code, name in DIVISION_ROWS
    }
    out: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=2).value
        if isinstance(label, str) and label in label_to_code:
            out[label_to_code[label]] = row
    return out


def _find_contractor_columns(
    ws, bids: list[NormalizedBid], col_start_fn=_col_start
) -> dict[str, int]:
    """Map each contractor name to the COST column it was written under (row 5).

    Reads the names back from the written sheet rather than recomputing column
    math, so a name-in-wrong-column write error surfaces as a missing/mismatched
    column (check 3 attribution). ``col_start_fn`` selects the sheet geometry:
    mirror stride 3 (default) or leveled stride 5 (``_lev_col_start``).
    """
    out: dict[str, int] = {}
    expected_cols = [col_start_fn(i) for i in range(len(bids))]
    for col in expected_cols:
        name = ws.cell(row=5, column=col).value
        if isinstance(name, str) and name:
            out[name] = col
    return out


def _expected_div_subtotals(bid: NormalizedBid) -> dict[str, Decimal]:
    """Aggregate a bid's expected per-CSI division subtotal (writer-independent)."""
    out: dict[str, Decimal] = {}
    for div in bid.divisions:
        amt = _expected_subtotal(div.subtotal_cell.state, div.subtotal_cell.amount)
        out[div.csi_code] = out.get(div.csi_code, Decimal("0")) + amt
    return out


def _check_sheet(
    ws,
    ordered_bids: list[NormalizedBid],
    sheet_name: str,
) -> tuple[list[AuditItem], dict[str, Decimal]]:
    """Run checks 1–3 on one data sheet against its own blessed values (§7.2).

    ``ordered_bids`` carry the expected values FOR THIS SHEET (mirror bids for
    Bid_Form, leveled bids for Leveled_Normalized). Returns the failure list and
    a ``{contractor_name: written_grand_total}`` map for the cross-sheet
    GT-equality invariant (§7.2.3).

    Per-sheet geometry (v0.3.0): Bid_Form keeps 3-column groups, col-A footer
    keys, and subtotals in the cost column. Leveled_Normalized uses 5-column
    FEB 26 groups (``_lev_col_start``), col-B footer DISPLAY labels
    (LEVELED_FOOTER_LABELS — col A is blank in its footer block), division
    subtotals in the COST SUBTOTALS column (+1), and grand-total/fee components
    on the COST column (the merged teal anchor).
    """
    failures: list[AuditItem] = []
    written_grand_totals: dict[str, Decimal] = {}

    leveled = sheet_name == "Leveled_Normalized"
    col_start_fn = _lev_col_start if leveled else _col_start
    sub_col_offset = LEVELED_CSUB_OFFSET if leveled else 0
    footer_row_fn = _find_footer_row_leveled if leveled else _find_footer_row

    name_to_col = _find_contractor_columns(ws, ordered_bids, col_start_fn)
    subtotal_rows = _find_subtotal_rows(ws)
    grand_total_row = footer_row_fn(ws, _GRAND_TOTAL_KEY)
    # Resolve the row of every grand-total component (construction + fee rows).
    component_rows = {
        key: footer_row_fn(ws, key) for key in _GRAND_TOTAL_COMPONENT_KEYS
    }

    for bid in ordered_bids:
        name = bid.contractor_name
        col = name_to_col.get(name)
        if col is None:
            failures.append(_fail(
                name,
                f"[{sheet_name}] Contractor column not found in the written sheet — "
                f"the name row (row 5) does not carry {name} under its expected "
                f"column. Cell mapping is broken; the figures under this contractor "
                f"cannot be trusted. Delivered with this column FLAGGED — do not rely "
                f"on any figure in it; re-run the matrix and confirm it ties out "
                f"before awarding.",
            ))
            continue
        col_letter = get_column_letter(col)

        # --- Check 1: per-contractor grand-total tie-out (core) ---
        blessed_gt = bid.footer.grand_total.amount
        if grand_total_row is None:
            failures.append(_fail(
                name,
                f"[{sheet_name}] GRAND_TOTAL footer row not found in the written sheet "
                f"— the grand total could not be located to verify it. Delivered with "
                f"this sheet FLAGGED — re-run the matrix and confirm it ties out "
                f"before relying on any total for an award.",
            ))
        else:
            written_gt = _as_decimal(ws.cell(row=grand_total_row, column=col).value)
            written_grand_totals[name] = written_gt
            expected_gt = (
                blessed_gt if (blessed_gt is not None
                               and bid.footer.grand_total.state == CellState.AMOUNT)
                else Decimal("0")
            )
            delta = abs(written_gt - expected_gt)
            if delta > TIEOUT_TOLERANCE:
                failures.append(_fail(
                    name,
                    f"[{sheet_name}] Grand-total tie-out FAILED: the written GRAND "
                    f"TOTAL cell ({col_letter}{grand_total_row} = {_fmt(written_gt)}) "
                    f"does not match the verified grand total "
                    f"({_fmt(expected_gt)}) — difference {_fmt(delta)}. This is a "
                    f"tool-rendering defect, not a bid finding. Delivered with this "
                    f"figure FLAGGED — verify the GRAND TOTAL against {name}'s "
                    f"submitted bid before relying on it for an award.",
                    value=f"written {_fmt(written_gt)} vs blessed {_fmt(expected_gt)} "
                          f"(Δ {_fmt(delta)})",
                ))

        # --- Check 2: footer arithmetic re-checked FROM the written cells ---
        # Re-sum EVERY rendered grand-total component (construction + GL +
        # Builders Risk + GC fee + Overhead & Profit + Other Fees + Bond) and
        # assert it equals the written GRAND TOTAL cell. The component set is the
        # single source of truth (normalized_models.GRAND_TOTAL_COMPONENT_KEYS),
        # the same one write_matrix's rendered footer and audit.py's
        # grand_total_component_sum use — so a firm that folds insurance
        # into Other Fees, or a bonded bid whose grand total INCLUDES bond, ties
        # out instead of falsely failing.
        _missing_components = [
            _COMPONENT_LABELS[comp]
            for comp, frow in component_rows.items()
            if frow is None
        ]
        if grand_total_row is None:
            _missing_components = _missing_components + [
                _COMPONENT_LABELS[_GRAND_TOTAL_KEY]
            ]
        if _missing_components:
            failures.append(_fail(
                name,
                f"[{sheet_name}] Footer arithmetic CANNOT be re-checked: expected "
                f"footer component row(s) {', '.join(_missing_components)} not found "
                f"— a footer label/row was dropped or overwritten. Delivered with the "
                f"footer FLAGGED — re-run the matrix and confirm the footer ties out "
                f"before relying on any total for an award.",
                value=f"missing footer row(s): {', '.join(_missing_components)}",
            ))
        else:
            summed = sum(
                (
                    _as_decimal(ws.cell(row=component_rows[key], column=col).value)
                    for key in _GRAND_TOTAL_COMPONENT_KEYS
                ),
                Decimal("0"),
            )
            w_grand = _as_decimal(ws.cell(row=grand_total_row, column=col).value)
            arith_delta = abs(summed - w_grand)
            # BIDDER-ERROR BRANCH (Marvin GOLD-DEV-6 ruling (1); Floyd
            # W2-1/W2-2): a bidder whose OWN stated grand total does not equal
            # its component composition (e.g. the recurring bond-on-top
            # presentation) produces the SAME delta in a faithfully written
            # sheet. Compute the bidder's own footer delta from the NORMALIZED
            # bid via the RENDERED composition — grand_total_component_sum
            # (memo other_fees suppressed to 0), the same single source of
            # truth the written footer rows consume — never raw footer fields
            # (W2-1). If the written delta matches the bidder's own delta
            # within tolerance, the sheet FAITHFULLY reproduces the bidder's
            # inconsistency: no POST_WRITE_TIEOUT_FAILURE — the pre-write
            # FOOTER_DISCREPANCY RED row and the R21 red on the GT cell
            # already tell that story in the RIGHT vocabulary (bidder error,
            # not tool defect). Suppression is per-bidder, per-sheet, check 2
            # ONLY (W2-2): any write corruption moves arith_delta away from
            # bidder_footer_delta and still fires here, and checks 1/3 guard
            # the GT and division cells independently.
            stated_gt = (
                Decimal(str(bid.footer.grand_total.amount))
                if (bid.footer.grand_total.amount is not None
                    and bid.footer.grand_total.state == CellState.AMOUNT)
                else Decimal("0")
            )
            bidder_footer_delta = abs(
                grand_total_component_sum(bid.footer) - stated_gt
            )
            faithful_bidder_error = (
                abs(arith_delta - bidder_footer_delta) <= TIEOUT_TOLERANCE
            )
            if arith_delta > TIEOUT_TOLERANCE and not faithful_bidder_error:
                failures.append(_fail(
                    name,
                    f"[{sheet_name}] Footer arithmetic FAILED in the written sheet: "
                    f"construction subtotal + GL + Builders Risk + GC fee + "
                    f"Overhead & Profit + Other Fees + Bond = {_fmt(summed)} does not "
                    f"equal the written GRAND TOTAL cell ({_fmt(w_grand)}) — difference "
                    f"{_fmt(arith_delta)}. A rendering error broke the footer sum. "
                    f"Delivered with this figure FLAGGED — verify {name}'s GRAND TOTAL "
                    f"and its footer components against the submitted bid before "
                    f"relying on them for an award.",
                    value=f"sum {_fmt(summed)} vs grand {_fmt(w_grand)} "
                          f"(Δ {_fmt(arith_delta)})",
                ))

        # --- Check 3: division subtotals — against THIS sheet's blessed map ---
        normalized_div_subtotal = _expected_div_subtotals(bid)
        for csi_code, _name in DIVISION_ROWS:
            sub_row = subtotal_rows.get(csi_code)
            if sub_row is None:
                failures.append(_fail(
                    name,
                    f"[{sheet_name}] Division {csi_code} SUBTOTAL row not found in the "
                    f"written sheet — a division row was dropped or its label was "
                    f"overwritten. Delivered with this division FLAGGED — re-run the "
                    f"matrix and confirm it ties out before relying on the {csi_code} "
                    f"figures for an award.",
                    division_csi=csi_code,
                ))
                continue
            sub_col = col + sub_col_offset
            sub_col_letter = get_column_letter(sub_col)
            written_sub = _as_decimal(ws.cell(row=sub_row, column=sub_col).value)
            expected_sub = normalized_div_subtotal.get(csi_code, Decimal("0"))
            sub_delta = abs(written_sub - expected_sub)
            if sub_delta > TIEOUT_TOLERANCE:
                failures.append(_fail(
                    name,
                    f"[{sheet_name}] Division subtotal tie-out FAILED for {csi_code}: "
                    f"written cell ({sub_col_letter}{sub_row} = {_fmt(written_sub)}) does "
                    f"not match the verified division subtotal ({_fmt(expected_sub)}) "
                    f"— difference {_fmt(sub_delta)}. The value was rendered into the "
                    f"wrong row/column or mis-written. Delivered with this figure "
                    f"FLAGGED — verify {name}'s {csi_code} subtotal against the "
                    f"submitted bid before relying on it for an award.",
                    value=f"written {_fmt(written_sub)} vs normalized "
                          f"{_fmt(expected_sub)} (Δ {_fmt(sub_delta)})",
                    division_csi=csi_code,
                ))

    return failures, written_grand_totals


def reconcile_written_matrix(
    output_path: str | Path,
    bids: list[NormalizedBid],
    audit_item_count: int,
    leveled_bids: Optional[list[NormalizedBid]] = None,
) -> list[AuditItem]:
    """Stage 6b — read the written .xlsx back and assert the tie-out invariants.

    Under Option C there are TWO data sheets to tie out (§7):
      * ``Bid_Form`` (mirror) ties to the AS-SUBMITTED values (``bids``).
      * ``Leveled_Normalized`` ties to the reclassified values (``leveled_bids``).
    Checks 1–3 run on each sheet with ITS OWN blessed map (so the mirror's
    DIV 11/DIV 13 are NOT compared to post-move values and vice-versa). A new
    invariant asserts ``Bid_Form`` GRAND TOTAL == ``Leveled_Normalized`` GRAND
    TOTAL per bidder (the move never changes a bidder's total). Check 4
    (audit-row parity) is unchanged — one AUDIT sheet with a View column.

    Parameters
    ----------
    output_path : the .xlsx that write_matrix just saved.
    bids : the mirror NormalizedBids (as-submitted blessed values).
    audit_item_count : the number of AuditItems generated (check 4).
    leveled_bids : the leveled NormalizedBids (reclassified blessed values). When
        None, the leveled sheet is tied to the mirror bids (no reclass case).

    Returns
    -------
    list[AuditItem] : RED POST_WRITE_TIEOUT_FAILURE items. Empty ⇒ ties out.
    """
    output_path = Path(output_path)
    failures: list[AuditItem] = []

    wb = openpyxl.load_workbook(output_path, data_only=True)

    ordered_mirror = _sort_bids(bids)
    # Leveled bids are written in the mirror's column order (by mirror leveled_total),
    # so order the leveled set to MATCH the mirror's contractor sequence.
    lev_by_name = {b.contractor_name: b for b in (leveled_bids or bids)}
    ordered_leveled = [
        lev_by_name.get(b.contractor_name, b) for b in ordered_mirror
    ]

    # --- Mirror sheet (Bid_Form) ties to as-submitted ---
    mirror_failures, mirror_gts = _check_sheet(
        wb["Bid_Form"], ordered_mirror, "Bid_Form"
    )
    failures.extend(mirror_failures)

    # --- Leveled sheet (Leveled_Normalized) ties to reclassified ---
    leveled_gts: dict[str, Decimal] = {}
    if "Leveled_Normalized" in wb.sheetnames:
        leveled_failures, leveled_gts = _check_sheet(
            wb["Leveled_Normalized"], ordered_leveled, "Leveled_Normalized"
        )
        failures.extend(leveled_failures)
    else:
        failures.append(_fail(
            "(all)",
            "Leveled_Normalized sheet not found in the written workbook — the "
            "leveled view was not produced. Delivered with this view FLAGGED — "
            "re-run the matrix; do not rely on a missing leveled comparison for an "
            "award.",
        ))

    # --- New invariant: mirror GT == leveled GT per bidder (§7.2.3) ---
    for bid in ordered_mirror:
        name = bid.contractor_name
        if name in mirror_gts and name in leveled_gts:
            gt_delta = abs(mirror_gts[name] - leveled_gts[name])
            if gt_delta > TIEOUT_TOLERANCE:
                failures.append(_fail(
                    name,
                    f"Mirror/leveled grand-total mismatch: Bid_Form GRAND TOTAL "
                    f"({_fmt(mirror_gts[name])}) != Leveled_Normalized GRAND TOTAL "
                    f"({_fmt(leveled_gts[name])}) — difference {_fmt(gt_delta)}. A "
                    f"normalization move leaked or double-counted dollars. Delivered "
                    f"with this figure FLAGGED — verify {name}'s GRAND TOTAL is "
                    f"identical on both sheets, against the submitted bid, before "
                    f"relying on it for an award.",
                    value=f"mirror {_fmt(mirror_gts[name])} vs leveled "
                          f"{_fmt(leveled_gts[name])} (Δ {_fmt(gt_delta)})",
                ))

    # --- Check 4: audit-row count parity (single AUDIT sheet, unchanged) ---
    written_audit_rows = _count_audit_rows(wb)
    if written_audit_rows != audit_item_count:
        failures.append(_fail(
            "(all)",
            f"Audit-row count parity FAILED: the tool generated "
            f"{audit_item_count} audit item(s) but {written_audit_rows} row(s) were "
            f"written to the AUDIT sheet. A board-facing flag may have been silently "
            f"lost. Delivered with the AUDIT tab FLAGGED — re-run the matrix and "
            f"confirm all flags are present before relying on this matrix for an "
            f"award.",
            value=f"generated {audit_item_count} vs written {written_audit_rows}",
        ))

    return failures


def _count_audit_rows(wb: openpyxl.Workbook) -> int:
    """Count the AuditItem data rows written to the AUDIT sheet.

    The writer lays items out below the column-header row (W-D S2-3: the key
    block now sits at the TOP of the tab, so the data region is located
    label-anchored from the header — ``find_audit_header_row`` — never a
    hardcoded row). We count contiguous non-empty Status cells (col A) below
    the header until the first blank — the data region.
    """
    if "AUDIT" not in wb.sheetnames:
        return 0
    ws = wb["AUDIT"]
    header_row = find_audit_header_row(ws)
    if header_row is None:
        return 0
    count = 0
    row = header_row + 1
    while ws.cell(row=row, column=1).value not in (None, ""):
        count += 1
        row += 1
    return count
