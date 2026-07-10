"""
FALKE Matrix Pipeline — Falke Leveling Rules (Leveled_Normalized)
==================================================================
Encodes Falke's Bid Leveling Program rules (Marvin's authoritative spec,
FALKE/03_Matrix/FALKE-LEVELING-RULES-SPEC.md — 34 rules, adoption-first) for
the Leveled_Normalized sheet:

  * Benchmark = MEDIAN of valid classified prices only (R9) — computed from
    cell STATE, never from written grid values (A6: a blank must never enter
    the median as 0.0).
  * Variance % per bidder vs benchmark (R11); Cyan ≤ ×0.80 / Yellow ≥ ×1.20
    (R12/R13, boundary INCLUSIVE per Q1 decided default).
  * Red for errors — missing pricing (R5), unapproved zero (R6), unapproved
    exclusion (R28), subtotal math (R20), grand-total math (R21).
  * Precedence Red > Cyan > Yellow > Neutral (R16).
  * Cyan/Yellow paint requires ≥3 valid bids (Q5/RISK-2 decided default);
    below that the benchmark still displays, no variance paint.
  * Confidence ladder (R29 + Q2): High = ≥4 valid & spread ≤30%;
    Low = <3 valid or spread >50%; else Medium.
  * Math tolerances max($5, 0.5%) for subtotal/grand-total checks (R19–R21,
    A4 — ⚠ RISK-1 noted in the spec, adopted as written).
  * Clarification-log language (R31) as cell comments; disclaimer (R34).

Ported from the Derick-approved sample builder
(FALKE/03_Matrix/falke_rules_sample_v2.py) in v0.3.0.
"""

from __future__ import annotations

from statistics import median
from typing import Optional

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from src.normalized_models import CellState, NormalizedBid

# ---------------------------------------------------------------------------
# Falke rule hues — ASSUMPTION (docx names colors without hex; Q9).
# ---------------------------------------------------------------------------

CYAN_HEX, YELLOW_HEX, RED_HEX = "00FFFF", "FFFF00", "FF0000"
CYAN_FILL = PatternFill("solid", fgColor=CYAN_HEX)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW_HEX)
RED_FILL = PatternFill("solid", fgColor=RED_HEX)

# ---------------------------------------------------------------------------
# Clarification-log language (R31, verbatim) + disclaimer (R34, verbatim)
# ---------------------------------------------------------------------------

MSG_RED = ("Please correct or clarify the pricing for this item. The current "
           "entry appears to be missing, incomplete, non-numeric or "
           "mathematically inconsistent.")
MSG_CYAN = ("Your price is more than 20% below the bid benchmark. Please "
            "confirm that the full scope, quantity, material, labor, equipment, "
            "access, logistics, OH&P and subcontractor costs are included.")
MSG_YELLOW = ("Your price is more than 20% above the bid benchmark. Please "
              "confirm whether this price includes added scope, premium "
              "materials, access constraints, overtime, contingency, duplicated "
              "scope or assumptions not carried by other bidders.")
DISCLAIMER = ("This bid-leveling analysis is intended to identify pricing "
              "variances, missing pricing, potential scope alignment issues and "
              "mathematical inconsistencies. Color-coded items should not be "
              "interpreted as final pricing errors without bidder clarification. "
              "All red, cyan and yellow items should be reviewed and clarified "
              "before making a recommendation.")
COMMENT_AUTHOR = "FALKE Leveling Rules"

# States that carry a VALID price into the benchmark median (R7/R9 + Q4:
# priced allowances count — they are contractual). EXPLICIT_ZERO is NOT valid
# (R6: zero = error unless approved); NULL_BLANK/EXCLUDED/BY_OWNER never are.
VALID_STATES = (CellState.AMOUNT, CellState.ALLOWANCE)

# Q5 / RISK-2 decided default: no cyan/yellow variance paint below 3 valid bids.
MIN_BIDS_FOR_PAINT = 3


def tol(expected: float, flat: float = 5.0) -> float:
    """R20/R21 math tolerance: max($5, 0.5%) — adopted as written (⚠ RISK-1)."""
    return max(flat, abs(expected) * 0.005)


def confidence(n_valid: int, spread_pct: Optional[float]) -> str:
    """R29 confidence, read as a ladder (Q2 decided default)."""
    if n_valid < 3 or (spread_pct is not None and spread_pct > 0.50):
        return "Low"
    if n_valid >= 4 and (spread_pct is not None and spread_pct <= 0.30):
        return "High"
    return "Medium"


def benchmark_stats(prices: list[float]) -> tuple[float, Optional[float]]:
    """R9/R10: (median benchmark, % spread) over the valid prices."""
    bench = float(median(prices))
    spread = ((max(prices) - min(prices)) / bench) if (len(prices) >= 2 and bench) else None
    return bench, spread


def variance_color(amount: float, bench: float) -> Optional[str]:
    """R12/R13 with the INCLUSIVE ±20% boundary (Q1 decided default)."""
    if amount <= bench * 0.80:
        return "cyan"
    if amount >= bench * 1.20:
        return "yellow"
    return None


def attach_comment(cell, text: str) -> None:
    """Attach the R31 clarification language as a cell comment."""
    cell.comment = Comment(text, COMMENT_AUTHOR, height=140, width=340)


class PaintTracker:
    """Per-sheet paint registry enforcing R16 red-first precedence + counts.

    A cell painted red is never downgraded; cyan beats yellow. Red paint also
    sets white bold text (readability on FF0000). Counts feed the R32 summary
    block and the pipeline's console report.
    """

    _ORDER = {"red": 0, "cyan": 1, "yellow": 2}
    _FILLS = {"red": RED_FILL, "cyan": CYAN_FILL, "yellow": YELLOW_FILL}

    def __init__(self, red_font=None):
        self._painted: dict[tuple[int, int], str] = {}
        self._red_font = red_font
        self.counts = {"red": 0, "cyan": 0, "yellow": 0, "neutral": 0,
                       "gate_suppressed": 0}
        self.by_bidder: dict[str, dict[str, int]] = {}

    def kind_at(self, cell) -> Optional[str]:
        return self._painted.get((cell.row, cell.column))

    def paint(self, cell, kind: str, bidder: str) -> None:
        prev = self._painted.get((cell.row, cell.column))
        if prev is not None and self._ORDER[prev] <= self._ORDER[kind]:
            return
        self._painted[(cell.row, cell.column)] = kind
        cell.fill = self._FILLS[kind]
        if kind == "red" and self._red_font is not None:
            cell.font = self._red_font
        self.counts[kind] += 1
        self.by_bidder.setdefault(bidder, {"red": 0, "cyan": 0, "yellow": 0})
        self.by_bidder[bidder][kind] += 1

    def count_neutral(self) -> None:
        self.counts["neutral"] += 1

    def count_gate_suppressed(self) -> None:
        self.counts["gate_suppressed"] += 1


# ---------------------------------------------------------------------------
# State classification (A6: benchmarks and paint read cell STATE, never the
# written grid — the writer prints tokens/blanks for non-priced states, and a
# blank must never enter the median as 0.0)
# ---------------------------------------------------------------------------

def div_status(bid: NormalizedBid, csi: str) -> tuple[str, Optional[float]]:
    """Classify one bidder's division: (kind, amount).

    kind ∈ {"priced", "not_comparable", "zero", "excluded", "by_owner",
    "missing"}. "priced" aggregates every VALID (AMOUNT/ALLOWANCE) subtotal
    for the CSI code (duplicate division codes aggregate, matching the
    writer); a subtotal composed entirely of Not-Comparable lines reports
    "not_comparable" with its amount (displayed, never benchmarked — ENC-2).
    Non-priced kinds are reported worst-first: excluded > by_owner > zero >
    missing.

    When the subtotal state alone reads "missing" OR "zero", the division's
    LINE states are consulted, so R28 (exclusion without approval) and the
    approved by-owner/not-applicable classifications keep their own language
    and treatment instead of collapsing into R5/R6:

    * blank subtotal (normalize renders a fully-excluded division's subtotal
      as NULL_BLANK, since excluded lines contribute no sum): any EXCLUDED
      line → "excluded"; else any BY_OWNER line → "by_owner"; only when no
      line is blank (conservative — a blank line keeps "missing").
    * stated-$0 subtotal (EXPLICIT_ZERO per REM-1 — a stated $0 is never a
      valid benchmark price): ALL lines EXCLUDED → "excluded"; ALL lines
      BY_OWNER → "by_owner"; otherwise stays "zero" (R6 language). Marvin's
      ruling — a recurring firm's pattern (typed $0.00 division TOTAL over
      'Excluded' lines) must read R28, not generic zero.
    """
    divs = [d for d in bid.divisions if d.csi_code == csi]
    if not divs:
        return "missing", None
    total, saw = 0.0, False
    kinds = set()
    for d in divs:
        st, amt = d.subtotal_cell.state, d.subtotal_cell.amount
        if st in VALID_STATES and amt is not None:
            total += float(amt)
            saw = True
        elif st == CellState.EXPLICIT_ZERO:
            kinds.add("zero")
        elif st == CellState.EXCLUDED:
            kinds.add("excluded")
        elif st == CellState.BY_OWNER_OTHERS:
            kinds.add("by_owner")
        else:
            kinds.add("missing")
    line_states = {
        c.state for d in divs for c in d.line_item_cells.values()
    }
    if saw:
        # ENC-2 at division level: a subtotal composed ENTIRELY of
        # Not-Comparable lines is itself not comparable — display it, keep it
        # out of the subtotal benchmark median.
        if line_states and line_states <= {CellState.NOT_COMPARABLE}:
            return "not_comparable", total
        return "priced", total
    if kinds == {"missing"}:
        # Blank subtotal — classify from the line states before calling it
        # missing (R28/R6/By-Owner keep their own language and treatment).
        if line_states and CellState.NULL_BLANK not in line_states:
            if CellState.EXCLUDED in line_states:
                return "excluded", None
            if CellState.BY_OWNER_OTHERS in line_states:
                return "by_owner", None
    if kinds == {"zero"}:
        # Stated-$0 subtotal (REM-1) — all-one-kind line consult.
        if line_states == {CellState.EXCLUDED}:
            return "excluded", None
        if line_states == {CellState.BY_OWNER_OTHERS}:
            return "by_owner", None
    for k in ("excluded", "by_owner", "zero", "missing"):
        if k in kinds:
            return k, None
    return "missing", None


def by_owner_token(bid: NormalizedBid, csi: str) -> str:
    """The verbatim classification token to DISPLAY for a by-owner division.

    ENC-1 (Marvin): a division classified "Not Applicable" must not render as
    "By Owner" — that would tell the board the owner carries the scope, a
    wrong story. Returns the first by-owner line/subtotal display (normalize
    populates it from ``by_others_verbatim``), falling back to "By Owner".
    """
    for d in bid.divisions:
        if d.csi_code != csi:
            continue
        for c in [d.subtotal_cell] + list(d.line_item_cells.values()):
            if c.state == CellState.BY_OWNER_OTHERS and c.display:
                token = c.display.strip()
                if token and token.upper() != "BY OTHERS":
                    return token
    return "By Owner"


def line_status(bid: NormalizedBid, csi: str, desc: str) -> tuple[str, Optional[float]]:
    """Classify one bidder's unified line item: (kind, amount)."""
    from src.write_matrix import _descriptions_match, _find_div

    d = _find_div(bid, csi)
    if d is None:
        return "missing", None
    cell = d.line_item_cells.get(desc)
    if cell is None:
        for key, c in d.line_item_cells.items():
            if _descriptions_match(desc, key):
                cell = c
                break
    if cell is None:
        return "missing", None
    if cell.state in VALID_STATES and cell.amount is not None:
        return "priced", float(cell.amount)
    if cell.state == CellState.EXPLICIT_ZERO:
        return "zero", 0.0
    if cell.state == CellState.NOT_COMPARABLE:
        # ENC-2: amount displayed, NEVER counted as a valid benchmark price.
        return ("not_comparable",
                float(cell.amount) if cell.amount is not None else None)
    if cell.state == CellState.EXCLUDED:
        return "excluded", None
    if cell.state == CellState.BY_OWNER_OTHERS:
        return "by_owner", None
    return "missing", None


def r20_math_fail(bid: NormalizedBid, csi: str) -> Optional[tuple[float, float, float]]:
    """R20: division subtotal vs line-item sum beyond max($5, 0.5%).

    Returns (subtotal, item_sum, delta) when the check FAILS, else None.
    Only fires when the bidder stated BOTH a valid subtotal and priced items.
    """
    divs = [d for d in bid.divisions if d.csi_code == csi]
    if not divs:
        return None
    sub = items = 0.0
    have_sub = have_items = False
    for d in divs:
        st, amt = d.subtotal_cell.state, d.subtotal_cell.amount
        if st in VALID_STATES and amt is not None:
            sub += float(amt)
            have_sub = True
        for c in d.line_item_cells.values():
            # The bidder's OWN arithmetic keeps a Not-Comparable amount
            # (ENC-2 fences it out of benchmarks only, never out of the
            # bidder's stated sums).
            if c.state in (CellState.AMOUNT, CellState.EXPLICIT_ZERO,
                           CellState.ALLOWANCE,
                           CellState.NOT_COMPARABLE) and c.amount is not None:
                items += float(c.amount)
                have_items = True
    if not (have_sub and have_items):
        return None
    delta = abs(sub - items)
    return (sub, items, delta) if delta > tol(sub) else None
