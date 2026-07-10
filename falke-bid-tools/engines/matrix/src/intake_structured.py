"""
FALKE Matrix Pipeline — Deterministic Structured Intake (.xlsx / .csv)
=======================================================================
Falke is moving from PDF bid submissions to structured spreadsheet files.
This module parses a per-bid .xlsx or .csv DETERMINISTICALLY — no vision, no
extraction agent — and emits a validated BidDocument JSON into the interim
dir, where the pipeline picks it up exactly like an agent-extracted bid.

Two supported layouts (anything else fails LOUDLY — never guess a layout):

  1. FALKE_FORM — the Falke standard bid form as a spreadsheet (the FEB 26
     'Ready to Use' structure): a header row carrying "CSI" and
     "BUILDING SYSTEM", per-bidder COST / COST SUBTOTALS columns, division
     blocks ("DIV XX ..." keys) with line items and "<NAME> SUBTOTAL" rows,
     then the footer (CONSTRUCTION COST SUBTOTAL, insurance/fee rows, fees
     SUBTOTAL, GRAND TOTAL CONSTRUCTION COST), an ALTERNATES section, and
     Notes/Qualifications/Exclusions text blocks.

  2. ROW_SCHEMA — the generic row schema from Falke's program instructions §2:
     one row per scope item with columns Item Code / Category (Trade) /
     Description / Quantity / Unit / one bidder price column / Status
     (Classification) / Notes. Exactly ONE bidder price column per file
     (per-bid submission contract).

Classification vocabulary (Falke §2 R3, mapped per Marvin's rules spec A5 —
approved classifications clear the R5/R6 error states):

  Base Bid       → plain priced amount
  Allowance      → is_allowance=True (no amount ⇒ missing pricing, R26)
  Alternate      → footer.alternates (never in the base comparison)
  Excluded       → is_excluded=True (RED unless user-approved, R28)
  By Owner       → is_by_owner_others=True (approved classification, R6)
  Not Applicable → is_by_owner_others=True, verbatim preserved (approved
                   classification per R6; the verbatim token is what the
                   leveled sheet DISPLAYS — ENC-1)
  Not Comparable → is_not_comparable=True: amount kept as submitted (R33) and
                   kept in the bidder's own subtotal math, but EXCLUDED from
                   every cross-bid benchmark median (R7/A5 — ENC-2); note +
                   extraction warning preserved

Numeric coercion per §3 (R4): strip currency symbols/commas/spaces,
parentheses ⇒ negative. A bare dash is MISSING pricing (Q11/RISK-5 decided
default: R5-class, not a zero). Spreadsheet error tokens (#REF! etc., R18)
⇒ missing + warning. A literal 0 typed in a price cell IS an explicit zero.
"""

from __future__ import annotations

import csv
import json
import re
import sys
import traceback
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl

from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)

STRUCTURED_SUFFIXES = (".xlsx", ".csv")

# Floyd C1: size ceiling before any parse — a zip bomb or runaway export must
# fail loudly per-file, never stall or crash the batch. Real bid forms are a
# few hundred KB; 20 MB is generous.
MAX_STRUCTURED_FILE_BYTES = 20 * 1024 * 1024

# Floyd C4: the on-disk ceiling guards COMPRESSED bytes only — an xlsx (a zip)
# under 20 MB on disk can decompress to gigabytes (demonstrated 7 MB → 482 MB;
# crafted bombs exceed 1000:1). Before openpyxl ever parses, the summed
# DECOMPRESSED member sizes (ZipInfo.file_size) must clear this ceiling. A real
# bid form's sheet XML is a few MB decompressed; 100 MB is generous.
MAX_DECOMPRESSED_XLSX_BYTES = 100 * 1024 * 1024


class IntakeError(Exception):
    """A structured bid file could not be parsed deterministically.

    Raised LOUDLY with a message naming the file and the reason — the pipeline
    surfaces it as a validation error and the skill layer hard-stops on any
    missing bidder. Never guess a layout.
    """


# ---------------------------------------------------------------------------
# Numeric coercion (R4) + price-token classification (R17 / R5 / R6)
# ---------------------------------------------------------------------------

# Spreadsheet formula/error tokens (R18).
_FORMULA_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#N/A", "#NAME?", "#NUM!",
                   "#NULL!"}

# Non-numeric pricing tokens (R17) that mean "missing unless classified" (R5).
_MISSING_TOKENS = {"n/a", "na", "tbd", "to be provided", "not priced",
                   "no price", "pending", "included but no value", "included",
                   "see above", "see notes", "per plans", "-", "–", "—"}


def _coerce_amount(value) -> tuple[Optional[Decimal], Optional[str]]:
    """Coerce a raw cell value to (Decimal | None, leftover_token | None).

    Returns (amount, None) for a clean number; (None, token) when the cell
    carried a non-numeric token (missing/classification text, R17); and
    (None, None) for a genuinely blank cell.
    """
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, str(value)
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)), None
    text = str(value).strip()
    if not text:
        return None, None
    if text in _FORMULA_ERRORS:
        return None, text
    cleaned = text.replace("$", "").replace(",", "").replace(" ", "")
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    if negative:
        cleaned = cleaned[1:-1]
    if cleaned in ("-", "–", "—", ""):
        # Dash = MISSING pricing (Q11/RISK-5 decided default), not a zero.
        return None, text
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None, text
    return (-amount if negative else amount), None


def _slug(contractor_name: str) -> str:
    """Same slug rule as the extraction agents (SKILL.md Step 5)."""
    s = contractor_name.lower().replace(" ", "_")
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s[:30] or "structured_bid"


def _norm(value) -> str:
    """Normalized header/label text for matching."""
    return str(value or "").strip().lower()


# ---------------------------------------------------------------------------
# Classification (Falke §2 R3 vocabulary → BidDocument flags)
# ---------------------------------------------------------------------------

def _apply_status(
    item: LineItem,
    status: str,
    warnings: list[str],
    where: str,
) -> str:
    """Apply a Status/Classification token to a LineItem IN PLACE.

    Returns the routing verdict: "line" (stays a division line item),
    "alternate" (goes to footer.alternates), or "line" for everything else.
    """
    s = _norm(status)
    if s in ("", "base bid", "included"):
        return "line"
    if s == "allowance":
        item.is_allowance = True
        return "line"
    if s == "alternate":
        return "alternate"
    if s == "excluded":
        item.is_excluded = True
        return "line"
    if s == "by owner":
        item.is_by_owner_others = True
        item.by_others_verbatim = str(status).strip()
        return "line"
    if s == "not applicable":
        # Approved classification (R6) — closest BidDocument flag is the
        # by-owner/others family; verbatim preserves the real term.
        item.is_by_owner_others = True
        item.by_others_verbatim = str(status).strip()
        return "line"
    if s == "not comparable":
        # ENC-2: dedicated flag — amount kept as submitted (R33), fenced out
        # of every cross-bid benchmark (R7/A5) downstream.
        item.is_not_comparable = True
        item.notes = ((item.notes + " | ") if item.notes else "") + \
            "Status: Not Comparable"
        warnings.append(
            f"{where}: status 'Not Comparable' — amount kept as submitted, "
            f"excluded from benchmark calculations (R7/R8)."
        )
        return "line"
    item.notes = ((item.notes + " | ") if item.notes else "") + \
        f"Status: {str(status).strip()}"
    warnings.append(
        f"{where}: unrecognized Status/Classification {str(status).strip()!r} "
        f"— recorded as a note, treated as unclassified."
    )
    return "line"


# ---------------------------------------------------------------------------
# Layout detection
# ---------------------------------------------------------------------------

def _find_falke_header(ws) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the Falke-form header row: returns (row, column-map) or None.

    The column map carries: csi, desc, cost, csub (COST SUBTOTALS).
    """
    for r in range(1, min(ws.max_row, 40) + 1):
        cols: dict[str, int] = {}
        bidder_groups = 0
        for c in range(1, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if v == "csi" and "csi" not in cols:
                cols["csi"] = c
            elif v == "building system" and "desc" not in cols:
                cols["desc"] = c
            elif v == "cost":
                if "cost" not in cols:
                    cols["cost"] = c
            elif v.startswith("cost") and "subtotal" in v:
                bidder_groups += 1
                if "csub" not in cols:
                    cols["csub"] = c
        if {"csi", "desc", "cost", "csub"} <= set(cols):
            if bidder_groups > 1:
                raise IntakeError(
                    f"this looks like a MULTI-BIDDER comparison matrix "
                    f"({bidder_groups} 'COST SUBTOTALS' groups) — structured "
                    f"intake takes one PER-BID form per file. Supply each "
                    f"bidder's own form separately."
                )
            return r, cols
    return None


def _find_row_schema_header(rows: list[list]) -> Optional[tuple[int, dict]]:
    """Locate the §2 row-schema header row in a grid of raw rows.

    Returns (row_index, header-map) or None. The header map carries the
    0-based column index of each recognized field plus the single bidder
    price column (name + index).
    """
    known = {
        "item code": "item_code",
        "category": "category", "trade": "category",
        "category / trade": "category", "category/trade": "category",
        "description": "description",
        "quantity": "quantity", "qty": "quantity",
        "unit": "unit",
        "status": "status", "classification": "status",
        "status / classification": "status", "status/classification": "status",
        "notes": "notes",
    }
    for idx, row in enumerate(rows[:20]):
        fields: dict[str, int] = {}
        price_cols: list[tuple[int, str]] = []
        for c, cell in enumerate(row):
            v = _norm(cell)
            if not v:
                continue
            if v in known:
                fields[known[v]] = c
            else:
                price_cols.append((c, str(cell).strip()))
        if "item_code" in fields and "description" in fields:
            if len(price_cols) != 1:
                names = [n for _, n in price_cols] or ["<none>"]
                raise IntakeError(
                    f"row-schema file must carry exactly ONE bidder price "
                    f"column (per-bid submission); found {len(price_cols)}: "
                    f"{', '.join(names)}. Split multi-bidder files into one "
                    f"file per bidder."
                )
            fields["_price_col"] = price_cols[0][0]
            fields["_bidder"] = price_cols[0][1]
            return idx, fields
    return None


# ---------------------------------------------------------------------------
# Layout 1 — Falke standard bid form (xlsx)
# ---------------------------------------------------------------------------

_DIV_KEY_RE = re.compile(r"^DIV\s*(\d{2})", re.IGNORECASE)

# Footer labels (matched on normalized text, most specific first).
_QUAL_HEADINGS = {
    "notes:": "notes", "notes": "notes",
    "qualifications:": "qualifications", "qualifications": "qualifications",
    "exclusions:": "exclusions", "exclusions": "exclusions",
    "assumptions:": "assumptions", "assumptions": "assumptions",
    "terms:": "terms", "terms": "terms",
}


def _parse_falke_form(path: Path, ws, header_row: int,
                      cols: dict[str, int], ws_raw=None) -> BidDocument:
    """Parse one bidder's Falke standard bid form spreadsheet.

    ``ws_raw`` is the same sheet loaded with ``data_only=False`` (formulas
    visible) — used ONLY to loud-fail on a formula in the bidder block (Floyd
    C3): under ``data_only=True`` an uncached formula name cell reads None and
    would silently misattribute the bid to the next string (the project name).
    """
    warnings: list[str] = []
    csi_c, desc_c, cost_c, csub_c = (cols["csi"], cols["desc"],
                                     cols["cost"], cols["csub"])

    # --- Bidder block above the header row: name / project / TOTAL GSF ---
    contractor_name: Optional[str] = None
    project_name: Optional[str] = None
    total_gsf: Optional[int] = None
    for r in range(1, header_row):
        # Floyd C3: a formula anywhere in the bidder block's value column is a
        # loud failure — never attribute a bid from computed/uncached cells.
        if ws_raw is not None:
            raw_v = ws_raw.cell(row=r, column=cost_c).value
            if isinstance(raw_v, str) and raw_v.startswith("="):
                raise IntakeError(
                    f"bidder block cell "
                    f"{ws_raw.cell(row=r, column=cost_c).coordinate} contains "
                    f"a FORMULA ({raw_v[:40]!r}) — the contractor name/GSF "
                    f"block must be literal values; refusing to attribute "
                    f"this bid from computed cells."
                )
        v = ws.cell(row=r, column=cost_c).value
        if v is None:
            continue
        row_labels = " ".join(
            _norm(ws.cell(row=r, column=c).value)
            for c in range(1, cost_c)
        )
        if isinstance(v, (int, float)) and "gsf" in row_labels:
            total_gsf = int(v)
        elif isinstance(v, str) and v.strip():
            if contractor_name is None:
                contractor_name = v.strip()
            elif project_name is None:
                project_name = v.strip()
    if not contractor_name:
        raise IntakeError(
            "Falke-form layout recognized but no contractor name found in the "
            "bidder block above the header row — cannot attribute this bid."
        )

    divisions: list[DivisionBid] = []
    footer = BidFooter()
    quals: dict[str, list[str]] = {}
    alternates: list[LineItem] = []

    cur_code: Optional[str] = None
    cur_name: Optional[str] = None
    cur_items: list[LineItem] = []
    mode = "divisions"          # divisions → footer → alternates → quals
    qual_field: Optional[str] = None

    def _close_division(subtotal: Optional[Decimal]) -> None:
        nonlocal cur_code, cur_name, cur_items
        if cur_code is None:
            return
        divisions.append(DivisionBid(
            csi_code=cur_code,
            division_name=cur_name or cur_code,
            cost_structure=(CostStructure.ITEMIZED if cur_items
                            else CostStructure.LUMP_SUM),
            division_subtotal=subtotal,
            line_items=cur_items,
        ))
        cur_code, cur_name, cur_items = None, None, []

    def _line_item(desc: str, raw_cell) -> LineItem:
        amount, token = _coerce_amount(raw_cell)
        item = LineItem(description=desc)
        if amount is not None:
            item.amount = amount
            if amount == 0:
                item.is_explicit_zero = True
        elif token is not None:
            t = _norm(token)
            if token in _FORMULA_ERRORS:
                item.notes = f"Price cell carried {token}"
                warnings.append(
                    f"{desc!r}: spreadsheet error token {token} in the price "
                    f"cell (R18) — treated as missing pricing."
                )
            elif "excluded" in t or t == "nic":
                item.is_excluded = True
            elif "by owner" in t or "by others" in t:
                item.is_by_owner_others = True
                item.by_others_verbatim = str(token).strip()
            elif t == "allowance":
                item.is_allowance = True
            elif t in _MISSING_TOKENS:
                item.notes = f"Price cell carried {str(token).strip()!r}"
            else:
                item.notes = f"Price cell carried {str(token).strip()!r}"
                warnings.append(
                    f"{desc!r}: non-numeric price token "
                    f"{str(token).strip()!r} (R17) — treated as missing "
                    f"pricing unless classified."
                )
        if "allowance" in desc.lower():
            item.is_allowance = True
        return item

    for r in range(header_row + 1, ws.max_row + 1):
        key = ws.cell(row=r, column=csi_c).value
        desc = ws.cell(row=r, column=desc_c).value
        desc_s = str(desc).strip() if isinstance(desc, str) else ""
        dn = _norm(desc)

        # Division header (col CSI carries "DIV XX ...").
        m = _DIV_KEY_RE.match(str(key).strip()) if isinstance(key, str) else None
        if m:
            _close_division(None)
            cur_code = f"DIV {m.group(1)} 00 00"
            cur_name = desc_s or cur_code
            mode = "divisions"
            continue

        if not desc_s:
            # Text continuation rows inside a qualifications block still carry
            # content in the cost column.
            if mode == "quals" and qual_field:
                cont = ws.cell(row=r, column=cost_c).value
                if isinstance(cont, str) and cont.strip():
                    quals.setdefault(qual_field, []).append(cont.strip())
            continue

        # Footer / section labels (most specific first).
        if "construction cost subtotal" in dn:
            _close_division(None)
            footer.construction_cost_subtotal, _ = _coerce_amount(
                ws.cell(row=r, column=cost_c).value)
            mode = "footer"
            continue
        if "grand total" in dn:
            footer.grand_total, _ = _coerce_amount(
                ws.cell(row=r, column=cost_c).value)
            mode = "footer"
            continue
        if mode in ("footer", "alternates", "quals"):
            if dn in _QUAL_HEADINGS:
                mode, qual_field = "quals", _QUAL_HEADINGS[dn]
                first = ws.cell(row=r, column=cost_c).value
                if isinstance(first, str) and first.strip():
                    quals.setdefault(qual_field, []).append(first.strip())
                continue
            if "alternate" in dn:
                mode = "alternates"
                continue
            if mode == "quals" and qual_field:
                text_v = ws.cell(row=r, column=cost_c).value
                if isinstance(text_v, str) and text_v.strip():
                    quals.setdefault(qual_field, []).append(text_v.strip())
                continue
            if mode == "alternates":
                amount, _tok = _coerce_amount(ws.cell(row=r, column=cost_c).value)
                if dn == "bond":
                    footer.bond = amount
                else:
                    alt = LineItem(description=desc_s, amount=amount)
                    if amount is not None and amount == 0:
                        alt.is_explicit_zero = True
                    alternates.append(alt)
                continue
            # Fee rows in the footer block.
            if dn.startswith("general liab"):
                footer.general_liability_insurance, _ = _coerce_amount(
                    ws.cell(row=r, column=cost_c).value)
            elif dn.startswith("builders risk"):
                footer.builders_risk_insurance, _ = _coerce_amount(
                    ws.cell(row=r, column=cost_c).value)
            elif dn.startswith("gc fee"):
                footer.gc_fee, _ = _coerce_amount(
                    ws.cell(row=r, column=cost_c).value)
            elif dn.startswith("overhead"):
                footer.overhead_and_profit, _ = _coerce_amount(
                    ws.cell(row=r, column=cost_c).value)
            elif dn == "subtotal":
                pass  # fees subtotal — derived, not stored
            elif dn == "bond":
                footer.bond, _ = _coerce_amount(
                    ws.cell(row=r, column=cost_c).value)
            else:
                warnings.append(
                    f"Unrecognized footer row {desc_s!r} ignored (value not "
                    f"captured) — verify against the source file."
                )
            continue

        # Inside a division block.
        if cur_code is not None:
            if dn.endswith("subtotal"):
                subtotal, _tok = _coerce_amount(
                    ws.cell(row=r, column=csub_c).value)
                _close_division(subtotal)
                continue
            cur_items.append(
                _line_item(desc_s, ws.cell(row=r, column=cost_c).value))

    _close_division(None)
    footer.alternates = alternates

    # Deterministic footer confidence: HIGH only when the stated components
    # compose to the stated grand total within $1.
    comp = [footer.construction_cost_subtotal,
            footer.general_liability_insurance,
            footer.builders_risk_insurance, footer.gc_fee,
            footer.overhead_and_profit, footer.bond]
    if footer.grand_total is not None and footer.construction_cost_subtotal is not None:
        total = sum((c for c in comp if c is not None), Decimal("0"))
        if abs(total - footer.grand_total) <= Decimal("1"):
            footer.grand_total_confidence = GrandTotalConfidence.HIGH
        else:
            footer.grand_total_confidence = GrandTotalConfidence.MEDIUM
            footer.confidence_flags = ["ARITHMETIC_DISCREPANCY"]
    else:
        footer.grand_total_confidence = GrandTotalConfidence.LOW
        footer.confidence_flags = ["GRAND_TOTAL_OR_SUBTOTAL_MISSING"]

    q = BidQualifications(**{k: "\n".join(v) for k, v in quals.items()})
    return BidDocument(
        contractor_name=contractor_name,
        project_name=project_name,
        total_gsf=total_gsf,
        form_type=FormType.FALKE_STANDARD,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=footer,
        qualifications=q,
        extraction_confidence=ExtractionConfidence.HIGH,
        extraction_warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Layout 2 — §2 row schema (csv or xlsx)
# ---------------------------------------------------------------------------

# Category/Trade → canonical division aliases beyond direct name matching
# (covers the docx §2 examples; anything unmapped flows through as-is and the
# engine's unrecognized-code machinery flags it RED — loud, never guessed).
_CATEGORY_ALIASES = {
    "demo": "DIV 02 00 00", "demolition": "DIV 02 00 00",
    "millwork": "DIV 06 00 00", "carpentry": "DIV 06 00 00",
    "waterproofing": "DIV 07 00 00", "roofing": "DIV 07 00 00",
    "drywall": "DIV 09 00 00", "painting": "DIV 09 00 00",
    "flooring": "DIV 09 00 00", "tile": "DIV 09 00 00",
    "fire alarm": "DIV 28 00 00",
    "fire suppression": "DIV 21 00 00", "sprinkler": "DIV 21 00 00",
}


def _resolve_category(category: str, item_code: str,
                      warnings: list[str]) -> tuple[str, str]:
    """Resolve a Category/Trade to (csi_code, division_name).

    Resolution order: explicit DIV token in the category or item code →
    canonical division-name match → alias map → PASS THROUGH VERBATIM (the
    engine's unrecognized-code audit flags it RED; we never guess).
    """
    from src.canon import CANONICAL_DIVISIONS

    for source in (category, item_code):
        m = _DIV_KEY_RE.match(str(source).strip())
        if m:
            code = f"DIV {m.group(1)} 00 00"
            name = next((d["division_name"] for d in CANONICAL_DIVISIONS
                         if d["csi_code"] == code), category or code)
            return code, name

    cat = _norm(category)
    if cat:
        for d in CANONICAL_DIVISIONS:
            name = d["division_name"].lower()
            if cat == name or cat in name or name in cat:
                return d["csi_code"], d["division_name"]
        for alias, code in _CATEGORY_ALIASES.items():
            if alias in cat:
                name = next(d["division_name"] for d in CANONICAL_DIVISIONS
                            if d["csi_code"] == code)
                return code, name

    label = category.strip() if category else (item_code or "UNSPECIFIED")
    warnings.append(
        f"Category/Trade {label!r} has no canonical division mapping — passed "
        f"through verbatim; the engine will flag it for estimator review."
    )
    return label, label


def _parse_row_schema(path: Path, rows: list[list]) -> BidDocument:
    """Parse one bidder's §2 row-schema file (one price column = one bidder)."""
    found = _find_row_schema_header(rows)
    if found is None:
        raise IntakeError(
            "no recognizable header row (need at least 'Item Code' and "
            "'Description' columns for the row-schema layout)."
        )
    hdr_idx, fields = found
    contractor_name = fields["_bidder"]
    price_c = fields["_price_col"]
    warnings: list[str] = []

    def _get(row: list, field: str) -> str:
        c = fields.get(field)
        if c is None or c >= len(row) or row[c] is None:
            return ""
        return str(row[c]).strip()

    # Ordered division buckets: csi_code → (name, [items])
    buckets: dict[str, tuple[str, list[LineItem]]] = {}
    alternates: list[LineItem] = []

    for row in rows[hdr_idx + 1:]:
        if not any(str(c).strip() for c in row if c is not None):
            continue
        item_code = _get(row, "item_code")
        desc = _get(row, "description")
        category = _get(row, "category")
        # Skip stray total rows (no item code + a total-like description).
        if not item_code and re.search(r"\btotal\b", desc, re.IGNORECASE):
            warnings.append(
                f"Row {desc!r} looks like a total row and was skipped "
                f"(totals are recomputed, never ingested)."
            )
            continue
        raw_price = row[price_c] if price_c < len(row) else None
        if not item_code and not desc and raw_price is None:
            continue
        if not item_code or not desc:
            warnings.append(
                f"Incomplete line (R24): item_code={item_code!r} "
                f"description={desc!r} — ingested with what is present; "
                f"flag for clarification."
            )

        amount, token = _coerce_amount(raw_price)
        item = LineItem(description=desc or item_code or "UNSPECIFIED")
        if amount is not None:
            item.amount = amount
            if amount == 0:
                item.is_explicit_zero = True
        elif token is not None:
            if token in _FORMULA_ERRORS:
                warnings.append(
                    f"{item.description!r}: spreadsheet error token {token} "
                    f"(R18) — treated as missing pricing."
                )
            item.notes = f"Price cell carried {str(token).strip()!r}"

        notes = _get(row, "notes")
        if notes:
            item.notes = ((item.notes + " | ") if item.notes else "") + notes
        qty, unit = _get(row, "quantity"), _get(row, "unit")
        if qty or unit:
            item.notes = ((item.notes + " | ") if item.notes else "") + \
                f"Qty: {qty or '?'} Unit: {unit or '?'}"
        where = f"{item_code or '?'} {desc or '?'}"
        verdict = _apply_status(item, _get(row, "status"), warnings, where)
        if "allowance" in desc.lower():
            item.is_allowance = True
        if item.is_allowance and notes and not item.allowance_basis:
            item.allowance_basis = notes
        if item.is_allowance and item.amount is None:
            warnings.append(
                f"{where}: allowance without an amount (R26) — treated as "
                f"missing pricing."
            )
        if verdict == "alternate":
            alternates.append(item)
            continue

        code, div_name = _resolve_category(category, item_code, warnings)
        buckets.setdefault(code, (div_name, []))[1].append(item)

    if not buckets and not alternates:
        raise IntakeError("no data rows found below the header row.")

    divisions: list[DivisionBid] = []
    grand = Decimal("0")
    for code, (div_name, items) in buckets.items():
        subtotal = sum(
            (i.amount for i in items
             if i.amount is not None
             and not i.is_by_owner_others and not i.is_excluded),
            Decimal("0"),
        )
        grand += subtotal
        divisions.append(DivisionBid(
            csi_code=code,
            division_name=div_name,
            cost_structure=CostStructure.ITEMIZED,
            division_subtotal=subtotal,
            line_items=items,
        ))

    warnings.append(
        "Row-schema file carries no footer: construction subtotal and grand "
        "total are DERIVED by summation of the classified priced items "
        "(Falke §15 'Total Bid Amount'); no fee/insurance rows stated."
    )
    footer = BidFooter(
        construction_cost_subtotal=grand,
        grand_total=grand,
        alternates=alternates,
        grand_total_confidence="LOW",
        confidence_flags=["GC_FEE_MISSING", "INSURANCE_NOT_STATED"],
    )
    return BidDocument(
        contractor_name=contractor_name,
        form_type=FormType.CONTRACTOR_OWN,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=footer,
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
        extraction_warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def parse_structured_bid(path: str | Path) -> BidDocument:
    """Parse one structured bid file (.xlsx or .csv) into a BidDocument.

    Deterministic — no vision, no model calls. Raises IntakeError with a
    clear, file-named message for anything outside the two supported layouts.
    """
    path = Path(path)
    try:
        # Floyd C1: size ceiling before any parser touches the file.
        size = path.stat().st_size
        if size > MAX_STRUCTURED_FILE_BYTES:
            raise IntakeError(
                f"file is {size / (1024 * 1024):.1f} MB — exceeds the "
                f"{MAX_STRUCTURED_FILE_BYTES // (1024 * 1024)} MB "
                f"structured-intake ceiling. Verify the file; a bid form "
                f"should be well under this."
            )

        if path.suffix.lower() == ".csv":
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = [row for row in csv.reader(f)]
            return _parse_row_schema(path, rows)

        if path.suffix.lower() == ".xlsx":
            # Floyd C4: reject a decompression bomb BEFORE openpyxl parses —
            # the on-disk ceiling above sees only compressed bytes.
            with zipfile.ZipFile(path) as zf:
                decompressed = sum(i.file_size for i in zf.infolist())
            if decompressed > MAX_DECOMPRESSED_XLSX_BYTES:
                raise IntakeError(
                    f"xlsx decompresses to "
                    f"{decompressed / (1024 * 1024):.0f} MB — exceeds the "
                    f"{MAX_DECOMPRESSED_XLSX_BYTES // (1024 * 1024)} MB "
                    f"decompressed ceiling (zip-bomb guard). A real bid "
                    f"form's sheet XML is a few MB; verify or re-export the "
                    f"file."
                )
            wb = openpyxl.load_workbook(path, data_only=True)
            # Second load with formulas visible — Floyd C3 (bidder-block
            # formula detection only; values still come from the cached wb).
            wb_raw = openpyxl.load_workbook(path, data_only=False)
            for ws in wb.worksheets:
                falke = _find_falke_header(ws)
                if falke is not None:
                    ws_raw = (wb_raw[ws.title]
                              if ws.title in wb_raw.sheetnames else None)
                    return _parse_falke_form(
                        path, ws, falke[0], falke[1], ws_raw=ws_raw)
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if _find_row_schema_header(rows) is not None:
                    return _parse_row_schema(path, rows)
            raise IntakeError(
                "unrecognized spreadsheet layout — expected either the Falke "
                "standard bid form (header row with 'CSI' / 'BUILDING SYSTEM' "
                "/ 'COST' / 'COST SUBTOTALS') or the Falke program §2 row "
                "schema ('Item Code' / 'Description' / one bidder price "
                "column). Convert the file to one of these layouts or submit "
                "the bid as a PDF for agent extraction."
            )

        raise IntakeError(
            f"unsupported structured-file extension {path.suffix!r} "
            f"(supported: {', '.join(STRUCTURED_SUFFIXES)})."
        )
    except IntakeError as e:
        raise IntakeError(f"{path.name}: {e}") from None


def run_structured_intake(interim_dir: str | Path) -> tuple[
    list[tuple[str, str]], list[tuple[str, str]]
]:
    """Parse every structured bid file in the interim dir to BidDocument JSON.

    Scans ``interim_dir`` for *.xlsx / *.csv (skipping Excel lock files),
    writes ``{slug}.json`` next to them, and returns (successes, failures) as
    (filename, slug.json | error-message) pairs. Failures are LOUD and
    per-file; they never abort the other files (Floyd C1 — a malformed zip,
    a non-UTF-8 csv, or ANY per-file exception becomes a reported failure,
    not a batch crash).

    Slug-collision protection (Floyd C2 — silent bidder loss is the failure
    class the whole tie-out apparatus exists to prevent):

    * two files in THIS run resolving to the same slug → the second fails
      LOUDLY, naming both source files; never overwritten.
    * a pre-existing ``{slug}.json`` (prior run / extraction agent) is only
      REFRESHED when it carries the SAME contractor_name (the normal
      re-run-the-pipeline flow, e.g. after an SF-gate exit 2); a different or
      unreadable contractor in the existing file fails LOUDLY instead.
    """
    interim_dir = Path(interim_dir)
    files = sorted(
        p for p in interim_dir.iterdir()
        if p.suffix.lower() in STRUCTURED_SUFFIXES
        and not p.name.startswith("~$")
    )
    successes: list[tuple[str, str]] = []
    failures: list[tuple[str, str]] = []
    slug_sources: dict[str, str] = {}  # slug → source filename (this run)
    for f in files:
        try:
            doc = parse_structured_bid(f)

            slug = _slug(doc.contractor_name)
            json_path = interim_dir / f"{slug}.json"
            if slug in slug_sources:
                raise IntakeError(
                    f"{f.name}: bidder {doc.contractor_name!r} resolves to "
                    f"the same output name ({slug}.json) as "
                    f"{slug_sources[slug]!r} from this intake run — writing "
                    f"it would silently drop a bidder. Rename one file's "
                    f"bidder or submit it as a PDF."
                )
            if json_path.exists():
                try:
                    existing = json.loads(json_path.read_text(encoding="utf-8"))
                    existing_name = existing.get("contractor_name")
                except Exception:
                    existing_name = None
                if existing_name != doc.contractor_name:
                    raise IntakeError(
                        f"{f.name}: output {slug}.json already exists for a "
                        f"DIFFERENT bidder ({existing_name!r} vs "
                        f"{doc.contractor_name!r}) — overwriting would "
                        f"silently drop a bidder. Clear the interim dir or "
                        f"rename one bidder's file."
                    )
                # Same contractor → refresh (normal pipeline re-run flow).
            json_path.write_text(doc.model_dump_json(indent=2),
                                 encoding="utf-8")
            slug_sources[slug] = f.name
            successes.append((f.name, json_path.name))
        except IntakeError as e:
            failures.append((f.name, str(e)))
        except Exception as e:  # Floyd C1: per-file, never abort the batch
            # Floyd B4 (advisory): an unexpected exception here may be an
            # ENGINE defect, not a bad client file — log the full traceback
            # to stderr so a code regression is never silently misdiagnosed
            # as dirty input.
            print(f"[intake_structured] unexpected exception parsing "
                  f"{f.name} — full traceback follows (possible ENGINE "
                  f"defect, not necessarily a bad file):", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            failures.append((
                f.name,
                f"{f.name}: unreadable/unparseable file "
                f"({type(e).__name__}: {e}) — fix or re-export the file, or "
                f"submit the bid as a PDF for agent extraction. If the file "
                f"opens cleanly in Excel, this may be an engine defect — a "
                f"full traceback was logged to stderr.",
            ))
    return successes, failures
