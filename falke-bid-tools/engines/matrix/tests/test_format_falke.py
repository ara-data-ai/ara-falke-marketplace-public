"""
FALKE Matrix — Falke house-format tests (Leveled_Normalized, v0.3.0)
=====================================================================
Verifies the house style extracted from Falke's manual matrix (see
FALKE/03_Matrix/FALKE-HOUSE-FORMAT-SPEC.md) is applied to the leveled sheet
at write time in the FEB 26 four-column geometry:

  * Avenir Book font, aqua A3EAF3 header/subtotal bands, gray A6A6A6 / white
    bidder block, teal 00A9CA totals with merged pairs; GT amount normalized
    to Avenir Book 12 bold white underlined (Derick's decision).
  * The leveled banner is the neutral house GRAY (rules-spec A1), merged
    full-width, text unchanged.
  * Bid_Form (mirror) is NOT house-formatted and keeps its legacy audit fills.
  * NO legacy ARA audit fill appears on the leveled sheet (rules-spec §4.4) —
    the Falke vocabulary (falke_rules hues) is the only paint there.
  * Written leveled values tie out (Stage 6b clean) — format is value-safe.

Run from the engine root:
    python3 -m pytest tests/test_format_falke.py -v
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from src.format_falke import (
    FALKE_ACCENT_RED,
    FALKE_AQUA,
    FALKE_FONT_NAME,
    FALKE_GRAY,
    FALKE_RAIL_GRAY,
    FALKE_TEAL,
)
from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)
from src.normalize import compute_cross_bid_stats, normalize_bid
from src.audit import audit_bids
from src.reconcile import reconcile_written_matrix
from src.run_config import RunInputs
from src.write_matrix import (
    LEVELED_CSUB_OFFSET,
    _lev_col_start,
    _lev_last_col,
    write_matrix,
)

# Legacy ARA audit palette (must never appear on the leveled sheet).
_LEGACY_AUDIT_HEXES = {"FFCCCC", "FFF2CC", "CCFFCC"}


def _doc(name: str, amount: Decimal, items=None,
         stated_subtotal=None) -> BidDocument:
    return BidDocument(
        contractor_name=name,
        form_type=FormType.FALKE_STANDARD,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=[
            DivisionBid(
                csi_code="DIV 01 00 00",
                division_name="General Requirements",
                cost_structure=CostStructure.ITEMIZED,
                division_subtotal=stated_subtotal or amount,
                classification_source=ClassificationSource.CONTRACTOR_NATIVE,
                contractor_native_code=None,
                line_items=items or [
                    LineItem(description="Project Management", amount=amount)
                ],
            )
        ],
        footer=BidFooter(
            construction_cost_subtotal=stated_subtotal or amount,
            gc_fee=None,
            grand_total=stated_subtotal or amount,
            alternates=[],
            grand_total_confidence=GrandTotalConfidence.LOW,
        ),
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _run_inputs() -> RunInputs:
    return RunInputs(
        project_name="Test Project",
        project_address="1 Test St",
        gross_sf=10_000.0,
        sf_basis_label="GSF",
        sf_source="explicit",
    )


def _write(tmpdir: str, docs=None, with_audit=False):
    docs = docs or [
        _doc("Alpha Builders", Decimal("100000")),
        _doc("Beta Construction", Decimal("120000")),
    ]
    bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
    items = audit_bids(bids) if with_audit else None
    out = Path(tmpdir) / "m.xlsx"
    write_matrix(bids, out, _run_inputs(), audit_items=items)
    return out, bids, items


def _hex(cell) -> str:
    rgb = cell.fill.fgColor.rgb
    return rgb[-6:].upper() if isinstance(rgb, str) else ""


def _find_row_by_col(ws, col: int, value: str) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    raise AssertionError(f"row with col{col}=={value!r} not found")


class TestFalkeHouseFormat:
    def test_leveled_sheet_carries_house_style(self):
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]

            # Column-header row 4: aqua band, Avenir Book bold, FEB 26 titles.
            hdr = ws.cell(row=4, column=1)  # "CSI"
            assert hdr.value == "CSI"
            assert _hex(hdr) == FALKE_AQUA
            assert hdr.font.name == FALKE_FONT_NAME
            assert hdr.font.bold is True
            first_cost = ws.cell(row=4, column=_lev_col_start(0))
            assert first_cost.value == "COST"
            assert ws.cell(
                row=4, column=_lev_col_start(0) + LEVELED_CSUB_OFFSET
            ).value == "COST \nSUBTOTALS"

            # Bidder block row 5: gray fill, white bold Avenir Book, merged.
            name_cell = ws.cell(row=5, column=_lev_col_start(0))
            assert _hex(name_cell) == FALKE_GRAY
            assert name_cell.font.name == FALKE_FONT_NAME
            assert name_cell.font.bold is True
            assert (name_cell.font.color.rgb or "")[-6:].upper() == "FFFFFF"

            # Division subtotal row: aqua on label + COST SUBTOTALS cell.
            sub_row = _find_row_by_col(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            assert _hex(ws.cell(row=sub_row, column=2)) == FALKE_AQUA
            csub = ws.cell(
                row=sub_row, column=_lev_col_start(0) + LEVELED_CSUB_OFFSET
            )
            assert _hex(csub) == FALKE_AQUA
            # R2: the subtotal amount lives in COST SUBTOTALS, not COST.
            assert isinstance(csub.value, (int, float))
            assert ws.cell(row=sub_row, column=_lev_col_start(0)).value is None

            # Teal total rows (col-B display labels; col A blank).
            gt_row = _find_row_by_col(ws, 2, "GRAND TOTAL CONSTRUCTION COST")
            ccs_row = _find_row_by_col(ws, 2, "CONSTRUCTION COST SUBTOTAL")
            assert ws.cell(row=gt_row, column=1).value is None
            assert ws.cell(row=ccs_row, column=1).value is None
            gt = ws.cell(row=gt_row, column=_lev_col_start(0))
            assert _hex(gt) == FALKE_TEAL
            assert gt.font.name == FALKE_FONT_NAME
            assert gt.font.bold is True
            assert gt.font.underline == "single"  # Derick's GT normalization
            assert (gt.font.color.rgb or "")[-6:].upper() == "FFFFFF"
            # Merged COST:COST SUBTOTALS pair.
            merges = {str(r) for r in ws.merged_cells.ranges}
            from openpyxl.utils import get_column_letter
            a = get_column_letter(_lev_col_start(0))
            b = get_column_letter(_lev_col_start(0) + LEVELED_CSUB_OFFSET)
            assert f"{a}{gt_row}:{b}{gt_row}" in merges

            # Body font is Avenir Book.
            body = ws.cell(row=sub_row - 1, column=2)
            assert body.font.name == FALKE_FONT_NAME

    def test_leveled_banner_is_neutral_gray(self):
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            banner = ws.cell(row=1, column=1)
            assert "does NOT match the submitted bids" in str(banner.value)
            assert _hex(banner) == FALKE_GRAY  # A1: yellow banner superseded
            # Disclaimer (R34) on row 3.
            assert "DISCLAIMER:" in str(ws.cell(row=3, column=1).value)

    def test_bid_form_sheet_is_not_formatted(self):
        # The house format is the leveled sheet's; the mirror keeps its
        # engineering rendering (machine keys, plain fonts, audit fills).
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            hdr = ws.cell(row=4, column=1)
            assert hdr.font.name != FALKE_FONT_NAME
            assert _hex(hdr) != FALKE_AQUA
            # Mirror footer keeps the col-A machine keys.
            assert _find_row_by_col(ws, 1, "GRAND_TOTAL")

    def test_no_legacy_audit_fill_on_leveled_sheet(self):
        # A stated subtotal that contradicts the line-item sum is an
        # ARITHMETIC_DISCREPANCY on the AUDIT register — and, under v0.3.0+,
        # the same defect paints the FALKE red (R20) on the leveled sheet.
        # W-D M-3/B4: the MIRROR carries NO ARA audit fills at all any more
        # (the mirror asserts nothing — every ARA diagnostic lives on AUDIT).
        docs = [
            _doc("Alpha Builders", Decimal("100000"),
                 items=[LineItem(description="Project Management",
                                 amount=Decimal("50000"))],
                 stated_subtotal=Decimal("100000")),  # items ≠ stated
            _doc("Beta Construction", Decimal("120000")),
        ]
        with tempfile.TemporaryDirectory() as d:
            out, _bids, items = _write(d, docs=docs, with_audit=True)
            assert items, "fixture must generate audit items"
            assert any(i.code.value == "ARITHMETIC_DISCREPANCY" for i in items)
            wb = openpyxl.load_workbook(out)

            # Mirror: NO ARA audit fill on any DATA cell (M-3/B4 — fills
            # removed). The sweep stops at the GRAND_TOTAL footer row: the
            # workbook LEGEND below it legitimately carries soft swatches.
            wsm = wb["Bid_Form"]
            m_gt = _find_row_by_col(wsm, 1, "GRAND_TOTAL")
            mirror_offenders = [
                c.coordinate
                for row in wsm.iter_rows(min_row=1, max_row=m_gt)
                for c in row
                if _hex(c) in _LEGACY_AUDIT_HEXES
            ]
            assert mirror_offenders == [], (
                f"ARA audit fills must be OFF the mirror (M-3/B4): "
                f"{mirror_offenders}"
            )

            # Leveled: NO legacy ARA audit hue on the data region (through
            # the GRAND TOTAL row — legend swatches below are documentation);
            # the R20 falke red covers the same defect.
            wsl = wb["Leveled_Normalized"]
            l_gt = _find_row_by_col(wsl, 2, "GRAND TOTAL CONSTRUCTION COST")
            offenders = [
                c.coordinate
                for row in wsl.iter_rows(min_row=1, max_row=l_gt)
                for c in row
                if _hex(c) in _LEGACY_AUDIT_HEXES
            ]
            assert offenders == [], (
                f"legacy ARA audit fills leaked onto Leveled_Normalized: "
                f"{offenders}"
            )
            sub_row = _find_row_by_col(wsl, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            alpha_col = next(
                _lev_col_start(i) for i in range(2)
                if wsl.cell(row=5, column=_lev_col_start(i)).value
                == "Alpha Builders"
            )
            assert _hex(wsl.cell(
                row=sub_row, column=alpha_col + LEVELED_CSUB_OFFSET
            )) == "FF0000", "R20 must paint the falke red on the leveled sheet"

    def test_leveled_values_tie_out_clean(self):
        # Value safety: the formatted, rules-painted leveled sheet still
        # passes the full Stage-6b tie-out (values written where the
        # reconciler reads them, per the shared geometry constants).
        with tempfile.TemporaryDirectory() as d:
            out, bids, _ = _write(d)
            failures = reconcile_written_matrix(out, bids, 0)
            assert failures == [], [f.message for f in failures]
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            gt_row = _find_row_by_col(ws, 2, "GRAND TOTAL CONSTRUCTION COST")
            by_name = {b.contractor_name: b for b in bids}
            for i in range(len(bids)):
                name = ws.cell(row=5, column=_lev_col_start(i)).value
                written = ws.cell(row=gt_row, column=_lev_col_start(i)).value
                assert written == float(by_name[name].footer.grand_total.amount)


def _rail_hex(side) -> str:
    rgb = side.color.rgb if side.color is not None else None
    return rgb.upper() if isinstance(rgb, str) else ""


class TestFalkeBorderRails:
    """v0.3.1 (house-format addendum §B): thin gray #7F7F7F left/right rails
    on EVERY cell of the table region — header row through GRAND TOTAL row,
    all columns, including BLANK cells."""

    def test_rails_on_every_cell_including_blanks(self):
        with tempfile.TemporaryDirectory() as d:
            out, bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            gt_row = _find_row_by_col(ws, 2, "GRAND TOTAL CONSTRUCTION COST")
            last = _lev_last_col(len(bids))

            # Merged ranges render borders on their OUTLINE only (openpyxl
            # propagates the anchor's l/r to the merge edges on save) — so
            # inside a merge, require left rail on the leftmost member and
            # right rail on the rightmost; interior edges never render.
            merge_of = {}
            for rng in ws.merged_cells.ranges:
                for row_cells in ws[rng.coord]:
                    for cell in row_cells:
                        merge_of[cell.coordinate] = rng

            def _has(side):
                return (side is not None and side.style == "thin"
                        and _rail_hex(side) == FALKE_RAIL_GRAY)

            missing = []
            for r in range(4, gt_row + 1):
                for c in range(1, last + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border
                    rng = merge_of.get(cell.coordinate)
                    if rng is None:
                        ok = _has(b.left) and _has(b.right)
                    else:
                        ok = True
                        if c == rng.min_col:
                            ok = ok and _has(b.left)
                        if c == rng.max_col:
                            ok = ok and _has(b.right)
                    if not ok:
                        missing.append(cell.coordinate)
            assert missing == [], (
                f"cells without gray l/r rails in table region: {missing[:20]}"
            )

    def test_blank_cost_cell_carries_rails(self):
        # The subtotal row's COST cell is deliberately BLANK (R2: amounts in
        # COST SUBTOTALS) — it must still carry the left/right rails.
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            sub_row = _find_row_by_col(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            blank = ws.cell(row=sub_row, column=_lev_col_start(0))
            assert blank.value is None
            assert blank.border.left.style == "thin"
            assert _rail_hex(blank.border.left) == FALKE_RAIL_GRAY
            assert blank.border.right.style == "thin"
            assert _rail_hex(blank.border.right) == FALKE_RAIL_GRAY
            # Spacer row below the subtotal: fully blank, still railed.
            spacer = ws.cell(row=sub_row + 1, column=_lev_col_start(0))
            assert spacer.value is None
            assert spacer.border.left.style == "thin"
            assert spacer.border.right.style == "thin"

    def test_top_bottom_edges_preserved(self):
        # The rail pass sets ONLY left/right; the accounting top-thin /
        # bottom-double on subtotal cells and the header box horizontals
        # survive the pass (addendum §B).
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            sub_row = _find_row_by_col(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            csub = ws.cell(row=sub_row,
                           column=_lev_col_start(0) + LEVELED_CSUB_OFFSET)
            assert csub.border.top.style == "thin"
            assert csub.border.bottom.style == "double"
            hdr = ws.cell(row=4, column=1)
            assert hdr.border.top.style == "thin"
            assert hdr.border.bottom.style == "thin"

    def test_subtotal_accents_are_falke_dark_red(self):
        # APPROVED (Derick 2026-07-05): subtotal accounting borders carry the
        # reference's dark red #D74648 (the CCS row's literal #D74547 is
        # normalized to the dominant hue — addendum §B ruling). Header and
        # GRAND TOTAL box horizontals stay BLACK (not covered by the
        # approval).
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]

            # Red accents: division subtotal + Fees Subtotal (COST SUBTOTALS
            # cell) and the CCS row (col-B label cell — its amount pair is
            # merged, the label carries the same SUBTOTAL_BORDER unmerged).
            div_row = _find_row_by_col(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            fees_row = _find_row_by_col(ws, 2, "Fees Subtotal")
            ccs_row = _find_row_by_col(ws, 2, "CONSTRUCTION COST SUBTOTAL")
            csub_col = _lev_col_start(0) + LEVELED_CSUB_OFFSET
            targets = [ws.cell(row=div_row, column=csub_col),
                       ws.cell(row=fees_row, column=csub_col),
                       ws.cell(row=ccs_row, column=2)]
            for c in targets:
                assert c.border.top.style == "thin", c.coordinate
                assert _rail_hex(c.border.top) == FALKE_ACCENT_RED, (
                    c.coordinate)
                assert c.border.bottom.style == "double", c.coordinate
                assert _rail_hex(c.border.bottom) == FALKE_ACCENT_RED, (
                    c.coordinate)

            # Header + GRAND TOTAL horizontals: still black (default color).
            gt_row = _find_row_by_col(ws, 2, "GRAND TOTAL CONSTRUCTION COST")
            for cell in (ws.cell(row=4, column=1),
                         ws.cell(row=gt_row, column=2)):
                assert cell.border.top.style == "thin"
                assert _rail_hex(cell.border.top) != FALKE_ACCENT_RED
                assert cell.border.bottom.style == "thin"
                assert _rail_hex(cell.border.bottom) != FALKE_ACCENT_RED

    def test_rails_stop_at_region_bounds(self):
        # Banner rows above and the notes/legend region below the GRAND
        # TOTAL row stay rail-free (reference stops rails at its GT row).
        with tempfile.TemporaryDirectory() as d:
            out, _bids, _ = _write(d)
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            gt_row = _find_row_by_col(ws, 2, "GRAND TOTAL CONSTRUCTION COST")
            banner = ws.cell(row=1, column=5)
            assert banner.border.left.style is None
            assert banner.border.right.style is None
            below = ws.cell(row=gt_row + 1, column=_lev_col_start(0))
            assert below.border.left.style is None
            assert below.border.right.style is None


class TestMirrorTokenRendering:
    """M-3 (W-D): the mirror renders what the bidder wrote — blank cells stay
    truly blank (never 0.00), EXCLUDED renders 'Excluded' italic, BY_OWNER
    renders the verbatim token italic, EXPLICIT_ZERO stays a real numeric 0,
    and no token/blank rendering ever trips Stage 6b (reconcile lockstep)."""

    @staticmethod
    def _token_docs():
        excluder = BidDocument(
            contractor_name="Excluder Co",
            form_type=FormType.FALKE_STANDARD,
            bid_document_input_type=InputType.DIGITAL_NATIVE,
            divisions=[
                DivisionBid(
                    csi_code="DIV 09 00 00", division_name="Finishes",
                    cost_structure=CostStructure.ITEMIZED,
                    division_subtotal=Decimal("0"),
                    line_items=[
                        LineItem(description="Corridor painting",
                                 is_excluded=True),
                        LineItem(description="Lobby flooring",
                                 is_excluded=True),
                    ],
                ),
                DivisionBid(
                    csi_code="DIV 02 00 00",
                    division_name="Existing Conditions",
                    cost_structure=CostStructure.ITEMIZED,
                    division_subtotal=None,
                    line_items=[
                        LineItem(description="Site demolition and clearing",
                                 is_by_owner_others=True,
                                 by_others_verbatim="Not Applicable"),
                    ],
                ),
                DivisionBid(
                    csi_code="DIV 03 00 00", division_name="Concrete",
                    cost_structure=CostStructure.ITEMIZED,
                    division_subtotal=Decimal("100000"),
                    line_items=[LineItem(description="Concrete repairs",
                                         amount=Decimal("100000"))],
                ),
            ],
            footer=BidFooter(
                construction_cost_subtotal=Decimal("100000"),
                grand_total=Decimal("100000"),
                grand_total_confidence=GrandTotalConfidence.LOW,
            ),
            qualifications=BidQualifications(),
            extraction_confidence=ExtractionConfidence.HIGH,
        )
        peer = _doc("Peer Builders", Decimal("120000"))
        return [excluder, peer]

    def _mirror(self, d):
        docs = self._token_docs()
        bids = compute_cross_bid_stats([normalize_bid(x) for x in docs])
        items = audit_bids(bids)
        out = Path(d) / "m.xlsx"
        write_matrix(bids, out, _run_inputs(), audit_items=items)
        wb = openpyxl.load_workbook(out)
        ws = wb["Bid_Form"]
        # Excluder's column.
        from src.write_matrix import _col_start
        col = next(_col_start(i) for i in range(2)
                   if ws.cell(row=5, column=_col_start(i)).value == "Excluder Co")
        return out, bids, items, ws, col

    def test_tokens_blanks_and_zero_render_per_state(self):
        with tempfile.TemporaryDirectory() as d:
            _out, _bids, _items, ws, col = self._mirror(d)

            # Subtotal cells: EXPLICIT_ZERO (stated $0) renders numeric 0.00.
            div09_sub = _find_row_by_col(ws, 2, "FINISHES SUBTOTAL")
            c = ws.cell(row=div09_sub, column=col)
            assert c.value == 0.0, "the bidder DID write $0 — numeric 0.00"

            # Excluded LINE cells render the italic 'Excluded' token.
            line_row = _find_row_by_col(ws, 2, "Corridor painting")
            lc = ws.cell(row=line_row, column=col)
            assert lc.value == "Excluded"
            assert lc.font.italic is True
            assert _hex(lc) not in _LEGACY_AUDIT_HEXES

            # BY_OWNER line renders the VERBATIM token, italic, no fill.
            bo_row = _find_row_by_col(ws, 2, "Site demolition and clearing")
            bc = ws.cell(row=bo_row, column=col)
            assert bc.value == "Not Applicable"
            assert bc.font.italic is True

            # BY_OWNER division's blank subtotal stays truly BLANK (NULL_BLANK
            # per the M-3 state table — the line carries the token).
            div02_sub = _find_row_by_col(ws, 2,
                                         "EXISTING CONDITIONS SUBTOTAL")
            assert ws.cell(row=div02_sub, column=col).value is None

            # A division absent from the bid: subtotal + $/SF truly blank —
            # never a fabricated 0.00 (Peer never bid DIV 07).
            div07_sub = _find_row_by_col(
                ws, 2, "THERMAL & MOISTURE PROTECTION SUBTOTAL")
            assert ws.cell(row=div07_sub, column=col).value is None
            assert ws.cell(row=div07_sub, column=col + 1).value is None

    def test_reconcile_lockstep_no_false_quarantine(self):
        """The token/blank mirror still ties out — Stage 6b zero failures."""
        with tempfile.TemporaryDirectory() as d:
            out, bids, items, _ws, _col = self._mirror(d)
            failures = reconcile_written_matrix(out, bids, len(items))
            assert failures == [], [f.message for f in failures]


class TestAuditSheetInformationArchitecture:
    """W-D B4/§2 + S2-3: the AUDIT key sits at the TOP of the tab, the View
    column speaks the actual sheet names, and both data sheets carry the ONE
    workbook legend."""

    def test_key_at_top_view_sheet_names_and_legend_both_sheets(self):
        from src.write_matrix import (
            AUDIT_HEADER_ROW,
            LEGEND_HEADER,
            LEGEND_PRECEDENCE,
            find_audit_header_row,
        )
        with tempfile.TemporaryDirectory() as d:
            out, _bids, items = _write(d, with_audit=True)
            assert items
            wb = openpyxl.load_workbook(out)
            wsa = wb["AUDIT"]

            # Key block ABOVE the header row.
            header_row = find_audit_header_row(wsa)
            assert header_row == AUDIT_HEADER_ROW
            top_text = " ".join(
                str(wsa.cell(row=r, column=1).value or "")
                for r in range(1, header_row)
            )
            assert "Total items audited:" in top_text
            assert "RED Critical:" in top_text
            assert "Board members: the decision view is Leveled_Normalized" \
                in top_text

            # View column: only sheet names / Both — never "As-Submitted",
            # never bare "Leveled".
            views = set()
            r = header_row + 1
            while wsa.cell(row=r, column=1).value not in (None, ""):
                views.add(wsa.cell(row=r, column=2).value)
                r += 1
            assert views <= {"Both", "Bid_Form", "Leveled_Normalized"}, views

            # ONE legend block, identical header + precedence, on BOTH data
            # sheets.
            for sheet in ("Bid_Form", "Leveled_Normalized"):
                ws = wb[sheet]
                _find_row_by_col(ws, 2, LEGEND_HEADER)
                _find_row_by_col(ws, 2, LEGEND_PRECEDENCE)

                # Marvin amendment (W-D, 2026-07-15): the ARA severities
                # render THREE legend rows — one per severity, each with its
                # OWN swatch in the single swatch column (col B), text in
                # col C, each ending "Appears only on the AUDIT tab."
                for text, hue in (
                    ("AUDIT rows filled soft red — ARA severity RED: "
                     "resolve before award. Appears only on the AUDIT tab.",
                     "FFCCCC"),
                    ("AUDIT rows filled soft yellow — ARA severity YELLOW: "
                     "review. Appears only on the AUDIT tab.", "FFF2CC"),
                    ("AUDIT rows filled soft green — ARA severity GREEN: "
                     "verified. Appears only on the AUDIT tab.", "CCFFCC"),
                ):
                    r = _find_row_by_col(ws, 3, text)
                    assert _hex(ws.cell(row=r, column=2)) == hue, (
                        f"[{sheet}] severity row must carry its OWN swatch "
                        f"{hue}, got {_hex(ws.cell(row=r, column=2))}"
                    )
