"""
FALKE Matrix — Falke leveling-rules tests (Leveled_Normalized, v0.3.0)
=======================================================================
Proves the encoded Falke program rules fire correctly on the written sheet:

  * A6 (Marvin's trap): a BLANK division never enters the benchmark median
    as 0.0 — benchmarks compute from classified cell STATE only.
  * R12/R13 with the INCLUSIVE ±20% boundary (Q1 decided default).
  * R16 red-first precedence (an R20 math error beats a yellow variance).
  * Q5/RISK-2: no cyan/yellow paint below 3 valid bids (benchmark still shows).
  * R6: an unapproved explicit zero paints RED.
  * R5: a blank division paints RED where peers priced it.

Run from the engine root:
    python3 -m pytest tests/test_falke_rules.py -v
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)
from src.normalize import compute_cross_bid_stats, normalize_bid
from src.run_config import RunInputs
from src.write_matrix import (
    LEVELED_CSUB_OFFSET,
    _lev_bench_col,
    _lev_col_start,
    write_matrix,
)

CYAN, YELLOW, RED = "00FFFF", "FFFF00", "FF0000"


def _div(code, name, subtotal=None, items=None,
         cost=CostStructure.LUMP_SUM):
    return DivisionBid(
        csi_code=code,
        division_name=name,
        cost_structure=cost,
        division_subtotal=subtotal,
        classification_source=ClassificationSource.CONTRACTOR_NATIVE,
        contractor_native_code=None,
        line_items=items or [],
    )


def _doc(name, divisions, grand_total):
    return BidDocument(
        contractor_name=name,
        form_type=FormType.FALKE_STANDARD,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=BidFooter(
            construction_cost_subtotal=grand_total,
            grand_total=grand_total,
            grand_total_confidence=GrandTotalConfidence.LOW,
        ),
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _concrete_bidder(name, amount):
    return _doc(name, [_div("DIV 03 00 00", "Concrete", Decimal(amount))],
                Decimal(amount))


def _run_inputs():
    return RunInputs(
        project_name="Rules Test",
        project_address="1 Test St",
        gross_sf=10_000.0,
        sf_basis_label="GSF",
        sf_source="explicit",
    )


def _write(tmp, docs):
    bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
    out = Path(tmp) / "m.xlsx"
    write_matrix(bids, out, _run_inputs())
    return openpyxl.load_workbook(out)["Leveled_Normalized"], len(docs)


def _hex(cell) -> str:
    rgb = cell.fill.fgColor.rgb
    return rgb[-6:].upper() if isinstance(rgb, str) else ""


def _row_of(ws, col, value):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    raise AssertionError(f"row with col{col}=={value!r} not found")


def _col_of(ws, n, name):
    for i in range(n):
        if ws.cell(row=5, column=_lev_col_start(i)).value == name:
            return _lev_col_start(i)
    raise AssertionError(f"bidder {name!r} not found on row 5")


class TestBenchmarkFromClassifiedStates:
    def test_blank_never_enters_the_median_as_zero(self):
        """Marvin's A6 trap: 3 priced bids (100k/110k/120k) + 1 BLANK bid.
        Benchmark must be the median of the THREE valid prices (110k) — if the
        blank leaked in as 0.0, the median would drop to 105k."""
        docs = [
            _concrete_bidder("Alpha", "100000"),
            _concrete_bidder("Bravo", "110000"),
            _concrete_bidder("Charlie", "120000"),
            # Delta: DIV 03 blank (NULL_BLANK), priced elsewhere so it has a GT.
            _doc("Delta", [
                _div("DIV 03 00 00", "Concrete", None,
                     items=[LineItem(description="Concrete", amount=None)],
                     cost=CostStructure.ITEMIZED),
                _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
            ], Decimal("50000")),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            bench = ws.cell(row=sub_row, column=_lev_bench_col(n)).value
            n_valid = ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value
            assert bench == 110000.0, (
                f"median must be 110k over the 3 valid prices, got {bench} "
                f"(a blank entered the median as 0.0)"
            )
            assert n_valid == 3
            # R5: Delta's blank DIV 03 paints RED (peers priced it).
            delta_csub = ws.cell(
                row=sub_row, column=_col_of(ws, n, "Delta") + LEVELED_CSUB_OFFSET
            )
            assert _hex(delta_csub) == RED
            assert delta_csub.comment is not None
            assert "R5" in delta_csub.comment.text

    def test_zero_subtotal_paints_red_not_neutral(self):
        """R6 at line level: an unapproved explicit zero paints RED on the
        line's COST cell."""
        docs = [
            _doc("Alpha", [_div(
                "DIV 01 00 00", "General Requirements", Decimal("0"),
                items=[LineItem(description="Final Cleaning",
                                amount=Decimal("0"), is_explicit_zero=True)],
                cost=CostStructure.ITEMIZED,
            )], Decimal("0")),
            _doc("Bravo", [_div(
                "DIV 01 00 00", "General Requirements", Decimal("10000"),
                items=[LineItem(description="Final Cleaning",
                                amount=Decimal("10000"))],
                cost=CostStructure.ITEMIZED,
            )], Decimal("10000")),
            _doc("Charlie", [_div(
                "DIV 01 00 00", "General Requirements", Decimal("12000"),
                items=[LineItem(description="Final Cleaning",
                                amount=Decimal("12000"))],
                cost=CostStructure.ITEMIZED,
            )], Decimal("12000")),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            line_row = _row_of(ws, 2, "Final Cleaning")
            zero_cell = ws.cell(row=line_row, column=_col_of(ws, n, "Alpha"))
            assert zero_cell.value == 0.0
            assert _hex(zero_cell) == RED
            assert "R6" in zero_cell.comment.text


class TestVarianceBoundaries:
    def _field(self):
        # Median of (80k, 100k, 100k, 120k) = 100k. Low bidder sits EXACTLY at
        # ×0.80, high bidder EXACTLY at ×1.20 — the Q1 inclusive boundary.
        return [
            _concrete_bidder("Low Bidder", "80000"),
            _concrete_bidder("Mid One", "100000"),
            _concrete_bidder("Mid Two", "100000"),
            _concrete_bidder("High Bidder", "120000"),
        ]

    def test_inclusive_20pct_boundaries_cyan_and_yellow(self):
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, self._field())
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")

            low = ws.cell(row=sub_row,
                          column=_col_of(ws, n, "Low Bidder") + LEVELED_CSUB_OFFSET)
            high = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "High Bidder") + LEVELED_CSUB_OFFSET)
            mid = ws.cell(row=sub_row,
                          column=_col_of(ws, n, "Mid One") + LEVELED_CSUB_OFFSET)
            assert _hex(low) == CYAN, "price at exactly ×0.80 must paint cyan (Q1)"
            assert _hex(high) == YELLOW, "price at exactly ×1.20 must paint yellow (Q1)"
            assert _hex(mid) == "A3EAF3", "within-range subtotal keeps the aqua band"
            assert "20% below" in low.comment.text     # R31 cyan language
            assert "20% above" in high.comment.text    # R31 yellow language

    def test_var_pct_written_in_separator_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, self._field())
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            var_col = _col_of(ws, n, "Low Bidder") + 4  # VAR % offset
            assert abs(ws.cell(row=sub_row, column=var_col).value - (-0.2)) < 1e-9


class TestRedFirstPrecedence:
    def test_r20_math_error_beats_yellow_variance(self):
        """A bidder whose stated subtotal (200k) both fails R20 (items sum to
        100k) and sits ≥×1.20 over the benchmark must paint RED, not yellow
        (R16: Red > Cyan > Yellow)."""
        docs = [
            _doc("Wrongmath", [_div(
                "DIV 03 00 00", "Concrete", Decimal("200000"),
                items=[LineItem(description="Concrete work",
                                amount=Decimal("100000"))],
                cost=CostStructure.LUMP_SUM,  # stated subtotal wins the cell
            )], Decimal("200000")),
            _concrete_bidder("Peer One", "100000"),
            _concrete_bidder("Peer Two", "100000"),
            _concrete_bidder("Peer Three", "100000"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Wrongmath") + LEVELED_CSUB_OFFSET)
            assert _hex(cell) == RED, "R16: red must beat the yellow variance"
            assert "R20" in cell.comment.text


class TestMinimumBidGate:
    def test_no_variance_paint_below_three_valid_bids(self):
        """RISK-2/Q5: with 2 valid bids (100k vs 160k) both sit beyond ±20% of
        the 130k midpoint-median — but paint is SUPPRESSED below 3 valid bids.
        The benchmark still displays."""
        docs = [
            _concrete_bidder("Alpha", "100000"),
            _concrete_bidder("Bravo", "160000"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            for name in ("Alpha", "Bravo"):
                cell = ws.cell(
                    row=sub_row, column=_col_of(ws, n, name) + LEVELED_CSUB_OFFSET
                )
                assert _hex(cell) == "A3EAF3", (
                    f"{name}: no cyan/yellow below 3 valid bids — got {_hex(cell)}"
                )
            assert ws.cell(row=sub_row, column=_lev_bench_col(n)).value == 130000.0
            # R29/Q2: <3 valid bids ⇒ Low confidence.
            assert ws.cell(row=sub_row, column=_lev_bench_col(n) + 3).value == "Low"


class TestExclusionAndSummary:
    def test_unapproved_exclusion_paints_red_and_summary_counts(self):
        docs = [
            _doc("Excluder", [
                _div("DIV 03 00 00", "Concrete", None,
                     items=[LineItem(description="Concrete work",
                                     is_excluded=True)],
                     cost=CostStructure.ITEMIZED),
                _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
            ], Decimal("50000")),
            _concrete_bidder("Peer One", "100000"),
            _concrete_bidder("Peer Two", "110000"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Excluder") + LEVELED_CSUB_OFFSET)
            assert cell.value == "Excluded"
            assert _hex(cell) == RED
            assert "R28" in cell.comment.text
            # R32 summary block: Excluder carries ≥1 red flag.
            red_row = _row_of(ws, 2, "Red Flags")
            red_count = ws.cell(
                row=red_row,
                column=_col_of(ws, n, "Excluder") + LEVELED_CSUB_OFFSET,
            ).value
            assert red_count and red_count >= 1
            # Workbook legend present (W-D: one block, both data sheets).
            _row_of(ws, 2, "LEGEND — READING THIS WORKBOOK "
                           "(every signal, all three tabs)")


class TestRem1StatedZeroSubtotal:
    """Marvin's REM-1 (gold-standard diff, 2026-07-03): a stated $0 division
    subtotal is NEVER a valid benchmark price (R6/R7)."""

    def test_marvin_acceptance_stated_zero_with_excluded_lines(self):
        """Marvin's acceptance test verbatim: A=8,675; B=34,081.85; C=55,700;
        D stated-$0 subtotal over excluded lines ⇒ benchmark 34,081.85 over
        3 valid bids, D's cell reads "Excluded" + RED + R28 comment, and D's
        zero is absent from the prices (with it, the median would be the
        21,378.43 midpoint)."""
        docs = [
            _concrete_bidder("Alpha", "8675"),
            _concrete_bidder("Bravo", "34081.85"),
            _concrete_bidder("Charlie", "55700"),
            _doc("Delta", [
                _div("DIV 03 00 00", "Concrete", Decimal("0"),
                     items=[LineItem(description="Concrete work",
                                     is_excluded=True)],
                     cost=CostStructure.ITEMIZED),
                _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
            ], Decimal("50000")),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            bench = ws.cell(row=sub_row, column=_lev_bench_col(n)).value
            n_valid = ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value
            assert bench == 34081.85, (
                f"benchmark must be 34,081.85 over the 3 valid prices, got "
                f"{bench} — a stated $0 subtotal entered the median (REM-1)"
            )
            assert n_valid == 3
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Delta") + LEVELED_CSUB_OFFSET)
            assert cell.value == "Excluded"
            assert _hex(cell) == RED
            assert "R28" in cell.comment.text

    def test_stated_zero_without_classified_lines_stays_r6_zero(self):
        """A stated-$0 LUMP_SUM subtotal with no classified lines (a recurring
        DIV 07 pattern) paints RED with the R6 zero language, writes 0.00,
        and still never enters the median."""
        docs = [
            _concrete_bidder("Alpha", "100000"),
            _concrete_bidder("Bravo", "110000"),
            _concrete_bidder("Charlie", "120000"),
            _doc("Delta", [
                _div("DIV 03 00 00", "Concrete", Decimal("0")),  # LUMP_SUM $0
                _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
            ], Decimal("50000")),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            assert ws.cell(row=sub_row, column=_lev_bench_col(n)).value == 110000.0
            assert ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value == 3
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Delta") + LEVELED_CSUB_OFFSET)
            assert cell.value == 0.0
            assert _hex(cell) == RED
            assert "R6" in cell.comment.text


class TestEnc1VerbatimClassificationDisplay:
    def test_not_applicable_division_displays_verbatim_token(self):
        """ENC-1 (Marvin): a division classified 'Not Applicable' must display
        the verbatim token, not 'By Owner' — the owner does not carry it."""
        docs = [
            _concrete_bidder("Alpha", "100000"),
            _concrete_bidder("Bravo", "110000"),
            _doc("Delta", [
                _div("DIV 03 00 00", "Concrete", None,
                     items=[LineItem(description="Concrete work",
                                     is_by_owner_others=True,
                                     by_others_verbatim="Not Applicable")],
                     cost=CostStructure.ITEMIZED),
                _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
            ], Decimal("50000")),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Delta") + LEVELED_CSUB_OFFSET)
            assert cell.value == "Not Applicable"
            assert _hex(cell) != RED  # approved classification, no error
            assert "Approved classification: Not Applicable" in cell.comment.text


class TestEnc2NotComparable:
    def _field(self):
        # A/B/C price the same line; D's is classified Not Comparable at 500k.
        def bidder(name, amount, nc=False):
            item = LineItem(description="Site Fencing",
                            amount=Decimal(amount), is_not_comparable=nc)
            return _doc(name, [_div(
                "DIV 01 00 00", "General Requirements", Decimal(amount),
                items=[item], cost=CostStructure.ITEMIZED,
            )], Decimal(amount))
        return [
            bidder("Alpha", "100000"),
            bidder("Bravo", "110000"),
            bidder("Charlie", "120000"),
            bidder("Delta", "500000", nc=True),
        ]

    def test_line_level_amount_displayed_but_out_of_benchmark(self):
        """ENC-2: the Not-Comparable amount is displayed as submitted with no
        paint and the R7/R8 comment, and the line benchmark is the median of
        the OTHER three (110k — with the 500k in, it would be 115k)."""
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, self._field())
            line_row = _row_of(ws, 2, "Site Fencing")
            bench = ws.cell(row=line_row, column=_lev_bench_col(n)).value
            n_valid = ws.cell(row=line_row, column=_lev_bench_col(n) + 2).value
            assert bench == 110000.0, (
                f"line benchmark must exclude the Not-Comparable 500k, got {bench}"
            )
            assert n_valid == 3
            cell = ws.cell(row=line_row, column=_col_of(ws, n, "Delta"))
            assert cell.value == 500000.0          # displayed as submitted
            assert cell.fill.patternType is None   # no paint
            assert "Not Comparable — excluded from benchmark (R7/R8)" in \
                cell.comment.text

    def test_division_composed_of_nc_lines_out_of_subtotal_benchmark(self):
        """ENC-2 one level up: a division whose lines are ALL Not Comparable
        must not push its (derived) subtotal into the subtotal median."""
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, self._field())
            sub_row = _row_of(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            bench = ws.cell(row=sub_row, column=_lev_bench_col(n)).value
            n_valid = ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value
            assert bench == 110000.0
            assert n_valid == 3
            csub = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Delta") + LEVELED_CSUB_OFFSET)
            assert csub.value == 500000.0          # displayed, aqua band kept
            assert _hex(csub) == "A3EAF3"
            assert "Not Comparable — excluded from benchmark (R7/R8)" in \
                csub.comment.text


class TestRem2DerivedSubtotalComment:
    def test_derived_subtotal_carries_disclosure_comment(self):
        """REM-2 (Marvin): a subtotal the engine DERIVED from line items (no
        stated subtotal on the form — a recurring scissor-lift pattern) gets
        an on-cell disclosure, with composition numbers when the bidder's
        stated Construction Cost Subtotal does not carry it."""
        derived = BidDocument(
            contractor_name="Derived Co",
            form_type=FormType.FALKE_STANDARD,
            bid_document_input_type=InputType.DIGITAL_NATIVE,
            divisions=[
                _div("DIV 01 00 00", "General Requirements", Decimal("100000")),
                # DIV 11: no stated subtotal — the engine derives 18,500.
                _div("DIV 11 00 00", "Equipment", None,
                     items=[LineItem(description="Scissor Lift",
                                     amount=Decimal("18500"))],
                     cost=CostStructure.ITEMIZED),
            ],
            footer=BidFooter(
                construction_cost_subtotal=Decimal("100000"),  # excludes 18.5k
                grand_total=Decimal("100000"),
                grand_total_confidence=GrandTotalConfidence.LOW,
            ),
            qualifications=BidQualifications(),
            extraction_confidence=ExtractionConfidence.HIGH,
        )
        docs = [derived, _concrete_bidder("Peer One", "90000")]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "EQUIPMENT SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Derived Co") + LEVELED_CSUB_OFFSET)
            assert cell.value == 18500.0
            assert cell.comment is not None, "derived subtotal must be disclosed"
            text = cell.comment.text
            assert "Subtotal DERIVED from priced line items" in text
            assert "$100,000.00" in text and "$118,500.00" in text, (
                "composition numbers (stated CCS vs division-subtotal sum) "
                f"must appear; got: {text}"
            )
            # A STATED subtotal gets no derived-disclosure comment.
            stated_row = _row_of(ws, 2, "GENERAL REQUIREMENTS SUBTOTAL")
            stated_cell = ws.cell(
                row=stated_row,
                column=_col_of(ws, n, "Derived Co") + LEVELED_CSUB_OFFSET)
            assert (stated_cell.comment is None
                    or "DERIVED" not in stated_cell.comment.text)


class TestEnc3CompositionCheck:
    """ENC-3 (S2-1, W-D): Σ displayed division subtotals vs stated CCS at
    max($5, 0.5%) — RED on the leveled CCS cell + [FALKE R21] comment + a real
    SUBTOTAL_COMPOSITION_DISCREPANCY AUDIT row (the $18,500 class)."""

    @staticmethod
    def _composition_doc(name, stated_ccs, div_amounts, derived=False):
        divisions = []
        for i, amount in enumerate(div_amounts):
            code, dname = ("DIV 03 00 00", "Concrete") if i == 0 else (
                "DIV 01 00 00", "General Requirements")
            if derived:
                divisions.append(_div(
                    code, dname, None,
                    items=[LineItem(description=f"{dname} work",
                                    amount=amount)],
                    cost=CostStructure.ITEMIZED))
            else:
                divisions.append(_div(code, dname, amount))
        return BidDocument(
            contractor_name=name,
            form_type=FormType.FALKE_STANDARD,
            bid_document_input_type=InputType.DIGITAL_NATIVE,
            divisions=divisions,
            footer=BidFooter(
                construction_cost_subtotal=stated_ccs,
                grand_total=stated_ccs,
                grand_total_confidence=GrandTotalConfidence.LOW,
            ),
            qualifications=BidQualifications(),
            extraction_confidence=ExtractionConfidence.HIGH,
        )

    def _run(self, docs):
        from src.audit import AuditCode, AuditStatus, audit_bids
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
        items = audit_bids(bids)
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            ccs_row = _row_of(ws, 2, "CONSTRUCTION COST SUBTOTAL")
            cells = {
                doc.contractor_name: ws.cell(
                    row=ccs_row, column=_col_of(ws, n, doc.contractor_name))
                for doc in docs
            }
            snapshot = {
                nm: (_hex(c), c.comment.text if c.comment else "")
                for nm, c in cells.items()
            }
        return items, snapshot

    def test_composition_failure_fires_red_cell_and_audit_row(self):
        from src.audit import AuditCode, AuditStatus
        docs = [
            # Stated CCS 130,000 but displayed divisions sum 150,000 (Δ 20,000).
            self._composition_doc("Composer Fail",
                                  Decimal("130000"),
                                  [Decimal("100000"), Decimal("50000")]),
            self._composition_doc("Composer Clean",
                                  Decimal("150000"),
                                  [Decimal("100000"), Decimal("50000")]),
        ]
        items, snapshot = self._run(docs)
        rows = [i for i in items
                if i.code == AuditCode.SUBTOTAL_COMPOSITION_DISCREPANCY]
        assert len(rows) == 1, [r.contractor_name for r in rows]
        row = rows[0]
        assert row.contractor_name == "Composer Fail"
        assert row.status == AuditStatus.RED
        assert row.view == "leveled"
        assert row.value == "$20,000"
        assert "does not compose" in row.message
        # Leveled CCS cell: RED + [FALKE R21] composition comment.
        hex_, comment = snapshot["Composer Fail"]
        assert hex_ == RED
        assert "[FALKE R21]" in comment
        assert "does not compose from the displayed division subtotals" in comment
        assert "delta $20,000.00" in comment
        # Clean bidder: no paint, no composition comment.
        hex_clean, comment_clean = snapshot["Composer Clean"]
        assert hex_clean != RED
        assert "does not compose" not in comment_clean

    def test_within_tolerance_does_not_fire(self):
        from src.audit import AuditCode
        docs = [
            # Δ 400 on a 100,000 CCS — within max($5, 0.5% × 100,000 = 500).
            self._composition_doc("Near Miss",
                                  Decimal("100000"),
                                  [Decimal("60400"), Decimal("40000")]),
            self._composition_doc("Peer", Decimal("90000"),
                                  [Decimal("50000"), Decimal("40000")]),
        ]
        items, snapshot = self._run(docs)
        assert not [i for i in items
                    if i.code == AuditCode.SUBTOTAL_COMPOSITION_DISCREPANCY]
        assert snapshot["Near Miss"][0] != RED

    def test_no_stated_ccs_skips_check(self):
        from src.audit import AuditCode
        doc = self._composition_doc("No CCS Co", None,
                                    [Decimal("100000"), Decimal("50000")])
        docs = [doc, self._composition_doc("Peer", Decimal("90000"),
                                           [Decimal("50000"), Decimal("40000")])]
        items, _snapshot = self._run(docs)
        assert not [i for i in items
                    if i.code == AuditCode.SUBTOTAL_COMPOSITION_DISCREPANCY
                    and i.contractor_name == "No CCS Co"]

    def test_rem2_pointer_names_the_register_row(self):
        """The REM-2 composition sentence now points at a row that EXISTS
        (same tolerance as ENC-3 — the dangling 'see AUDIT' fix)."""
        from src.audit import AuditCode, audit_bids
        derived = self._composition_doc(
            "Derived Gap Co", Decimal("130000"),
            [Decimal("100000"), Decimal("50000")], derived=True)
        docs = [derived,
                self._composition_doc("Peer", Decimal("90000"),
                                      [Decimal("50000"), Decimal("40000")])]
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
        items = audit_bids(bids)
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Derived Gap Co")
                           + LEVELED_CSUB_OFFSET)
            assert cell.comment is not None
            assert ("see the SUBTOTAL_COMPOSITION_DISCREPANCY row on the "
                    "AUDIT tab") in cell.comment.text
        # ...and the row it names is really on the register.
        assert [i for i in items
                if i.code == AuditCode.SUBTOTAL_COMPOSITION_DISCREPANCY
                and i.contractor_name == "Derived Gap Co"]


class TestMedianMembershipContract:
    """M-2 (W-D): ONE median-membership rule everywhere — the leveled
    benchmark block and the audit-side medians/gap-strings consume the SAME
    set (falke_rules.median_membership: priced AND amount>0 AND not
    R20-failed). Contract-tests the consumer sites against each other."""

    def _field(self):
        # Four clean pricers + one R20-failed bidder (stated 508k, lines 500k)
        # + one absent bidder (draws the scope-gap string).
        failed = _doc("Failed Math Co", [
            _div("DIV 03 00 00", "Concrete", Decimal("508000"),
                 items=[LineItem(description="Failed Math concrete work",
                                 amount=Decimal("500000"))],
                 cost=CostStructure.ITEMIZED),
        ], Decimal("508000"))
        absent = _doc("Absent Co", [
            _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        ], Decimal("50000"))
        return [
            _concrete_bidder("P310", "310000"),
            _concrete_bidder("P450", "450000"),
            _concrete_bidder("P500", "500000"),
            _concrete_bidder("P520", "520000"),
            failed, absent,
        ]

    def test_r20_failed_subtotal_leaves_every_median(self):
        from src.audit import AuditCode, audit_bids
        from src import falke_rules as fr
        docs = self._field()
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])

        # The ONE membership rule: failed bidder is out, four members remain.
        assert fr.median_membership(
            next(b for b in bids if b.contractor_name == "Failed Math Co"),
            "DIV 03 00 00") is None
        members = [fr.median_membership(b, "DIV 03 00 00") for b in bids]
        assert sorted(m for m in members if m is not None) == [
            310000.0, 450000.0, 500000.0, 520000.0]

        # Consumer 1 — the written benchmark block.
        items = audit_bids(bids)
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            assert ws.cell(row=sub_row, column=_lev_bench_col(n)).value == 475000.0
            assert ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value == 4
            # Failed bidder keeps its R20 red and its written VAR% vs the
            # fenced median: (508,000-475,000)/475,000.
            fcol = _col_of(ws, n, "Failed Math Co")
            fcell = ws.cell(row=sub_row, column=fcol + LEVELED_CSUB_OFFSET)
            assert _hex(fcell) == RED
            var = ws.cell(row=sub_row, column=fcol + 4).value
            assert abs(var - 0.069474) < 1e-4

        # Consumer 2 — the audit gap-value string cites the SAME median.
        gap = next(i for i in items
                   if i.code == AuditCode.SCOPE_GAP_IMPLICIT
                   and i.contractor_name == "Absent Co"
                   and i.division_csi == "DIV 03 00 00")
        assert gap.value == "$475,000", (
            f"audit gap string must cite the fenced sheet median, got {gap.value}")

    def test_nc_composed_subtotal_fenced_from_audit_median(self):
        from src.audit import AuditCode, audit_bids
        nc = _doc("NC Co", [
            _div("DIV 03 00 00", "Concrete", None,
                 items=[LineItem(description="Owner option concrete package",
                                 amount=Decimal("145000"),
                                 is_not_comparable=True)],
                 cost=CostStructure.ITEMIZED),
        ], Decimal("145000"))
        absent = _doc("Absent Co", [
            _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        ], Decimal("50000"))
        docs = [_concrete_bidder("P190", "190000"),
                _concrete_bidder("P200", "200000"),
                _concrete_bidder("P250", "250000"), nc, absent]
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
        items = audit_bids(bids)
        gap = next(i for i in items
                   if i.code == AuditCode.SCOPE_GAP_IMPLICIT
                   and i.contractor_name == "Absent Co"
                   and i.division_csi == "DIV 03 00 00")
        assert gap.value == "$200,000", (
            f"NC-composed subtotal must be fenced (GOLD-DEV-4), got {gap.value}")

    def test_zero_and_negative_subtotals_are_not_members(self):
        from src import falke_rules as fr
        neg = _doc("Credit Co", [
            _div("DIV 03 00 00", "Concrete", Decimal("-5000")),
        ], Decimal("-5000"))
        zero = _doc("Zero Co", [
            _div("DIV 03 00 00", "Concrete", Decimal("0")),
        ], Decimal("0"))
        for doc in (neg, zero):
            bid = normalize_bid(doc)
            assert fr.median_membership(bid, "DIV 03 00 00") is None, (
                f"{doc.contractor_name} must not enter any median")


class TestCreditSemantics:
    """W-D ruling 5 (Floyd C-W4-3): a net-negative division subtotal is
    LEGAL, preserved, rendered accounting-negative — never clamped, never
    dropped. R22 red + NEGATIVE_UNCLASSIFIED RED row; fenced from medians;
    the bidder's own arithmetic keeps it."""

    def _docs(self):
        credit = _doc("Credit Co", [
            _div("DIV 03 00 00", "Concrete", Decimal("-5000")),
            _div("DIV 01 00 00", "General Requirements", Decimal("100000")),
        ], Decimal("95000"))
        return [credit,
                _concrete_bidder("Peer One", "200000"),
                _concrete_bidder("Peer Two", "220000"),
                _concrete_bidder("Peer Three", "240000")]

    def test_leveled_r22_red_and_fenced_benchmark(self):
        docs = self._docs()
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            col = _col_of(ws, n, "Credit Co")
            csub = ws.cell(row=sub_row, column=col + LEVELED_CSUB_OFFSET)
            # Rendered negative — never clamped, never blank.
            assert csub.value == -5000.0
            assert _hex(csub) == RED
            assert "[FALKE R22]" in csub.comment.text
            assert ("Negative value without deductive-alternate or "
                    "approved-credit classification — R22.") in csub.comment.text
            # House accounting format renders negatives in parentheses.
            assert "(" in csub.number_format
            # Benchmark: credit fenced (M-2 rule 2) — median over 3 peers.
            assert ws.cell(row=sub_row, column=_lev_bench_col(n)).value == 220000.0
            assert ws.cell(row=sub_row, column=_lev_bench_col(n) + 2).value == 3
            # No VAR% for the unclassified credit.
            var_cell = ws.cell(row=sub_row, column=col + 4)
            assert var_cell.value is None

    def test_audit_negative_unclassified_row(self):
        from src.audit import AuditCode, AuditStatus, audit_bids
        bids = compute_cross_bid_stats([normalize_bid(d) for d in self._docs()])
        items = audit_bids(bids)
        rows = [i for i in items if i.code == AuditCode.NEGATIVE_UNCLASSIFIED]
        assert len(rows) == 1
        row = rows[0]
        assert row.contractor_name == "Credit Co"
        assert row.division_csi == "DIV 03 00 00"
        assert row.status == AuditStatus.RED
        assert row.view == "both"
        assert row.value == "$-5,000"
        assert "Net negative division subtotal ($-5,000)" in row.message
        assert "deductive alternate or approved credit (R22)" in row.message

    def test_mirror_renders_credit_and_ties_out(self):
        from src.audit import audit_bids
        from src.reconcile import reconcile_written_matrix
        from src.write_matrix import _col_start
        docs = self._docs()
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
        items = audit_bids(bids)
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            from src.write_matrix import write_matrix as wm
            wm(bids, out, _run_inputs(), audit_items=items)
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            sub_row = _row_of(ws, 2, "CONCRETE SUBTOTAL")
            col = next(_col_start(i) for i in range(len(docs))
                       if ws.cell(row=5, column=_col_start(i)).value == "Credit Co")
            c = ws.cell(row=sub_row, column=col)
            assert c.value == -5000.0, "mirror renders the credit as submitted"
            assert "(" in c.number_format, "accounting-negative on the mirror"
            # Stage 6b lockstep: negative subtotal ties out, zero failures.
            failures = reconcile_written_matrix(out, bids, len(items))
            assert failures == [], [f.message for f in failures]


class TestEnc5LineTokens:
    """ENC-5 (W-D): line-level EXCLUDED / BY_OWNER cells inside a division
    render the bidder's classification token on the leveled sheet — italic,
    NO paint (severity lives on the AUDIT register)."""

    def _docs(self):
        mixed = _doc("Mixed Co", [
            _div("DIV 09 00 00", "Finishes",
                 items=[
                     LineItem(description="Wall finishes package",
                              amount=Decimal("180000")),
                     LineItem(description="Corridor painting scope",
                              is_excluded=True),
                     LineItem(description="Owner appliance package",
                              is_by_owner_others=True,
                              by_others_verbatim="NIC — By Others"),
                 ],
                 subtotal=Decimal("180000")),
        ], Decimal("180000"))
        peer = _doc("Peer Co", [
            _div("DIV 09 00 00", "Finishes",
                 items=[LineItem(description="Complete interior scope",
                                 amount=Decimal("200000"))],
                 subtotal=Decimal("200000")),
        ], Decimal("200000"))
        return [mixed, peer]

    def test_line_tokens_render_italic_unpainted(self):
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, self._docs())
            col = _col_of(ws, n, "Mixed Co")

            excl_row = _row_of(ws, 2, "Corridor painting scope")
            ec = ws.cell(row=excl_row, column=col)
            assert ec.value == "Excluded"
            assert ec.font.italic is True
            assert _hex(ec) not in (RED, CYAN, YELLOW)

            byo_row = _row_of(ws, 2, "Owner appliance package")
            bc = ws.cell(row=byo_row, column=col)
            assert bc.value == "NIC — By Others", "verbatim token (ENC-1)"
            assert bc.font.italic is True
            assert _hex(bc) not in (RED, CYAN, YELLOW)

            # The priced line and the division subtotal are untouched.
            priced_row = _row_of(ws, 2, "Wall finishes package")
            assert ws.cell(row=priced_row, column=col).value == 180000.0
            sub_row = _row_of(ws, 2, "FINISHES SUBTOTAL")
            assert ws.cell(row=sub_row,
                           column=col + LEVELED_CSUB_OFFSET).value == 180000.0


class TestRem2WithinToleranceDisclosure:
    """W-D refinement (real-set finding): a composition delta INSIDE Falke's
    max($5, 0.5%) tolerance keeps its REM-2 on-cell disclosure (the
    real-world $18,500-on-$3.9M within-tolerance shape) but says it is within tolerance
    (RISK-1) instead of pointing at an AUDIT row that doesn't exist."""

    def test_within_tolerance_delta_disclosed_without_dangling_pointer(self):
        from src.audit import AuditCode, audit_bids
        # Stated CCS 3,903,200; displayed divisions sum 3,921,700
        # (delta 18,500 < tol = 19,516) — DIV 11 subtotal is DERIVED.
        derived = BidDocument(
            contractor_name="Baywater Class Co",
            form_type=FormType.FALKE_STANDARD,
            bid_document_input_type=InputType.DIGITAL_NATIVE,
            divisions=[
                _div("DIV 03 00 00", "Concrete", Decimal("3903200")),
                _div("DIV 11 00 00", "Equipment", None,
                     items=[LineItem(description="Scissor lift rental",
                                     amount=Decimal("18500"))],
                     cost=CostStructure.ITEMIZED),
            ],
            footer=BidFooter(
                construction_cost_subtotal=Decimal("3903200"),
                grand_total=Decimal("3903200"),
                grand_total_confidence=GrandTotalConfidence.LOW,
            ),
            qualifications=BidQualifications(),
            extraction_confidence=ExtractionConfidence.HIGH,
        )
        docs = [derived, _concrete_bidder("Peer One", "3900000")]
        bids = compute_cross_bid_stats([normalize_bid(d) for d in docs])
        items = audit_bids(bids)
        # Within tolerance: NO ENC-3 register row…
        assert not [i for i in items
                    if i.code == AuditCode.SUBTOTAL_COMPOSITION_DISCREPANCY]
        with tempfile.TemporaryDirectory() as tmp:
            ws, n = _write(tmp, docs)
            sub_row = _row_of(ws, 2, "EQUIPMENT SUBTOTAL")
            cell = ws.cell(row=sub_row,
                           column=_col_of(ws, n, "Baywater Class Co")
                           + LEVELED_CSUB_OFFSET)
            text = cell.comment.text
            # …but the on-cell disclosure SURVIVES, honestly worded.
            assert "delta $18,500.00" in text
            assert "within Falke's max($5, 0.5%) math tolerance" in text
            assert "RISK-1" in text
            assert "SUBTOTAL_COMPOSITION_DISCREPANCY" not in text
            # CCS cell carries no ENC-3 red.
            ccs_row = _row_of(ws, 2, "CONSTRUCTION COST SUBTOTAL")
            ccs = ws.cell(row=ccs_row,
                          column=_col_of(ws, n, "Baywater Class Co"))
            assert _hex(ccs) != RED
