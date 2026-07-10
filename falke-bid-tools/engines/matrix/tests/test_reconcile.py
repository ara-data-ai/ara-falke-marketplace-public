"""
FALKE Matrix — Stage 6b post-write reconciliation tests
=======================================================
Proves the closed-loop tie-out (`reconcile_written_matrix`) actually FIRES and,
under LOUD QUARANTINE (Derick's decision, Marvin's disclosure spec), DELIVERS the
file with the defect flagged rather than refusing:

  * HAPPY PATH — a correctly-written matrix ties out cleanly (no failures).
  * FAILURE INJECTION (one per check) — deliberately corrupt the written .xlsx
    and assert Stage 6b returns a RED POST_WRITE_TIEOUT_FAILURE with the right
    (reworded) message:
      check 1 — overwrite a GRAND TOTAL cell with a wrong number
      check 2 — overwrite a footer component so the sum ≠ grand total
      check 3 — move a division subtotal value to the wrong number
      check 4 — drop an AUDIT-sheet row
  * PIPELINE QUARANTINE — run_pipeline through a corrupted-write monkeypatch and
    assert it DELIVERS the file, exits with the distinct code 3, and stamps the
    RED banner + RED cell mark + RED AUDIT row + QUARANTINE line.
  * QUARANTINE DISCLOSURE — apply_quarantine renders the banner ({N} count and
    singular/plural + structural "one or more" fallback), the in-place cell marks
    (grand-total + division), and the reworded AUDIT strings (no "refusing to
    deliver"); a clean run renders NONE of it.

Run from the engine root:
    python3 -m pytest tests/test_reconcile.py -v
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from src.audit import AuditCode, AuditStatus, audit_bids
from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
)
from src.normalize import compute_cross_bid_stats, normalize_bid
from src.reconcile import (
    _GRAND_TOTAL_KEY,
    _GL_KEY,
    _find_footer_row,
    _find_subtotal_rows,
    reconcile_written_matrix,
)
from src.run_config import RunInputs
from src.write_matrix import _col_start, write_matrix


# ---------------------------------------------------------------------------
# Fixtures — a small, tie-able two-bidder field
# ---------------------------------------------------------------------------

def _footer(
    construction, gl, br, gc_fee, grand_total,
    overhead_and_profit=None, other_fees_subtotal=None, bond=None,
):
    return BidFooter(
        construction_cost_subtotal=construction,
        general_liability_insurance=gl,
        builders_risk_insurance=br,
        gc_fee=gc_fee,
        overhead_and_profit=overhead_and_profit,
        other_fees_subtotal=other_fees_subtotal,
        bond=bond,
        grand_total=grand_total,
        grand_total_confidence=GrandTotalConfidence.LOW,
    )


def _div(code, name, subtotal):
    return DivisionBid(
        csi_code=code,
        division_name=name,
        cost_structure=CostStructure.LUMP_SUM,
        division_subtotal=subtotal,
        classification_source=ClassificationSource.CONTRACTOR_NATIVE,
        contractor_native_code=None,
        line_items=[],
    )


def _doc(name, divisions, footer):
    return BidDocument(
        contractor_name=name,
        form_type=FormType.FALKE_STANDARD,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=footer,
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _run_inputs(gsf=10_000.0):
    return RunInputs(
        project_name="Harbor View Tower",
        project_address="100 Test Ave, Test City FL 00000",
        gross_sf=gsf,
        sf_basis_label="GSF",
        sf_source="explicit",
    )


def _build_field():
    """Two bidders, each with a clean footer that sums to its grand total."""
    # Bidder A: 50k + 400k construction = 450k; +5k GL +5k BR +40k GC = 500k grand
    a = _doc("Apex Restoration", [
        _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        _div("DIV 03 00 00", "Concrete", Decimal("400000")),
    ], _footer(Decimal("450000"), Decimal("5000"), Decimal("5000"),
               Decimal("40000"), Decimal("500000")))
    # Bidder B: 60k + 360k = 420k; +4k +4k +32k = 460k grand
    b = _doc("Beacon Builders", [
        _div("DIV 01 00 00", "General Requirements", Decimal("60000")),
        _div("DIV 03 00 00", "Concrete", Decimal("360000")),
    ], _footer(Decimal("420000"), Decimal("4000"), Decimal("4000"),
               Decimal("32000"), Decimal("460000")))
    return compute_cross_bid_stats([normalize_bid(a), normalize_bid(b)])


def _write(tmp_path: Path, bids):
    out = tmp_path / "matrix.xlsx"
    audit_items = audit_bids(bids)
    write_matrix(bids, out, _run_inputs(), audit_items=audit_items)
    return out, audit_items


# ---------------------------------------------------------------------------
# HAPPY PATH
# ---------------------------------------------------------------------------

def test_happy_path_ties_out_cleanly(tmp_path):
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)
    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures == [], (
        "A correctly-written matrix must tie out with zero failures; got: "
        + "; ".join(f.message for f in failures)
    )


# ---------------------------------------------------------------------------
# CHECK 1 — corrupt a GRAND TOTAL cell
# ---------------------------------------------------------------------------

def test_check1_wrong_grand_total_fires(tmp_path):
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    gt_row = _find_footer_row(ws, _GRAND_TOTAL_KEY)
    # Apex is the low bid (500k) — but ordering doesn't matter; corrupt col C.
    ws.cell(row=gt_row, column=_col_start(0)).value = 999999.0
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "A wrong grand-total cell must produce a tie-out failure."
    assert all(f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE for f in failures)
    assert all(f.status == AuditStatus.RED for f in failures)
    assert any("Grand-total tie-out FAILED" in f.message for f in failures)


# ---------------------------------------------------------------------------
# CHECK 2 — break footer arithmetic (sum ≠ grand total)
# ---------------------------------------------------------------------------

def test_check2_footer_arithmetic_break_fires(tmp_path):
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    gl_row = _find_footer_row(ws, _GL_KEY)
    # Inflate GL on col C by 100k WITHOUT touching the grand total → sum breaks.
    ws.cell(row=gl_row, column=_col_start(0)).value = 105000.0
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "A footer that no longer sums must produce a tie-out failure."
    # Check 1 (grand total unchanged) still passes for that bidder; check 2 fires.
    assert any("Footer arithmetic FAILED" in f.message for f in failures)
    assert all(f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE for f in failures)


# ---------------------------------------------------------------------------
# CHECK 2 (missing-row) — a dropped footer-component label is a STRUCTURAL
# quarantine failure: delivered with the footer FLAGGED, pipeline exits 3
# ---------------------------------------------------------------------------

def test_check2_missing_footer_row_red_quarantines_and_pipeline_exits_3(
    tmp_path, monkeypatch, capsys
):
    """A dropped footer-component row is a RED structural quarantine failure.

    An absent footer label is a write defect, so the re-sum cannot run. Under
    loud-quarantine (Derick's decision) the engine no longer refuses to deliver —
    it DELIVERS the file with the footer FLAGGED and exits with the distinct code
    3. First proves the RED POST_WRITE_TIEOUT_FAILURE fires on a direct injection,
    then proves the same defect delivers + exits 3 end-to-end (the file exists).
    """
    import json

    from src import pipeline as pipeline_mod

    # --- Part A: direct injection — assert the RED code fires ---
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)
    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    gl_row = _find_footer_row(ws, _GL_KEY)
    ws.delete_rows(gl_row, 1)  # drop the GL_INSURANCE label/row from col A
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "A missing footer-component row must produce a tie-out failure."
    assert any(
        "Footer arithmetic CANNOT be re-checked" in f.message
        and "GL insurance" in f.message
        for f in failures
    )
    assert all(f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE for f in failures)
    assert all(f.status == AuditStatus.RED for f in failures)

    # --- Part B: end-to-end — the same defect hard-stops the pipeline (exit 2) ---
    interim = tmp_path / "interim"
    interim.mkdir()
    bid_json = {
        "contractor_name": "Apex Restoration",
        "form_type": "FALKE_STANDARD",
        "bid_document_input_type": "DIGITAL_NATIVE",
        "extraction_confidence": "HIGH",
        "divisions": [{
            "csi_code": "DIV 03 00 00",
            "division_name": "Concrete",
            "cost_structure": "LUMP_SUM",
            "division_subtotal": "400000",
            "classification_source": "CONTRACTOR_NATIVE",
            "line_items": [],
        }],
        "footer": {
            "construction_cost_subtotal": "400000",
            "general_liability_insurance": "5000",
            "builders_risk_insurance": "5000",
            "gc_fee": "40000",
            "grand_total": "450000",
            "grand_total_confidence": "LOW",
        },
        "qualifications": {},
    }
    (interim / "apex.json").write_text(json.dumps(bid_json))

    config = tmp_path / "project.yaml"
    config.write_text(
        "project_name: Harbor View Tower\n"
        "project_address: 100 Test Ave, Test City FL 00000\n"
        "gross_sf: 10000\n"
        "sf_basis_label: GSF\n"
    )
    pipe_out = tmp_path / "out.xlsx"

    real_write = pipeline_mod.write_matrix

    def deleting_write(*args, **kwargs):
        result = real_write(*args, **kwargs)
        out_path = kwargs.get("output_path") or args[1]
        w = openpyxl.load_workbook(out_path)
        s = w["Bid_Form"]
        s.delete_rows(_find_footer_row(s, _GL_KEY), 1)  # drop a footer label/row
        w.save(out_path)
        return result

    monkeypatch.setattr(pipeline_mod, "write_matrix", deleting_write)

    with pytest.raises(SystemExit) as exc:
        pipeline_mod.run_pipeline(
            interim_dir=interim,
            out_path=pipe_out,
            project_config=config,
            sf_basis=10000.0,
        )
    assert exc.value.code == 3, (
        "Pipeline must exit(3) — delivered with verification failures — on a "
        "missing footer row (no longer exit 2 / refuse)."
    )
    # The file IS delivered (loud quarantine), not withheld.
    assert pipe_out.exists(), "Quarantine must DELIVER the file, not withhold it."
    captured = capsys.readouterr().out
    assert "POST-WRITE TIE-OUT FAILED" in captured
    assert "DELIVERED WITH" in captured
    assert "refusing to deliver" not in captured.lower()
    # A structural failure can't be cleanly enumerated → "one or more" fallback.
    wb = openpyxl.load_workbook(pipe_out)
    banner = wb["Bid_Form"].cell(row=2, column=1).value
    assert "one or more figures" in banner


# ---------------------------------------------------------------------------
# CHECK 3 — move a division subtotal to a wrong value (row/column mis-write)
# ---------------------------------------------------------------------------

def test_check3_division_subtotal_miswrite_fires(tmp_path):
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    sub_rows = _find_subtotal_rows(ws)
    concrete_row = sub_rows["DIV 03 00 00"]
    # Corrupt the Concrete subtotal on col C (should be 400k) to a wrong number.
    ws.cell(row=concrete_row, column=_col_start(0)).value = 123456.0
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "A mis-written division subtotal must produce a failure."
    assert any(
        "Division subtotal tie-out FAILED" in f.message
        and f.division_csi == "DIV 03 00 00"
        for f in failures
    )
    assert all(f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE for f in failures)


# ---------------------------------------------------------------------------
# CHECK 4 — drop an AUDIT-sheet row (silent loss of a board-facing flag)
# ---------------------------------------------------------------------------

def test_check4_dropped_audit_row_fires(tmp_path):
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["AUDIT"]
    # Delete the first data row (row 5) — simulates truncation/overflow loss.
    ws.delete_rows(5, 1)
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "A dropped audit row must produce a tie-out failure."
    assert any("Audit-row count parity FAILED" in f.message for f in failures)
    assert all(f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE for f in failures)


# ---------------------------------------------------------------------------
# PIPELINE GATE — the failure DELIVERS with loud quarantine (exit 3)
# ---------------------------------------------------------------------------

def test_pipeline_quarantines_and_delivers_on_tieout_failure(
    tmp_path, monkeypatch, capsys
):
    """run_pipeline must DELIVER + loud-quarantine + exit(3) on a corrupt write.

    We let the real write_matrix run, then corrupt the saved grand-total cell
    before Stage 6b reads it — proving the gate fires end-to-end and that the
    engine now delivers the file with a RED banner, a RED cell mark, and a RED
    AUDIT line (instead of refusing to deliver). One corrupt grand total → an
    enumerable single flagged figure (banner {N} == 1).
    """
    import json

    from src import pipeline as pipeline_mod

    # --- Build a project config + interim dir of one valid bid JSON ---
    interim = tmp_path / "interim"
    interim.mkdir()
    bid_json = {
        "contractor_name": "Apex Restoration",
        "form_type": "FALKE_STANDARD",
        "bid_document_input_type": "DIGITAL_NATIVE",
        "extraction_confidence": "HIGH",
        "divisions": [
            {
                "csi_code": "DIV 03 00 00",
                "division_name": "Concrete",
                "cost_structure": "LUMP_SUM",
                "division_subtotal": "400000",
                "classification_source": "CONTRACTOR_NATIVE",
                "line_items": [],
            }
        ],
        "footer": {
            "construction_cost_subtotal": "400000",
            "general_liability_insurance": "5000",
            "builders_risk_insurance": "5000",
            "gc_fee": "40000",
            "grand_total": "450000",
            "grand_total_confidence": "LOW",
        },
        "qualifications": {},
    }
    (interim / "apex.json").write_text(json.dumps(bid_json))

    config = tmp_path / "project.yaml"
    config.write_text(
        "project_name: Harbor View Tower\n"
        "project_address: 100 Test Ave, Test City FL 00000\n"
        "gross_sf: 10000\n"
        "sf_basis_label: GSF\n"
    )
    out = tmp_path / "out.xlsx"

    # Wrap write_matrix so that AFTER the real save we corrupt the grand total.
    real_write = pipeline_mod.write_matrix

    def corrupting_write(*args, **kwargs):
        result = real_write(*args, **kwargs)
        out_path = kwargs.get("output_path") or args[1]
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Bid_Form"]
        gt_row = _find_footer_row(ws, _GRAND_TOTAL_KEY)
        ws.cell(row=gt_row, column=_col_start(0)).value = 1.0  # obviously wrong
        wb.save(out_path)
        return result

    monkeypatch.setattr(pipeline_mod, "write_matrix", corrupting_write)

    with pytest.raises(SystemExit) as exc:
        pipeline_mod.run_pipeline(
            interim_dir=interim,
            out_path=out,
            project_config=config,
            sf_basis=10000.0,
        )
    assert exc.value.code == 3, (
        "Pipeline must exit(3) — delivered with verification failures — on a "
        "post-write tie-out failure (no longer exit 2 / refuse)."
    )
    captured = capsys.readouterr().out
    assert "POST-WRITE TIE-OUT FAILED" in captured
    assert "DELIVERED WITH" in captured
    assert "refusing to deliver" not in captured.lower()

    # The file IS delivered and loud-quarantined end-to-end.
    assert out.exists(), "Quarantine must DELIVER the file, not withhold it."
    wb = openpyxl.load_workbook(out)

    # RED banner on BOTH board-facing sheets, with the enumerable count {N} == 1.
    for sheet in ("Bid_Form", "Leveled_Normalized"):
        ws = wb[sheet]
        assert "AUTOMATED CHECK FAILED" in str(ws.cell(row=1, column=1).value), (
            f"{sheet} must carry the RED quarantine banner on row 1."
        )
        line2 = str(ws.cell(row=2, column=1).value)
        assert "1 figure on this sheet does not" in line2, (
            f"{sheet} banner must enumerate the single flagged figure ({{N}}==1, "
            f"singular)."
        )

    # The failing GRAND_TOTAL cell carries the RED fill + verify-against-source
    # comment (Bid_Form, the bid that was corrupted — single bidder → col D=4).
    ws = wb["Bid_Form"]
    gt_row = _find_footer_row(ws, _GRAND_TOTAL_KEY)
    cell = ws.cell(row=gt_row, column=_col_start(0))
    assert cell.comment is not None, "Failing GRAND_TOTAL cell must carry a comment."
    assert "does not reconcile to source" in cell.comment.text
    assert cell.fill.patternType is not None, "Failing cell must be RED-filled."

    # The AUDIT sheet carries the reworded RED message + the QUARANTINE line.
    wa = wb["AUDIT"]
    audit_text = "\n".join(
        str(c.value) for row in wa.iter_rows() for c in row if c.value
    )
    assert "Delivered with this figure FLAGGED" in audit_text, (
        "AUDIT must carry the reworded 'Delivered with this figure FLAGGED' text."
    )
    assert "refusing to deliver" not in audit_text.lower()
    assert "QUARANTINE:" in audit_text, "AUDIT must carry the QUARANTINE summary line."


def test_pipeline_passes_clean_run(tmp_path):
    """A clean run must reach Stage 7 without raising (gate does not false-fire)."""
    import json

    from src import pipeline as pipeline_mod

    interim = tmp_path / "interim"
    interim.mkdir()
    for nm, gt in (("apex", "450000"), ("beacon", "460000")):
        bid_json = {
            "contractor_name": f"{nm.title()} Restoration",
            "form_type": "FALKE_STANDARD",
            "bid_document_input_type": "DIGITAL_NATIVE",
            "extraction_confidence": "HIGH",
            "divisions": [{
                "csi_code": "DIV 03 00 00",
                "division_name": "Concrete",
                "cost_structure": "LUMP_SUM",
                "division_subtotal": "360000",
                "classification_source": "CONTRACTOR_NATIVE",
                "line_items": [],
            }],
            "footer": {
                "construction_cost_subtotal": "360000",
                "general_liability_insurance": "5000",
                "builders_risk_insurance": "5000",
                "gc_fee": "20000",
                "grand_total": "390000",
                "grand_total_confidence": "LOW",
            },
            "qualifications": {},
        }
        (interim / f"{nm}.json").write_text(json.dumps(bid_json))

    config = tmp_path / "project.yaml"
    config.write_text(
        "project_name: Harbor View Tower\n"
        "project_address: 100 Test Ave, Test City FL 00000\n"
        "gross_sf: 10000\n"
        "sf_basis_label: GSF\n"
    )
    out = tmp_path / "out.xlsx"

    # Should complete without SystemExit.
    pipeline_mod.run_pipeline(
        interim_dir=interim,
        out_path=out,
        project_config=config,
        sf_basis=10000.0,
    )
    assert out.exists()


# ---------------------------------------------------------------------------
# INSURANCE-FOLDED FOOTER — insurance folded into other_fees, OH&P broken out, no GL/BR
# ---------------------------------------------------------------------------
#
# Regression for a real-world bug: a recurring firm states NO GL / Builders Risk
# split — its insurance lands in `other_fees_subtotal` and it breaks out
# `overhead_and_profit` separately. The pre-fix Stage 6b summed only
# construction + GL + BR + GC fee, so it OMITTED OH&P + other-fees and
# false-failed that bidder by ~$400k even though the footer ties out cleanly.
# The synthetic fixtures missed this because none carried OH&P + other-fees.

def _build_insurance_folded_field():
    """One insurance-folded bidder + one ordinary bidder, both tie to grand total.

    Folded: 2,700,000 construction + 190,000 GC fee + 300,000 OH&P +
    110,000 other-fees (insurance) = 3,300,000 grand total.
    GL and BR are absent (None) — this firm does not split insurance.
    (Fully synthetic round figures; no client data.)
    """
    folded = _doc("Summit Builders Group Inc", [
        _div("DIV 01 00 00", "General Requirements", Decimal("700000")),
        _div("DIV 03 00 00", "Concrete", Decimal("2000000")),
    ], _footer(
        Decimal("2700000"), None, None, Decimal("190000"),
        Decimal("3300000"),
        overhead_and_profit=Decimal("300000"),
        other_fees_subtotal=Decimal("110000"),
    ))
    # An ordinary GL/BR-splitting bidder so the field has ≥2 bids.
    beacon = _doc("Beacon Builders", [
        _div("DIV 01 00 00", "General Requirements", Decimal("60000")),
        _div("DIV 03 00 00", "Concrete", Decimal("360000")),
    ], _footer(Decimal("420000"), Decimal("4000"), Decimal("4000"),
               Decimal("32000"), Decimal("460000")))
    return compute_cross_bid_stats([normalize_bid(folded), normalize_bid(beacon)])


def test_insurance_folded_footer_ties_out_and_audit_passes(tmp_path):
    """An insurance-folded footer (OH&P + other-fees, no GL/BR) must tie out at
    Stage 6b AND not raise a FOOTER_DISCREPANCY — it would have false-failed pre-fix."""
    bids = _build_insurance_folded_field()
    out, audit_items = _write(tmp_path, bids)

    # Stage 6b: no post-write tie-out failure.
    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures == [], (
        "An insurance-folded footer (insurance in other_fees, OH&P broken out, no "
        "GL/BR) must tie out with zero Stage-6b failures; got: "
        + "; ".join(f.message for f in failures)
    )

    # Pre-write audit: the folded bidder must NOT carry a FOOTER_DISCREPANCY.
    folded_footer_reds = [
        a for a in audit_items
        if a.contractor_name == "Summit Builders Group Inc"
        and a.code == AuditCode.FOOTER_DISCREPANCY
    ]
    assert folded_footer_reds == [], (
        "The folded footer composes to its grand total (construction + GC fee + "
        "OH&P + other-fees) — audit must not raise FOOTER_DISCREPANCY."
    )

    # The written footer must RENDER OH&P and Other Fees as labeled rows that
    # compose to the grand total (board visibility).
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Bid_Form"]
    folded_col = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=5, column=col).value == "Summit Builders Group Inc":
            folded_col = col
            break
    assert folded_col is not None

    def _component(key):
        r = _find_footer_row(ws, key)
        assert r is not None, f"footer row {key} must be rendered"
        return Decimal(str(ws.cell(row=r, column=folded_col).value or 0))

    composed = (
        _component("CONSTRUCTION_SUBTOTAL")
        + _component("GL_INSURANCE")
        + _component("BUILDERS_RISK")
        + _component("GC_FEE")
        + _component("OVERHEAD_PROFIT")
        + _component("OTHER_FEES")
    )
    grand = _component("GRAND_TOTAL")
    assert abs(composed - grand) <= Decimal("1"), (
        f"Rendered footer components must compose to the grand total: "
        f"{composed} vs {grand}"
    )
    assert _component("OVERHEAD_PROFIT") == Decimal("300000")
    assert _component("OTHER_FEES") == Decimal("110000")


# ---------------------------------------------------------------------------
# BOND-INCLUSIVE FOOTER — bond is an additive grand-total component (Marvin)
# ---------------------------------------------------------------------------
#
# Regression for the bond split-brain bug (the Atlas case): a bonded bid states
# its bond INSIDE the grand total. Pre-fix, audit.py composed bond into the grand
# total (GREEN at FOOTER_DISCREPANCY) but Stage 6b's local component tuple OMITTED
# bond, so it re-summed short by exactly the bond amount and false-failed RED
# (GREEN-at-audit / RED-at-reconcile split-brain). The two gates now derive their
# component set from the SAME single source of truth (GRAND_TOTAL_COMPONENT_KEYS,
# bond included) and therefore cannot disagree.

def _build_bonded_field():
    """One bonded bidder (Atlas) whose stated grand total INCLUDES bond, plus an
    ordinary bidder so the field has >=2 bids.

    Atlas: 997,000 construction + 139,700 GC fee + 38,000 GL + 21,000 Builders
    Risk + 18,600 bond = 1,214,300 grand total (bond inside the grand total).
    """
    atlas = _doc("Atlas Restoration", [
        _div("DIV 01 00 00", "General Requirements", Decimal("197000")),
        _div("DIV 03 00 00", "Concrete", Decimal("800000")),
    ], _footer(
        Decimal("997000"), Decimal("38000"), Decimal("21000"),
        Decimal("139700"), Decimal("1214300"),
        bond=Decimal("18600"),
    ))
    beacon = _doc("Beacon Builders", [
        _div("DIV 01 00 00", "General Requirements", Decimal("60000")),
        _div("DIV 03 00 00", "Concrete", Decimal("360000")),
    ], _footer(Decimal("420000"), Decimal("4000"), Decimal("4000"),
               Decimal("32000"), Decimal("460000")))
    return compute_cross_bid_stats([normalize_bid(atlas), normalize_bid(beacon)])


def test_bonded_grand_total_ties_out_at_stage_6b(tmp_path):
    """A bond-inclusive grand total (bond rolled INTO the total) must tie out at
    Stage 6b with NO POST_WRITE_TIEOUT_FAILURE — it false-RED'd by exactly the
    bond amount pre-fix, when Stage 6b's component set omitted bond."""
    bids = _build_bonded_field()
    out, audit_items = _write(tmp_path, bids)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures == [], (
        "A bonded bid whose grand total INCLUDES bond must tie out with zero "
        "Stage-6b failures (bond is an additive grand-total component); got: "
        + "; ".join(f.message for f in failures)
    )
    # And specifically: no tie-out failure attributable to the bonded bidder.
    assert not any(
        f.contractor_name == "Atlas Restoration"
        and f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE
        for f in failures
    )


def test_bonded_bid_audit_and_reconcile_agree_no_split_brain(tmp_path):
    """The exact split-brain regression guard: a bond-inclusive bid that passes
    the pre-write audit FOOTER_DISCREPANCY check GREEN must ALSO pass the Stage-6b
    tie-out GREEN. The two gates derive bond treatment from the same single source
    of truth, so they cannot disagree (no GREEN-at-audit / RED-at-reconcile)."""
    bids = _build_bonded_field()
    out, audit_items = _write(tmp_path, bids)

    # Gate 1 — pre-write audit: the bonded bid must NOT carry a FOOTER_DISCREPANCY
    # (bond is composed into the grand total → footer reconciles GREEN).
    atlas_footer_reds = [
        a for a in audit_items
        if a.contractor_name == "Atlas Restoration"
        and a.code == AuditCode.FOOTER_DISCREPANCY
    ]
    assert atlas_footer_reds == [], (
        "A bonded bid whose stated grand total includes bond reconciles at the "
        "audit footer check — it must not raise FOOTER_DISCREPANCY."
    )

    # Gate 2 — Stage 6b read-back: the SAME bid must also tie out GREEN. If gate 1
    # is GREEN and gate 2 is RED, the two gates used a different bond treatment —
    # the split-brain bug. They must agree.
    failures = reconcile_written_matrix(out, bids, len(audit_items))
    atlas_tieout_reds = [
        f for f in failures
        if f.contractor_name == "Atlas Restoration"
        and f.code == AuditCode.POST_WRITE_TIEOUT_FAILURE
    ]
    assert atlas_tieout_reds == [], (
        "Split-brain guard: a bonded bid that passes the audit footer check GREEN "
        "must also pass the Stage-6b tie-out GREEN — the two gates share one bond "
        "treatment and cannot disagree; got: "
        + "; ".join(f.message for f in atlas_tieout_reds)
    )


def test_bonded_footer_renders_bond_inside_additive_fees_subtotal(tmp_path):
    """write_matrix renders Bond as an ADDITIVE footer row that rolls into Fees
    Subtotal, and construction + Fees Subtotal == GRAND TOTAL for a bonded bid
    (bond inside the total, not a bottom 'Bond (Alternate)' memo)."""
    bids = _build_bonded_field()
    out, _audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Bid_Form"]
    atlas_col = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=5, column=col).value == "Atlas Restoration":
            atlas_col = col
            break
    assert atlas_col is not None

    def _component(key):
        r = _find_footer_row(ws, key)
        assert r is not None, f"footer row {key} must be rendered"
        return Decimal(str(ws.cell(row=r, column=atlas_col).value or 0))

    # Bond is rendered as the stated bond on its additive row.
    assert _component("BOND") == Decimal("18600")
    # Fees Subtotal includes bond (it is the sum of all additive fee components).
    assert _component("FEES_SUBTOTAL") == (
        Decimal("38000") + Decimal("21000") + Decimal("139700") + Decimal("18600")
    )
    # And construction + Fees Subtotal composes to the GRAND TOTAL.
    composed = _component("CONSTRUCTION_SUBTOTAL") + _component("FEES_SUBTOTAL")
    grand = _component("GRAND_TOTAL")
    assert abs(composed - grand) <= Decimal("1"), (
        f"construction + Fees Subtotal must equal the grand total for a bonded "
        f"bid: {composed} vs {grand}"
    )


# ---------------------------------------------------------------------------
# Structure-first footer composition (Marvin's fiduciary condition)
# ---------------------------------------------------------------------------
# other_fees_subtotal is classified ADDITIVE vs MEMO from the fee-row STRUCTURE,
# never by back-solving from the contractor's stated grand total. When the
# stated total is itself wrong, composition must NOT silently mirror it — the
# audit must raise FOOTER_DISCREPANCY (RED).

from src.normalized_models import (  # noqa: E402
    GRAND_TOTAL_COMPONENT_KEYS,
    grand_total_component_amounts,
)


def test_component_keys_constant_covers_every_decider_emitted_key():
    """Drift-prevention invariant: the single-source-of-truth constant
    GRAND_TOTAL_COMPONENT_KEYS must cover EVERY component the value-decider
    (grand_total_component_amounts) can emit. This is the guard for the exact bug
    class we just fixed: bond drifted because the decider emitted a component the
    Stage-6b component set did not list. A future new additive component that is
    added to the decider but forgotten here would break this superset check —
    loudly — instead of silently re-summing short the way bond did.

    Build a footer that populates every additive component (construction, GL,
    Builders Risk, GC fee, Overhead & Profit, Other Fees, Bond) so the decider
    emits its full key set, then assert the constant is a superset of those keys.
    """
    footer = normalize_bid(_doc("All Components Co", [
        _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        _div("DIV 03 00 00", "Concrete", Decimal("400000")),
    ], _footer(
        Decimal("450000"), Decimal("5000"), Decimal("5000"),
        Decimal("40000"), Decimal("528000"),
        overhead_and_profit=Decimal("20000"),
        other_fees_subtotal=Decimal("10000"),
        bond=Decimal("8000"),
    ))).footer
    emitted = set(grand_total_component_amounts(footer).keys())
    assert set(GRAND_TOTAL_COMPONENT_KEYS) >= emitted, (
        "GRAND_TOTAL_COMPONENT_KEYS must cover every key the decider can emit; "
        f"decider emitted {sorted(emitted - set(GRAND_TOTAL_COMPONENT_KEYS))} "
        "that the single-source-of-truth constant does not list — a new component "
        "drifted exactly the way bond did."
    )


def _footer_reds(name, bid):
    items = audit_bids([bid])
    return [
        a for a in items
        if a.contractor_name == name and a.code == AuditCode.FOOTER_DISCREPANCY
    ]


def test_other_fees_additive_when_insurance_absent():
    """Insurance-folded pattern: GL/BR blank, insurance lives in other_fees ⇒
    ADDITIVE, decided from structure alone (independent of the stated grand total)."""
    folded = normalize_bid(_doc("Summit Group", [
        _div("DIV 03 00 00", "Concrete", Decimal("2700000")),
    ], _footer(
        Decimal("2700000"), None, None, Decimal("190000"),
        Decimal("3300000"),
        overhead_and_profit=Decimal("300000"),
        other_fees_subtotal=Decimal("110000"),
    )))
    comps = grand_total_component_amounts(folded.footer)
    assert comps.get("OTHER_FEES") == Decimal("110000"), (
        "other_fees must be ADDITIVE when GL/BR are absent (insurance can only "
        "live in other_fees)."
    )
    # And it ties out → no FOOTER_DISCREPANCY.
    assert _footer_reds("Summit Group", folded) == []


def test_other_fees_memo_when_it_rolls_up_already_counted_fees():
    """other_fees that equals (GL+BR+GC) is a 'Total Fees' MEMO line ⇒ EXCLUDE,
    even if including it would happen to match a stated grand total."""
    # GL 5k + BR 5k + GC 40k = 50k. other_fees = 50k (memo of those rows).
    # Construction 450k + 50k fees = 500k grand total.
    bid = normalize_bid(_doc("Memo Co", [
        _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        _div("DIV 03 00 00", "Concrete", Decimal("400000")),
    ], _footer(
        Decimal("450000"), Decimal("5000"), Decimal("5000"),
        Decimal("40000"), Decimal("500000"),
        other_fees_subtotal=Decimal("50000"),
    )))
    comps = grand_total_component_amounts(bid.footer)
    assert "OTHER_FEES" not in comps, (
        "other_fees equal to (GL+BR+GC) is a memo roll-up — must be EXCLUDED, "
        "not double-counted."
    )
    assert _footer_reds("Memo Co", bid) == []


def test_wrong_stated_total_does_not_silently_compose__raises_footer_discrepancy():
    """The fiduciary case: a footer whose STRUCTURE composes to 500k but whose
    contractor-stated grand total is a (deliberately wrong) 507,473. Structure
    must NOT be fudged to match the wrong number — audit raises FOOTER_DISCREPANCY
    (RED) and the composition stays the structure-derived 500k."""
    # Construction 450k + GL 5k + BR 5k + GC 40k = 500k by structure.
    # other_fees = 7,473 — NOT a roll-up of already-counted rows, and GL/BR are
    # PRESENT, so it is NOT additive either. Stated GT (507,473) would only
    # reconcile if we back-solved other_fees in; we must refuse.
    bid = normalize_bid(_doc("Acme Restoration", [
        _div("DIV 01 00 00", "General Requirements", Decimal("50000")),
        _div("DIV 03 00 00", "Concrete", Decimal("400000")),
    ], _footer(
        Decimal("450000"), Decimal("5000"), Decimal("5000"),
        Decimal("40000"), Decimal("507473"),
        other_fees_subtotal=Decimal("7473"),
    )))
    comps = grand_total_component_amounts(bid.footer)
    assert "OTHER_FEES" not in comps, (
        "Unexplained other_fees with GL/BR present must NOT be pulled in just to "
        "close the gap to a (wrong) stated grand total."
    )
    assert sum(comps.values(), Decimal("0")) == Decimal("500000"), (
        "Composition must stay structure-derived (500k), not bent to the stated "
        "507,473."
    )
    reds = _footer_reds("Acme Restoration", bid)
    assert reds, (
        "Structure does not tie to the stated grand total — audit MUST raise "
        "FOOTER_DISCREPANCY (RED) rather than silently composing to the wrong total."
    )
    assert reds[0].status == AuditStatus.RED


# ---------------------------------------------------------------------------
# LOUD QUARANTINE — banner, in-place cell marks, AUDIT rewording (Marvin's spec)
# ---------------------------------------------------------------------------

from src.write_matrix import apply_quarantine  # noqa: E402


def _corrupt_grand_total(out: Path, col_index: int = 0, value: float = 1.0) -> None:
    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    gt_row = _find_footer_row(ws, _GRAND_TOTAL_KEY)
    ws.cell(row=gt_row, column=_col_start(col_index)).value = value
    wb.save(out)


def test_clean_run_renders_no_quarantine_artifacts(tmp_path):
    """A matrix that ties out gets NO banner, NO cell marks, NO AUDIT quarantine —
    the banner must never cry wolf (Marvin §2)."""
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)
    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures == []  # clean → apply_quarantine is never called

    wb = openpyxl.load_workbook(out)
    for sheet in ("Bid_Form", "Leveled_Normalized"):
        ws = wb[sheet]
        assert "AUTOMATED CHECK FAILED" not in str(ws.cell(row=1, column=1).value)
        assert not any(c.comment for row in ws.iter_rows() for c in row)
    audit_text = "\n".join(
        str(c.value) for row in wb["AUDIT"].iter_rows() for c in row if c.value
    )
    assert "QUARANTINE" not in audit_text
    assert "POST_WRITE_TIEOUT_FAILURE" not in audit_text


def test_quarantine_banner_grand_total_mark_and_audit_rewording(tmp_path):
    """apply_quarantine stamps the RED banner on BOTH sheets ({N}==1 singular), a
    RED cell mark + verify comment on the failing GRAND_TOTAL, and the reworded
    AUDIT rows + QUARANTINE line — with no 'refusing to deliver' anywhere."""
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)
    _corrupt_grand_total(out, col_index=0, value=1.0)  # corrupt the low bidder GT

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures, "Corrupt GT must produce tie-out failures."
    n = apply_quarantine(out, failures, bids, leveled_bids=bids)
    assert n == 1, "A single corrupt grand total is one enumerable flagged figure."

    wb = openpyxl.load_workbook(out)

    # Banner on both board-facing sheets, line 1 RED + bold, {N}==1 singular.
    for sheet in ("Bid_Form", "Leveled_Normalized"):
        ws = wb[sheet]
        c1 = ws.cell(row=1, column=1)
        assert "AUTOMATED CHECK FAILED" in str(c1.value)
        assert c1.font.bold is True
        assert c1.fill.patternType is not None  # RED fill present
        assert "1 figure on this sheet does not" in str(ws.cell(row=2, column=1).value)
        assert "marked in their cells" in str(ws.cell(row=3, column=1).value)

    # On the leveled sheet the RED quarantine banner sits ABOVE the yellow
    # normalization banner (which shifted down to rows 4–5).
    wsl = wb["Leveled_Normalized"]
    assert "ESTIMATOR-NORMALIZED VIEW" in str(wsl.cell(row=4, column=1).value)

    # RED cell mark + verify comment on the failing GRAND_TOTAL (Bid_Form, low bid col D).
    ws = wb["Bid_Form"]
    gt_row = _find_footer_row(ws, _GRAND_TOTAL_KEY)
    cell = ws.cell(row=gt_row, column=_col_start(0))
    assert cell.fill.patternType is not None
    assert cell.comment is not None
    assert "does not reconcile to source — verify" in cell.comment.text
    assert "Check this figure against" in cell.comment.text

    # AUDIT: reworded RED row present, QUARANTINE summary line present, no refusal.
    audit_text = "\n".join(
        str(c.value) for row in wb["AUDIT"].iter_rows() for c in row if c.value
    )
    assert "POST_WRITE_TIEOUT_FAILURE" in audit_text
    assert "Delivered with this figure FLAGGED" in audit_text
    assert "QUARANTINE:" in audit_text
    assert "1 figure(s) failed the tool's self-check" in audit_text
    assert "refusing to deliver" not in audit_text.lower()


def test_quarantine_banner_is_merged_full_width_with_row_height(tmp_path):
    """The RED quarantine banner must render as a MERGED full-width block on BOTH
    sheets — col A through the sheet's last used column — with an explicit row
    height, so the long L2/L3 paragraphs flow horizontally instead of stacking in
    a tall, narrow column-A block (the v0.2.0 banner-formatting defect). v0.3.0:
    the two sheets have DIFFERENT widths (mirror stride 3; leveled stride 5 +
    benchmark block)."""
    from openpyxl.utils import get_column_letter

    from src.write_matrix import _lev_last_col

    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)
    _corrupt_grand_total(out, col_index=0, value=1.0)
    failures = reconcile_written_matrix(out, bids, len(audit_items))
    apply_quarantine(out, failures, bids, leveled_bids=bids)

    # Last populated column per sheet, computed (never hardcoded).
    last_col_of = {
        "Bid_Form": get_column_letter(_col_start(len(bids) - 1) + 2),
        "Leveled_Normalized": get_column_letter(_lev_last_col(len(bids))),
    }

    wb = openpyxl.load_workbook(out)
    for sheet, last_col in last_col_of.items():
        ws = wb[sheet]
        ranges = {str(r) for r in ws.merged_cells.ranges}
        for row in (1, 2, 3):
            assert f"A{row}:{last_col}{row}" in ranges, (
                f"{sheet} banner row {row} must be merged A→{last_col}, "
                f"got merges {sorted(ranges)}"
            )
            height = ws.row_dimensions[row].height
            assert height and height >= 15.0, (
                f"{sheet} banner row {row} needs an explicit row height; got {height}"
            )
            assert ws.cell(row=row, column=1).alignment.wrap_text is True

    # When BOTH banners fire on the leveled sheet, the GRAY normalization banner
    # (shifted to rows 4–5 below the RED block) must ALSO stay merged full-width
    # + sized (the quarantine pass shifts merges/heights with the values).
    wsl = wb["Leveled_Normalized"]
    lev_last = last_col_of["Leveled_Normalized"]
    gray_ranges = {str(r) for r in wsl.merged_cells.ranges}
    assert "ESTIMATOR-NORMALIZED VIEW" in str(wsl.cell(row=4, column=1).value)
    for row in (4, 5):
        assert f"A{row}:{lev_last}{row}" in gray_ranges, (
            f"Leveled gray banner row {row} must stay merged A→{lev_last} "
            f"after the quarantine row-shift, got {sorted(gray_ranges)}"
        )
        h = wsl.row_dimensions[row].height
        assert h and h >= 15.0, f"Leveled gray banner row {row} lost its height: {h}"


def test_leveled_gray_banner_is_merged_full_width_with_row_height(tmp_path):
    """The GRAY estimator-normalized banner on Leveled_Normalized (always
    rendered, no quarantine needed; restyled from yellow to the neutral house
    gray per rules-spec A1 in v0.3.0) must be a MERGED full-width block — col A
    through the sheet's last used column — with an explicit row height, so its
    long text flows horizontally instead of stacking in a tall, narrow column-A
    block (the v0.2.1 second-banner defect)."""
    from openpyxl.utils import get_column_letter

    from src.format_falke import FALKE_GRAY
    from src.write_matrix import _lev_last_col

    bids = _build_field()
    out, _audit_items = _write(tmp_path, bids)  # clean run — no quarantine

    last_col = get_column_letter(_lev_last_col(len(bids)))

    wb = openpyxl.load_workbook(out)
    ws = wb["Leveled_Normalized"]
    # The gray banner sits in rows 1–2 (no RED banner above it on a clean run).
    assert "ESTIMATOR-NORMALIZED VIEW" in str(ws.cell(row=1, column=1).value)
    rgb = ws.cell(row=1, column=1).fill.fgColor.rgb
    assert isinstance(rgb, str) and rgb[-6:].upper() == FALKE_GRAY, (
        f"Leveled banner must carry the neutral house gray (A1), got {rgb}"
    )
    ranges = {str(r) for r in ws.merged_cells.ranges}
    for row in (1, 2):
        assert f"A{row}:{last_col}{row}" in ranges, (
            f"Leveled gray banner row {row} must be merged A→{last_col}, "
            f"got merges {sorted(ranges)}"
        )
        height = ws.row_dimensions[row].height
        assert height and height >= 15.0, (
            f"Leveled gray banner row {row} needs an explicit row height; got {height}"
        )
        assert ws.cell(row=row, column=1).alignment.wrap_text is True


def test_quarantine_division_subtotal_mark(tmp_path):
    """A division-subtotal tie-out failure marks the (contractor, division)
    SUBTOTAL cell — not the grand total (Marvin §3)."""
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    concrete_row = _find_subtotal_rows(ws)["DIV 03 00 00"]
    ws.cell(row=concrete_row, column=_col_start(0)).value = 123456.0
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert any(f.division_csi == "DIV 03 00 00" for f in failures)
    apply_quarantine(out, failures, bids, leveled_bids=bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    concrete_row = _find_subtotal_rows(ws)["DIV 03 00 00"]
    sub_cell = ws.cell(row=concrete_row, column=_col_start(0))
    assert sub_cell.comment is not None, "Failing division SUBTOTAL cell must be marked."
    assert "does not reconcile to source" in sub_cell.comment.text


def test_quarantine_structural_failure_uses_one_or_more_fallback(tmp_path):
    """A STRUCTURAL failure (dropped footer row) can't be cleanly enumerated, so
    the banner falls back to 'one or more figures' and never under-counts (§2)."""
    bids = _build_field()
    out, audit_items = _write(tmp_path, bids)

    wb = openpyxl.load_workbook(out)
    ws = wb["Bid_Form"]
    ws.delete_rows(_find_footer_row(ws, _GL_KEY), 1)  # drop a footer component row
    wb.save(out)

    failures = reconcile_written_matrix(out, bids, len(audit_items))
    assert failures
    n = apply_quarantine(out, failures, bids, leveled_bids=bids)
    assert n == -1, "Structural failure → non-enumerable count sentinel (-1)."

    wb = openpyxl.load_workbook(out)
    line2 = str(wb["Bid_Form"].cell(row=2, column=1).value)
    assert "one or more figures on this sheet do not" in line2
