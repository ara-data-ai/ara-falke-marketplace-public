"""
FALKE Matrix — Structured intake tests (.xlsx / .csv, v0.3.0)
==============================================================
Proves the deterministic structured intake (intake_structured.py):

  * FALKE_FORM layout — a synthetic Falke standard bid-form xlsx parses to a
    correct BidDocument (bidder block, divisions, explicit-zero vs blank,
    footer, alternates/bond, qualifications text, HIGH confidence when the
    footer composes).
  * ROW_SCHEMA layout — the §2 row schema (csv) parses with the full
    classification vocabulary mapped onto BidDocument flags.
  * Malformed / unrecognized files fail LOUDLY (IntakeError naming the file);
    a multi-bidder matrix and a multi-price-column file are rejected.
  * run_structured_intake writes validated {slug}.json files into the interim
    dir, and the pipeline consumes them end-to-end (xlsx → matrix, exit 0).

Run from the engine root:
    python3 -m pytest tests/test_intake_structured.py -v
"""

from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from src.intake_structured import (
    IntakeError,
    parse_structured_bid,
    run_structured_intake,
)
from src.models import BidDocument, CostStructure, FormType


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

def _write_falke_form(path: Path, name="Acme Restoration Inc") -> Path:
    """Synthetic Falke standard bid form (FEB 26 'Ready to Use' structure)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B5"] = "BIDs Comparison Matrix"
    ws["B6"] = "Lobby Renovation Project"
    ws["B7"] = "1 Test Ave, Test City FL"
    ws["E8"] = name
    ws["E9"] = "Lobby Renovation Project"
    ws["C10"] = "TOTAL GSF"
    ws["E10"] = 13622
    # Header row (FEB 26 titles).
    ws["B12"] = "CSI"
    ws["C12"] = "BUILDING SYSTEM"
    ws["E12"] = "COST"
    ws["F12"] = "COST \nSUBTOTALS"
    ws["G12"] = "$/SF"
    ws["H12"] = "$/SXFX \nSUBTOTALS"
    # DIV 01.
    ws["B14"] = "DIV 01 00 00"
    ws["C14"] = "General Requirements"
    ws["C15"] = "Project Management"
    ws["E15"] = 100000
    ws["C16"] = "Final Cleaning"
    ws["E16"] = 0                      # literal zero typed → explicit zero
    ws["C17"] = "Temporary Toilets"    # blank price → missing (NOT zero)
    ws["C18"] = "GENERAL REQUIREMENTS SUBTOTAL"
    ws["F18"] = 100000
    # DIV 03.
    ws["B20"] = "DIV 03 00 00"
    ws["C20"] = "Concrete"
    ws["C21"] = "Concrete repairs"
    ws["E21"] = 50000
    ws["C22"] = "CONCRETE SUBTOTAL"
    ws["F22"] = 50000
    # Footer (FEB 26 labels — incl. the reference's own 'Liabilty' typo).
    ws["C24"] = "CONSTRUCTION COST SUBTOTAL"
    ws["E24"] = 150000
    ws["C25"] = "General Liabilty insurance"
    ws["E25"] = 3000
    ws["C26"] = "Builders Risk Insurance"
    ws["E26"] = 2000
    ws["C27"] = "GC Fee"
    ws["E27"] = 15000
    ws["C28"] = "SUBTOTAL"
    ws["F28"] = 20000
    ws["C29"] = "GRAND TOTAL CONSTRUCTION COST"
    ws["E29"] = 170000
    # Alternates + bond + qualifications text.
    ws["C31"] = "ALTERNATES (Add / Deduct):"
    ws["C32"] = "Bond"
    ws["E32"] = 0
    ws["C33"] = "Deduct: skip painting"
    ws["E33"] = -5000
    ws["C35"] = "Exclusions:"
    ws["E35"] = "No asbestos work"
    ws["E36"] = "No permits included"
    wb.save(path)
    return path


_CSV_ROWS = """Item Code,Category,Description,Quantity,Unit,Acme Builders,Status,Notes
001,Demo,Demolition of lobby,,LS,25000,Base Bid,
002,Concrete,Concrete repair,100,SF,"$12,500.00",,
003,Concrete,Core drill,,,0,Base Bid,
004,Finishes,Painting,,,15000,Allowance,fixture schedule incomplete
005,Finishes,Premium paint upgrade,,,5000,Alternate,
006,Millwork,Lobby reception desk,,,8000,Excluded,
007,MEP,Fire pump,,,-,By Owner,
008,Electrical,Lighting package,,,#REF!,Base Bid,
009,Landscaping,Planters,,,3000,Not Applicable,
010,Concrete,Sidewalk repair,,,7000,Not Comparable,
,,TOTAL,,,75500,,
"""


def _write_csv(path: Path) -> Path:
    path.write_text(_CSV_ROWS, encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# FALKE_FORM layout
# ---------------------------------------------------------------------------

class TestFalkeFormLayout:
    def _doc(self, tmp_path) -> BidDocument:
        return parse_structured_bid(_write_falke_form(tmp_path / "acme.xlsx"))

    def test_bidder_block_and_identity(self, tmp_path):
        doc = self._doc(tmp_path)
        assert doc.contractor_name == "Acme Restoration Inc"
        assert doc.total_gsf == 13622
        assert doc.form_type == FormType.FALKE_STANDARD

    def test_divisions_lines_and_zero_vs_blank(self, tmp_path):
        doc = self._doc(tmp_path)
        assert [d.csi_code for d in doc.divisions] == [
            "DIV 01 00 00", "DIV 03 00 00"
        ]
        d01 = doc.divisions[0]
        assert d01.division_subtotal == Decimal("100000")
        assert d01.cost_structure == CostStructure.ITEMIZED
        by_desc = {i.description: i for i in d01.line_items}
        # Literal 0 typed → explicit zero; blank cell → missing (NOT zero).
        assert by_desc["Final Cleaning"].is_explicit_zero is True
        assert by_desc["Final Cleaning"].amount == Decimal("0")
        assert by_desc["Temporary Toilets"].amount is None
        assert by_desc["Temporary Toilets"].is_explicit_zero is False

    def test_footer_alternates_and_confidence(self, tmp_path):
        doc = self._doc(tmp_path)
        f = doc.footer
        assert f.construction_cost_subtotal == Decimal("150000")
        assert f.general_liability_insurance == Decimal("3000")
        assert f.builders_risk_insurance == Decimal("2000")
        assert f.gc_fee == Decimal("15000")
        assert f.grand_total == Decimal("170000")
        assert f.bond == Decimal("0")
        # 150k + 3k + 2k + 15k + 0 == 170k → HIGH confidence, no flags.
        assert f.grand_total_confidence.value == "HIGH"
        assert f.confidence_flags == []
        assert len(f.alternates) == 1
        assert f.alternates[0].amount == Decimal("-5000")

    def test_qualifications_text_captured(self, tmp_path):
        doc = self._doc(tmp_path)
        assert "No asbestos work" in (doc.qualifications.exclusions or "")
        assert "No permits included" in (doc.qualifications.exclusions or "")

    def test_round_trips_through_biddocument_validation(self, tmp_path):
        doc = self._doc(tmp_path)
        BidDocument.model_validate(json.loads(doc.model_dump_json()))

    def test_multi_bidder_matrix_rejected(self, tmp_path):
        p = _write_falke_form(tmp_path / "matrix.xlsx")
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        ws["I12"] = "COST"
        ws["J12"] = "COST \nSUBTOTALS"  # second bidder group
        wb.save(p)
        with pytest.raises(IntakeError, match="MULTI-BIDDER"):
            parse_structured_bid(p)


# ---------------------------------------------------------------------------
# ROW_SCHEMA layout (§2)
# ---------------------------------------------------------------------------

class TestRowSchemaLayout:
    def _doc(self, tmp_path) -> BidDocument:
        return parse_structured_bid(_write_csv(tmp_path / "acme.csv"))

    def test_bidder_from_price_column_header(self, tmp_path):
        doc = self._doc(tmp_path)
        assert doc.contractor_name == "Acme Builders"
        assert doc.form_type == FormType.CONTRACTOR_OWN

    def test_category_resolution(self, tmp_path):
        doc = self._doc(tmp_path)
        codes = {d.csi_code: d for d in doc.divisions}
        assert "DIV 02 00 00" in codes            # Demo alias
        assert "DIV 03 00 00" in codes            # Concrete name match
        assert "DIV 09 00 00" in codes            # Finishes name match
        assert "DIV 06 00 00" in codes            # Millwork alias
        assert "DIV 26 00 00" in codes            # Electrical name match
        # Unmappable categories pass through VERBATIM (engine flags them).
        assert "MEP" in codes
        assert "Landscaping" in codes
        assert any("MEP" in w for w in doc.extraction_warnings)

    def test_classification_vocabulary_mapping(self, tmp_path):
        doc = self._doc(tmp_path)
        items = {i.description: i
                 for d in doc.divisions for i in d.line_items}
        # R4 coercion: "$12,500.00" → 12500.
        assert items["Concrete repair"].amount == Decimal("12500.00")
        # Literal 0 → explicit zero.
        assert items["Core drill"].is_explicit_zero is True
        # Allowance → flag + basis from notes.
        assert items["Painting"].is_allowance is True
        assert items["Painting"].allowance_basis == "fixture schedule incomplete"
        # Alternate → routed to footer.alternates, not a division line.
        assert "Premium paint upgrade" not in items
        assert doc.footer.alternates[0].description == "Premium paint upgrade"
        # Excluded → flag; excluded amount stays OUT of the division subtotal.
        assert items["Lobby reception desk"].is_excluded is True
        d06 = next(d for d in doc.divisions if d.csi_code == "DIV 06 00 00")
        assert d06.division_subtotal == Decimal("0")
        # By Owner + dash price (Q11: dash = missing, never zero).
        fp = items["Fire pump"]
        assert fp.is_by_owner_others is True
        assert fp.amount is None and fp.is_explicit_zero is False
        # Not Applicable → by-owner family with verbatim preserved.
        pl = items["Planters"]
        assert pl.is_by_owner_others is True
        assert pl.by_others_verbatim == "Not Applicable"
        # Not Comparable → dedicated flag (ENC-2): amount kept + note +
        # warning; fenced out of benchmarks downstream, never deducted.
        sw = items["Sidewalk repair"]
        assert sw.is_not_comparable is True
        assert sw.amount == Decimal("7000")
        assert "Not Comparable" in (sw.notes or "")
        assert any("Not Comparable" in w for w in doc.extraction_warnings)
        # #REF! (R18) → missing + warning.
        lp = items["Lighting package"]
        assert lp.amount is None
        assert any("#REF!" in w for w in doc.extraction_warnings)
        # Qty/Unit preserved in notes (no BidDocument field for them).
        assert "Qty: 100" in (items["Concrete repair"].notes or "")

    def test_totals_derived_and_total_row_skipped(self, tmp_path):
        doc = self._doc(tmp_path)
        # Demo 25000 + Concrete (12500+0+7000) + Finishes 15000 (allowance)
        # + Millwork 0 (excluded out) + MEP/Electrical/Landscaping 0 = 59500.
        assert doc.footer.grand_total == Decimal("59500")
        assert doc.footer.construction_cost_subtotal == Decimal("59500")
        assert doc.footer.grand_total_confidence.value == "LOW"
        assert any("total row" in w.lower() for w in doc.extraction_warnings)
        assert any("DERIVED by summation" in w for w in doc.extraction_warnings)

    def test_round_trips_through_biddocument_validation(self, tmp_path):
        doc = self._doc(tmp_path)
        BidDocument.model_validate(json.loads(doc.model_dump_json()))

    def test_multiple_price_columns_rejected(self, tmp_path):
        p = tmp_path / "two.csv"
        p.write_text(
            "Item Code,Description,Acme,Beta\n001,Demo,100,200\n",
            encoding="utf-8",
        )
        with pytest.raises(IntakeError, match="exactly ONE bidder price"):
            parse_structured_bid(p)


# ---------------------------------------------------------------------------
# Loud failure + directory intake + pipeline end-to-end
# ---------------------------------------------------------------------------

class TestLoudFailureAndPipeline:
    def test_unrecognized_xlsx_layout_fails_loudly(self, tmp_path):
        p = tmp_path / "mystery.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "Some random export"
        wb.save(p)
        with pytest.raises(IntakeError) as exc:
            parse_structured_bid(p)
        assert "mystery.xlsx" in str(exc.value)
        assert "unrecognized spreadsheet layout" in str(exc.value)

    def test_unsupported_extension_fails_loudly(self, tmp_path):
        p = tmp_path / "bid.txt"
        p.write_text("hello")
        with pytest.raises(IntakeError, match="unsupported"):
            parse_structured_bid(p)

    def test_run_structured_intake_writes_slug_json(self, tmp_path):
        _write_falke_form(tmp_path / "acme bid.xlsx")
        bad = tmp_path / "mystery.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "not a bid"
        wb.save(bad)

        successes, failures = run_structured_intake(tmp_path)
        assert [(s[0], s[1]) for s in successes] == [
            ("acme bid.xlsx", "acme_restoration_inc.json")
        ]
        assert len(failures) == 1 and failures[0][0] == "mystery.xlsx"
        raw = json.loads((tmp_path / "acme_restoration_inc.json").read_text())
        BidDocument.model_validate(raw)

    def test_pipeline_end_to_end_from_xlsx(self, tmp_path, capsys):
        """xlsx bid in the interim dir → deterministic intake → full matrix
        (no extraction agent anywhere), tie-out clean."""
        from src import pipeline as pipeline_mod

        interim = tmp_path / "interim"
        interim.mkdir()
        _write_falke_form(interim / "acme.xlsx")
        config = tmp_path / "project.yaml"
        config.write_text(
            "project_name: Harbor View Tower\n"
            "project_address: 100 Test Ave, Test City FL 00000\n"
            "gross_sf: 13622\n"
            "sf_basis_label: GSF\n"
        )
        out = tmp_path / "out.xlsx"
        pipeline_mod.run_pipeline(
            interim_dir=interim,
            out_path=out,
            project_config=config,
            sf_basis=13622.0,
        )
        assert out.exists()
        captured = capsys.readouterr().out
        assert "Structured intake — 1 parsed, 0 failed" in captured
        assert "Tie-out OK" in captured
        wb = openpyxl.load_workbook(out)
        assert {"Bid_Form", "Leveled_Normalized"} <= set(wb.sheetnames)


# ---------------------------------------------------------------------------
# Floyd gate conditions C1 / C2 / C3 (FLOYD-VV-GATE-v030.md)
# ---------------------------------------------------------------------------

class TestFloydC1BatchRobustness:
    def test_malformed_xlsx_and_bad_encoding_csv_do_not_abort_batch(self, tmp_path):
        """C1: a corrupt-zip .xlsx and a non-UTF-8 .csv must each become a
        per-file failure — the good file in the same batch still succeeds."""
        _write_falke_form(tmp_path / "good.xlsx")
        (tmp_path / "corrupt.xlsx").write_bytes(b"this is not a zip archive")
        (tmp_path / "latin1.csv").write_bytes(
            "Item Code,Description,Acme\n001,façade repair,100\n"
            .encode("latin-1")
        )

        successes, failures = run_structured_intake(tmp_path)
        assert [s[0] for s in successes] == ["good.xlsx"], (
            "the good file must survive the batch"
        )
        failed_names = {f[0] for f in failures}
        assert failed_names == {"corrupt.xlsx", "latin1.csv"}
        for _name, msg in failures:
            assert "unreadable/unparseable" in msg or "IntakeError" in msg

    def test_oversized_file_fails_loudly(self, tmp_path, monkeypatch):
        """C1: the size ceiling rejects a runaway/zip-bomb file before any
        parser touches it (ceiling lowered for the test)."""
        import src.intake_structured as intake_mod

        monkeypatch.setattr(intake_mod, "MAX_STRUCTURED_FILE_BYTES", 1024)
        big = tmp_path / "big.csv"
        big.write_text("Item Code,Description,Acme\n" + "x" * 2048)
        with pytest.raises(IntakeError, match="ceiling"):
            parse_structured_bid(big)


class TestFloydC2SlugCollision:
    def test_colliding_bidder_names_never_silently_overwrite(self, tmp_path):
        """C2 (Floyd's exact case): 'A.B.C. Inc' and 'ABC Inc' both slug to
        abc_inc — the second file must FAIL loudly naming both sources, and
        the first bidder's JSON must survive untouched."""
        (tmp_path / "a first.csv").write_text(
            "Item Code,Description,A.B.C. Inc\n001,Demo,100\n")
        (tmp_path / "b second.csv").write_text(
            "Item Code,Description,ABC Inc\n001,Demo,200\n")

        successes, failures = run_structured_intake(tmp_path)
        assert [s for s in successes] == [("a first.csv", "abc_inc.json")]
        assert len(failures) == 1
        fname, msg = failures[0]
        assert fname == "b second.csv"
        assert "a first.csv" in msg and "abc_inc.json" in msg
        # The surviving JSON is the FIRST bidder's.
        raw = json.loads((tmp_path / "abc_inc.json").read_text())
        assert raw["contractor_name"] == "A.B.C. Inc"

    def test_same_contractor_rerun_refreshes_instead_of_failing(self, tmp_path):
        """The pipeline runs intake on every invocation (e.g. the SF-gate
        exit-2 → --sf-confirmed re-run): a pre-existing JSON for the SAME
        contractor is refreshed, not treated as a collision."""
        _write_falke_form(tmp_path / "acme.xlsx")
        s1, f1 = run_structured_intake(tmp_path)
        s2, f2 = run_structured_intake(tmp_path)  # re-run, same dir
        assert f1 == [] and f2 == []
        assert s1 == s2 == [("acme.xlsx", "acme_restoration_inc.json")]

    def test_existing_json_for_different_bidder_never_overwritten(self, tmp_path):
        """A pre-existing {slug}.json from a DIFFERENT bidder (e.g. a prior
        extraction agent) is never overwritten — loud per-file failure."""
        (tmp_path / "abc_inc.json").write_text(
            json.dumps({"contractor_name": "Another Firm Entirely"}))
        (tmp_path / "abc.csv").write_text(
            "Item Code,Description,ABC Inc\n001,Demo,100\n")

        successes, failures = run_structured_intake(tmp_path)
        assert successes == []
        assert len(failures) == 1 and "DIFFERENT bidder" in failures[0][1]
        raw = json.loads((tmp_path / "abc_inc.json").read_text())
        assert raw["contractor_name"] == "Another Firm Entirely"  # untouched


class TestFloydC3FormulaInNameCell:
    def test_formula_in_bidder_block_fails_loudly(self, tmp_path):
        """C3: a formula in the bidder-block name cell must raise a loud
        IntakeError — never fall through and attribute the bid to the project
        name (an uncached formula reads None under data_only=True)."""
        p = _write_falke_form(tmp_path / "formula.xlsx")
        wb = openpyxl.load_workbook(p)
        wb.active["E8"] = "=A1&B1"   # name cell becomes a formula
        wb.save(p)
        with pytest.raises(IntakeError, match="FORMULA"):
            parse_structured_bid(p)


class TestFloydC4DecompressionBomb:
    def test_zip_bomb_rejected_before_parse(self, tmp_path, monkeypatch):
        """C4: a VALID xlsx whose on-disk size clears the 20 MB ceiling but
        whose DECOMPRESSED size exceeds the decompressed ceiling must be
        rejected loudly BEFORE openpyxl parses (ceiling lowered for the test;
        the mechanism — summed ZipInfo.file_size — is what's under test)."""
        import zipfile

        import src.intake_structured as intake_mod

        monkeypatch.setattr(intake_mod, "MAX_DECOMPRESSED_XLSX_BYTES",
                            1024 * 1024)  # 1 MB decompressed cap
        bomb = tmp_path / "bomb.xlsx"
        with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as zf:
            # 4 MB of zeros compresses to ~4 KB on disk — tiny compressed,
            # far over the decompressed cap.
            zf.writestr("xl/worksheets/sheet1.xml", b"0" * (4 * 1024 * 1024))
        assert bomb.stat().st_size < intake_mod.MAX_STRUCTURED_FILE_BYTES
        with pytest.raises(IntakeError, match="decompresse[sd].*zip-bomb"):
            parse_structured_bid(bomb)

    def test_zip_bomb_is_per_file_in_batch(self, tmp_path, monkeypatch):
        """C4 + C1 contract: the bomb is a loud per-file failure; the good
        file in the same batch still succeeds."""
        import zipfile

        import src.intake_structured as intake_mod

        monkeypatch.setattr(intake_mod, "MAX_DECOMPRESSED_XLSX_BYTES",
                            1024 * 1024)
        _write_falke_form(tmp_path / "good.xlsx")
        with zipfile.ZipFile(tmp_path / "bomb.xlsx", "w",
                             zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/worksheets/sheet1.xml", b"0" * (4 * 1024 * 1024))

        successes, failures = run_structured_intake(tmp_path)
        assert [s[0] for s in successes] == ["good.xlsx"]
        assert len(failures) == 1
        assert failures[0][0] == "bomb.xlsx"
        assert "zip-bomb" in failures[0][1]


class TestFloydB4EngineDefectObservability:
    def test_unexpected_exception_logs_traceback_and_names_engine_defect(
        self, tmp_path, monkeypatch, capsys
    ):
        """B4 (advisory): an unexpected exception (engine regression) is still
        a per-file failure, but the message flags a possible ENGINE defect and
        the full traceback goes to stderr — never silently misdiagnosed as a
        bad client file."""
        import src.intake_structured as intake_mod

        _write_falke_form(tmp_path / "good.xlsx")
        (tmp_path / "bug.csv").write_text(
            "Item Code,Description,Acme\n001,Demo,100\n")

        real = intake_mod.parse_structured_bid

        def buggy(path):
            if Path(path).name == "bug.csv":
                raise AttributeError("simulated engine regression")
            return real(path)

        monkeypatch.setattr(intake_mod, "parse_structured_bid", buggy)
        successes, failures = intake_mod.run_structured_intake(tmp_path)
        assert [s[0] for s in successes] == ["good.xlsx"]
        assert len(failures) == 1 and failures[0][0] == "bug.csv"
        assert "engine defect" in failures[0][1]
        err = capsys.readouterr().err
        assert "AttributeError: simulated engine regression" in err
        assert "Traceback" in err
