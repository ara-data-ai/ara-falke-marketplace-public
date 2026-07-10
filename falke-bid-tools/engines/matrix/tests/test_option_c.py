"""
FALKE Matrix — Option C acceptance tests (faithful mirror + leveled view)
=========================================================================
Encodes Marvin's §8 acceptance gate for Option C against a synthetic
ACME fixture carrying the reference numbers from the build spec:

    Mirror (Bid_Form, as-submitted):
        DIV 01 = $287,340.80   (native — dumpsters NOT added)
        DIV 11 = $54,959.40    (Dumpsters stay)
        DIV 13 = $242,238.26   (Flooring(Labor) stays)
    Leveled (Leveled_Normalized, moves applied):
        DIV 01 = $342,300.20   (= 287,340.80 + 54,959.40)
        DIV 11 = 0,  DIV 13 = 0,  DIV 09 += flooring

Covers: mirror-foots-to-bid, in-place marker present, leveled-applies-moves,
cross-bid-only-in-leveled, no-phantom-gap, Stage-6b dual-view, and the new
mirror-GT == leveled-GT invariant.

Run from the engine root:
    python3 -m pytest tests/test_option_c.py -v
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from src.audit import AuditCode, audit_bids
from src.firm_config import Firm, KnownFirmsConfig, Reclassification
from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)
from src.normalize import build_normalized_view, compute_cross_bid_stats, normalize_bid
from src.reconcile import reconcile_written_matrix
from src.run_config import RunInputs
from src.write_matrix import write_matrix

def _synthetic_firm_config() -> KnownFirmsConfig:
    """A synthetic, ACTIVE known-firm config (no real names) carrying the two
    reclass quirks this acceptance suite exercises. Installed into the module
    firm cache by the autouse fixture below so every ``normalize_bid(doc)`` in
    this file drives the identical name-agnostic reclass path a real overlay
    would — without any real firm name in the tree."""
    return KnownFirmsConfig(firms=[Firm(
        firm_id="acme",
        match=["acme"],
        reclassifications=[
            Reclassification(
                rule_id="ACME_FLOORING_LABOR",
                from_division="DIV 13 00 00", to_division="DIV 09 00 00",
                when_description_contains_all=["flooring", "labor"]),
            Reclassification(
                rule_id="ACME_DUMPSTER",
                from_division="DIV 11 00 00", to_division="DIV 01 00 00",
                when_description_contains_all=["dumpster"]),
        ],
    )])


@pytest.fixture(autouse=True)
def _install_synthetic_firm(monkeypatch):
    """Point the module firm-config cache at the synthetic overlay for every
    test here (auto-restored after each test)."""
    monkeypatch.setattr(
        "src.normalize._KNOWN_FIRMS_CACHE", _synthetic_firm_config())


# Reference figures (build spec §8 / brief).
DIV01_NATIVE = Decimal("287340.80")
DUMPSTERS = Decimal("54959.40")
FLOORING = Decimal("242238.26")
DIV09_BASE = Decimal("300000.00")
DIV01_LEVELED = DIV01_NATIVE + DUMPSTERS  # 342300.20
FIRM_GRAND = DIV01_NATIVE + DUMPSTERS + FLOORING + DIV09_BASE  # 884538.46


def _footer(grand_total):
    return BidFooter(
        construction_cost_subtotal=grand_total,
        gc_fee=None,
        grand_total=grand_total,
        alternates=[],
        grand_total_confidence=GrandTotalConfidence.LOW,
    )


def _div(code, name, items=None, subtotal=None, cost=CostStructure.ITEMIZED):
    return DivisionBid(
        csi_code=code,
        division_name=name,
        cost_structure=cost,
        division_subtotal=subtotal,
        classification_source=ClassificationSource.CONTRACTOR_NATIVE,
        contractor_native_code=None,
        line_items=items or [],
    )


def _doc(name, divisions, grand_total):
    return BidDocument(
        contractor_name=name,
        form_type=FormType.FALKE_STANDARD,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=_footer(grand_total),
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _firm_doc() -> BidDocument:
    """ACME as-submitted: Dumpsters in DIV 11, Flooring(Labor) in DIV 13."""
    return _doc("Acme Restoration LLC", [
        _div("DIV 01 00 00", "General Requirements",
             items=[LineItem(description="Project Management", amount=DIV01_NATIVE)],
             subtotal=DIV01_NATIVE),
        _div("DIV 09 00 00", "Finishes",
             items=[LineItem(description="Painting", amount=DIV09_BASE)],
             subtotal=DIV09_BASE),
        _div("DIV 11 00 00", "Equipment",
             items=[LineItem(description="Dumpsters", amount=DUMPSTERS)],
             subtotal=DUMPSTERS),
        _div("DIV 13 00 00", "Special Construction",
             items=[LineItem(description="Flooring (Labor)", amount=FLOORING)],
             subtotal=FLOORING),
    ], FIRM_GRAND)


def _firm_doc_image_scan() -> BidDocument:
    """ACME as above, but flagged IMAGE_SCAN so IMAGE_OCR_UNCERTAINTY fires on
    every priced division — including the reclass-TOUCHED DIV 01, where the OCR
    row would carry the LEVELED subtotal ($342,300) vs the mirror's $287,341."""
    doc = _firm_doc()
    doc.bid_document_input_type = InputType.IMAGE_SCAN
    return doc


def _peer_doc() -> BidDocument:
    """A non-known peer with flooring legitimately in DIV 09 (for cross-bid)."""
    return _doc("Coastal Restoration LLC", [
        _div("DIV 01 00 00", "General Requirements",
             items=[LineItem(description="Project Management", amount=Decimal("300000"))],
             subtotal=Decimal("300000")),
        _div("DIV 09 00 00", "Finishes",
             items=[LineItem(description="Flooring + Painting", amount=Decimal("560000"))],
             subtotal=Decimal("560000")),
    ], Decimal("860000"))


def _run_inputs():
    return RunInputs(
        project_name="Sample Tower Condominiums",
        project_address="100 Example Dr, Anytown FL",
        gross_sf=100_000.0,
        sf_basis_label="balcony SF",
        sf_source="explicit",
    )


def _sub(bid, code):
    d = next((d for d in bid.divisions if d.csi_code == code), None)
    return d.subtotal_cell.amount if (d and d.subtotal_cell.amount is not None) else Decimal("0")


# ---------------------------------------------------------------------------
# §8.1 — Mirror foots to the bid
# ---------------------------------------------------------------------------

class TestMirrorFootsToBid:
    def test_acme_mirror_subtotals_are_as_submitted(self):
        bid = normalize_bid(_firm_doc())
        assert _sub(bid, "DIV 01 00 00") == DIV01_NATIVE
        assert _sub(bid, "DIV 11 00 00") == DUMPSTERS
        assert _sub(bid, "DIV 13 00 00") == FLOORING
        # Grand total unchanged.
        assert bid.footer.grand_total.amount == FIRM_GRAND


# ---------------------------------------------------------------------------
# §8.2 — Recommendation is visible (in-place marker)
# ---------------------------------------------------------------------------

class TestMarkerPresent:
    def test_normalization_note_column_on_bid_form(self):
        bids = [normalize_bid(_firm_doc()), normalize_bid(_peer_doc())]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(b, d) for b, d in
             zip(bids, [_firm_doc(), _peer_doc()])]
        )
        items = audit_bids(leveled)
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "m.xlsx"
            write_matrix(bids, out, _run_inputs(), audit_items=items, leveled_bids=leveled)
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            assert ws.cell(row=4, column=3).value == "Normalization Note"
            notes = " ".join(
                str(ws.cell(row=r, column=3).value or "")
                for r in range(1, ws.max_row + 1)
            )
        assert "Dumpsters" in notes
        assert "normalize → DIV 01 (General Requirements)" in notes
        assert "applied in Leveled_Normalized" in notes
        assert "Flooring (Labor)" in notes and "DIV 09 (Finishes)" in notes


# ---------------------------------------------------------------------------
# §8.3 — Leveled view exists, labeled, moves applied, GT identical
# ---------------------------------------------------------------------------

class TestLeveledView:
    def test_leveled_applies_moves_and_keeps_grand_total(self):
        mirror = normalize_bid(_firm_doc())
        leveled = build_normalized_view(mirror, _firm_doc())
        assert _sub(leveled, "DIV 01 00 00") == DIV01_LEVELED
        assert _sub(leveled, "DIV 11 00 00") == Decimal("0")
        assert _sub(leveled, "DIV 13 00 00") == Decimal("0")
        assert _sub(leveled, "DIV 09 00 00") == DIV09_BASE + FLOORING
        # GT identical to the mirror.
        assert leveled.footer.grand_total.amount == mirror.footer.grand_total.amount

    def test_leveled_sheet_has_banner(self):
        bids = [normalize_bid(_firm_doc())]
        leveled = [build_normalized_view(bids[0], _firm_doc())]
        items = audit_bids(compute_cross_bid_stats(leveled))
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "m.xlsx"
            write_matrix(bids, out, _run_inputs(), audit_items=items, leveled_bids=leveled)
            wb = openpyxl.load_workbook(out)
            assert "Leveled_Normalized" in wb.sheetnames
            ws = wb["Leveled_Normalized"]
            banner = " ".join(
                str(ws.cell(row=r, column=1).value or "") for r in (1, 2)
            )
        assert "ESTIMATOR-NORMALIZED VIEW" in banner
        assert "does NOT match the submitted bids" in banner


# ---------------------------------------------------------------------------
# §8.4 — No phantom gap
# ---------------------------------------------------------------------------

class TestNoPhantomGap:
    def test_no_scope_gap_on_acme_vacated_divisions(self):
        docs = [_firm_doc(), _peer_doc()]
        mirrors = [normalize_bid(d) for d in docs]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(m, d) for m, d in zip(mirrors, docs)]
        )
        items = audit_bids(leveled)
        gaps = [
            i for i in items
            if i.code == AuditCode.SCOPE_GAP_IMPLICIT
            and i.contractor_name == "Acme Restoration LLC"
            and i.division_csi in ("DIV 11 00 00", "DIV 13 00 00")
        ]
        assert gaps == [], f"phantom SCOPE_GAP_IMPLICIT raised: {[g.division_csi for g in gaps]}"
        # Cell-level: no SCOPE_GAP_IMPLICIT flag on those leveled divisions either.
        acme_lev = next(b for b in leveled if b.contractor_name == "Acme Restoration LLC")
        for code in ("DIV 11 00 00", "DIV 13 00 00"):
            div = next((d for d in acme_lev.divisions if d.csi_code == code), None)
            if div is not None:
                assert "SCOPE_GAP_IMPLICIT" not in div.subtotal_cell.flags

    def test_genuine_blank_for_other_bidder_still_flags(self):
        """A peer who genuinely leaves a high-median division blank still flags."""
        big = _doc("Big Spender LLC", [
            _div("DIV 09 00 00", "Finishes",
                 items=[LineItem(description="Finishes", amount=Decimal("500000"))],
                 subtotal=Decimal("500000")),
        ], Decimal("500000"))
        # Peer with DIV 09 blank (NULL) — genuine gap, no reclass.
        blank = _doc("Gap Co", [
            _div("DIV 09 00 00", "Finishes",
                 items=[LineItem(description="Finishes", amount=None)],
                 subtotal=None),
            _div("DIV 01 00 00", "General Requirements",
                 items=[LineItem(description="GR", amount=Decimal("400000"))],
                 subtotal=Decimal("400000")),
        ], Decimal("400000"))
        docs = [big, blank]
        mirrors = [normalize_bid(d) for d in docs]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(m, d) for m, d in zip(mirrors, docs)]
        )
        items = audit_bids(leveled)
        gaps = [
            i for i in items
            if i.code == AuditCode.SCOPE_GAP_IMPLICIT and i.contractor_name == "Gap Co"
            and i.division_csi == "DIV 09 00 00"
        ]
        assert gaps, "a genuine blank should still raise SCOPE_GAP_IMPLICIT"


# ---------------------------------------------------------------------------
# §8.5 — Cross-bid stats only on leveled
# ---------------------------------------------------------------------------

class TestCrossBidOnlyOnLeveled:
    def test_cross_bid_codes_tagged_leveled(self):
        docs = [_firm_doc(), _peer_doc()]
        mirrors = [normalize_bid(d) for d in docs]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(m, d) for m, d in zip(mirrors, docs)]
        )
        items = audit_bids(leveled)
        cross = {AuditCode.SCOPE_GAP_IMPLICIT, AuditCode.CROSS_BID_HIGH_VARIANCE,
                 AuditCode.GC_FEE_OUTLIER, AuditCode.GC_FEE_NORMAL}
        # Division-subtotal-bearing intra-bid codes on a reclass-TOUCHED division
        # carry the leveled subtotal, so they are also tagged "leveled" to keep
        # the mirror honest (Floyd/Marvin §8). ACME's reclass touches DIV 01
        # (to_division) and DIV 09 (to_division) plus DIV 11/DIV 13 (from_division).
        subtotal_bearing = {AuditCode.ARITHMETIC_VERIFIED,
                            AuditCode.ARITHMETIC_DISCREPANCY,
                            AuditCode.LUMP_SUM_DIVISION}
        acme_touched = {"DIV 01 00 00", "DIV 09 00 00", "DIV 11 00 00", "DIV 13 00 00"}
        for i in items:
            if i.code in cross:
                assert i.view == "leveled", f"{i.code} should be leveled-only, got {i.view}"
            elif (
                i.code in subtotal_bearing
                and i.contractor_name == "Acme Restoration LLC"
                and i.division_csi in acme_touched
            ):
                assert i.view == "leveled", (
                    f"{i.code} on a reclass-touched division should be leveled-only, "
                    f"got {i.view}"
                )
            else:
                assert i.view == "both", f"{i.code} should be 'both', got {i.view}"


# ---------------------------------------------------------------------------
# §8.6 — AUDIT reframe
# ---------------------------------------------------------------------------

class TestAuditReframe:
    def test_known_firm_reclassified_recommendation_framing(self):
        bid = normalize_bid(_firm_doc())
        recl = [f for f in bid.summary_flags if f.flag_type == "KNOWN_FIRM_RECLASSIFIED"]
        assert recl
        for f in recl:
            assert "recommends normalizing" in f.message
            assert "shown IN PLACE on the Bid_Form" in f.message
            assert "Leveled_Normalized" in f.message
            # Points at the FROM division on the mirror.
            assert f.division_csi in ("DIV 11 00 00", "DIV 13 00 00")


# ---------------------------------------------------------------------------
# §8.7 — Stage 6b ties out BOTH sheets (incl. mirror-GT == leveled-GT)
# ---------------------------------------------------------------------------

class TestStage6bDualView:
    def _write(self, tmp):
        docs = [_firm_doc(), _peer_doc()]
        mirrors = [normalize_bid(d) for d in docs]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(m, d) for m, d in zip(mirrors, docs)]
        )
        items = audit_bids(leveled)
        out = Path(tmp) / "m.xlsx"
        write_matrix(mirrors, out, _run_inputs(), audit_items=items, leveled_bids=leveled)
        return out, mirrors, leveled, items

    def test_dual_view_tieout_passes(self):
        with tempfile.TemporaryDirectory() as tmp:
            out, mirrors, leveled, items = self._write(tmp)
            failures = reconcile_written_matrix(
                output_path=out, bids=mirrors,
                audit_item_count=len(items), leveled_bids=leveled,
            )
        assert failures == [], [f.message for f in failures]

    def test_tieout_catches_leaked_grand_total(self):
        """If a leveled GT diverges from the mirror GT, the new invariant fires."""
        with tempfile.TemporaryDirectory() as tmp:
            out, mirrors, leveled, items = self._write(tmp)
            # Corrupt the leveled sheet's GRAND TOTAL cell for ACME. v0.3.0
            # geometry: col A is blank in the leveled footer block — the row is
            # anchored on the col-B DISPLAY label, and the amount lives on the
            # merged COST anchor under the bidder block (row-5 name match).
            wb = openpyxl.load_workbook(out)
            ws = wb["Leveled_Normalized"]
            gt_row = next(
                r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "GRAND TOTAL CONSTRUCTION COST"
            )
            col = next(c for c in range(4, ws.max_column + 1)
                       if ws.cell(row=5, column=c).value == "Acme Restoration LLC")
            ws.cell(row=gt_row, column=col).value = 999999.0
            wb.save(out)
            failures = reconcile_written_matrix(
                output_path=out, bids=mirrors,
                audit_item_count=len(items), leveled_bids=leveled,
            )
        assert any("Mirror/leveled grand-total mismatch" in f.message
                   or "Grand-total tie-out FAILED" in f.message
                   for f in failures)


# ---------------------------------------------------------------------------
# §8.8 — Mirror honesty: no "Both" audit row carries a non-mirror dollar value
# ---------------------------------------------------------------------------

class TestMirrorAuditHonesty:
    """Acceptance invariant (Floyd/Marvin §8): NO ``view="both"`` AUDIT row may
    carry a dollar value that differs from the corresponding Bid_Form
    (as-submitted) cell. A board cross-reading the AUDIT and Bid_Form tabs must
    never see a "Both" row asserting a value the mirror does not show.

    Pre-fix this FAILED: ARITHMETIC_VERIFIED for ACME DIV 01 was tagged "Both"
    while carrying the LEVELED subtotal ($342,300) vs the mirror's $287,341 — and
    DIV 09 ($542,238 vs $300,000). The re-tag (touched divisions → "leveled")
    removes the contradiction.
    """

    def _write(self, tmp, acme=_firm_doc):
        docs = [acme(), _peer_doc()]
        mirrors = [normalize_bid(d) for d in docs]
        leveled = compute_cross_bid_stats(
            [build_normalized_view(m, d) for m, d in zip(mirrors, docs)]
        )
        items = audit_bids(leveled)
        out = Path(tmp) / "m.xlsx"
        write_matrix(mirrors, out, _run_inputs(), audit_items=items, leveled_bids=leveled)
        return out, items

    @staticmethod
    def _dollars(s):
        """Parse a board dollar string like '$342,300' to an int, else None."""
        if not s or not isinstance(s, str):
            return None
        t = s.strip()
        if not t.startswith("$"):
            return None
        body = t[1:].replace(",", "")
        try:
            return int(round(float(body)))
        except ValueError:
            return None

    @staticmethod
    def _mirror_cells(out):
        """Map (contractor_name, csi_code) → as-submitted Bid_Form subtotal value."""
        wb = openpyxl.load_workbook(out)
        ws = wb["Bid_Form"]
        # Contractor name → its cost column (names live on row 5).
        name_col = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=5, column=c).value
            if isinstance(v, str) and v.strip():
                name_col[v.strip()] = c
        # CSI code (col A) → subtotal value is on the "... SUBTOTAL" row that
        # follows. We locate each division header (col A = CSI), then walk down to
        # the row whose col B ends with "SUBTOTAL".
        out_map = {}
        for r in range(1, ws.max_row + 1):
            csi = ws.cell(row=r, column=1).value
            if not (isinstance(csi, str) and csi.startswith("DIV ")):
                continue
            sub_row = None
            for rr in range(r + 1, ws.max_row + 1):
                # Stop if we hit the next division header.
                nxt = ws.cell(row=rr, column=1).value
                if isinstance(nxt, str) and nxt.startswith("DIV "):
                    break
                label = ws.cell(row=rr, column=2).value
                if isinstance(label, str) and label.upper().endswith("SUBTOTAL"):
                    sub_row = rr
                    break
            if sub_row is None:
                continue
            for name, col in name_col.items():
                val = ws.cell(row=sub_row, column=col).value
                if isinstance(val, (int, float)):
                    out_map[(name, csi.strip())] = int(round(float(val)))
        return out_map

    def _offenders_for(self, acme):
        with tempfile.TemporaryDirectory() as tmp:
            out, items = self._write(tmp, acme=acme)
            mirror = self._mirror_cells(out)

        offenders = []
        for it in items:
            if it.view != "both":
                continue
            if it.division_csi is None:
                continue
            dollars = self._dollars(it.value)
            if dollars is None:
                continue
            key = (it.contractor_name, it.division_csi)
            if key not in mirror:
                continue
            if dollars != mirror[key]:
                offenders.append(
                    f"{it.code.value} {it.contractor_name} {it.division_csi}: "
                    f"audit(Both)={dollars} vs Bid_Form={mirror[key]}"
                )
        return offenders

    def test_no_both_audit_row_contradicts_mirror(self):
        # Runs on BOTH a DIGITAL_NATIVE fixture (exercises ARITHMETIC_*/LUMP_SUM)
        # and an IMAGE_SCAN fixture (exercises IMAGE_OCR_UNCERTAINTY, the emit
        # site Floyd's completeness probe caught). On the image-scan ACME, the
        # reclass-touched DIV 01 OCR row carries the leveled $342,300 — it MUST be
        # tagged 'leveled', not 'both', or it contradicts the $287,341 mirror cell.
        for label, acme in (("digital_native", _firm_doc),
                              ("image_scan", _firm_doc_image_scan)):
            offenders = self._offenders_for(acme)
            assert not offenders, (
                f"[{label}] view='both' AUDIT row(s) carry a value that differs "
                "from the as-submitted Bid_Form cell:\n  " + "\n  ".join(offenders)
            )

    def test_invariant_would_catch_the_pre_fix_contradiction(self):
        """Sanity: the ACME DIV 01/DIV 09 rows that triggered the bug are the
        ones now tagged 'leveled' (proving the invariant has teeth — without the
        re-tag they would be 'both' and would fail the assertion above)."""
        with tempfile.TemporaryDirectory() as tmp:
            _, items = self._write(tmp)
        for csi, leveled_val in (("DIV 01 00 00", "$342,300"),
                                 ("DIV 09 00 00", "$542,238")):
            row = next(
                i for i in items
                if i.code == AuditCode.ARITHMETIC_VERIFIED
                and i.contractor_name == "Acme Restoration LLC"
                and i.division_csi == csi
            )
            assert row.view == "leveled", (
                f"{csi} ARITHMETIC_VERIFIED must be leveled, got {row.view}"
            )
            assert row.value == leveled_val, (
                f"{csi} carries the leveled subtotal {leveled_val}, got {row.value}"
            )
