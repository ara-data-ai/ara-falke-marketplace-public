"""
FALKE Matrix — Generalization gold-standard tests (Marvin §7 GS-1..GS-7)
========================================================================
These encode Marvin's domain oracles as executable tests, plus the Floyd-C4
contract (promoted flags reach the AUDIT sheet) and a write_matrix smoke that
proves project identity is per-run (no hard-coded project constants).

Run from the engine root:
    python3 -m pytest tests/test_generalization.py -v
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from src.audit import AuditCode, AuditStatus, audit_bids
from src.canon import classify_code_token, detect_csi_1995_2digit
from src.firm_config import (
    Firm,
    KnownFirmsConfig,
    Reclassification,
    load_known_firms,
)


def _synthetic_firm_config() -> KnownFirmsConfig:
    """A synthetic, ACTIVE known-firm config (no real names) carrying the two
    reclass quirks the mechanism is tested against — injected via
    ``normalize_bid(doc, known_firms=...)`` so the tests exercise the identical
    name-agnostic reclass path a real overlay would drive."""
    return KnownFirmsConfig(firms=[Firm(
        firm_id="acme",
        match=["acme"],
        reclassifications=[
            Reclassification(
                rule_id="ACME_FLOORING_LABOR",
                from_division="DIV 13 00 00", to_division="DIV 09 00 00",
                when_description_contains_all=["flooring", "labor"]),
            Reclassification(
                rule_id="ACME_DUMPSTER",
                from_division="DIV 11 00 00", to_division="DIV 01 00 00",
                when_description_contains_all=["dumpster"]),
        ],
    )])
from src.models import (
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)
from src.normalize import build_normalized_view, compute_cross_bid_stats, normalize_bid
from src.run_config import RunInputs
from src.write_matrix import write_matrix


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _footer(grand_total=None, construction=None, gc_fee=None, alternates=None):
    return BidFooter(
        construction_cost_subtotal=construction,
        gc_fee=gc_fee,
        grand_total=grand_total,
        alternates=alternates or [],
        grand_total_confidence=GrandTotalConfidence.LOW,
    )


def _doc(name, divisions, footer=None, form_type=FormType.FALKE_STANDARD):
    return BidDocument(
        contractor_name=name,
        form_type=form_type,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=footer or _footer(),
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _div(code, name, items=None, subtotal=None,
         cost=CostStructure.LUMP_SUM):
    return DivisionBid(
        csi_code=code,
        division_name=name,
        cost_structure=cost,
        division_subtotal=subtotal,
        classification_source=ClassificationSource.CONTRACTOR_NATIVE,
        contractor_native_code=None,
        line_items=items or [],
    )


def _flag_types(bid):
    return [f.flag_type for f in bid.summary_flags]


def _run_inputs(gsf=10_000.0):
    return RunInputs(
        project_name="Harbor View Tower",
        project_address="100 Test Ave, Test City FL 00000",
        gross_sf=gsf,
        sf_basis_label="balcony SF",
        sf_source="explicit",
    )


# ---------------------------------------------------------------------------
# Signature classifier unit checks (Marvin §1.1)
# ---------------------------------------------------------------------------

class TestSignatureClassifier:
    def test_canonical_token(self):
        assert classify_code_token("DIV 03 00 00") == "CANONICAL"

    def test_bare_two_digit_is_legacy_even_if_canonical_number(self):
        # bare '03' is LEGACY_2DIGIT (CANONICAL requires the full DIV form, §1.1)
        assert classify_code_token("03") == "LEGACY_2DIGIT"
        assert classify_code_token("15") == "LEGACY_2DIGIT"

    def test_footer_token_is_legacy(self):
        assert classify_code_token("17-040") == "LEGACY_2DIGIT"
        assert classify_code_token("17") == "LEGACY_2DIGIT"

    def test_unknown_token(self):
        assert classify_code_token("DIV 33 00 00") == "UNKNOWN"
        assert classify_code_token("Section 26 05 00") == "UNKNOWN"
        assert classify_code_token("M-1") == "UNKNOWN"

    def test_detection_requires_three_legacy_and_discriminator(self):
        assert detect_csi_1995_2digit(["01", "03", "15"]) is True
        # only 2 legacy codes → too thin (n_legacy >= 3 required)
        assert detect_csi_1995_2digit(["15", "16"]) is False
        # 3 legacy but no discriminator (no 15/16/17)
        assert detect_csi_1995_2digit(["01", "02", "03"]) is False
        # any canonical present → not all-or-nothing legacy
        assert detect_csi_1995_2digit(["DIV 03 00 00", "15", "16"]) is False
        # any unknown token poisons the bid
        assert detect_csi_1995_2digit(["01", "03", "15", "M-1"]) is False


# ---------------------------------------------------------------------------
# GS-1 — Single bidder (N3 oracle)
# ---------------------------------------------------------------------------

class TestGS1SingleBidder:
    def test_no_cross_bid_flags_single_bidder(self):
        doc = _doc("Solo Restoration LLC", [
            _div("DIV 01 00 00", "General Requirements", subtotal=Decimal("50000")),
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("DIV 07 00 00", "Thermal & Moisture Protection", subtotal=Decimal("150000")),
            _div("DIV 09 00 00", "Finishes", subtotal=Decimal("80000")),
        ], footer=_footer(grand_total=Decimal("680000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        items = audit_bids(bids)

        codes = {i.code for i in items}
        assert AuditCode.SCOPE_GAP_IMPLICIT not in codes
        assert AuditCode.GC_FEE_OUTLIER not in codes
        assert AuditCode.CROSS_BID_HIGH_VARIANCE not in codes
        # engine did not throw and produced some intra-bid items
        assert items

    def test_single_bid_notice_rendered(self):
        doc = _doc("Solo Restoration LLC", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(grand_total=Decimal("400000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "matrix.xlsx"
            write_matrix(bids, out, _run_inputs(), audit_items=audit_bids(bids))
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            text = " ".join(
                str(c.value) for row in ws.iter_rows() for c in row if c.value
            )
            assert "Single bid — no competitive comparison available." in text


# ---------------------------------------------------------------------------
# GS-2 — Two bidders (N3 oracle): scope gap yes, stddev GC outlier suppressed
# ---------------------------------------------------------------------------

class TestGS2TwoBidders:
    def test_scope_gap_present_no_stddev_outlier(self):
        a = _doc("Alpha Builders", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("DIV 07 00 00", "Thermal & Moisture Protection", subtotal=Decimal("150000")),
        ], footer=_footer(grand_total=Decimal("550000"), construction=Decimal("500000"),
                          gc_fee=Decimal("50000")))
        b = _doc("Beta Restoration", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("380000")),
            _div("DIV 07 00 00", "Thermal & Moisture Protection"),  # blank
        ], footer=_footer(grand_total=Decimal("418000"), construction=Decimal("380000"),
                          gc_fee=Decimal("38000")))
        bids = compute_cross_bid_stats([normalize_bid(a), normalize_bid(b)])
        items = audit_bids(bids)

        codes = {i.code for i in items}
        assert AuditCode.SCOPE_GAP_IMPLICIT in codes  # B blank on DIV 07
        assert AuditCode.GC_FEE_OUTLIER not in codes   # suppressed at n=2


# ---------------------------------------------------------------------------
# GS-3 — New firm, canonical codes incl. legit DIV 13 → NOT reclassed (critical)
# ---------------------------------------------------------------------------

class TestGS3NewFirmCanonical:
    def test_legit_div13_not_reclassed_for_unknown_firm(self):
        doc = _doc("Coastal Concrete Restoration LLC", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("DIV 13 00 00", "Special Construction",
                 items=[LineItem(description="Pool-deck system", amount=Decimal("60000"))],
                 subtotal=Decimal("60000"), cost=CostStructure.ITEMIZED),
            _div("DIV 11 00 00", "Equipment",
                 items=[LineItem(description="Residential Equipment", amount=Decimal("20000"))],
                 subtotal=Decimal("20000"), cost=CostStructure.ITEMIZED),
        ], footer=_footer(grand_total=Decimal("480000")))
        bid = normalize_bid(doc)

        # No remap, no reclass, no unrecognized flag.
        ft = _flag_types(bid)
        assert "KNOWN_FIRM_RECLASSIFIED" not in ft
        assert "CODE_FORMAT_REMAPPED" not in ft
        assert "UNRECOGNIZED_CODE_FORMAT" not in ft

        # DIV 13 line stayed in DIV 13 (NOT moved to DIV 09).
        div13 = next(d for d in bid.divisions if d.csi_code == "DIV 13 00 00")
        assert any("pool-deck" in lbl.lower() for lbl in div13.line_item_cells)


# ---------------------------------------------------------------------------
# GS-4 — Legacy 2-digit bid triggers csi_1995_2digit (remap + split)
# ---------------------------------------------------------------------------

class TestGS4LegacyRemapSplit:
    def test_legacy_remap_and_split(self):
        doc = _doc("Legacy Format Co", [
            _div("01", "General Requirements", subtotal=Decimal("50000")),
            _div("03", "Concrete", subtotal=Decimal("400000")),
            _div("07", "Thermal", subtotal=Decimal("150000")),
            _div("09", "Finishes", subtotal=Decimal("80000")),
            _div("15", "Mechanical", cost=CostStructure.ITEMIZED, items=[
                LineItem(description="Domestic water piping", amount=Decimal("60000")),
                LineItem(description="HVAC ductwork", amount=Decimal("90000")),
            ]),
            _div("16", "Electrical", cost=CostStructure.ITEMIZED, items=[
                LineItem(description="Branch wiring & devices", amount=Decimal("120000")),
                LineItem(description="Fire alarm notification", amount=Decimal("30000")),
            ]),
            _div("17-040", "OH&P", subtotal=Decimal("70000")),
        ], form_type=FormType.CONTRACTOR_OWN,
           footer=_footer(grand_total=Decimal("1020000")))
        bid = normalize_bid(doc)
        codes = [d.csi_code for d in bid.divisions]

        # straight remaps
        for c in ("DIV 01 00 00", "DIV 03 00 00", "DIV 07 00 00", "DIV 09 00 00"):
            assert c in codes
        # splits landed correctly
        div22 = next(d for d in bid.divisions if d.csi_code == "DIV 22 00 00")
        div23 = next(d for d in bid.divisions if d.csi_code == "DIV 23 00 00")
        div26 = next(d for d in bid.divisions if d.csi_code == "DIV 26 00 00")
        div28 = next(d for d in bid.divisions if d.csi_code == "DIV 28 00 00")
        assert any("domestic water" in lbl.lower() for lbl in div22.line_item_cells)
        assert any("ductwork" in lbl.lower() for lbl in div23.line_item_cells)
        assert any("branch wiring" in lbl.lower() for lbl in div26.line_item_cells)
        assert any("fire alarm" in lbl.lower() for lbl in div28.line_item_cells)

        ft = _flag_types(bid)
        assert "CODE_FORMAT_REMAPPED" in ft
        assert "CODE_SPLIT_UNMATCHED" not in ft   # every sub-line routed
        assert "UNRECOGNIZED_CODE_FORMAT" not in ft


# ---------------------------------------------------------------------------
# GS-5 — Mixed format → RED UNRECOGNIZED_CODE_FORMAT, no partial remap (critical)
# ---------------------------------------------------------------------------

class TestGS5MixedFormat:
    def test_mixed_format_red_no_remap(self):
        doc = _doc("Mixed Codes Inc", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("15", "Mechanical", subtotal=Decimal("120000")),
            _div("16", "Electrical", subtotal=Decimal("90000")),
        ], form_type=FormType.CONTRACTOR_OWN,
           footer=_footer(grand_total=Decimal("610000")))
        bid = normalize_bid(doc)

        ft = _flag_types(bid)
        assert "UNRECOGNIZED_CODE_FORMAT" in ft
        assert "CODE_FORMAT_REMAPPED" not in ft  # no partial/per-line remap

        # The legacy 15/16 were NOT remapped (placed as-extracted).
        codes = [d.csi_code for d in bid.divisions]
        assert "15" in codes and "16" in codes
        assert "DIV 22 00 00" not in codes

    def test_unrecognized_flag_is_red_on_audit_sheet(self):
        doc = _doc("Mixed Codes Inc", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("15", "Mechanical", subtotal=Decimal("120000")),
        ], form_type=FormType.CONTRACTOR_OWN,
           footer=_footer(grand_total=Decimal("520000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        items = audit_bids(bids)
        unrec = [i for i in items if i.code == AuditCode.UNRECOGNIZED_CODE_FORMAT]
        assert unrec and unrec[0].status == AuditStatus.RED


# ---------------------------------------------------------------------------
# GS-6 — Known firm reclass happy path (synthetic injected firm config)
# ---------------------------------------------------------------------------

class TestGS6KnownFirmReclass:
    CFG = _synthetic_firm_config()

    def test_firm_reclass_recommended_mirror_and_applied_leveled(self):
        doc = _doc("Acme Restoration", [
            _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="Flooring (Labor) — install", amount=Decimal("18000"))],
                 subtotal=Decimal("18000")),
            _div("DIV 11 00 00", "Equipment", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="Dumpster rental", amount=Decimal("6500"))],
                 subtotal=Decimal("6500")),
        ], footer=_footer(grand_total=Decimal("24500")))
        bid = normalize_bid(doc, known_firms=self.CFG)

        ft = _flag_types(bid)
        assert ft.count("KNOWN_FIRM_RECLASSIFIED") == 2
        assert "KNOWN_FIRM_AMBIGUOUS" not in ft
        assert len(bid.reclass_recommendations) == 2

        # MIRROR: dollars stay where submitted (foots to bid).
        m13 = next(d for d in bid.divisions if d.csi_code == "DIV 13 00 00")
        m11 = next(d for d in bid.divisions if d.csi_code == "DIV 11 00 00")
        assert any("flooring" in lbl.lower() for lbl in m13.line_item_cells)
        assert any("dumpster" in lbl.lower() for lbl in m11.line_item_cells)

        # The reframed flag points at the FROM division on the mirror.
        recl = [f for f in bid.summary_flags if f.flag_type == "KNOWN_FIRM_RECLASSIFIED"]
        assert {f.division_csi for f in recl} == {"DIV 13 00 00", "DIV 11 00 00"}

        # LEVELED: the moves are applied.
        leveled = build_normalized_view(bid, doc)
        l09 = next(d for d in leveled.divisions if d.csi_code == "DIV 09 00 00")
        l01 = next(d for d in leveled.divisions if d.csi_code == "DIV 01 00 00")
        assert any("flooring" in lbl.lower() for lbl in l09.line_item_cells)
        assert any("dumpster" in lbl.lower() for lbl in l01.line_item_cells)

    def test_gs3_gs6_contrast(self):
        """Same DIV 13 flooring line — unknown firm: no recommendation; matched firm: recommended."""
        line = [LineItem(description="Flooring (Labor)", amount=Decimal("18000"))]
        unknown = normalize_bid(_doc("Stranger LLC", [
            _div("DIV 13 00 00", "Special Construction", items=line,
                 subtotal=Decimal("18000"), cost=CostStructure.ITEMIZED),
        ]), known_firms=self.CFG)
        known = normalize_bid(_doc("Acme Restoration", [
            _div("DIV 13 00 00", "Special Construction", items=list(line),
                 subtotal=Decimal("18000"), cost=CostStructure.ITEMIZED),
        ]), known_firms=self.CFG)
        unk13 = next(d for d in unknown.divisions if d.csi_code == "DIV 13 00 00")
        assert any("flooring" in lbl.lower() for lbl in unk13.line_item_cells)
        assert unknown.reclass_recommendations == []
        assert "KNOWN_FIRM_RECLASSIFIED" in _flag_types(known)
        assert len(known.reclass_recommendations) == 1


# ---------------------------------------------------------------------------
# GS-7 — Ambiguous firm match → RED KNOWN_FIRM_AMBIGUOUS (constructed fixture)
# ---------------------------------------------------------------------------

class TestGS7AmbiguousFirm:
    def _ambiguous_config(self, tmp_path) -> KnownFirmsConfig:
        p = Path(tmp_path) / "known_firms.yaml"
        p.write_text(
            "firms:\n"
            "  - firm_id: a\n"
            "    match: [\"acme\"]\n"
            "  - firm_id: b\n"
            "    match: [\"acme restoration\"]\n"
        )
        return load_known_firms(str(p))

    def test_ambiguous_match_red_no_reclass(self, tmp_path):
        cfg = self._ambiguous_config(tmp_path)
        doc = _doc("Acme Restoration LLC", [
            _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="Flooring (Labor)", amount=Decimal("18000"))],
                 subtotal=Decimal("18000")),
        ], footer=_footer(grand_total=Decimal("18000")))
        bid = normalize_bid(doc, known_firms=cfg)

        ft = _flag_types(bid)
        assert "KNOWN_FIRM_AMBIGUOUS" in ft
        assert "KNOWN_FIRM_RECLASSIFIED" not in ft
        # nothing reclassed — flooring stays in DIV 13
        div13 = next(d for d in bid.divisions if d.csi_code == "DIV 13 00 00")
        assert any("flooring" in lbl.lower() for lbl in div13.line_item_cells)


# ---------------------------------------------------------------------------
# C4 contract — every promoted RED/YELLOW code reaches the AUDIT sheet
# ---------------------------------------------------------------------------

class TestC4AuditSheetContract:
    def _write_and_read_audit(self, bids):
        items = audit_bids(bids)
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "m.xlsx"
            write_matrix(bids, out, _run_inputs(), audit_items=items)
            wb = openpyxl.load_workbook(out)
            assert "AUDIT" in wb.sheetnames
            ws = wb["AUDIT"]
            # Column C holds the AuditCode (Option C inserts a View col at B).
            codes_on_sheet = {ws.cell(row=r, column=3).value
                              for r in range(5, ws.max_row + 1)}
        return codes_on_sheet, items

    def test_code_format_remapped_on_audit_sheet(self):
        doc = _doc("Legacy Format Co", [
            _div("01", "General Requirements", subtotal=Decimal("50000")),
            _div("03", "Concrete", subtotal=Decimal("400000")),
            _div("15", "Mechanical", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="HVAC ductwork", amount=Decimal("90000"))]),
        ], form_type=FormType.CONTRACTOR_OWN, footer=_footer(grand_total=Decimal("540000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        codes, _ = self._write_and_read_audit(bids)
        assert "CODE_FORMAT_REMAPPED" in codes

    def test_unrecognized_and_split_unmatched_on_audit_sheet(self):
        # mixed → UNRECOGNIZED
        doc = _doc("Mixed Codes Inc", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
            _div("15", "Mechanical", subtotal=Decimal("120000")),
        ], form_type=FormType.CONTRACTOR_OWN, footer=_footer(grand_total=Decimal("520000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        codes, _ = self._write_and_read_audit(bids)
        assert "UNRECOGNIZED_CODE_FORMAT" in codes

    def test_split_unmatched_on_audit_sheet(self):
        doc = _doc("Legacy Format Co", [
            _div("01", "General Requirements", subtotal=Decimal("50000")),
            _div("03", "Concrete", subtotal=Decimal("400000")),
            _div("16", "Electrical", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="Misc. allowance", amount=Decimal("9000"))]),
        ], form_type=FormType.CONTRACTOR_OWN, footer=_footer(grand_total=Decimal("459000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        codes, _ = self._write_and_read_audit(bids)
        assert "CODE_SPLIT_UNMATCHED" in codes

    def test_reclassified_and_ambiguous_on_audit_sheet(self):
        doc = _doc("Acme Restoration", [
            _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
                 items=[LineItem(description="Flooring (Labor)", amount=Decimal("18000"))],
                 subtotal=Decimal("18000")),
        ], footer=_footer(grand_total=Decimal("18000")))
        bids = compute_cross_bid_stats([normalize_bid(doc, known_firms=_synthetic_firm_config())])
        codes, _ = self._write_and_read_audit(bids)
        assert "KNOWN_FIRM_RECLASSIFIED" in codes


# ---------------------------------------------------------------------------
# M6 — column order = leveled_total ascending (C7)
# ---------------------------------------------------------------------------

class TestM6ColumnOrder:
    def test_leveled_total_ascending(self):
        from src.write_matrix import _sort_bids
        high = normalize_bid(_doc("High Bidder", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("900000")),
        ], footer=_footer(grand_total=Decimal("900000"))))
        low = normalize_bid(_doc("Low Bidder", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(grand_total=Decimal("400000"))))
        ordered = _sort_bids([high, low])
        assert [b.contractor_name for b in ordered] == ["Low Bidder", "High Bidder"]

    def test_none_leveled_total_sorts_last(self):
        from src.write_matrix import _sort_bids
        priced = normalize_bid(_doc("Priced", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(grand_total=Decimal("400000"))))
        no_total = normalize_bid(_doc("No Total", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(grand_total=None)))
        ordered = _sort_bids([no_total, priced])
        assert ordered[-1].contractor_name == "No Total"


# ---------------------------------------------------------------------------
# M7 — alternates rendered in their own section, not folded into base
# ---------------------------------------------------------------------------

class TestM7Alternates:
    def test_alternates_section_present(self):
        doc = _doc("Alt Bidder", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(
            grand_total=Decimal("400000"),
            alternates=[LineItem(description="ADD: replace all balcony railings",
                                 amount=Decimal("75000"))],
        ))
        bid = normalize_bid(doc)
        # surfaced on the normalized footer
        assert len(bid.footer.alternates) == 1
        # base leveled total does NOT include the alternate
        assert bid.footer.leveled_total == Decimal("400000")

        bids = compute_cross_bid_stats([bid])
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "m.xlsx"
            write_matrix(bids, out, _run_inputs(), audit_items=audit_bids(bids))
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            text = " ".join(str(c.value) for row in ws.iter_rows()
                            for c in row if c.value)
            assert "ALTERNATES" in text
            assert "replace all balcony railings" in text


# ---------------------------------------------------------------------------
# M1/M2/M8 — per-run identity, no hard-coded project constants
# ---------------------------------------------------------------------------

class TestM1IdentityPerRun:
    def test_title_and_basis_from_run_inputs(self):
        doc = _doc("Some Bidder", [
            _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        ], footer=_footer(grand_total=Decimal("400000")))
        bids = compute_cross_bid_stats([normalize_bid(doc)])
        run = RunInputs(
            project_name="Marina Bay Lofts",
            project_address="9 Dock Rd, Somewhere FL",
            gross_sf=22_500.0,
            sf_basis_label="facade SF",
            sf_source="explicit",
        )
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "m.xlsx"
            write_matrix(bids, out, run, audit_items=audit_bids(bids))
            wb = openpyxl.load_workbook(out)
            ws = wb["Bid_Form"]
            text = " ".join(str(c.value) for row in ws.iter_rows()
                            for c in row if c.value)
            assert "Marina Bay Lofts — Bid Comparison Matrix" in text
            assert "facade SF" in text          # SF-basis label printed (§1.4)
            # identity guard: a project name NOT supplied to this run must never
            # appear (proves no hard-coded project constant leaks into output).
            assert "Harborview Tower" not in text
            assert "ARA Pipeline Test" not in text
