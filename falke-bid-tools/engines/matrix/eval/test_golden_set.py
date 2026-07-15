"""GOLDEN-SET v1 harness — executes Marvin's cell-level contract (T1-5 / G(a)).

Implements the harness contract in eval/golden/gold-project-v1/README.md:

  1. Overlay swap discipline: back up any existing
     config/known_firms.local.yaml, install gold-project-v1/overlay/
     known_firms.local.yaml in its place (swap, never merge), restore after.
     A MISSING gold overlay is a loud RuntimeError; a silently-skipped overlay
     would also fail loudly via the Beacon expectations (DIV 01/DIV 11 +
     KNOWN_FIRM_RECLASSIFIED) — never a silent pass.
  2. Main run: python3 -m src.pipeline on interim/ with --sf-basis 100000,
     assert exit 0, then expectations.yaml cell by cell using the
     locator_contract (label-anchored, same idiom as reconcile.py — row-5
     names / "CSI"-anchored, col-B subtotal + leveled-footer labels, col-A
     mirror footer keys, benchmark block right of the last group).
  3. Bidder-footer-error run (bond-on-top variant): interim-quarantine/,
     assert exit 0 + file delivered + expectations-quarantine.yaml — since
     the GOLD-DEV-6 fix, a faithfully-reproduced BIDDER footer inconsistency
     no longer quarantines: R21 red + FOOTER_DISCREPANCY tell the story, and
     the harness asserts the quarantine chain's ABSENCE (clean_run block).
  4. FAULT-INJECTION step (Marvin GOLD-DEV-6 ruling (4), Floyd W2-5): corrupt
     one GT cell (+ one commented division subtotal, for the compose proof)
     in a COPY of the main-run workbook, call reconcile_written_matrix +
     apply_quarantine directly, and assert the full quarantine disclosure
     chain (banner both sheets, composed cell marks, RED AUDIT rows,
     QUARANTINE summary line). The pipeline-level exit-3 gate remains
     regression-covered by tests/test_reconcile.py.
  5. Report EVERY mismatch (no fail-fast); nonzero on any mismatch.

Tolerances (defined at the top of expectations.yaml): dollars $0.01,
fractions 1e-4, comments substring, tokens exact.

Adjudication protocol (Marvin): a first-run diff is REPORTED, never silently
"fixed" on either side — Marvin adjudicates. Expectation entries tagged
`defect:` encode current behavior for registered GOLD-DEV findings.

Run in the suite:      python3 -m pytest eval/test_golden_set.py -v
Standalone runner:     python3 eval/test_golden_set.py
(both from the engine root, engines/matrix/)
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
import tempfile
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace
from typing import Optional

import openpyxl
import pytest
import yaml

ENGINE_ROOT = Path(__file__).resolve().parent.parent          # engines/matrix
GOLD = ENGINE_ROOT / "eval" / "golden" / "gold-project-v1"
CONFIG_DIR = ENGINE_ROOT / "config"
OVERLAY_SRC = GOLD / "overlay" / "known_firms.local.yaml"
OVERLAY_DST = CONFIG_DIR / "known_firms.local.yaml"
OVERLAY_BAK = CONFIG_DIR / "known_firms.local.yaml.golden-backup"

sys.path.insert(0, str(ENGINE_ROOT))

# Locator constants — the SAME single-source geometry reconcile.py reads back
# (locator_contract: label-anchored, never absolute coordinates).
from src.write_matrix import (  # noqa: E402
    DIVISION_ROWS,
    LEVELED_CSUB_OFFSET,
    LEVELED_VAR_OFFSET,
    LEVELED_FOOTER_LABELS,
    _col_start,
    _lev_bench_col,
    _lev_col_start,
)

DIV_NAME = dict(DIVISION_ROWS)
FALKE_HEX = {"red": "FF0000", "cyan": "00FFFF", "yellow": "FFFF00"}
# Quarantine cell marks use write_matrix.RED_FILL (soft red FFCCCC); accept
# either red-family hex for "RED fill" assertions on quarantine marks.
QUARANTINE_RED_HEXES = {"FFCCCC", "FF0000"}
# W-D B4/§2: View values are the SHEET NAMES they point at (the
# "As-Submitted" third dialect is dead).
VIEW_LABEL = {"both": "Both", "leveled": "Leveled_Normalized",
              "mirror": "Bid_Form"}

FOOTER_KEY_ORDER = [
    "CONSTRUCTION_SUBTOTAL", "GL_INSURANCE", "BUILDERS_RISK", "GC_FEE",
    "OVERHEAD_PROFIT", "OTHER_FEES", "BOND", "FEES_SUBTOTAL", "GRAND_TOTAL",
]

SUMMARY_LABELS = {
    "total_bid": "Total Bid Amount (submitted)",
    "adjusted_total": "Adjusted Total (= leveled total, Q10)",
    "red": "Red Flags",
    "cyan": "Cyan Flags",
    "yellow": "Yellow Flags",
    "avg_variance": "Average Variance",
}

DOLLAR_TOL = 0.01
FRACTION_TOL = 1e-4


# ---------------------------------------------------------------------------
# Diff collector — report EVERY mismatch, no fail-fast (contract §4)
# ---------------------------------------------------------------------------

class Diffs:
    def __init__(self) -> None:
        self.items: list[str] = []

    def add(self, ctx: str, msg: str) -> None:
        self.items.append(f"{ctx}: {msg}")

    def check(self, cond: bool, ctx: str, msg: str) -> bool:
        if not cond:
            self.add(ctx, msg)
        return cond

    def report(self) -> str:
        return "\n".join(f"  [{i+1:02d}] {m}" for i, m in enumerate(self.items))


# ---------------------------------------------------------------------------
# Low-level cell helpers
# ---------------------------------------------------------------------------

def _fill_hex(cell) -> Optional[str]:
    f = cell.fill
    if f is None or f.patternType != "solid":
        return None
    rgb = getattr(f.fgColor, "rgb", None)
    if not isinstance(rgb, str):
        return None
    return rgb[-6:].upper()


def _has_falke_paint(cell) -> bool:
    return _fill_hex(cell) in set(FALKE_HEX.values())


def _num(cell) -> Optional[float]:
    v = cell.value
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _blankish(cell) -> bool:
    return cell.value in (None, "")


def _comment_text(cell) -> str:
    return cell.comment.text if cell.comment is not None else ""


def _close(actual: Optional[float], expected: float, tol: float) -> bool:
    return actual is not None and abs(actual - expected) <= tol


# ---------------------------------------------------------------------------
# Label-anchored locators (same idiom as reconcile.py — locator_contract)
# ---------------------------------------------------------------------------

def _find_name_row(ws) -> int:
    """Row of contractor names = row below the col-A "CSI" header (survives
    the quarantine banner row-shift on Leveled_Normalized); fallback row 5."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "CSI":
            return row + 1
    return 5


def _subtotal_rows(ws) -> dict[str, int]:
    """Map canonical CSI code → its '{NAME UPPER} SUBTOTAL' row (col B)."""
    label_to_code = {f"{n.upper()} SUBTOTAL": c for c, n in DIVISION_ROWS}
    out: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if isinstance(v, str) and v in label_to_code:
            out[label_to_code[v]] = row
    return out


def _division_header_row(ws, csi: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == csi:
            return row
    return None


def _row_col_b_equals(ws, text: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == text:
            return row
    return None


def _row_col_b_contains(ws, text: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if isinstance(v, str) and text in v:
            return row
    return None


def _row_col_a_equals(ws, text: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == text:
            return row
    return None


def _mirror_line_row(ws, sub_rows: dict, csi: str, desc: str) -> Optional[int]:
    """Find a mirror line row (col B == desc) inside a division block
    (col-A division header → its SUBTOTAL row)."""
    head = _row_col_a_equals(ws, csi)
    sub = sub_rows.get(csi)
    if head is None or sub is None:
        return None
    for row in range(head + 1, sub):
        if ws.cell(row=row, column=2).value == desc:
            return row
    return None


class LeveledGeom:
    """Leveled_Normalized geometry, read back label-anchored per bidder."""

    def __init__(self, ws, bidder_order: list[str], diffs: Diffs) -> None:
        self.ws = ws
        self.n = len(bidder_order)
        self.name_row = _find_name_row(ws)
        self.idx: dict[str, int] = {}
        actual = []
        for i in range(self.n):
            nm = ws.cell(row=self.name_row, column=_lev_col_start(i)).value
            actual.append(nm)
            if isinstance(nm, str):
                self.idx[nm] = i
        diffs.check(
            actual == bidder_order, "leveled.bidder_order",
            f"expected {bidder_order}, sheet row {self.name_row} carries {actual}",
        )
        # Best-effort: map any expected bidder the sheet order missed.
        for i, nm in enumerate(bidder_order):
            self.idx.setdefault(nm, i)
        self.sub_rows = _subtotal_rows(ws)
        self.bench_col = _lev_bench_col(self.n)

    def cost(self, name: str) -> int:
        return _lev_col_start(self.idx[name])

    def csub(self, name: str) -> int:
        return _lev_col_start(self.idx[name]) + LEVELED_CSUB_OFFSET

    def var(self, name: str) -> int:
        return _lev_col_start(self.idx[name]) + LEVELED_VAR_OFFSET

    def footer_row(self, key: str) -> Optional[int]:
        return _row_col_b_equals(self.ws, LEVELED_FOOTER_LABELS[key])


class MirrorGeom:
    """Bid_Form mirror geometry (stride 3, col-A footer machine keys)."""

    def __init__(self, ws, bidder_order: list[str], diffs: Diffs) -> None:
        self.ws = ws
        self.n = len(bidder_order)
        self.name_row = _find_name_row(ws)
        self.idx: dict[str, int] = {}
        actual = []
        for i in range(self.n):
            nm = ws.cell(row=self.name_row, column=_col_start(i)).value
            actual.append(nm)
            if isinstance(nm, str):
                self.idx[nm] = i
        diffs.check(
            actual == bidder_order, "mirror.bidder_order",
            f"expected {bidder_order}, sheet row {self.name_row} carries {actual}",
        )
        for i, nm in enumerate(bidder_order):
            self.idx.setdefault(nm, i)
        self.sub_rows = _subtotal_rows(ws)

    def col(self, name: str) -> int:
        return _col_start(self.idx[name])

    def footer_row(self, key: str) -> Optional[int]:
        return _row_col_a_equals(self.ws, key)


# ---------------------------------------------------------------------------
# Shared assertion pieces
# ---------------------------------------------------------------------------

def _check_paint(cell, paint_exp, ctx: str, d: Diffs) -> None:
    """paint: none → NO falke fill; {color, rule} → exact falke hue."""
    hex_ = _fill_hex(cell)
    if paint_exp in (None, "none"):
        d.check(
            hex_ not in set(FALKE_HEX.values()), ctx,
            f"expected NO Falke paint, cell carries fill {hex_}",
        )
    elif isinstance(paint_exp, dict):
        want = FALKE_HEX[paint_exp["color"]]
        d.check(
            hex_ == want, ctx,
            f"expected {paint_exp['color']} ({want}, rule "
            f"{paint_exp.get('rule')}), cell fill is {hex_}",
        )


def _check_comment(cell, exp: dict, ctx: str, d: Diffs) -> None:
    text = _comment_text(cell)
    if "comment_contains" in exp:
        want = exp["comment_contains"]
        d.check(want in text, ctx,
                f"comment missing {want!r} (comment is {text!r})")
    for want in exp.get("comment_contains_all", []):
        d.check(want in text, ctx,
                f"comment missing {want!r} (comment is {text!r})")


def _check_division_cell(g: LeveledGeom, csi: str, bidder: str,
                         exp: dict, d: Diffs) -> None:
    ctx = f"leveled.divisions[{csi}][{bidder}]"
    sub_row = g.sub_rows.get(csi)
    if not d.check(sub_row is not None, ctx, "SUBTOTAL row not found"):
        return
    csub = g.ws.cell(row=sub_row, column=g.csub(bidder))
    var_cell = g.ws.cell(row=sub_row, column=g.var(bidder))
    kind = exp["kind"]

    if kind in ("priced", "not_comparable"):
        d.check(_close(_num(csub), float(exp["value"]), DOLLAR_TOL), ctx,
                f"value: expected {exp['value']}, written {csub.value!r}")
    elif kind == "zero":
        d.check(_close(_num(csub), float(exp.get("value", 0.0)), DOLLAR_TOL),
                ctx, f"zero: expected 0.00, written {csub.value!r}")
    elif kind in ("excluded", "by_owner"):
        token = exp.get("token", "Excluded" if kind == "excluded" else None)
        d.check(csub.value == token, ctx,
                f"token: expected {token!r} exactly, written {csub.value!r}")
    elif kind in ("missing", "missing_or_blank"):
        d.check(_blankish(csub) or _num(csub) == 0.0 and kind == "missing_or_blank",
                ctx, f"expected blank subtotal cell, written {csub.value!r}")

    # VAR % — raw fraction where expected, blank where absent.
    if "var" in exp:
        d.check(_close(_num(var_cell), float(exp["var"]), FRACTION_TOL), ctx,
                f"VAR%: expected {exp['var']}, written {var_cell.value!r}")
    elif kind in ("missing", "zero", "excluded", "by_owner",
                  "not_comparable", "missing_or_blank"):
        d.check(_blankish(var_cell), ctx,
                f"VAR%: expected blank, written {var_cell.value!r}")

    _check_paint(csub, exp.get("paint"), ctx + ".paint", d)
    _check_comment(csub, exp, ctx + ".comment", d)


def _check_benchmark(g: LeveledGeom, csi: str, exp: dict, d: Diffs) -> None:
    ctx = f"leveled.benchmarks[{csi}]"
    sub_row = g.sub_rows.get(csi)
    if not d.check(sub_row is not None, ctx, "SUBTOTAL row not found"):
        return
    b = g.ws.cell(row=sub_row, column=g.bench_col)
    s = g.ws.cell(row=sub_row, column=g.bench_col + 1)
    nv = g.ws.cell(row=sub_row, column=g.bench_col + 2)
    cf = g.ws.cell(row=sub_row, column=g.bench_col + 3)
    if exp.get("benchmark_cells") == "empty":
        for cell, lab in ((b, "BENCHMARK"), (s, "% SPREAD"),
                          (nv, "VALID BIDS"), (cf, "CONFIDENCE")):
            d.check(_blankish(cell), ctx,
                    f"{lab}: expected empty, written {cell.value!r}")
        return
    d.check(_close(_num(b), float(exp["median"]), DOLLAR_TOL), ctx,
            f"median: expected {exp['median']}, written {b.value!r}")
    d.check(_close(_num(s), float(exp["spread"]), FRACTION_TOL), ctx,
            f"spread: expected {exp['spread']}, written {s.value!r}")
    d.check(nv.value == exp["n_valid"], ctx,
            f"n_valid: expected {exp['n_valid']}, written {nv.value!r}")
    d.check(cf.value == exp["confidence"], ctx,
            f"confidence: expected {exp['confidence']!r}, written {cf.value!r}")


def _line_row_in_block(g: LeveledGeom, csi: str, desc: str) -> Optional[int]:
    """Find a line row (col B == desc) strictly inside a division block."""
    head = _division_header_row(g.ws, csi)
    sub = g.sub_rows.get(csi)
    if head is None or sub is None:
        return None
    for row in range(head + 1, sub):
        if g.ws.cell(row=row, column=2).value == desc:
            return row
    return None


def _check_line_row(g: LeveledGeom, csi: str, entry: dict, d: Diffs) -> None:
    desc = entry["desc"]
    ctx = f"leveled.line_rows[{csi}][{desc!r}]"
    row = _line_row_in_block(g, csi, desc)
    if not d.check(row is not None, ctx, "line row not found in division block"):
        return
    if "bidder" in entry and "token" in entry:
        # ENC-5 (W-D): line-level classification token — exact text, italic,
        # no Falke paint; every OTHER bidder's cell stays blank.
        cell = g.ws.cell(row=row, column=g.cost(entry["bidder"]))
        d.check(cell.value == entry["token"], ctx,
                f"token: expected {entry['token']!r} exactly, "
                f"written {cell.value!r}")
        if entry.get("italic"):
            d.check(bool(cell.font.italic), ctx, "token must render italic")
        d.check(not _has_falke_paint(cell), ctx,
                f"token cell must carry NO paint, fill is {_fill_hex(cell)}")
        for name in g.idx:
            if name == entry["bidder"]:
                continue
            other = g.ws.cell(row=row, column=g.cost(name))
            d.check(_blankish(other), ctx,
                    f"expected NO value for {name}, carries {other.value!r}")
    elif "bidder" in entry and "value" in entry:
        cell = g.ws.cell(row=row, column=g.cost(entry["bidder"]))
        d.check(_close(_num(cell), float(entry["value"]), DOLLAR_TOL), ctx,
                f"COST: expected {entry['value']}, written {cell.value!r}")
        if "paint" in entry:
            _check_paint(cell, entry["paint"], ctx + ".paint", d)
        _check_comment(cell, entry, ctx + ".comment", d)
    else:
        # note / renders: nothing — row present, NO value for any bidder.
        for name in g.idx:
            cell = g.ws.cell(row=row, column=g.cost(name))
            d.check(_blankish(cell), ctx,
                    f"expected NO value for any bidder, {name} carries "
                    f"{cell.value!r}")


def _check_footer(g: LeveledGeom, footer_exp: dict, d: Diffs) -> None:
    rows = {k: g.footer_row(k) for k in FOOTER_KEY_ORDER}
    for k, r in rows.items():
        d.check(r is not None, f"leveled.footer[{k}]", "footer row not found")
    for bidder, values in footer_exp.items():
        for k, want in zip(FOOTER_KEY_ORDER, values):
            r = rows.get(k)
            if r is None:
                continue
            col = g.csub(bidder) if k == "FEES_SUBTOTAL" else g.cost(bidder)
            cell = g.ws.cell(row=r, column=col)
            d.check(
                _close(_num(cell), float(want), DOLLAR_TOL),
                f"leveled.footer[{bidder}][{k}]",
                f"expected {want}, written {cell.value!r}",
            )


def _check_summary_block(g: LeveledGeom, exp: dict, d: Diffs) -> None:
    rows = {key: _row_col_b_equals(g.ws, label)
            for key, label in SUMMARY_LABELS.items()}
    for key, r in rows.items():
        d.check(r is not None, f"leveled.summary[{key}]",
                f"summary row {SUMMARY_LABELS[key]!r} not found")
    for bidder, vals in exp.items():
        for key, want in vals.items():
            r = rows.get(key)
            if r is None:
                continue
            cell = g.ws.cell(row=r, column=g.csub(bidder))
            ctx = f"leveled.summary[{bidder}][{key}]"
            if key in ("red", "cyan", "yellow"):
                d.check(cell.value == want, ctx,
                        f"expected {want}, written {cell.value!r}")
            elif key == "avg_variance":
                d.check(_close(_num(cell), float(want), FRACTION_TOL), ctx,
                        f"expected {want}, written {cell.value!r}")
            else:
                d.check(_close(_num(cell), float(want), DOLLAR_TOL), ctx,
                        f"expected {want}, written {cell.value!r}")


def _check_alternates(g: LeveledGeom, alternates_exp: list, d: Diffs) -> None:
    header = _row_col_b_contains(g.ws, "Bid Alternates")
    if not d.check(header is not None, "leveled.alternates",
                   "alternates header not found"):
        return
    for alt in alternates_exp:
        ctx = f"leveled.alternates[{alt['bidder']}][{alt['desc_contains']}]"
        found = None
        for row in range(header + 1, header + 30):
            v = g.ws.cell(row=row, column=2).value
            if isinstance(v, str) and alt["desc_contains"] in v \
                    and v.startswith(alt["bidder"]):
                found = row
                break
        if not d.check(found is not None, ctx, "alternate row not found"):
            continue
        cell = g.ws.cell(row=found, column=g.cost(alt["bidder"]))
        d.check(_close(_num(cell), float(alt["value"]), DOLLAR_TOL), ctx,
                f"expected {alt['value']}, written {cell.value!r}")


def _check_presence(ws, presence: dict, d: Diffs) -> None:
    row1 = str(ws.cell(row=1, column=1).value or "")
    row3 = str(ws.cell(row=3, column=1).value or "")
    d.check(presence["banner_row1_contains"] in row1, "leveled.banner_row1",
            f"expected to contain {presence['banner_row1_contains']!r}, "
            f"row 1 is {row1!r}")
    d.check(presence["disclaimer_row3_contains"] in row3,
            "leveled.disclaimer_row3",
            f"expected to contain {presence['disclaimer_row3_contains']!r}, "
            f"row 3 is {row3!r}")
    for key in ("legend_header_contains", "assumptions_header_contains",
                "alternates_header_contains", "summary_header_contains"):
        want = presence[key]
        d.check(_row_col_b_contains(ws, want) is not None,
                f"leveled.presence[{key}]", f"no col-B row contains {want!r}")


# ---------------------------------------------------------------------------
# AUDIT sheet
# ---------------------------------------------------------------------------

def _read_audit_rows(ws) -> list[dict]:
    """Read the AUDIT data region — label-anchored below the column-header
    row (the W-D key block sits at the TOP of the tab, and an inserted
    QUARANTINE line can shift absolute rows)."""
    from src.write_matrix import find_audit_header_row

    out = []
    header = find_audit_header_row(ws)
    if header is None:
        return out
    row = header + 1
    while ws.cell(row=row, column=1).value not in (None, ""):
        out.append({
            "status": ws.cell(row=row, column=1).value,
            "view": ws.cell(row=row, column=2).value,
            "code": ws.cell(row=row, column=3).value,
            "contractor": ws.cell(row=row, column=4).value,
            "division": ws.cell(row=row, column=5).value or "",
            "line_item": ws.cell(row=row, column=6).value or "",
            "value": ws.cell(row=row, column=7).value,
            "message": ws.cell(row=row, column=8).value or "",
            "row": row,
        })
        row += 1
    return out


def _audit_key(r: dict) -> tuple:
    return (r["contractor"], r["division"], r["code"], r["status"], r["view"])


def _check_audit_register(actual_rows: list[dict], exp_rows_by_contractor: dict,
                          d: Diffs, ctx_prefix: str = "audit") -> None:
    """Set-based register comparison (README: assert as a SET keyed
    (contractor, division, code, status, view); value exact where given)."""
    remaining = list(actual_rows)
    for contractor, rows in exp_rows_by_contractor.items():
        for spec in rows:
            division, code, status, view, value = (
                spec[0] or "", spec[1], spec[2], VIEW_LABEL[spec[3]], spec[4],
            )
            key = (contractor, division, code, status, view)
            ctx = f"{ctx_prefix}[{contractor}][{division or 'bid-level'}][{code}]"
            match = None
            for r in remaining:
                if _audit_key(r) != key:
                    continue
                if value is None:
                    if r["value"] in (None, ""):
                        match = r
                        break
                elif r["value"] == value:
                    match = r
                    break
            if match is not None:
                remaining.remove(match)
                continue
            # Key match with wrong value → report the value specifically.
            key_only = [r for r in remaining if _audit_key(r) == key]
            if key_only:
                r = key_only[0]
                remaining.remove(r)
                d.add(ctx, f"Value column: expected {value!r}, "
                           f"written {r['value']!r} (row {r['row']})")
            else:
                d.add(ctx, f"expected register row missing "
                           f"(status={status}, view={view}, value={value!r})")
    for r in remaining:
        d.add(f"{ctx_prefix}.unexpected",
              f"row {r['row']}: {r['status']} | {r['view']} | {r['code']} | "
              f"{r['contractor']} | {r['division']} | value={r['value']!r}")


def _check_audit_totals(actual_rows: list[dict], totals: dict, d: Diffs,
                        ctx: str = "audit.totals") -> None:
    got = {
        "RED": sum(1 for r in actual_rows if r["status"] == "RED"),
        "YELLOW": sum(1 for r in actual_rows if r["status"] == "YELLOW"),
        "GREEN": sum(1 for r in actual_rows if r["status"] == "GREEN"),
        "rows": len(actual_rows),
    }
    for k in ("RED", "YELLOW", "GREEN", "rows"):
        d.check(got[k] == totals[k], ctx,
                f"{k}: expected {totals[k]}, sheet has {got[k]}")


# ---------------------------------------------------------------------------
# Pipeline execution + overlay swap discipline
# ---------------------------------------------------------------------------

@contextmanager
def gold_overlay():
    """Install the gold overlay as config/known_firms.local.yaml (swap, never
    merge); back up + restore any pre-existing local overlay. Loud on a
    missing source overlay (contract §1)."""
    if not OVERLAY_SRC.exists():
        raise RuntimeError(
            f"GOLDEN-SET overlay missing: {OVERLAY_SRC} — the gold run MUST "
            f"execute with exactly this overlay (README §Harness 1). Refusing "
            f"to run without it; a silently-absent overlay would invalidate "
            f"the Beacon reclass expectations."
        )
    had_existing = OVERLAY_DST.exists()
    try:
        if had_existing:
            shutil.move(str(OVERLAY_DST), str(OVERLAY_BAK))
        shutil.copyfile(str(OVERLAY_SRC), str(OVERLAY_DST))
        if not OVERLAY_DST.exists():   # belt and braces — never run silent
            raise RuntimeError(f"overlay install failed: {OVERLAY_DST}")
        yield
    finally:
        if OVERLAY_DST.exists():
            OVERLAY_DST.unlink()
        if had_existing and OVERLAY_BAK.exists():
            shutil.move(str(OVERLAY_BAK), str(OVERLAY_DST))


def run_gold_pipeline(interim_dir: Path, out_path: Path) -> SimpleNamespace:
    """README §Harness 2/3 — the exact documented invocation, subprocess so
    the exit code is asserted for real."""
    env = dict(os.environ)
    env["PYTHONPATH"] = str(ENGINE_ROOT)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [sys.executable, "-m", "src.pipeline",
         "--interim-dir", str(interim_dir),
         "--project-config", str(GOLD / "project.yaml"),
         "--sf-basis", "100000",
         "--out", str(out_path)],
        cwd=str(ENGINE_ROOT), env=env, capture_output=True, text=True,
        timeout=300,
    )
    return SimpleNamespace(
        exit_code=proc.returncode,
        stdout=proc.stdout,
        stderr=proc.stderr,
        out_path=out_path,
    )


def _load_yaml(path: Path) -> dict:
    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


# ---------------------------------------------------------------------------
# MAIN-RUN checker (expectations.yaml)
# ---------------------------------------------------------------------------

def check_main_run(run: SimpleNamespace, exp: dict) -> Diffs:
    d = Diffs()
    d.check(run.exit_code == exp["run"]["expected_exit_code"], "run.exit_code",
            f"expected {exp['run']['expected_exit_code']}, got {run.exit_code}"
            f"\n--- stderr tail ---\n{run.stderr[-2000:]}")
    if not d.check(run.out_path.exists(), "run.output",
                   f"{run.out_path} was not written"):
        return d

    wb = openpyxl.load_workbook(run.out_path)
    d.check(wb.sheetnames == exp["run"]["expected_sheets"], "run.sheets",
            f"expected {exp['run']['expected_sheets']}, got {wb.sheetnames}")

    bidder_order = exp["bidder_order"]
    lev_exp = exp["leveled_sheet"]
    ws = wb["Leveled_Normalized"]
    g = LeveledGeom(ws, bidder_order, d)

    _check_presence(ws, lev_exp["presence"], d)

    for csi, bexp in lev_exp["benchmarks"].items():
        _check_benchmark(g, csi, bexp, d)

    for csi, cells in lev_exp["divisions"].items():
        if "all_bidders" in cells:
            for bidder in bidder_order:
                _check_division_cell(g, csi, bidder, cells["all_bidders"], d)
        else:
            for bidder, cexp in cells.items():
                _check_division_cell(g, csi, bidder, cexp, d)

    for csi, entries in lev_exp.get("line_rows", {}).items():
        for entry in entries:
            _check_line_row(g, csi, entry, d)

    _check_footer(g, lev_exp["footer"], d)

    # Grand-total paint: none for all bidders (R21 clean)
    gt_row = g.footer_row("GRAND_TOTAL")
    if d.check(gt_row is not None, "leveled.grand_total", "GT row not found") \
            and lev_exp.get("grand_total_paint") == "none_for_all_bidders":
        for bidder in bidder_order:
            cell = ws.cell(row=gt_row, column=g.cost(bidder))
            d.check(not _has_falke_paint(cell),
                    f"leveled.grand_total_paint[{bidder}]",
                    f"expected no paint, fill is {_fill_hex(cell)}")

    _check_alternates(g, lev_exp["alternates"], d)
    _check_summary_block(g, lev_exp["summary_block"], d)

    # Paint totals — PaintTracker counts printed on the console report.
    pt = lev_exp["paint_totals"]
    m = re.search(
        r"red=(\d+) cyan=(\d+) yellow=(\d+) neutral=(\d+) "
        r"paint_suppressed_lt3_bids=(\d+)", run.stdout)
    if d.check(m is not None, "paint_totals",
               "console paint-count line not found in stdout"):
        got = dict(zip(("red", "cyan", "yellow", "neutral", "gate_suppressed"),
                       (int(x) for x in m.groups())))
        for k in ("red", "cyan", "yellow", "neutral", "gate_suppressed"):
            d.check(got[k] == pt[k], f"paint_totals.{k}",
                    f"expected {pt[k]}, console reports {got[k]}")

    # Workbook LEGEND — ONE identical block on BOTH data sheets (W-D B4/§2),
    # with the precedence line.
    for sheet in exp.get("legend_present_on", []):
        ws_l = wb[sheet]
        header_row = _row_col_b_contains(
            ws_l, exp["leveled_sheet"]["presence"]["legend_header_contains"])
        d.check(header_row is not None, f"legend[{sheet}].header",
                "workbook legend header not found in col B")
        prec = exp.get("legend_precedence_contains")
        if prec:
            d.check(_row_col_b_contains(ws_l, prec) is not None,
                    f"legend[{sheet}].precedence",
                    f"no col-B row contains {prec!r}")

    # AUDIT register — full set + totals.
    audit_rows = _read_audit_rows(wb["AUDIT"])
    _check_audit_totals(audit_rows, exp["audit_sheet"]["totals"], d)
    _check_audit_register(audit_rows, exp["audit_sheet"]["rows"], d)

    # Mirror (Bid_Form) — minimal v1 assertions.
    mws = wb["Bid_Form"]
    mg = MirrorGeom(mws, bidder_order, d)
    mgt_row = mg.footer_row("GRAND_TOTAL")
    if d.check(mgt_row is not None, "mirror.grand_total",
               "GRAND_TOTAL footer key not found in col A"):
        for bidder, want in exp["mirror_sheet"]["grand_totals"].items():
            cell = mws.cell(row=mgt_row, column=mg.col(bidder))
            d.check(_close(_num(cell), float(want), DOLLAR_TOL),
                    f"mirror.grand_totals[{bidder}]",
                    f"expected {want}, written {cell.value!r}")
    rf = exp["mirror_sheet"]["reclass_faithfulness"]
    for csi, want, label in (
        ("DIV 01 00 00", rf["beacon_div01_subtotal"], "beacon_div01_subtotal"),
        ("DIV 11 00 00", rf["beacon_div11_subtotal"], "beacon_div11_subtotal"),
    ):
        r = mg.sub_rows.get(csi)
        if not d.check(r is not None, f"mirror.{label}",
                       f"{csi} SUBTOTAL row not found"):
            continue
        cell = mws.cell(row=r, column=mg.col("Beacon Shoreline Builders Inc."))
        d.check(_close(_num(cell), float(want), DOLLAR_TOL), f"mirror.{label}",
                f"expected {want}, written {cell.value!r}")
    note_want = rf["beacon_div11_note_contains"]
    note_found = any(
        isinstance(mws.cell(row=r, column=3).value, str)
        and note_want in mws.cell(row=r, column=3).value
        for r in range(1, mws.max_row + 1)
    )
    d.check(note_found, "mirror.beacon_div11_note",
            f"no col-C Normalization Note contains {note_want!r}")
    r22 = mg.sub_rows.get("DIV 22 00 00")
    if d.check(r22 is not None, "mirror.echo_div22", "DIV 22 SUBTOTAL not found"):
        cell = mws.cell(row=r22, column=mg.col(
            "Eastline Mechanical & Restoration Corp"))
        d.check(_close(_num(cell),
                       float(exp["mirror_sheet"]["echo_div22_subtotal"]),
                       DOLLAR_TOL),
                "mirror.echo_div22_subtotal",
                f"expected {exp['mirror_sheet']['echo_div22_subtotal']}, "
                f"written {cell.value!r}")

    # --- Mirror v1.1 (W-D M-3): blanks-as-blanks, verbatim tokens, no ARA
    # fills on data cells ---
    tok = exp["mirror_sheet"].get("tokens")
    if tok:
        for e in tok.get("blank_subtotals", []):
            ctx = f"mirror.tokens.blank[{e['bidder']}][{e['division']}]"
            r = mg.sub_rows.get(e["division"])
            if not d.check(r is not None, ctx, "SUBTOTAL row not found"):
                continue
            cell = mws.cell(row=r, column=mg.col(e["bidder"]))
            d.check(_blankish(cell), ctx,
                    f"expected truly BLANK subtotal, written {cell.value!r}")
            sf = mws.cell(row=r, column=mg.col(e["bidder"]) + 1)
            d.check(_blankish(sf), ctx,
                    f"expected blank $/SF beside a blank subtotal, "
                    f"written {sf.value!r}")
        for e in tok.get("line_tokens", []):
            ctx = (f"mirror.tokens.line[{e['bidder']}][{e['division']}]"
                   f"[{e['desc']!r}]")
            row = _mirror_line_row(mws, mg.sub_rows, e["division"], e["desc"])
            if not d.check(row is not None, ctx,
                           "line row not found in division block"):
                continue
            cell = mws.cell(row=row, column=mg.col(e["bidder"]))
            d.check(cell.value == e["token"], ctx,
                    f"expected token {e['token']!r} exactly, "
                    f"written {cell.value!r}")
            if e.get("italic"):
                d.check(bool(cell.font.italic), ctx,
                        "token must render italic")
        for e in tok.get("explicit_zero_subtotals", []):
            ctx = f"mirror.tokens.zero[{e['bidder']}][{e['division']}]"
            r = mg.sub_rows.get(e["division"])
            if not d.check(r is not None, ctx, "SUBTOTAL row not found"):
                continue
            cell = mws.cell(row=r, column=mg.col(e["bidder"]))
            d.check(_close(_num(cell), float(e["value"]), DOLLAR_TOL), ctx,
                    f"EXPLICIT_ZERO must stay numeric "
                    f"{e['value']}, written {cell.value!r}")
        if tok.get("no_ara_fills"):
            ara_hexes = {"FFCCCC", "FFF2CC", "CCFFCC"}
            gt_r = mgt_row or mws.max_row
            offenders = [
                c.coordinate
                for row in mws.iter_rows(min_row=1, max_row=gt_r)
                for c in row
                if c.column != 3 and _fill_hex(c) in ara_hexes
            ]
            d.check(offenders == [], "mirror.tokens.no_ara_fills",
                    f"ARA fills on mirror data cells: {offenders}")
    return d


# ---------------------------------------------------------------------------
# BIDDER-FOOTER-ERROR-RUN checker (expectations-quarantine.yaml, clean exit 0
# since the GOLD-DEV-6 bidder-error branch landed)
# ---------------------------------------------------------------------------

def check_quarantine_run(run: SimpleNamespace, exp: dict) -> Diffs:
    d = Diffs()
    d.check(run.exit_code == exp["run"]["expected_exit_code"], "quar.exit_code",
            f"expected {exp['run']['expected_exit_code']}, got {run.exit_code}"
            f"\n--- stderr tail ---\n{run.stderr[-2000:]}")
    delivered = run.out_path.exists()
    d.check(delivered == exp["run"]["file_delivered"], "quar.file_delivered",
            f"expected file_delivered={exp['run']['file_delivered']}, "
            f"exists={delivered}")
    if not delivered:
        return d

    wb = openpyxl.load_workbook(run.out_path)
    d.check(wb.sheetnames == exp["run"]["expected_sheets"], "quar.sheets",
            f"expected {exp['run']['expected_sheets']}, got {wb.sheetnames}")
    bidder_order = exp["bidder_order"]
    lev_exp = exp["leveled_sheet"]
    ws = wb["Leveled_Normalized"]
    g = LeveledGeom(ws, bidder_order, d)

    for csi, bexp in lev_exp["benchmarks"].items():
        _check_benchmark(g, csi, bexp, d)
    for csi, cells in lev_exp["divisions"].items():
        for bidder, cexp in cells.items():
            _check_division_cell(g, csi, bidder, cexp, d)
    _check_footer(g, lev_exp["footer"], d)
    _check_summary_block(g, lev_exp["summary_block"], d)

    # Grand-total paint per bidder — the R21 red + comment must SURVIVE on
    # Quayside's GT cell (D2 fixed: no quarantine mark overwrites it).
    gt_row = g.footer_row("GRAND_TOTAL")
    if d.check(gt_row is not None, "quar.leveled.grand_total",
               "GT row not found"):
        for bidder, pexp in lev_exp["grand_total_paint"].items():
            cell = ws.cell(row=gt_row, column=g.cost(bidder))
            ctx = f"quar.grand_total_paint[{bidder}]"
            _check_paint(cell, pexp.get("paint"), ctx, d)
            _check_comment(cell, pexp, ctx + ".comment", d)

    # Clean-run ABSENCE assertions — the quarantine chain must NOT fire on a
    # faithfully-reproduced bidder footer error (GOLD-DEV-6 fixed).
    clean = exp["clean_run"]
    for sheet in ("Bid_Form", "Leveled_Normalized"):
        row1 = str(wb[sheet].cell(row=1, column=1).value or "")
        d.check(clean["no_quarantine_banner_contains"] not in row1,
                f"quar.clean_run.banner[{sheet}]",
                f"quarantine banner present on a bidder-error run: {row1!r}")

    audit_rows = _read_audit_rows(wb["AUDIT"])
    tieout_rows = [r for r in audit_rows
                   if r["code"] == "POST_WRITE_TIEOUT_FAILURE"]
    if clean["no_tieout_rows"]:
        d.check(tieout_rows == [], "quar.clean_run.tieout_rows",
                f"{len(tieout_rows)} POST_WRITE_TIEOUT_FAILURE row(s) on a "
                f"bidder-error run (expected none)")
    _check_audit_totals(audit_rows, exp["audit_sheet"]["totals"], d,
                        ctx="quar.audit.totals")
    _check_audit_register(audit_rows, exp["audit_sheet"]["rows"], d,
                          ctx_prefix="quar.audit")

    wsa = wb["AUDIT"]
    found = any(
        isinstance(wsa.cell(row=r, column=1).value, str)
        and clean["no_quarantine_summary_line_contains"]
        in wsa.cell(row=r, column=1).value
        for r in range(1, wsa.max_row + 1)
    )
    d.check(not found, "quar.clean_run.summary_line",
            "a QUARANTINE summary line is present on a bidder-error run")
    return d


# ---------------------------------------------------------------------------
# FAULT-INJECTION checker — the quarantine chain proven on a TRUE tool defect
# (Marvin GOLD-DEV-6 ruling (4); Floyd W2-5: this step lands in the same PR
# that flips the variant to exit 0, so the disclosure-chain regression
# coverage never drops. Pipeline exit-3 mapping itself stays covered by
# tests/test_reconcile.py::test_pipeline_quarantines_and_delivers_on_tieout_failure.)
# ---------------------------------------------------------------------------

def _rebuild_gold_bids():
    """Rebuild the main-set NormalizedBids in-process so the fault-injection
    step can call reconcile_written_matrix/apply_quarantine directly. Loads
    the known-firms config EXPLICITLY (the gold overlay must be installed —
    call inside gold_overlay()); no module-cache pollution."""
    import json

    from src.audit import audit_bids
    from src.firm_config import load_known_firms
    from src.models import BidDocument
    from src.normalize import (
        build_normalized_view,
        compute_cross_bid_stats,
        normalize_bid,
    )

    cfg = load_known_firms()
    mirrors, leveled = [], []
    for p in sorted((GOLD / "interim").glob("*.json")):
        doc = BidDocument.model_validate(json.loads(p.read_text("utf-8")))
        mirror = normalize_bid(doc, known_firms=cfg)
        mirrors.append(mirror)
        leveled.append(build_normalized_view(mirror, doc))
    leveled = compute_cross_bid_stats(leveled)
    audit_items = audit_bids(leveled)
    return mirrors, leveled, audit_items


def check_fault_injection(main_run: SimpleNamespace, main_exp: dict,
                          exp: dict) -> Diffs:
    from src.reconcile import reconcile_written_matrix
    from src.write_matrix import apply_quarantine

    d = Diffs()
    fi = exp["fault_injection"]
    if not d.check(main_run.out_path.exists(), "fault.source",
                   "main-run workbook missing — cannot fault-inject"):
        return d
    work = main_run.out_path.with_name("gold-v1-fault-injected.xlsx")
    shutil.copyfile(main_run.out_path, work)
    bidder_order = main_exp["bidder_order"]

    with gold_overlay():
        mirrors, leveled, audit_items = _rebuild_gold_bids()

    # Sanity: the rebuilt bids tie out CLEAN against the untouched copy —
    # proves the in-process rebuild matches the subprocess run before we
    # corrupt anything.
    pre = reconcile_written_matrix(work, mirrors, len(audit_items),
                                   leveled_bids=leveled)
    if not d.check(pre == [], "fault.pre_corruption",
                   f"rebuilt bids must tie out clean pre-corruption; got "
                   f"{len(pre)} failure(s): "
                   f"{[f.message[:80] for f in pre[:3]]}"):
        return d

    # Corrupt (1) the first bidder's GRAND TOTAL on Bid_Form (a true tool
    # defect on a GT cell) and (2) a COMMENTED division-subtotal cell on
    # Leveled_Normalized (Delta DIV 03 carries the R20 red + comment) so the
    # compose-not-overwrite behavior is proven on a real prior flag.
    wb = openpyxl.load_workbook(work)
    wsm = wb["Bid_Form"]
    gt_row = _row_col_a_equals(wsm, "GRAND_TOTAL")
    if not d.check(gt_row is not None, "fault.corrupt",
                   "Bid_Form GRAND_TOTAL row not found"):
        return d
    wsm.cell(row=gt_row, column=_col_start(fi["corrupt"]["bidder_index"])
             ).value = float(fi["corrupt"]["value"])

    wsl = wb["Leveled_Normalized"]
    gl = LeveledGeom(wsl, bidder_order, Diffs())  # geometry only
    div_bidder = "Delta Gulf Contracting LLC"
    div_csi = "DIV 03 00 00"
    sub_cell = wsl.cell(row=gl.sub_rows[div_csi], column=gl.csub(div_bidder))
    prior_comment = _comment_text(sub_cell)
    d.check("R20" in prior_comment, "fault.compose_target",
            f"expected an existing R20 comment on {div_bidder} {div_csi} "
            f"(compose proof target), found {prior_comment!r}")
    sub_cell.value = float(sub_cell.value) + 77777.0
    wb.save(work)

    # Direct reconcile: the corruption must fire — including check 2 on the
    # corrupted bidder (arith_delta moves AWAY from the bidder's own footer
    # delta, proving the bidder-error branch does not blind the checker).
    failures = reconcile_written_matrix(work, mirrors, len(audit_items),
                                        leveled_bids=leveled)
    d.check(any("Grand-total tie-out FAILED" in f.message for f in failures),
            "fault.check1", "GT corruption did not trip check 1")
    d.check(any("Footer arithmetic FAILED" in f.message for f in failures),
            "fault.check2",
            "GT corruption did not trip check 2 (bidder-error branch must "
            "not suppress a REAL write defect)")
    d.check(any("Division subtotal tie-out FAILED" in f.message
                and f.division_csi == div_csi for f in failures),
            "fault.check3", "division corruption did not trip check 3")
    if not failures:
        return d

    apply_quarantine(output_path=work, failures=failures, bids=mirrors,
                     leveled_bids=leveled)

    wb2 = openpyxl.load_workbook(work)
    # 1. RED banner, both data sheets.
    for sheet in ("Bid_Form", "Leveled_Normalized"):
        row1 = wb2[sheet].cell(row=1, column=1).value
        row3 = str(wb2[sheet].cell(row=3, column=1).value or "")
        d.check(row1 == fi["banner_line1_exact"],
                f"fault.banner[{sheet}].line1",
                f"expected exact {fi['banner_line1_exact']!r}, got {row1!r}")
        d.check(fi["banner_line3_contains"] in row3,
                f"fault.banner[{sheet}].line3",
                f"expected to contain {fi['banner_line3_contains']!r}, "
                f"got {row3!r}")

    # 2. Quarantine mark on the corrupted GT cell (Bid_Form).
    mg = MirrorGeom(wb2["Bid_Form"], bidder_order, Diffs())
    gt_row2 = mg.footer_row("GRAND_TOTAL")
    corrupt_bidder = bidder_order[fi["corrupt"]["bidder_index"]]
    gt_cell = wb2["Bid_Form"].cell(row=gt_row2, column=mg.col(corrupt_bidder))
    d.check(_fill_hex(gt_cell) in set(fi["quarantine_fill_hexes"]),
            "fault.gt_mark.fill",
            f"expected quarantine fill, got {_fill_hex(gt_cell)}")
    d.check(fi["mark_comment_contains"] in _comment_text(gt_cell),
            "fault.gt_mark.comment",
            f"comment missing {fi['mark_comment_contains']!r} "
            f"(comment is {_comment_text(gt_cell)!r})")

    # 3. COMPOSED mark on the previously-commented division cell: quarantine
    # fill wins visually, the prior R20 story survives below the separator.
    gl2 = LeveledGeom(wb2["Leveled_Normalized"], bidder_order, Diffs())
    sub2 = wb2["Leveled_Normalized"].cell(
        row=gl2.sub_rows[div_csi], column=gl2.csub(div_bidder))
    d.check(_fill_hex(sub2) in set(fi["quarantine_fill_hexes"]),
            "fault.compose.fill",
            f"expected quarantine fill on the composed mark, "
            f"got {_fill_hex(sub2)}")
    text2 = _comment_text(sub2)
    d.check(fi["mark_comment_contains"] in text2, "fault.compose.quarantine",
            f"composed comment missing the quarantine text ({text2!r})")
    d.check("--- prior flag on this cell ---" in text2,
            "fault.compose.separator",
            f"composed comment missing the prior-flag separator ({text2!r})")
    d.check("R20" in text2, "fault.compose.prior_story",
            f"the prior R20 story was ERASED by the quarantine mark "
            f"({text2!r})")

    # 4. Appended RED AUDIT rows + the QUARANTINE summary line.
    audit_rows = _read_audit_rows(wb2["AUDIT"])
    tieout_rows = [r for r in audit_rows
                   if r["code"] == "POST_WRITE_TIEOUT_FAILURE"]
    d.check(len(tieout_rows) == len(failures), "fault.audit_rows",
            f"expected {len(failures)} POST_WRITE_TIEOUT_FAILURE rows, "
            f"found {len(tieout_rows)}")
    d.check(all(r["status"] == "RED" for r in tieout_rows),
            "fault.audit_rows.status", "tie-out rows must be RED")
    wsa = wb2["AUDIT"]
    found = any(
        isinstance(wsa.cell(row=r, column=1).value, str)
        and fi["audit_summary_line_contains"] in wsa.cell(row=r, column=1).value
        for r in range(1, wsa.max_row + 1)
    )
    d.check(found, "fault.summary_line",
            f"no AUDIT col-A line contains "
            f"{fi['audit_summary_line_contains']!r}")
    return d


# ---------------------------------------------------------------------------
# pytest wiring — one overlay swap, both runs, module-scoped
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def gold_runs(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("gold-v1")
    with gold_overlay():
        main = run_gold_pipeline(GOLD / "interim", tmp / "gold-v1.xlsx")
        quar = run_gold_pipeline(GOLD / "interim-quarantine",
                                 tmp / "gold-v1-quarantine.xlsx")
    return {"main": main, "quar": quar}


def test_golden_main_run(gold_runs):
    exp = _load_yaml(GOLD / "expectations.yaml")
    d = check_main_run(gold_runs["main"], exp)
    assert not d.items, (
        f"\nGOLDEN-SET main run: {len(d.items)} mismatch(es) vs "
        f"expectations.yaml (Marvin adjudicates — do not silently fix either "
        f"side):\n{d.report()}"
    )


def test_golden_quarantine_run(gold_runs):
    exp = _load_yaml(GOLD / "expectations-quarantine.yaml")
    d = check_quarantine_run(gold_runs["quar"], exp)
    assert not d.items, (
        f"\nGOLDEN-SET bidder-footer-error run: {len(d.items)} mismatch(es) "
        f"vs expectations-quarantine.yaml (Marvin adjudicates):\n{d.report()}"
    )


def test_golden_fault_injection(gold_runs):
    """The quarantine disclosure chain, proven on a TRUE tool defect —
    a corrupted copy of the main-run workbook (Floyd W2-5: this coverage
    lands in the same change that flips the variant run to exit 0)."""
    main_exp = _load_yaml(GOLD / "expectations.yaml")
    exp = _load_yaml(GOLD / "expectations-quarantine.yaml")
    d = check_fault_injection(gold_runs["main"], main_exp, exp)
    assert not d.items, (
        f"\nGOLDEN-SET fault injection: {len(d.items)} mismatch(es) "
        f"(Marvin adjudicates):\n{d.report()}"
    )


# ---------------------------------------------------------------------------
# Standalone runner — python3 eval/test_golden_set.py
# ---------------------------------------------------------------------------

def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="gold-v1-"))
    with gold_overlay():
        main_run = run_gold_pipeline(GOLD / "interim", tmp / "gold-v1.xlsx")
        quar_run = run_gold_pipeline(GOLD / "interim-quarantine",
                                     tmp / "gold-v1-quarantine.xlsx")
    main_exp = _load_yaml(GOLD / "expectations.yaml")
    quar_exp = _load_yaml(GOLD / "expectations-quarantine.yaml")
    d_main = check_main_run(main_run, main_exp)
    d_quar = check_quarantine_run(quar_run, quar_exp)
    d_fault = check_fault_injection(main_run, main_exp, quar_exp)
    print("=" * 70)
    print("GOLDEN-SET v1 harness — standalone report")
    print("=" * 70)
    print(f"Main run:               exit {main_run.exit_code} — "
          f"{len(d_main.items)} mismatch(es)")
    if d_main.items:
        print(d_main.report())
    print(f"Bidder-error run:       exit {quar_run.exit_code} — "
          f"{len(d_quar.items)} mismatch(es)")
    if d_quar.items:
        print(d_quar.report())
    print(f"Fault-injection step:   {len(d_fault.items)} mismatch(es)")
    if d_fault.items:
        print(d_fault.report())
    print(f"Workbooks kept for inspection under: {tmp}")
    return 1 if (d_main.items or d_quar.items or d_fault.items) else 0


if __name__ == "__main__":
    sys.exit(main())
