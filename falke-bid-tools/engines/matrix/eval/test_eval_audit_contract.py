"""
FALKE Matrix — Adversarial Audit-Contract Eval (Floyd's V&V evidence pack)
==========================================================================
Author: Boris (ARA Applied AI Architect) — Phase D of the matrix-generalization
sprint. This is the EVAL, not Christine's unit tests. It is the evidence Floyd
uses at the V&V gate.

What it proves (the C4 contract):
    For an ADVERSARIAL bid set that exercises every code-format / firm-quirk
    failure mode, the correct RED/YELLOW AuditCode actually LANDS on the written
    AUDIT sheet — i.e. it survives the full path
        normalize_bid → compute_cross_bid_stats → audit_bids → write_matrix
    and is read back OUT OF THE .xlsx FILE, not merely off the in-memory item
    list. (Christine's tests check the in-memory items; this eval closes the loop
    through the actual file the board receives.)

Method (Boris's grader hierarchy — strongest verifiable check first):
    Code-graded, exact assertion. Each case names the contractor, the AuditCode
    it must produce, the required STATUS (RED/YELLOW), and one or more NEGATIVE
    codes that must NOT appear for that contractor (the no-silent-mislevel and
    no-partial-remap guarantees). We open the saved workbook, read the AUDIT
    sheet (col A=Status, col B=Code, col C=Contractor per _write_audit_sheet),
    and verify each expectation.

Coverage (the brief's required minimum, mapped to Marvin's gold cases §7):
    GS-5  mixed canonical+legacy   → RED  UNRECOGNIZED_CODE_FORMAT, no partial remap
    GS-4  clean legacy-2-digit     → YELLOW CODE_FORMAT_REMAPPED (csi_1995_2digit)
    GS-7  name collides >1 profile → RED  KNOWN_FIRM_AMBIGUOUS
    GS-3  new firm, legit DIV 13   → NOT reclassed (no KNOWN_FIRM_RECLASSIFIED)
    NEW   code-15 AND code-16 lump → RED  CODE_SPLIT_UNMATCHED ×2 (code-16 bug fix)
    GS-1  single bidder            → no cross-bid flags + single-bid notice
    GS-6  Acme reclass (support) → YELLOW KNOWN_FIRM_RECLASSIFIED (positive foil to GS-3)

It is named ``test_*`` so pytest collects it at the directory level (Floyd's CI
gate can run ``pytest eval/``), while ``main()`` keeps it runnable standalone.

Run standalone (prints a PASS/FAIL summary, exit 0 / non-0):
    PYTHONPATH=<engine root> python3 eval/test_eval_audit_contract.py
Or under pytest (each case is a collectable test):
    PYTHONPATH=<engine root> python3 -m pytest eval/ -v
"""

from __future__ import annotations

import sys
import tempfile
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

# Allow `python3 eval/test_eval_audit_contract.py` from the engine root.
_ENGINE_ROOT = Path(__file__).resolve().parent.parent
if str(_ENGINE_ROOT) not in sys.path:
    sys.path.insert(0, str(_ENGINE_ROOT))

from src.audit import audit_bids  # noqa: E402
from src.firm_config import (  # noqa: E402
    Firm,
    KnownFirmsConfig,
    Reclassification,
    load_known_firms,
)
from src.models import (  # noqa: E402
    BidDocument,
    BidFooter,
    BidQualifications,
    ClassificationSource,
    CostStructure,
    DivisionBid,
    ExtractionConfidence,
    FormType,
    GrandTotalConfidence,
    InputType,
    LineItem,
)
from src.normalize import compute_cross_bid_stats, normalize_bid  # noqa: E402
from src.run_config import RunInputs  # noqa: E402
from src.write_matrix import write_matrix  # noqa: E402


# ---------------------------------------------------------------------------
# Bid-fixture builders (adversarial inputs)
# ---------------------------------------------------------------------------

def _footer(grand_total=None, construction=None, gc_fee=None, alternates=None):
    return BidFooter(
        construction_cost_subtotal=construction,
        gc_fee=gc_fee,
        grand_total=grand_total,
        alternates=alternates or [],
        grand_total_confidence=GrandTotalConfidence.LOW,
    )


def _doc(name, divisions, footer=None, form_type=FormType.FALKE_STANDARD):
    return BidDocument(
        contractor_name=name,
        form_type=form_type,
        bid_document_input_type=InputType.DIGITAL_NATIVE,
        divisions=divisions,
        footer=footer or _footer(),
        qualifications=BidQualifications(),
        extraction_confidence=ExtractionConfidence.HIGH,
    )


def _div(code, name, items=None, subtotal=None, cost=CostStructure.LUMP_SUM):
    return DivisionBid(
        csi_code=code,
        division_name=name,
        cost_structure=cost,
        division_subtotal=subtotal,
        classification_source=ClassificationSource.CONTRACTOR_NATIVE,
        contractor_native_code=None,
        line_items=items or [],
    )


def _run_inputs(gsf=10_000.0):
    """Per-run identity — a synthetic project, to prove generalization."""
    return RunInputs(
        project_name="Adversarial Eval Tower",
        project_address="1 V&V Way, Eval City FL 00000",
        gross_sf=gsf,
        sf_basis_label="balcony SF",
        sf_source="explicit",
    )


# --- The adversarial bidders (one per failure mode) ------------------------

def _bid_gs5_mixed():
    """GS-5: BOTH a canonical DIV 03 and bare legacy 15/16 on one bid.
    Expect RED UNRECOGNIZED_CODE_FORMAT; NO partial remap (no DIV 22 etc.)."""
    return _doc("Mixed Codes Inc", [
        _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        _div("15", "Mechanical", subtotal=Decimal("120000")),
        _div("16", "Electrical", subtotal=Decimal("90000")),
    ], form_type=FormType.CONTRACTOR_OWN,
       footer=_footer(grand_total=Decimal("610000")))


def _bid_gs4_clean_legacy():
    """GS-4: clean legacy-2-digit schedule → csi_1995_2digit remap, all routed.
    Expect YELLOW CODE_FORMAT_REMAPPED; NO RED, NO CODE_SPLIT_UNMATCHED."""
    return _doc("Legacy Format Co", [
        _div("01", "General Requirements", subtotal=Decimal("50000")),
        _div("03", "Concrete", subtotal=Decimal("400000")),
        _div("07", "Thermal", subtotal=Decimal("150000")),
        _div("09", "Finishes", subtotal=Decimal("80000")),
        _div("15", "Mechanical", cost=CostStructure.ITEMIZED, items=[
            LineItem(description="Domestic water piping", amount=Decimal("60000")),
            LineItem(description="HVAC ductwork", amount=Decimal("90000")),
        ]),
        _div("16", "Electrical", cost=CostStructure.ITEMIZED, items=[
            LineItem(description="Branch wiring & devices", amount=Decimal("120000")),
            LineItem(description="Fire alarm notification", amount=Decimal("30000")),
        ]),
        _div("17-040", "OH&P", subtotal=Decimal("70000")),
    ], form_type=FormType.CONTRACTOR_OWN,
       footer=_footer(grand_total=Decimal("1020000")))


def _bid_split_unmatched():
    """code-16 bug-fix proof: a legacy bid where code 15 AND code 16 each carry a
    lump-sum line whose description matches NO routing keyword. BOTH must fall to
    RED CODE_SPLIT_UNMATCHED. The old inline code-16 branch could never emit this
    flag (everything non-alarm routed to DIV 26); this case proves it now can.

    Other lines are present so the bid is a clean legacy schedule (passes the §1
    detection boolean: ≥3 legacy codes, a 15/16/17 discriminator, no canonical,
    no unknown) — otherwise it would be UNRECOGNIZED, not split."""
    return _doc("Vague Split Co", [
        _div("01", "General Requirements", subtotal=Decimal("50000")),
        _div("03", "Concrete", subtotal=Decimal("300000")),
        # Both descriptions are deliberately keyword-free: they match NO §2.2
        # routing term for either trade, so each must fall to CODE_SPLIT_UNMATCHED.
        # (Avoid words like "mechanical"/"electrical" — those ARE routing keywords.)
        _div("15", "Mechanical", cost=CostStructure.ITEMIZED, items=[
            LineItem(description="Misc. allowance — see clarification", amount=Decimal("11000")),
        ]),
        _div("16", "Electrical", cost=CostStructure.ITEMIZED, items=[
            LineItem(description="Misc. allowance", amount=Decimal("9000")),
        ]),
    ], form_type=FormType.CONTRACTOR_OWN,
       footer=_footer(grand_total=Decimal("370000")))


def _bid_gs3_new_firm_div13():
    """GS-3: a NEW firm (not in known_firms.yaml) with a legitimate DIV 13
    Special-Construction line. Must NOT be reclassed (no KNOWN_FIRM_RECLASSIFIED,
    no UNRECOGNIZED). Critical negative test from the reclass side."""
    return _doc("Coastal Concrete Restoration LLC", [
        _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
        _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
             items=[LineItem(description="Pool-deck system", amount=Decimal("60000"))],
             subtotal=Decimal("60000")),
        _div("DIV 11 00 00", "Equipment", cost=CostStructure.ITEMIZED,
             items=[LineItem(description="Residential equipment", amount=Decimal("20000"))],
             subtotal=Decimal("20000")),
    ], footer=_footer(grand_total=Decimal("480000")))


def _bid_gs6_firm():
    """GS-6 (positive foil to GS-3): Acme's SAME DIV 13 flooring line IS moved.
    Expect YELLOW KNOWN_FIRM_RECLASSIFIED (×2: flooring 13→09, dumpster 11→01)."""
    return _doc("Acme Restoration", [
        _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
             items=[LineItem(description="Flooring (Labor) — install", amount=Decimal("18000"))],
             subtotal=Decimal("18000")),
        _div("DIV 11 00 00", "Equipment", cost=CostStructure.ITEMIZED,
             items=[LineItem(description="Dumpster rental", amount=Decimal("6500"))],
             subtotal=Decimal("6500")),
    ], footer=_footer(grand_total=Decimal("24500")))


# ---------------------------------------------------------------------------
# AUDIT-sheet reader (reads back OUT of the written .xlsx)
# ---------------------------------------------------------------------------

def _audit_rows_from_file(path: Path) -> list[tuple[str, str, str]]:
    """Read (status, code, contractor) tuples from the saved AUDIT sheet.

    Per _write_audit_sheet (W-D S2-3 layout): the key block sits at the TOP;
    data rows follow the label-anchored column-header row. Col A=Status,
    B=View, C=Code, D=Contractor. Returns one tuple per audit row written.
    """
    from src.write_matrix import find_audit_header_row

    wb = openpyxl.load_workbook(path)
    if "AUDIT" not in wb.sheetnames:
        return []
    ws = wb["AUDIT"]
    header = find_audit_header_row(ws)
    if header is None:
        return []
    rows: list[tuple[str, str, str]] = []
    for r in range(header + 1, ws.max_row + 1):
        status = ws.cell(row=r, column=1).value
        code = ws.cell(row=r, column=3).value
        contractor = ws.cell(row=r, column=4).value
        # Skip anything that isn't a real audit row (blank tail, etc.).
        if not code or not contractor:
            continue
        rows.append((str(status), str(code), str(contractor)))
    return rows


def _bid_form_text(path: Path) -> str:
    wb = openpyxl.load_workbook(path)
    ws = wb["Bid_Form"]
    return " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value
    )


# ---------------------------------------------------------------------------
# Eval expectations
# ---------------------------------------------------------------------------

@dataclass
class Expect:
    """One adversarial-case expectation, graded against the written file."""
    case_id: str
    contractor: str
    # (status, code) pairs that MUST appear for this contractor on the AUDIT sheet
    must_have: list[tuple[str, str]] = field(default_factory=list)
    # codes that must NOT appear for this contractor (the negative guarantees)
    must_not_have: list[str] = field(default_factory=list)
    # exact count requirement: code -> n (e.g. CODE_SPLIT_UNMATCHED must appear 2×)
    exact_count: dict[str, int] = field(default_factory=dict)
    # optional substring that must be present in the Bid_Form sheet text
    bid_form_contains: Optional[str] = None


# The adversarial multi-bid set. GS-5, GS-4, split, GS-3, GS-6 run TOGETHER so
# cross-bid stats are real (n=5); GS-1 and GS-7 run in their own isolated runs
# (single-bidder / custom firm config) below.
_MULTI_BIDS = [
    _bid_gs5_mixed(),
    _bid_gs4_clean_legacy(),
    _bid_split_unmatched(),
    _bid_gs3_new_firm_div13(),
    _bid_gs6_firm(),
]

_MULTI_EXPECTATIONS = [
    Expect(
        case_id="GS-5 mixed canonical+legacy",
        contractor="Mixed Codes Inc",
        must_have=[("RED", "UNRECOGNIZED_CODE_FORMAT")],
        must_not_have=["CODE_FORMAT_REMAPPED"],   # no partial/per-line remap
    ),
    Expect(
        case_id="GS-4 clean legacy-2-digit remap",
        contractor="Legacy Format Co",
        must_have=[("YELLOW", "CODE_FORMAT_REMAPPED")],
        must_not_have=["UNRECOGNIZED_CODE_FORMAT", "CODE_SPLIT_UNMATCHED"],
    ),
    Expect(
        case_id="code-15 AND code-16 lump → CODE_SPLIT_UNMATCHED (code-16 fix)",
        contractor="Vague Split Co",
        must_have=[("RED", "CODE_SPLIT_UNMATCHED")],
        # The code-16 bug fix: BOTH the 15 line and the 16 line are unroutable.
        exact_count={"CODE_SPLIT_UNMATCHED": 2},
        must_not_have=["UNRECOGNIZED_CODE_FORMAT"],
    ),
    Expect(
        case_id="GS-3 new firm legit DIV 13 → NOT reclassed",
        contractor="Coastal Concrete Restoration LLC",
        must_not_have=[
            "KNOWN_FIRM_RECLASSIFIED",
            "CODE_FORMAT_REMAPPED",
            "UNRECOGNIZED_CODE_FORMAT",
        ],
    ),
    Expect(
        case_id="GS-6 firm reclass (positive foil)",
        contractor="Acme Restoration",
        must_have=[("YELLOW", "KNOWN_FIRM_RECLASSIFIED")],
        exact_count={"KNOWN_FIRM_RECLASSIFIED": 2},  # flooring 13→09, dumpster 11→01
        must_not_have=["KNOWN_FIRM_AMBIGUOUS"],
    ),
]


def _ambiguous_firm_config(tmp_path: Path) -> KnownFirmsConfig:
    """GS-7 fixture: two entries whose match terms both hit one name."""
    p = tmp_path / "ambiguous_known_firms.yaml"
    p.write_text(
        "firms:\n"
        "  - firm_id: a\n"
        "    match: [\"acme\"]\n"
        "  - firm_id: b\n"
        "    match: [\"acme restoration\"]\n"
    )
    return load_known_firms(str(p))


def _synthetic_firm_config() -> KnownFirmsConfig:
    """A synthetic, ACTIVE known-firm config (no real names) carrying the GS-6
    reclass quirks — injected into the multi-bid run so the positive foil fires
    the identical name-agnostic reclass path a real overlay would drive."""
    return KnownFirmsConfig(firms=[Firm(
        firm_id="acme",
        match=["acme"],
        reclassifications=[
            Reclassification(
                rule_id="ACME_FLOORING_LABOR",
                from_division="DIV 13 00 00", to_division="DIV 09 00 00",
                when_description_contains_all=["flooring", "labor"]),
            Reclassification(
                rule_id="ACME_DUMPSTER",
                from_division="DIV 11 00 00", to_division="DIV 01 00 00",
                when_description_contains_all=["dumpster"]),
        ],
    )])


# ---------------------------------------------------------------------------
# The eval runner
# ---------------------------------------------------------------------------

@dataclass
class CaseResult:
    case_id: str
    passed: bool
    detail: str


def _grade(expect: Expect, rows: list[tuple[str, str, str]],
           bid_form_text: str) -> CaseResult:
    """Code-grade one expectation against the AUDIT rows read from the file."""
    for_contractor = [(s, c) for (s, c, name) in rows if name == expect.contractor]
    codes_present = [c for (_s, c) in for_contractor]
    failures: list[str] = []

    for (status, code) in expect.must_have:
        if (status, code) not in for_contractor:
            failures.append(
                f"expected {status} {code} on AUDIT sheet for "
                f"'{expect.contractor}' — got {sorted(set(for_contractor))}"
            )

    for code in expect.must_not_have:
        if code in codes_present:
            failures.append(
                f"{code} must NOT appear for '{expect.contractor}' but it did"
            )

    for code, n in expect.exact_count.items():
        actual = codes_present.count(code)
        if actual != n:
            failures.append(
                f"{code} expected ×{n} for '{expect.contractor}', got ×{actual}"
            )

    if expect.bid_form_contains and expect.bid_form_contains not in bid_form_text:
        failures.append(
            f"Bid_Form sheet missing required text: {expect.bid_form_contains!r}"
        )

    if failures:
        return CaseResult(expect.case_id, False, "; ".join(failures))
    return CaseResult(expect.case_id, True, "ok")


def run_eval() -> list[CaseResult]:
    """Run every adversarial case end-to-end and grade against the written file."""
    results: list[CaseResult] = []

    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)

        # --- Run 1: the multi-bidder adversarial pack (n=5) ---
        # inject a synthetic active firm config so the GS-6 positive-foil reclass
        # fires (name-agnostic path); the other four bidders don't match it.
        _firm_cfg = _synthetic_firm_config()
        multi_norm = compute_cross_bid_stats(
            [normalize_bid(doc, known_firms=_firm_cfg) for doc in _MULTI_BIDS]
        )
        multi_items = audit_bids(multi_norm)
        multi_out = tmp / "multi.xlsx"
        write_matrix(multi_norm, multi_out, _run_inputs(), audit_items=multi_items)
        multi_rows = _audit_rows_from_file(multi_out)
        multi_text = _bid_form_text(multi_out)
        for expect in _MULTI_EXPECTATIONS:
            results.append(_grade(expect, multi_rows, multi_text))

        # --- Run 2: GS-1 single bidder (isolated; cross-bid must be silent) ---
        single_norm = compute_cross_bid_stats([normalize_bid(_doc(
            "Solo Restoration LLC", [
                _div("DIV 01 00 00", "General Requirements", subtotal=Decimal("50000")),
                _div("DIV 03 00 00", "Concrete", subtotal=Decimal("400000")),
                _div("DIV 07 00 00", "Thermal & Moisture Protection", subtotal=Decimal("150000")),
                _div("DIV 09 00 00", "Finishes", subtotal=Decimal("80000")),
            ], footer=_footer(grand_total=Decimal("680000"))))])
        single_items = audit_bids(single_norm)
        single_out = tmp / "single.xlsx"
        write_matrix(single_norm, single_out, _run_inputs(), audit_items=single_items)
        single_rows = _audit_rows_from_file(single_out)
        single_text = _bid_form_text(single_out)
        results.append(_grade(Expect(
            case_id="GS-1 single bidder → no cross-bid flags + notice",
            contractor="Solo Restoration LLC",
            must_not_have=[
                "SCOPE_GAP_IMPLICIT",
                "GC_FEE_OUTLIER",
                "CROSS_BID_HIGH_VARIANCE",
            ],
            bid_form_contains="Single bid — no competitive comparison available.",
        ), single_rows, single_text))

        # --- Run 3: GS-7 ambiguous firm match (custom colliding config) ---
        amb_cfg = _ambiguous_firm_config(tmp)
        amb_norm = compute_cross_bid_stats([normalize_bid(_doc(
            "Acme Restoration LLC", [
                _div("DIV 13 00 00", "Special Construction", cost=CostStructure.ITEMIZED,
                     items=[LineItem(description="Flooring (Labor)", amount=Decimal("18000"))],
                     subtotal=Decimal("18000")),
            ], footer=_footer(grand_total=Decimal("18000"))),
            known_firms=amb_cfg)])
        amb_items = audit_bids(amb_norm)
        amb_out = tmp / "ambiguous.xlsx"
        write_matrix(amb_norm, amb_out, _run_inputs(), audit_items=amb_items)
        amb_rows = _audit_rows_from_file(amb_out)
        amb_text = _bid_form_text(amb_out)
        results.append(_grade(Expect(
            case_id="GS-7 name collides >1 profile → KNOWN_FIRM_AMBIGUOUS",
            contractor="Acme Restoration LLC",
            must_have=[("RED", "KNOWN_FIRM_AMBIGUOUS")],
            must_not_have=["KNOWN_FIRM_RECLASSIFIED"],
        ), amb_rows, amb_text))

    return results


# ---------------------------------------------------------------------------
# pytest entry points (one test per case) — keeps this eval CI-collectable
# ---------------------------------------------------------------------------

_CACHED: Optional[list[CaseResult]] = None


def _results() -> list[CaseResult]:
    global _CACHED
    if _CACHED is None:
        _CACHED = run_eval()
    return _CACHED


def _make_pytest_funcs():
    """Generate one test_* function per expected case at import time."""
    case_ids = [e.case_id for e in _MULTI_EXPECTATIONS] + [
        "GS-1 single bidder → no cross-bid flags + notice",
        "GS-7 name collides >1 profile → KNOWN_FIRM_AMBIGUOUS",
    ]
    g = globals()
    for i, cid in enumerate(case_ids):
        def _t(cid=cid):
            res = next(r for r in _results() if r.case_id == cid)
            assert res.passed, res.detail
        g[f"test_case_{i:02d}"] = _t


_make_pytest_funcs()


# ---------------------------------------------------------------------------
# Standalone runner
# ---------------------------------------------------------------------------

def main() -> int:
    results = run_eval()
    print("=" * 74)
    print("FALKE Matrix — Adversarial Audit-Contract Eval (Floyd V&V evidence)")
    print("=" * 74)
    passed = 0
    for r in results:
        tag = "PASS" if r.passed else "FAIL"
        print(f"  [{tag}] {r.case_id}")
        if not r.passed:
            print(f"         → {r.detail}")
        passed += 1 if r.passed else 0
    total = len(results)
    print("-" * 74)
    print(f"  {passed}/{total} cases passed.")
    contract_proven = "yes" if passed == total else "no"
    print(f"  C4 contract proven (each RED/YELLOW reaches the AUDIT sheet): "
          f"{contract_proven}")
    print("=" * 74)
    return 0 if passed == total else 1


if __name__ == "__main__":
    sys.exit(main())
