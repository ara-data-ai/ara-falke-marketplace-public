#!/usr/bin/env python3
"""Compare a candidate create-matrix output xlsx against the golden xlsx.

Fingerprints both files on: per-bidder grand totals (Leveled_Normalized
GRAND TOTAL row, COST columns), division count, section-subtotal row count,
and AUDIT RED/YELLOW/GREEN status counts. Exit 0 = PASS, 1 = FAIL.

Usage: python3 compare.py <candidate.xlsx> [<golden.xlsx>]
       (golden defaults to the single .xlsx sitting next to this script)
"""
import sys
from pathlib import Path

import openpyxl


def fingerprint(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    lev = wb["Leveled_Normalized"]
    hdr = next(r for r in range(1, lev.max_row + 1) if lev.cell(r, 1).value == "CSI")
    cost_cols = [c for c in range(1, lev.max_column + 1)
                 if str(lev.cell(hdr, c).value or "").strip() == "COST"]
    bidders = [str(lev.cell(hdr + 1, c).value).strip() for c in cost_cols]
    gt = next(r for r in range(hdr, lev.max_row + 1)
              if str(lev.cell(r, 2).value or "").startswith("GRAND TOTAL"))
    totals = {b: round(float(lev.cell(gt, c).value or 0), 2)
              for b, c in zip(bidders, cost_cols)}
    divisions = sum(1 for r in range(hdr, lev.max_row + 1)
                    if str(lev.cell(r, 1).value or "").startswith("DIV "))
    sections = sum(1 for r in range(hdr, lev.max_row + 1)
                   if str(lev.cell(r, 2).value or "").strip().upper().endswith("SUBTOTAL"))
    aud = wb["AUDIT"]
    statuses = [str(aud.cell(r, 1).value or "").strip() for r in range(1, aud.max_row + 1)]
    audit = {s: statuses.count(s) for s in ("RED", "YELLOW", "GREEN")}
    return {"per-bidder grand totals": totals, "division count": divisions,
            "section-subtotal rows": sections, "audit status counts": audit}


def main():
    if len(sys.argv) not in (2, 3):
        sys.exit(__doc__)
    if len(sys.argv) == 3:
        golden = Path(sys.argv[2])
    else:
        found = sorted(Path(__file__).resolve().parent.glob("*.xlsx"))
        if len(found) != 1:
            sys.exit(f"expected exactly 1 golden .xlsx next to compare.py, found "
                     f"{len(found)} — pass the golden path explicitly")
        golden = found[0]
    cand, gold = fingerprint(sys.argv[1]), fingerprint(golden)
    failed = False
    for key, want in gold.items():
        ok = cand[key] == want
        failed |= not ok
        print(f"{'PASS' if ok else 'FAIL'}  {key}: candidate={cand[key]}  golden={want}")
    print("RESULT:", "FAIL" if failed else "PASS")
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
