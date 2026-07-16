"""Scorecard run-pack PARSER — the single-upload input path (P1-4, verdict c;
Marvin's ratification §3, §8, §9).

One workbook in (``--inputs pack.xlsx``), everything the render needs out:
the band + trade lines, the framework, the scores, the aliases, the exclusions,
and the declarations that let the card disclose how the evaluation plan was set.

Operator cost drops from 4 uploads + 3 templates-from-nowhere + ~8 answers to
1 upload, 2 confirmations (SF, baseline — unchanged BY DESIGN), 1 save location.

WHAT THIS MODULE MUST NEVER DO
------------------------------
Satisfy a gate. Not one cell in the pack may. The SF-basis confirmation and the
baseline confirmation are answered by a human, in the conversation, per run:
``--inputs`` does NOT imply ``--sf-confirmed`` and does NOT imply
``--baseline-confirmed``. The pack carries DATA; the gates consume DECISIONS,
and those are different things (R1 / §5).

That is enforced structurally rather than by discipline: the schema contains no
confirmation field, and ``_read_settings`` rejects any unrecognized label with
exit 2 naming it (R5). Someone hand-adding ``sf_confirmed: yes`` to be helpful
gets a hard stop, not silence. A field that does not exist cannot be
auto-satisfied and cannot be built in later by a maintainer chasing convenience.

Net effect on the operator: nothing changes. They still confirm SF, still
confirm the baseline. What changes is that they stop re-typing eight firm names
into a grid that hard-stops on a typo. THE FRICTION WE REMOVE IS DATA ENTRY. THE
FRICTION WE KEEP IS JUDGMENT.

THE INTEGRITY POSTURE (§8.2)
----------------------------
    Schema and identity are HARD STOPS. Provenance and lineage are CONFIRMABLE.

A schema mismatch means the parser cannot be trusted to read the file. A project
mismatch means the wrong building. Neither is a judgment call. But "is this pack
from the matrix run you mean?" is exactly the kind of question the existing
suggest-and-confirm idiom was built to ask a human — so run_id is EVIDENCE, not
the gate. The roster is the gate (§8.4).
"""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from . import pack_schema as ps
from .baseline_parser import parse_baseline_sheet
from .matrix import normalize_name
from .scoring_inputs import parse_framework_table, parse_scores_table


class PackError(ValueError):
    """Any pack problem the CLI reports as ``[STOP] ...`` and exits 2."""


@dataclass
class ParsedPack:
    """Everything the run needs from one pack, plus the facts the audit judges."""

    path: str
    pack_format_version: str
    producer: str
    matrix_format_version: str
    matrix_run_id: str
    matrix_file_name: str
    emitted_at: str
    project_name: str
    project_address: str
    sf_basis_value: Optional[float]
    sf_basis_label: str
    # Baseline
    band_low: float
    band_high: float
    band_mid: float
    baseline_lines: List[Dict[str, Any]]
    baseline_provenance: Dict[str, Any]
    # Framework (the evaluation PLAN) + its declaration
    framework: List[Dict[str, Any]]
    framework_basis: str
    framework_lock_date: str
    framework_ruling_note: str
    # Scores (the evaluation RECORD)
    category_scores: Dict[str, Dict[str, Optional[float]]]
    scoring_completed_date: str
    # Settings conventions
    bid_opening_date: str
    addenda_through: str
    aliases: Dict[str, str]
    matrix_exclusions: List[Tuple[str, str]]
    additional_exclusions: List[Tuple[str, str]]
    # Standing-framework reference (AUTHORITATIVE values, from the doc props)
    standing_version: str
    standing_effective_date: str
    standing_hash: str
    # Binding outcome (§8.3) — set by bind_pack_to_matrix
    binding: Dict[str, Any] = field(default_factory=dict)

    @property
    def standing_available(self) -> bool:
        """False = the W8 bootstrap: no standing reference was on file, so the
        drift check may WARN but must claim nothing about policy drift."""
        return bool(self.standing_hash)

    @property
    def framework_hash(self) -> str:
        """Semantic hash of the framework AS SUPPLIED in this pack."""
        return ps.framework_semantic_hash(
            [(r["short_label"], r["weight"]) for r in self.framework])


# ---------------------------------------------------------------------------
# low-level, label-addressed readers (R2)
# ---------------------------------------------------------------------------

def _fmt(value) -> str:
    """Cell -> trimmed string. Dates render ISO so the coherence checks in
    audit.py compare like with like regardless of how Excel typed the cell."""
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_date(text: str) -> Optional[_dt.date]:
    """'2026-06-01' -> date, anything unparseable -> None."""
    if not text:
        return None
    try:
        return _dt.datetime.strptime(str(text).strip()[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _require_date(path: str, sheet: str, label: str, text: str) -> None:
    """The two-clock audit (W3/W4) compares dates, so a date that cannot be
    parsed is not a date — it is an unenforceable claim.

    Stopping here at the input gate is the honest option: the alternative is a
    check that silently declines to run on exactly the runs where someone typed
    something creative into a clock field.
    """
    if as_date(text) is None:
        raise PackError(
            f"'{path}' ({sheet} sheet): '{label}' must be a date in YYYY-MM-DD "
            f"form; got {text!r}. These dates are what the coherence checks "
            f"compare — a date the tool cannot read is a date it cannot audit."
        )


def _blank_row(ws, row: int, width: int) -> bool:
    return all(_fmt(ws.cell(row=row, column=c).value) == ""
               for c in range(1, width + 1))


def _unknown_field(path: str, sheet: str, row: int, label,
                   allowed: Tuple[str, ...]) -> PackError:
    return PackError(
        f"'{path}' ({sheet} sheet, row {row}): unrecognized field {label!r}. "
        f"The run pack has a fixed set of fields and this is not one of them — "
        f"it was either renamed, added by hand, or typed below a blank row in a "
        f"table block (a blank row ends the block). Expected one of: "
        f"{', '.join(allowed)}. Re-emit the pack from the matrix run rather "
        f"than editing its structure."
    )


def _read_scalar_block(ws, path: str, allowed: Tuple[str, ...],
                       *, stop_labels: Tuple[str, ...] = ()) -> Dict[str, Any]:
    """Read the ``Field | Value`` scalar block, rejecting unknown labels (R5).

    Used on the sheets that have exactly ONE scalar block followed by ONE table
    (Baseline / Framework / Scores). Settings, which interleaves a scalar block
    with three table blocks, uses _walk_settings instead — see the note there
    about why "stop at the first table" is not sufficient for R5.
    """
    wanted = {ps.norm_label(f): f for f in allowed}
    stops = {ps.norm_label(s) for s in stop_labels}
    out: Dict[str, Any] = {}
    for row in range(2, ws.max_row + 1):
        key = ps.norm_label(ws.cell(row=row, column=1).value)
        if not key:
            continue
        if key in stops:
            break
        if key not in wanted:
            raise _unknown_field(path, ws.title, row,
                                 ws.cell(row=row, column=1).value, allowed)
        out[wanted[key]] = ws.cell(row=row, column=2).value
    return out


def _walk_settings(ws, path: str) -> Dict[str, Any]:
    """ONE structural pass over the whole Settings sheet (R2 + R5).

    Why a full walk rather than "read scalars until the first table block":
    that shortcut leaves R5 with a hole big enough to drive the whole ruling
    through. An ``sf_confirmed | yes`` row appended BELOW the last table block
    is not in the scalar region (the scan already stopped) and not in a table's
    data region (a blank row ended it) — so it would be silently ignored, which
    is the one outcome R5 exists to prevent. R1 is only structural if EVERY row
    on this sheet is accounted for.

    The walk: a fully-blank row ends whatever block is open and returns to
    scalar mode; in scalar mode, column A must be a known scalar field or a
    known table label, or it is a loud stop. So a stray row anywhere on the
    sheet lands in scalar mode and is rejected by name.
    """
    scalars: Dict[str, Any] = {}
    tables: Dict[str, List[Tuple[str, ...]]] = {
        label: [] for label, _h in ps.SETTINGS_TABLE_BLOCKS}

    scalar_fields = {ps.norm_label(f): f for f in ps.SETTINGS_SCALAR_FIELDS}
    table_headers = {ps.norm_label(label): headers
                     for label, headers in ps.SETTINGS_TABLE_BLOCKS}
    table_names = {ps.norm_label(label): label
                   for label, _h in ps.SETTINGS_TABLE_BLOCKS}

    open_table: Optional[str] = None
    expect_header = False
    width = 3

    for row in range(2, ws.max_row + 1):  # row 1 is the sheet title
        if _blank_row(ws, row, width):
            open_table, expect_header = None, False
            continue

        raw = ws.cell(row=row, column=1).value
        key = ps.norm_label(raw)

        if open_table is not None:
            headers = table_headers[ps.norm_label(open_table)]
            if expect_header:
                got = [ps.norm_label(ws.cell(row=row, column=c).value)
                       for c in range(1, len(headers) + 1)]
                want = [ps.norm_label(h) for h in headers]
                if got != want:
                    raise PackError(
                        f"'{path}' (Settings sheet, row {row}): the "
                        f"'{open_table}' block must be followed by its header "
                        f"row ({' | '.join(headers)}); found {got}. Do not "
                        f"reshape the pack."
                    )
                expect_header = False
                continue
            tables[open_table].append(
                tuple(_fmt(ws.cell(row=row, column=c).value)
                      for c in range(1, len(headers) + 1)))
            continue

        # --- scalar mode: every row must be a known field or a known block ---
        if key in table_names:
            open_table = table_names[key]
            expect_header = True
            continue
        if key not in scalar_fields:
            raise _unknown_field(path, "Settings", row, raw,
                                 ps.SETTINGS_SCALAR_FIELDS)
        scalars[scalar_fields[key]] = ws.cell(row=row, column=2).value

    return {"scalars": scalars, "tables": tables}


def _find_header_row(ws, path: str, headers: Tuple[str, ...],
                     *, first_only: bool = False) -> int:
    """Locate a table by scanning column A for its first header cell (R2)."""
    first = ps.norm_label(headers[0])
    want = [ps.norm_label(h) for h in headers]
    for row in range(1, ws.max_row + 1):
        if ps.norm_label(ws.cell(row=row, column=1).value) != first:
            continue
        if first_only:
            return row
        got = [ps.norm_label(ws.cell(row=row, column=c).value)
               for c in range(1, len(headers) + 1)]
        if got == want:
            return row
    raise PackError(
        f"'{path}' ({ws.title} sheet): could not find the table header row "
        f"({' | '.join(headers)}). Do not reshape the pack."
    )


# ---------------------------------------------------------------------------
# the pack's stamp — AUTHORITATIVE (§8.1)
# ---------------------------------------------------------------------------

def read_pack_stamp(wb) -> Dict[str, str]:
    try:
        props = {p.name: p.value for p in wb.custom_doc_props.props}
    except Exception:
        return {}
    return {k: ("" if v is None else str(v)) for k, v in props.items()}


# ---------------------------------------------------------------------------
# the parse
# ---------------------------------------------------------------------------

def parse_pack(path: str) -> ParsedPack:
    """Parse a run pack. Raises PackError (-> [STOP], exit 2) on any problem."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise PackError("openpyxl is required to read the run pack "
                        "(pip install openpyxl).") from exc
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise PackError(f"Cannot open run pack '{path}': {exc}") from exc

    missing = [s for s in ps.PACK_SHEETS if s not in wb.sheetnames]
    if missing:
        raise PackError(
            f"'{path}' is not a scorecard run pack — it is missing the "
            f"{', '.join(missing)} tab(s). Found: "
            f"{', '.join(wb.sheetnames) or '(none)'}. The pack is emitted by "
            f"create-matrix beside the matrix; if you meant to supply the "
            f"individual baseline/framework/scores files, use their own flags."
        )

    stamp = read_pack_stamp(wb)

    # ---- I1 / I2 FIRST: schema and producer are hard stops (§8.2) ----------
    # Checked before anything else is read, because a schema mismatch means this
    # parser cannot be trusted to interpret the cells it is about to read.
    _check_pack_schema(path, stamp, wb)

    settings = _read_settings(wb[ps.SHEET_SETTINGS], path)
    baseline = _read_baseline(wb[ps.SHEET_BASELINE], path)
    framework, decl = _read_framework(wb[ps.SHEET_FRAMEWORK], path)
    # Settings is read FIRST on purpose: an excluded bidder still has a Scores
    # row (the column was filled at matrix time, before the ruling existed), and
    # nobody should have to score a bidder they just ruled out.
    scores, scoring_date = _read_scores(
        wb[ps.SHEET_SCORES], path, framework,
        skip_firms=[f for f, _r in settings["additional_exclusions"]])

    return ParsedPack(
        path=path,
        pack_format_version=_fmt(settings["scalars"].get(ps.S_PACK_FORMAT_VERSION)),
        producer=_fmt(settings["scalars"].get(ps.S_PRODUCER)),
        matrix_format_version=_fmt(
            settings["scalars"].get(ps.S_MATRIX_FORMAT_VERSION)),
        matrix_run_id=_fmt(settings["scalars"].get(ps.S_MATRIX_RUN_ID)),
        matrix_file_name=_fmt(settings["scalars"].get(ps.S_MATRIX_FILE_NAME)),
        emitted_at=_fmt(settings["scalars"].get(ps.S_EMITTED_AT)),
        project_name=_fmt(settings["scalars"].get(ps.S_PROJECT_NAME)),
        project_address=_fmt(settings["scalars"].get(ps.S_PROJECT_ADDRESS)),
        sf_basis_value=settings["sf_basis_value"],
        sf_basis_label=_fmt(settings["scalars"].get(ps.S_SF_BASIS_LABEL)),
        band_low=baseline["band_low"],
        band_high=baseline["band_high"],
        band_mid=baseline["band_mid"],
        baseline_lines=baseline["lines"],
        baseline_provenance=baseline["provenance"],
        framework=framework,
        framework_basis=decl["basis"],
        framework_lock_date=decl["lock_date"],
        framework_ruling_note=decl["ruling_note"],
        category_scores=scores,
        scoring_completed_date=scoring_date,
        bid_opening_date=_fmt(settings["scalars"].get(ps.S_BID_OPENING_DATE)),
        addenda_through=_fmt(settings["scalars"].get(ps.S_ADDENDA_THROUGH)),
        aliases=settings["aliases"],
        matrix_exclusions=settings["matrix_exclusions"],
        additional_exclusions=settings["additional_exclusions"],
        # The doc properties are AUTHORITATIVE; the Settings rows echo them for
        # the archived-artifact reader. _read_settings has already proven the two
        # agree, so reading the property here is the same value with a stronger
        # provenance — and it is the ONLY anchor for the standing hash, which is
        # not re-derivable from the matrix.
        standing_version=stamp.get(ps.PACK_STAMP_STANDING_VERSION_PROP, ""),
        standing_effective_date=stamp.get(ps.PACK_STAMP_STANDING_DATE_PROP, ""),
        standing_hash=stamp.get(ps.PACK_STAMP_STANDING_HASH_PROP, ""),
    )


def _check_pack_schema(path: str, stamp: Dict[str, str], wb) -> None:
    """I1 + I2 — schema and producer identity. Both exit 2."""
    raw_version = stamp.get(ps.PACK_STAMP_FORMAT_PROP)
    if not raw_version:
        raise PackError(
            f"'{path}' carries no run-pack stamp — it was not emitted by "
            f"create-matrix. A pack's schema version is what tells this "
            f"scorecard it can read the file at all; refusing to guess. Re-emit "
            f"the pack from the matrix run, or supply the individual "
            f"baseline/framework/scores files instead."
        )
    try:
        version = ps.parse_pack_version(raw_version)
    except ValueError as exc:
        raise PackError(f"'{path}': {exc}") from exc

    lo, hi = ps.SUPPORTED_PACK_FORMAT
    if not (lo <= version <= hi):
        raise PackError(
            f"'{path}': run-pack format {raw_version} is outside this "
            f"scorecard's supported range "
            f"{lo[0]}.{lo[1]}–{hi[0]}.{hi[1]}. The pack and the scorecard ship "
            f"in the same plugin at the same version, so this means one of them "
            f"is stale: update the plugin, or re-emit the pack with the "
            f"create-matrix you are running now."
        )

    # I2 — refuse to parse an undeclared producer (same rule as the matrix stamp)
    producer = stamp.get(ps.PACK_STAMP_PRODUCER_PROP, "")
    if producer != "falke-bid-tools/matrix":
        raise PackError(
            f"'{path}' is stamped by unknown producer {producer!r}. Refusing to "
            f"parse a run pack from an undeclared producer."
        )


def _read_settings(ws, path: str) -> Dict[str, Any]:
    """The identity-and-binding tab (§3.1), with unknown-key rejection (R5)."""
    walked = _walk_settings(ws, path)
    scalars = walked["scalars"]
    tables = walked["tables"]

    for required in ps.SETTINGS_REQUIRED_OPERATOR_FIELDS:
        if _fmt(scalars.get(required)) == "":
            raise PackError(
                f"'{path}' (Settings sheet): '{required}' is blank and it is "
                f"required. It is the clock this evaluation is measured against "
                f"— the matrix cannot know it, so you must supply it. Enter it "
                f"as YYYY-MM-DD and re-run."
            )
        _require_date(path, "Settings", required, _fmt(scalars.get(required)))

    addenda = _fmt(scalars.get(ps.S_ADDENDA_THROUGH))
    if addenda:
        _require_date(path, "Settings", ps.S_ADDENDA_THROUGH, addenda)

    sf_raw = scalars.get(ps.S_SF_BASIS_VALUE)
    sf_value: Optional[float]
    try:
        sf_value = float(sf_raw) if sf_raw not in (None, "") else None
    except (TypeError, ValueError):
        raise PackError(
            f"'{path}' (Settings sheet): '{ps.S_SF_BASIS_VALUE}' must be "
            f"numeric; got {sf_raw!r}."
        )

    aliases: Dict[str, str] = {}
    for matrix_name, display_name in tables[ps.T_DISPLAY_ALIASES]:
        if not matrix_name and not display_name:
            continue
        if not matrix_name or not display_name:
            raise PackError(
                f"'{path}' (Settings sheet, {ps.T_DISPLAY_ALIASES}): a row has "
                f"only one of Matrix Name / Display Name "
                f"({matrix_name!r} / {display_name!r}). An alias needs both, or "
                f"neither."
            )
        aliases[matrix_name] = display_name

    matrix_exclusions = _read_exclusion_block(
        path, ps.T_MATRIX_EXCLUSIONS, tables[ps.T_MATRIX_EXCLUSIONS],
        require_reason=False)
    additional = _read_exclusion_block(
        path, ps.T_ADDITIONAL_EXCLUSIONS, tables[ps.T_ADDITIONAL_EXCLUSIONS],
        require_reason=True)

    return {
        "scalars": scalars,
        "sf_basis_value": sf_value,
        "aliases": aliases,
        "matrix_exclusions": matrix_exclusions,
        "additional_exclusions": additional,
    }


def _read_exclusion_block(path: str, label: str, rows, *,
                          require_reason: bool) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    for firm, reason in rows:
        if not firm and not reason:
            continue
        if not firm:
            raise PackError(
                f"'{path}' (Settings sheet, {label}): a row carries a reason "
                f"({reason!r}) but no firm name."
            )
        # NO SILENT DROPS, EVER (Marvin §1.5). Setting a bidder aside is a
        # ruling, and a ruling without a recorded reason is not reviewable by
        # the board — or by a losing bidder's attorney reading the award file in
        # 2029 (R7).
        if require_reason and not reason:
            raise PackError(
                f"'{path}' (Settings sheet, {label}): '{firm}' is excluded with "
                f"no reason given. A reason is mandatory — an exclusion is a "
                f"ruling, and the award file has to say why it was made."
            )
        out.append((firm, reason))
    return out


def _read_baseline(ws, path: str) -> Dict[str, Any]:
    """The cost yardstick + the P1-6 provenance declaration (§3.2)."""
    scalars = _read_scalar_block(
        ws, path, ps.BASELINE_SCALAR_FIELDS,
        stop_labels=(ps.BASELINE_TRADE_HEADERS[0],))

    band_low, band_high, band_mid, lines = parse_baseline_sheet(ws, path)

    provenance = _fmt(scalars.get(ps.B_PROVENANCE)).lower()
    if not provenance:
        raise PackError(
            f"'{path}' (Baseline sheet): '{ps.B_PROVENANCE}' is blank and it is "
            f"required. How the baseline was arrived at is a fact only you know "
            f"— the tool does not know it and will not guess it. Enter one of: "
            f"{' | '.join(ps.PROVENANCE_VALUES)}."
        )
    if provenance not in ps.PROVENANCE_VALUES:
        raise PackError(
            f"'{path}' (Baseline sheet): '{ps.B_PROVENANCE}' must be one of "
            f"{' | '.join(ps.PROVENANCE_VALUES)}; got "
            f"{_fmt(scalars.get(ps.B_PROVENANCE))!r}."
        )
    estimator = _fmt(scalars.get(ps.B_ESTIMATOR))
    if not estimator:
        raise PackError(
            f"'{path}' (Baseline sheet): '{ps.B_ESTIMATOR}' is blank and it is "
            f"required. The award file carries the estimator of record; name "
            f"the person and their firm."
        )

    return {
        "band_low": band_low,
        "band_high": band_high,
        "band_mid": band_mid,
        "lines": lines,
        # Parsed and RECORDED in scorecard_run.json at P1-4; the
        # declaration-keyed document language, the fingerprint-contradiction
        # gate, the circularity rule, and the SB 4-D / HB 913 attestation
        # enforcement are all P1-6 and are deliberately NOT built here (§6.2).
        "provenance": {
            "provenance": provenance,
            "estimator_of_record": estimator,
            "basis_date": _fmt(scalars.get(ps.B_BASIS_DATE)),
            "basis_documents": _fmt(scalars.get(ps.B_BASIS_DOCUMENTS)),
            "mi_sirs_derived": _fmt(scalars.get(ps.B_MI_SIRS_DERIVED)),
            "mi_sirs_conflict": _fmt(scalars.get(ps.B_MI_SIRS_CONFLICT)),
        },
    }


def _read_framework(ws, path: str) -> Tuple[List[Dict[str, Any]], Dict[str, str]]:
    """The evaluation PLAN + its declaration (§3.3, §4.2)."""
    scalars = _read_scalar_block(
        ws, path, ps.FRAMEWORK_SCALAR_FIELDS,
        stop_labels=(ps.FRAMEWORK_HEADERS[0],))

    basis = _fmt(scalars.get(ps.F_BASIS)).lower()
    if not basis:
        raise PackError(
            f"'{path}' (Framework sheet): '{ps.F_BASIS}' is blank and it is "
            f"required. Declare which plan these weights are: "
            f"{' | '.join(ps.FRAMEWORK_BASIS_VALUES)}."
        )
    if basis not in ps.FRAMEWORK_BASIS_VALUES:
        raise PackError(
            f"'{path}' (Framework sheet): '{ps.F_BASIS}' must be one of "
            f"{' | '.join(ps.FRAMEWORK_BASIS_VALUES)}; got "
            f"{_fmt(scalars.get(ps.F_BASIS))!r}."
        )

    ruling_note = _fmt(scalars.get(ps.F_RULING_NOTE))
    # W7 — an input-gate stop, not an audit finding: a required cell is empty.
    if basis in ps.FRAMEWORK_BASES_REQUIRING_NOTE and not ruling_note:
        raise PackError(
            f"'{path}' (Framework sheet): '{ps.F_BASIS}' is '{basis}', so "
            f"'{ps.F_RULING_NOTE}' is required. Declaring anything other than "
            f"'{ps.BASIS_STANDING}' means the board is being shown a plan that "
            f"is not the standing one — record why in one line (e.g. 'board "
            f"approved revised weights on <date>'). It goes in the award file."
        )

    lock_date = _fmt(scalars.get(ps.F_LOCK_DATE))
    # `project-specific` CLAIMS a pre-opening lock, so it must state when (§4.2).
    # A claim of "we locked this before bids opened" with no date is not a
    # declaration — it is an assertion the award file cannot check.
    if basis == ps.BASIS_PROJECT_SPECIFIC and not lock_date:
        raise PackError(
            f"'{path}' (Framework sheet): '{ps.F_BASIS}' is "
            f"'{ps.BASIS_PROJECT_SPECIFIC}', so '{ps.F_LOCK_DATE}' is required. "
            f"That declaration says these weights were set for this project "
            f"BEFORE bids were opened — give the date it happened (YYYY-MM-DD) "
            f"so the record can stand on its own."
        )
    if lock_date:
        _require_date(path, "Framework", ps.F_LOCK_DATE, lock_date)

    header_row = _find_header_row(ws, path, ps.FRAMEWORK_HEADERS)
    framework = parse_framework_table(ws, path, header_row=header_row)

    return framework, {
        "basis": basis,
        "lock_date": lock_date,
        "ruling_note": ruling_note,
    }


def _read_scores(ws, path: str, framework, *,
                 skip_firms=None) -> Tuple[Dict[str, Dict[str, Optional[float]]],
                                          str]:
    """The evaluation RECORD (§3.4)."""
    scalars = _read_scalar_block(
        ws, path, ps.SCORES_SCALAR_FIELDS,
        stop_labels=(ps.SCORES_FIRM_HEADER,))

    scoring_date = _fmt(scalars.get(ps.SC_SCORING_COMPLETED_DATE))

    header_row = _find_header_row(ws, path, (ps.SCORES_FIRM_HEADER,),
                                 first_only=True)
    scores = parse_scores_table(ws, path, framework, header_row=header_row,
                                skip_firms=skip_firms)

    # ---- The completion declaration, conditional on the record's state -----
    # Marvin corrected his own P1-4 §3.4 here: he made this field
    # unconditionally required on the assumption coverage is always 100% — the
    # very dead-pathway state F3 identified — and restoring the pathway would
    # have turned it into a self-contradiction generator firing on every honest
    # provisional run. A control that fires on the honest case gets clicked past.
    #
    # The field is not a date. It is a DECLARATION: *the evaluation record is
    # closed as of this date.* A closed record with blanks is incoherent; an
    # open record has no completion date BECAUSE IT IS NOT COMPLETE. So the
    # requirement keys on the state of the record, exactly as severity keys on
    # the declaration — and the honest path is the lazy path: mid-evaluation the
    # operator does nothing, leaves the cell alone, and gets the provisional
    # card. The declaration costs a cell only at the moment it becomes true.
    blanks = sum(1 for cells in scores.values()
                 for v in cells.values() if v is None)

    if blanks == 0 and not scoring_date:
        raise PackError(
            f"'{path}' (Scores sheet): every category is scored, so the "
            f"evaluation record is complete — declare when it was completed. "
            f"Enter '{ps.SC_SCORING_COMPLETED_DATE}' as YYYY-MM-DD and re-run."
        )
    if blanks > 0 and scoring_date:
        # exit 2, NOT a BLOCKER — his sharpened doctrine (§5.4): a
        # self-contradicting declaration is a BLOCKER when the tool can still
        # build the RIGHT document and the contradiction is a fact to disclose
        # ON it (W2). It is an exit-2 input gate when the contradiction is about
        # WHICH DOCUMENT TO BUILD — the cells say provisional, the date says
        # final, and the tool has no basis to pick. You cannot render-then-flag
        # your way out of not knowing what to render.
        raise PackError(
            f"'{path}' (Scores sheet): the tab declares scoring completed on "
            f"{scoring_date}, but {blanks} category score(s) are blank. Either "
            f"finish them, or clear the date and render provisionally."
        )
    if scoring_date:
        _require_date(path, "Scores", ps.SC_SCORING_COMPLETED_DATE, scoring_date)
    return scores, scoring_date


# ---------------------------------------------------------------------------
# Re-derivation source for the pack's Matrix Exclusions block (R3)
# ---------------------------------------------------------------------------

MATRIX_AUDIT_SHEET = "AUDIT"
MATRIX_AUDIT_HEADERS = ("Status", "View", "Code", "Contractor")
INPUT_EXCLUDED_CODE = "INPUT_EXCLUDED"


def read_matrix_input_exclusions(matrix_path: str) -> Optional[List[Tuple[str, str]]]:
    """Re-read the matrix's OWN exclusion rulings from its AUDIT sheet.

    The matrix pipeline writes one RED ``INPUT_EXCLUDED`` row per bid it dropped
    (parse / schema / intake / normalize failure) — the F1 rule that no drop is
    ever silent. The pack echoes those rulings; this is where the parser proves
    the echo is honest instead of trusting the sheet lock (R3).

    Returns None when the workbook has no AUDIT sheet (a legacy or hand-built
    matrix) — there is nothing to re-derive against, so the caller skips the
    check rather than manufacturing a mismatch from a missing sheet.

    Located by header-row scan, never by row index (R2): the AUDIT sheet's key
    block above the headers grows whenever a status tier is added.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        return None
    try:
        wb = openpyxl.load_workbook(matrix_path, data_only=True, read_only=True)
    except Exception:
        return None
    if MATRIX_AUDIT_SHEET not in wb.sheetnames:
        return None

    rows = list(wb[MATRIX_AUDIT_SHEET].iter_rows(values_only=True))
    want = [ps.norm_label(h) for h in MATRIX_AUDIT_HEADERS]
    header_idx = None
    for idx, row in enumerate(rows):
        if not row:
            continue
        got = [ps.norm_label(v) for v in list(row)[:len(want)]]
        if got == want:
            header_idx = idx
            break
    if header_idx is None:
        return None

    out: List[Tuple[str, str]] = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 4:
            continue
        if _fmt(row[2]) != INPUT_EXCLUDED_CODE:
            continue
        firm = _fmt(row[3])
        message = _fmt(row[7]) if len(row) > 7 else ""
        if firm:
            out.append((firm, message))
    return out


# ---------------------------------------------------------------------------
# The binding table (§8.3) — pack <-> matrix
# ---------------------------------------------------------------------------

def bind_pack_to_matrix(
    pack: ParsedPack,
    parsed_matrix,
    *,
    matrix_stamp: Optional[Dict[str, str]],
    matrix_project_name: str,
    matrix_project_address: str,
    matrix_exclusions: Optional[List[Tuple[str, str]]] = None,
) -> List[str]:
    """Bind a pack to the matrix it claims to come from. Raises PackError on the
    hard stops (I3 / I6 / I8); returns log lines for the confirmable tiers.

    ``pack.binding`` is populated with the outcome for scorecard_run.json and
    for audit C22.

    | I1 | pack schema out of range          | exit 2  (in _check_pack_schema) |
    | I2 | unknown pack producer             | exit 2  (in _check_pack_schema) |
    | I3 | project identity mismatch         | exit 2, ALWAYS. No warning tier |
    | I4 | run_id matches                    | clean; logged                   |
    | I5 | run_id differs, roster reconciles | WARN + operator confirmation    |
    | I6 | run_id differs AND roster differs | exit 2                          |
    | I7 | run_id absent (hand-built/legacy) | suggest-and-confirm + WARN      |
    | I8 | a producer field fails re-derive  | exit 2, naming the field        |
    """
    log: List[str] = []
    matrix_stamp = matrix_stamp or {}
    matrix_run_id = (matrix_stamp.get("run_id") or "").strip()

    # ---- I3: project identity. Exit 2, always — no warning tier. -----------
    # Cross-project contamination puts Building A's bidders on Building B's
    # card. There is no legitimate case.
    _assert_identity(pack, "project name", pack.project_name,
                     matrix_project_name)
    _assert_identity(pack, "project address", pack.project_address,
                     matrix_project_address)

    # ---- I8: producer fields re-derived from the matrix (R3) ---------------
    # The xlsx lock is advisory UI and trivially removed, so the lock proves
    # nothing. This does.
    _assert_producer_field(pack, ps.S_PRODUCER, pack.producer,
                           matrix_stamp.get("producer", ""))
    _assert_producer_field(pack, ps.S_MATRIX_FORMAT_VERSION,
                           pack.matrix_format_version,
                           matrix_stamp.get("format_version", ""))

    roster = [b.raw_name for b in parsed_matrix.blocks]
    _assert_roster(pack, roster)
    if matrix_exclusions is not None:
        _assert_matrix_exclusions(pack, matrix_exclusions)

    # ---- I4 / I5 / I6 / I7: the run_id tiers -------------------------------
    if not matrix_run_id:
        # I7 — the matrix predates the run_id stamp, or was hand-built. The
        # legacy path must exist (§9.1: legitimate INDEFINITELY), so this is
        # suggest-and-confirm provenance, not a refusal.
        #
        # Floyd F-3: when the matrix has NO identity to compare against, I3
        # (cross-project, "exit 2, ALWAYS — there is no legitimate case") is
        # silently unable to run, and only this WARN stands between Building B's
        # pack and Building A's card. Marvin's I7 assumed "hand-built or
        # legacy"; a same-version matrix produced before the identity stamp is a
        # case his table never contemplated, and Falke has such matrices on disk
        # today — picking the wrong one from a folder is a plausible slip.
        # So the degradation is NAMED, not left to be inferred. Same principle
        # as C-1: say the thing you know.
        identity_known = bool((matrix_stamp.get("project_name") or "").strip())
        i3_ran = identity_known and bool(pack.project_name)
        pack.binding = {
            "tier": "I7",
            "run_id_pack": pack.matrix_run_id,
            "run_id_matrix": None,
            "status": "unstamped-matrix",
            "confirmed_required": True,
            "cross_project_check_ran": i3_ran,
        }
        line = ("PACK BINDING (I7): this matrix carries no run identity — it "
                "predates the stamp or was built by hand. The pack's firm "
                "roster reconciles against it, but the inputs cannot be proven "
                "pipeline-originated. Recorded in the run log. WARN.")
        if not i3_ran:
            line += (" This matrix also carries no project identity, so the "
                     "cross-project check (I3) COULD NOT RUN: nothing here "
                     "verifies that this pack belongs to this building. Confirm "
                     "the matrix and the pack are the same project before "
                     "relying on this card.")
        log.append(line)
        return log

    if matrix_run_id == pack.matrix_run_id:
        pack.binding = {
            "tier": "I4",
            "run_id_pack": pack.matrix_run_id,
            "run_id_matrix": matrix_run_id,
            "status": "bound",
            "confirmed_required": False,
        }
        log.append(f"PACK BINDING (I4): pack is bound to matrix run "
                   f"{matrix_run_id} — clean.")
        return log

    # run_id differs. The roster already reconciled above (I6 would have raised),
    # so this is I5: a legitimate corrected-matrix re-run.
    #
    # WHY THIS IS NOT A HARD STOP (§8.4): a bidder's extraction is wrong, the
    # matrix is re-run, a new run_id is minted — and the operator has already
    # scored eight bidders in the old pack. Hard-refusing means they re-key
    # everything, which recreates the exact failure class the pack exists to
    # kill. The substantive question is never "do the IDs match?" — it is "do the
    # firms match?". The run_id is EVIDENCE, not the gate.
    pack.binding = {
        "tier": "I5",
        "run_id_pack": pack.matrix_run_id,
        "run_id_matrix": matrix_run_id,
        "status": "different-run-reconciled",
        "confirmed_required": True,
    }
    log.append(
        f"PACK BINDING (I5): this pack was built from a different matrix run "
        f"(pack {pack.matrix_run_id or '(none)'} vs matrix {matrix_run_id}). "
        f"The scored-firm roster, the project identity, and the SF all "
        f"reconcile, so this is most likely a corrected-matrix re-run — but "
        f"confirm it is the matrix you mean. WARN.")
    return log


def _assert_identity(pack: ParsedPack, what: str, pack_value: str,
                     matrix_value: str) -> None:
    if not matrix_value:
        # The matrix predates identity stamping; I7's roster reconcile carries
        # the provenance question instead. Never invent a mismatch from silence.
        return
    if normalize_name(pack_value) != normalize_name(matrix_value):
        raise PackError(
            f"Run pack / matrix {what} mismatch — the pack says "
            f"{pack_value!r} and the matrix says {matrix_value!r}. These are "
            f"different projects. Scoring one building's bidders against "
            f"another building's card is not a case with a legitimate reading, "
            f"so this stops here: supply the pack that was emitted alongside "
            f"THIS matrix."
        )


def _assert_producer_field(pack: ParsedPack, label: str, pack_value: str,
                           matrix_value: str) -> None:
    if not matrix_value:
        return
    if str(pack_value).strip() != str(matrix_value).strip():
        raise PackError(
            f"Run pack field '{label}' does not match the matrix: pack says "
            f"{pack_value!r}, matrix says {matrix_value!r}. This field is "
            f"filled by the producer and re-derived here; a mismatch means the "
            f"pack was edited. Re-emit it from the matrix run."
        )


def _assert_roster(pack: ParsedPack, roster: List[str]) -> None:
    """I6 — the roster IS the gate (§8.4).

    Names the difference in BOTH directions, because "which firm is missing"
    and "which firm is unexpected" are different repairs.
    """
    # The pack's roster is the scored firms PLUS the ones the operator ruled out
    # on the Settings tab — those rows are in the workbook and were deliberately
    # not scored. Excluding a bidder must not read as a roster break; that is the
    # difference between a reasoned ruling and a deleted row, and §3.4 turns on
    # exactly that distinction.
    pack_firms = {normalize_name(f): f for f in pack.category_scores}
    for firm, _reason in pack.additional_exclusions:
        pack_firms.setdefault(normalize_name(firm), firm)
    matrix_firms = {normalize_name(f): f for f in roster}

    in_pack_only = [pack_firms[k] for k in pack_firms if k not in matrix_firms]
    in_matrix_only = [matrix_firms[k] for k in matrix_firms if k not in pack_firms]
    if not in_pack_only and not in_matrix_only:
        return

    parts = ["The run pack's scored-firm roster does not match this matrix."]
    if in_pack_only:
        parts.append("In the pack but not in the matrix: "
                     + ", ".join(sorted(in_pack_only)) + ".")
    if in_matrix_only:
        parts.append("In the matrix but not in the pack: "
                     + ", ".join(sorted(in_matrix_only)) + ".")
    parts.append(
        "The Firm column is producer-filled and must stay exactly as emitted — "
        "deleting a row does not exclude a bidder (use Additional Exclusions "
        "on the Settings tab, with a reason). If the matrix was re-run with a "
        "different bidder field, re-emit the pack from that run.")
    raise PackError(" ".join(parts))


def _assert_matrix_exclusions(pack: ParsedPack,
                              matrix_exclusions: List[Tuple[str, str]]) -> None:
    """I8 for the Matrix Exclusions block — re-derived from the matrix's own
    AUDIT sheet, never trusted from the pack (R3).

    The matrix run already made these rulings with logged reasons. An operator
    may ADD exclusions; an operator may not quietly delete the record of one the
    matrix made.
    """
    pack_names = {normalize_name(f) for f, _r in pack.matrix_exclusions}
    matrix_names = {normalize_name(f) for f, _r in matrix_exclusions}
    if pack_names == matrix_names:
        return
    dropped = sorted(matrix_names - pack_names)
    added = sorted(pack_names - matrix_names)
    parts = ["The run pack's 'Matrix Exclusions' block does not match this "
             "matrix's own exclusion rulings."]
    if dropped:
        parts.append(
            "Recorded by the matrix but missing from the pack: "
            + ", ".join(f for f, _r in matrix_exclusions
                        if normalize_name(f) in dropped) + ".")
    if added:
        parts.append("Present in the pack but not ruled by the matrix: "
                     + ", ".join(f for f, _r in pack.matrix_exclusions
                                 if normalize_name(f) in added) + ".")
    parts.append(
        "That block is the matrix run's record and is not editable here — "
        "reversing a matrix ruling means re-running the matrix, where the "
        "ruling lives.")
    raise PackError(" ".join(parts))


# ---------------------------------------------------------------------------
# Settings-derived run inputs
# ---------------------------------------------------------------------------

def resolve_pack_exclusions(pack: ParsedPack,
                            roster: List[str]) -> Optional[List[str]]:
    """The operator's Additional Exclusions, validated against the roster.

    Validating the firm name here is correct on its own merits (an exclusion
    naming nobody is a typo the operator wants to know about) AND it is what
    closes R5's last hole: a hand-added ``sf_confirmed | yes`` row appended
    below a table block would otherwise be read as table data rather than as an
    unknown scalar key. It is not a firm, so it stops here.
    """
    if not pack.additional_exclusions:
        return None
    known = {normalize_name(n) for n in roster}
    names = []
    for firm, _reason in pack.additional_exclusions:
        if normalize_name(firm) not in known:
            raise PackError(
                f"'{pack.path}' (Settings sheet, {ps.T_ADDITIONAL_EXCLUSIONS}): "
                f"'{firm}' is not a bidder in this matrix. The firms in this "
                f"matrix are: {', '.join(sorted(roster))}. Exclusions must name "
                f"a bidder that is actually in the scored field."
            )
        names.append(firm)
    return names or None


def apply_aliases_to_scores(
    category_scores: Dict[str, Dict[str, Optional[float]]],
    aliases: Optional[Dict[str, str]],
) -> Dict[str, Dict[str, Optional[float]]]:
    """Re-key the pack's scores from MATRIX names to DISPLAY names.

    The pack's Firm column is producer-filled with the matrix's own raw names
    (§3.4 — that IS the re-keying killer), and the Settings alias block maps
    ``Matrix Name -> Display Name`` (§3.1). But everything downstream of the
    parse — the scored-field crosscheck, the overrides lookup, ranking, render —
    keys off the DISPLAYED name, because that is the contract the individual
    --category-scores file has always had (the operator types display names into
    it by hand).

    So the two channels meet here, at the pack boundary, and the pack is the one
    that adapts: this is a pure re-keying, and the alternative — teaching the
    scored-field crosscheck to accept either name — would put a second naming
    contract inside the engine, which is how the alias/duplicate machinery got
    delicate in the first place.

    Note this makes the pack STRICTLY better than the JSON path it replaces:
    with a pack, an aliased firm's name is never typed by a human at all.
    """
    if not aliases:
        return category_scores
    display_for = {normalize_name(k): v for k, v in aliases.items()}
    return {display_for.get(normalize_name(firm), firm): scores
            for firm, scores in category_scores.items()}


def resolve_pack_aliases(pack: ParsedPack,
                         roster: List[str]) -> Optional[Dict[str, str]]:
    """The operator's Display Aliases, validated against the roster."""
    if not pack.aliases:
        return None
    known = {normalize_name(n) for n in roster}
    for matrix_name in pack.aliases:
        if normalize_name(matrix_name) not in known:
            raise PackError(
                f"'{pack.path}' (Settings sheet, {ps.T_DISPLAY_ALIASES}): "
                f"'{matrix_name}' is not a bidder in this matrix. The firms in "
                f"this matrix are: {', '.join(sorted(roster))}. The left column "
                f"must match a firm exactly as it appears on the Scores tab."
            )
    return dict(pack.aliases)
