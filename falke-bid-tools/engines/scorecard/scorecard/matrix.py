"""Generic bid-comparison matrix parser (Marvin logic-spec §0, §1).

Safety-critical: an error here silently corrupts every downstream section, so
detection is GENERIC (no hard-coded columns/rows) and every assumption is
measured, then reported in the run log. Project ground-truth facts are used by
the tests to VALIDATE this parser, never baked into it.

Detection contract:
  - Bidder name row: detected as the row whose cells name firms AND whose
    following row(s) carry the COST / COST SUBTOTALS / $/SF / $/SF SUBTOTALS
    sub-header quartet.
  - Block width + stride: MEASURED from the repeating quartet and the gap to
    the next named header (NOT assumed to be 4 / 5).
  - Grand total: the row matching 'GRAND TOTAL CONSTRUCTION COST'
    (case/space-insensitive), OR the producer's col-A machine key GRAND_TOTAL,
    OR an exact-label match from GT_EXACT_LABELS; fenced fallback = lowest row
    below the markup adders whose label contains the WORD 'total' (word-
    boundary — 'subtotal' never matches) AND which carries a numeric in at
    least one bidder column (a legend/prose row has none). NEVER
    'CONSTRUCTION COST SUBTOTAL' (pre-markup).
  - Sheet selection (Marvin P0-7 ruling): explicit config/CLI value wins; else
    the default is 'Leveled_Normalized' when present; a producer workbook
    missing it HARD-STOPS; a single-sheet legacy workbook consumes its only
    sheet. Never a workbook-ordering accident.
  - Duplicates: detected by normalized firm name; default keep left-most.
  - Completeness: count populated CSI division subtotals per block; flag blocks
    missing divisions that >= N peers populate (Marvin §1.4) — FLAG, never
    auto-drop.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from .errors import (GrandTotalNotFoundError, MatrixStructureError,
                     ProducerVersionError)


# ----------------------------------------------------------------------------
# sheet selection (Marvin P0-7 ruling) + producer-version contract (Floyd f)
# ----------------------------------------------------------------------------
# The producer's two data sheets. The board scorecard consumes the LEVELED
# (apples-to-apples) view by default; the mirror is for reconciliation /
# dispute / debug runs only (explicit --sheet), never a silent default.
LEVELED_SHEET = "Leveled_Normalized"
MIRROR_SHEET = "Bid_Form"

# sheet_mode values recorded on the parse + card disclosure (narratives.py)
MODE_LEVELED = "leveled"
MODE_MIRROR = "mirror"
MODE_LEGACY = "legacy"

# Producer-version contract (Floyd consolidated ruling, verdict f). The matrix
# writer stamps producer + format version as workbook custom document
# properties; the parser checks the stamp against this supported range.
# RELEASE.md step 2 carries the tripwire: a minor+ (format) bump on the matrix
# engine must revisit this range in the same commit.
STAMP_PRODUCER_PROP = "falke_bid_tools.producer"
STAMP_FORMAT_PROP = "falke_bid_tools.format_version"
PRODUCER_KEY = "falke-bid-tools/matrix"
# supported format-version range: >= (0,3) and < (0,5). v0.3-era workbooks
# (and the untracked v0.3 fixtures/eval fixtures) predate the stamp entirely —
# a MISSING stamp is logged and treated as pre-stamp/legacy, never an error.
SUPPORTED_PRODUCER = {PRODUCER_KEY: ((0, 3), (0, 5))}


def resolve_sheet(sheetnames: List[str],
                  configured: Optional[str] = None) -> Tuple[str, str]:
    """Resolve which sheet the scorecard consumes, per Marvin's P0-7 ruling.

    Returns (sheet_name, sheet_mode). Rules (in order):
      1. An EXPLICIT value (config matrix.sheet_name / CLI --sheet) wins; a
         missing explicit sheet is a hard stop.
      2. Default = 'Leveled_Normalized' when present (the apples-to-apples
         decision view; grand totals are producer-verified identical to the
         mirror).
      3. A producer-shaped workbook (carries 'Bid_Form' among multiple sheets)
         WITHOUT the leveled view is a HARD STOP naming what was expected —
         never a silent first-sheet fallback (Marvin hard rule 2). Re-run with
         an explicit --sheet to make a non-default read a logged choice.
      4. A single-sheet workbook (legacy single-sheet format) consumes its only
         sheet, disclosed as legacy.
      5. Anything else (multiple sheets, none recognized) is a hard stop
         requiring an explicit --sheet.

    sheet_mode: 'leveled' (the leveled view), 'mirror' (a non-leveled sheet
    chosen while a leveled view exists in the workbook — NOT apples-to-apples
    at division level), or 'legacy' (no leveled view exists).
    """
    if configured:
        if configured not in sheetnames:
            raise MatrixStructureError(
                f"Sheet {configured!r} not in workbook {sheetnames}.")
        chosen = configured
    elif LEVELED_SHEET in sheetnames:
        chosen = LEVELED_SHEET
    elif MIRROR_SHEET in sheetnames and len(sheetnames) > 1:
        raise MatrixStructureError(
            f"Expected sheet {LEVELED_SHEET!r} is missing from this "
            f"producer-format workbook (sheets: {sheetnames}). The board "
            f"scorecard consumes the leveled/normalized view by default "
            f"(Marvin P0-7 ruling) and never silently falls back to another "
            f"sheet. Re-generate the matrix with a current create-matrix "
            f"version, or pass --sheet {MIRROR_SHEET} to make an "
            f"as-submitted read an explicit, logged choice."
        )
    elif len(sheetnames) == 1:
        chosen = sheetnames[0]
    else:
        raise MatrixStructureError(
            f"Workbook has multiple sheets {sheetnames} and none is the "
            f"default {LEVELED_SHEET!r}. Pass --sheet <name> to select one "
            f"explicitly — the consumed sheet is never a workbook-ordering "
            f"accident (Marvin P0-7 ruling)."
        )
    if chosen == LEVELED_SHEET:
        mode = MODE_LEVELED
    elif LEVELED_SHEET in sheetnames:
        mode = MODE_MIRROR
    else:
        mode = MODE_LEGACY
    return chosen, mode


def read_producer_stamp(wb) -> Optional[Dict[str, str]]:
    """Read the producer stamp (custom document properties) from a workbook.

    Returns {'producer': ..., 'format_version': ...} or None when unstamped
    (pre-v0.4.1 producer output, legacy single-sheet workbooks, hand-built files).
    Readable in read_only mode (verified openpyxl >= 3.1).
    """
    try:
        props = {p.name: p.value for p in wb.custom_doc_props.props}
    except Exception:  # pragma: no cover - openpyxl<3.1 has no custom props
        return None
    producer = props.get(STAMP_PRODUCER_PROP)
    version = props.get(STAMP_FORMAT_PROP)
    if producer is None and version is None:
        return None
    return {"producer": producer, "format_version": version}


def _parse_version(v: str) -> Tuple[int, int]:
    parts = str(v or "").strip().split(".")
    try:
        return int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
    except (ValueError, IndexError):
        raise ProducerVersionError(
            f"Unparseable producer format version {v!r} in workbook stamp.")


def check_producer_stamp(stamp: Optional[Dict[str, str]]) -> Optional[str]:
    """Enforce SUPPORTED_PRODUCER against a workbook stamp.

    Returns a log line (or None when unstamped); raises ProducerVersionError
    when the stamp names a producer/version this scorecard does not support.
    """
    if stamp is None:
        return None
    producer = stamp.get("producer")
    version = stamp.get("format_version")
    if producer not in SUPPORTED_PRODUCER:
        raise ProducerVersionError(
            f"Workbook is stamped by unknown producer {producer!r} "
            f"(format {version!r}); this scorecard supports "
            f"{sorted(SUPPORTED_PRODUCER)}. Refusing to parse a workbook "
            f"from an undeclared producer."
        )
    lo, hi = SUPPORTED_PRODUCER[producer]
    v = _parse_version(version)
    if not (lo <= v < hi):
        raise ProducerVersionError(
            f"Workbook producer format {version!r} ({producer}) is outside "
            f"this scorecard's supported range "
            f">={lo[0]}.{lo[1]},<{hi[0]}.{hi[1]}. A NEWER matrix format may "
            f"carry changes this parser has not been validated against — "
            f"update the scorecard (SUPPORTED_PRODUCER) or regenerate the "
            f"matrix with a supported create-matrix version."
        )
    return (f"producer stamp: {producer} format {version} — inside supported "
            f"range >={lo[0]}.{lo[1]},<{hi[0]}.{hi[1]}.")


# Structural signature of the producer's Stage-6b RED quarantine banner
# (POST_WRITE_TIEOUT_FAILURE — write_matrix._QUARANTINE_BANNER_LINE_1). A
# workbook carrying it failed the producer's own final self-check and must
# never be consumed silently (Marvin P0-7 hard rule 5 -> audit C17 BLOCKER).
QUARANTINE_SIGNATURE = "automated check failed"
QUARANTINE_SCAN_ROWS = 6


# ----------------------------------------------------------------------------
# normalization helpers
# ----------------------------------------------------------------------------
def normalize_name(name: str) -> str:
    """Case-fold + strip punctuation/space for duplicate detection.

    Dotted acronyms, spacing, and case all fold away, so e.g.
    'A.C.M.E' -> 'acme'; 'Mc Bride' -> 'mcbride'; 'Borealis' -> 'borealis'.
    """
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def display_name(name: str) -> str:
    """Display normalization (Marvin §1.5): collapse a dotted single-letter
    acronym to its run of capitals (e.g. 'A.C.M.E' -> 'ACME'), join a leading
    'Mc ' to the following word ('Mc Bride' -> 'McBride'), and collapse spaces.
    Generic rules — no firm name is hard-coded in engine source."""
    s = (name or "").strip()
    # dotted single-letter acronym (e.g. 'A.C.M.E', 'A.C.M.E.') -> 'ACME'
    if re.fullmatch(r"(?:[A-Za-z]\.){2,}[A-Za-z]?\.?", s):
        return re.sub(r"[^A-Za-z]", "", s).upper()
    # 'Mc <Word>' -> 'McWord' (Scots/Irish surname prefix join)
    s = re.sub(r"\bMc\s+([A-Z])", r"Mc\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_label(s) -> str:
    """Case/space-insensitive label key for row matching.

    Collapses ANY run of whitespace (incl. the double-spaces seen in the real
    the validation matrix) to a single space, strips, lower-cases.
    """
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


# Canonical sub-header buckets. The repeating per-bidder quartet is identified
# by these four buckets rather than by literal label text, because the source
# matrix carries typos and irregular spacing (e.g. "$/SXFX  SUBTOTALS",
# "COST  SUBTOTALS" with double spaces). We classify on a NORMALIZED PREFIX so
# the matcher tolerates that variance (Marvin §0/§1: detect, don't hard-code).
SUB_COST = "cost"
SUB_COST_SUBTOTAL = "cost_subtotal"
SUB_PSF = "psf"
SUB_PSF_SUBTOTAL = "psf_subtotal"
QUARTET_BUCKETS = (SUB_COST, SUB_COST_SUBTOTAL, SUB_PSF, SUB_PSF_SUBTOTAL)


def classify_subheader(value) -> Optional[str]:
    """Map a raw cell value to one of the four quartet buckets, or None.

    Tolerant of the real-matrix variance:
      - 'COST'                 -> SUB_COST
      - 'COST  SUBTOTALS'      -> SUB_COST_SUBTOTAL
      - '$/SF' / '$/SXFX'      -> SUB_PSF
      - '$/SF SUBTOTALS' / '$/SXFX  SUBTOTALS' -> SUB_PSF_SUBTOTAL
    Match is on a normalized prefix ('cost' / '$/s'), with 'subtotal' presence
    distinguishing the running-subtotal column from the per-line column.
    """
    n = _norm_label(value)
    if not n:
        return None
    has_subtotal = "subtotal" in n
    # $/SF family — the source typos this as "$/SXFX", and the
    # create-matrix engine stamps the confirmed SF-basis label into the header
    # ("$/GSF", "$/balcony SF"), so ANY "$/..." header is a per-SF column.
    if n.startswith("$/"):
        return SUB_PSF_SUBTOTAL if has_subtotal else SUB_PSF
    if n.startswith("cost"):
        return SUB_COST_SUBTOTAL if has_subtotal else SUB_COST
    return None


# ----------------------------------------------------------------------------
# in-memory worksheet grid (performance)
# ----------------------------------------------------------------------------
class Grid:
    """One-shot, in-memory snapshot of a worksheet's used range.

    Performance contract: openpyxl random cell access (``ws.cell(row, col)`` /
    ``ws[f"{col}{row}"]``) is O(n) per call on a read-only workbook, so the
    round-3 logic (per-block window scans, positional backfill, span-based
    division counting) iterating ~162 cols x ~215 rows x 10 blocks several times
    turned into a ~46-minute parse. We read the whole used range exactly ONCE via
    ``ws.iter_rows(values_only=True)`` (a single sequential pass) into a 2-D list
    of cell VALUES, then every detector reads from this grid instead of touching
    openpyxl again. This is a pure access-pattern swap: the values returned are
    identical to ``ws.cell(row, col).value``.

    Indexing contract (off-by-one safety): ``cell(row, col)`` takes openpyxl's
    1-BASED (row, col) — exactly the same arguments the call sites already pass
    to ``ws.cell(row=..., column=...)`` — and internally maps to the 0-based grid
    via ``rows[row - 1][col - 1]``. Out-of-range (row/col < 1 or beyond the
    materialized range) returns None, matching openpyxl's behavior for empty/
    unfilled cells, so detection logic that probed slightly past the used range
    keeps the same result.
    """

    def __init__(self, ws):
        # single sequential pass over the used range; values_only avoids
        # constructing Cell objects.
        #
        # ABSOLUTE ALIGNMENT (Grid index-slip fix). In read_only mode,
        # ``ws.iter_rows(values_only=True)`` with NO bounds yields rows over the
        # worksheet's stored <dimension>, whose min_row/min_col are NOT
        # guaranteed to be 1/1 (a sheet whose used range starts below row 1 or
        # right of column A reports min_col>1). The unbounded iterator then puts
        # the dimension's first column at tuple index 0, so ``cell(row, col)``'s
        # 1-based->0-based map silently slips by (min_col-1) columns and
        # (min_row-1) rows — every subtotal read (Crest col F, Dorne col K) lands
        # one or more columns off and returns None, the false-0/20 root cause.
        # Pinning the iterator to min_row=1/min_col=1 (and openpyxl's own
        # max_*) forces tuple index 0 == column A and the first tuple == row 1,
        # so the Grid's 1-based indexing is exact. read_only stays fast (still a
        # single sequential pass).
        m_row = ws.max_row or 0
        m_col = ws.max_column or 0
        if m_row and m_col:
            self._rows: List[Tuple] = list(ws.iter_rows(
                min_row=1, min_col=1, max_row=m_row, max_col=m_col,
                values_only=True))
        else:
            self._rows = list(ws.iter_rows(values_only=True))
        self._nrows = len(self._rows)
        self._ncols = max((len(r) for r in self._rows), default=0)

    @property
    def max_row(self) -> int:
        return self._nrows

    @property
    def max_col(self) -> int:
        return self._ncols

    def cell(self, row: int, col: int):
        """Value at 1-based (row, col); None if outside the materialized range.

        Mirrors ``ws.cell(row=row, column=col).value`` so call sites swap with
        no argument changes. ``row``/``col`` are openpyxl 1-based; the grid is
        0-based, hence the ``- 1`` on each axis (the single place this offset
        lives, so it can't drift across the many former cell-access sites)."""
        if row < 1 or col < 1 or row > self._nrows:
            return None
        r = self._rows[row - 1]
        c = col - 1
        if c >= len(r):
            return None
        return r[c]


# ----------------------------------------------------------------------------
# data classes
# ----------------------------------------------------------------------------
@dataclass
class BidderBlock:
    raw_name: str
    name: str                  # display-normalized
    norm: str                  # duplicate key
    start_col: int             # 1-based COST column
    cols: Dict[str, int]       # subheader -> 1-based column
    grand_total: Optional[float] = None
    populated_divisions: int = 0
    included: bool = True
    drop_reason: Optional[str] = None
    flags: List[str] = field(default_factory=list)

    @property
    def start_col_letter(self) -> str:
        return get_column_letter(self.start_col)


@dataclass
class ParsedMatrix:
    sheet_name: str
    header_row: int
    block_width: int
    block_stride: int
    grand_total_row: int
    grand_total_label: str
    gsf_value: Optional[float]          # detected; reported but NEVER used for $/SF
    gsf_row: Optional[int]
    blocks: List[BidderBlock]           # ALL detected blocks (incl. dropped)
    division_rows: List[Tuple[int, str]]  # (row, label) of CSI division subtotal rows
    log: List[str] = field(default_factory=list)
    # the CONSTRUCTION COST SUBTOTAL row (pre-markup). NEVER the compared total
    # (Marvin §1.2), but it IS a per-bidder subtotal the QA fingerprint test must
    # scan (Harbor's 3,000,000 lives here, not on a division row). None if the
    # matrix carries no such labeled row.
    construction_subtotal_row: Optional[int] = None
    # HOW the consumed sheet was chosen (Marvin P0-7): 'leveled' | 'mirror' |
    # 'legacy'. Drives the mandatory on-card disclosure line.
    sheet_mode: str = MODE_LEGACY
    # producer stamp read from the workbook custom properties (None = unstamped
    # pre-stamp/legacy workbook). {'producer':..., 'format_version':...}
    producer_stamp: Optional[Dict[str, str]] = None
    # True when the consumed sheet carries the producer's Stage-6b RED
    # quarantine banner (POST_WRITE_TIEOUT_FAILURE). The workbook failed the
    # producer's own self-check — audit C17 blocks on this flag.
    quarantine_flag: bool = False
    # the in-memory Grid snapshot used for parsing. Carried on the parsed result
    # so downstream readers (the QA fingerprint test) reuse the SAME absolute,
    # read-only grid instead of reopening the workbook and doing read_only
    # ws.cell() random access (unsupported/broken in read_only mode — the
    # fingerprint 0-hits root cause). Not part of the dataclass equality / repr.
    grid: Optional["Grid"] = field(default=None, repr=False, compare=False)

    @property
    def included_blocks(self) -> List[BidderBlock]:
        return [b for b in self.blocks if b.included]


def apply_display_aliases(parsed: "ParsedMatrix",
                          aliases: Optional[Dict[str, str]]) -> List[str]:
    """Rewrite each block's DISPLAY name via an optional alias map.

    Output/gold-card naming (Marvin §1.5): the matrix carries full legal firm
    names ('Acme Restoration', 'Granite Remodel Group') but the board scorecard
    uses short names ('Acme', 'Granite'). This optional, default-empty map
    lets the run output the short names WITHOUT touching duplicate detection
    (which stays on the raw normalized name) or the audit trail (raw_name is
    preserved and logged). The alias is the SINGLE place display naming is
    overridden; everything downstream (overrides lookup, ranking,
    render) keys off the resulting ``block.name`` so the names line up end to end.

    ``aliases`` keys may be the raw matrix name, the current display name, or a
    normalized form (matched via normalize_name) so a caller can supply whichever
    is convenient. Returns log lines for each rewrite.
    """
    log: List[str] = []
    if not aliases:
        return log
    # build a normalized-key lookup so 'Acme Restoration', 'acme
    # restoration', and 'AcmeRestoration' all resolve.
    norm_map: Dict[str, str] = {}
    for k, v in aliases.items():
        if k is None or v is None or not str(k).strip() or not str(v).strip():
            continue
        norm_map[normalize_name(str(k))] = str(v).strip()
    if not norm_map:
        return log

    def _lookup(block_norm: str) -> Optional[str]:
        # exact normalized match first
        if block_norm in norm_map:
            return norm_map[block_norm]
        # tolerate firm-suffix drift ("...Inc.", "...Builders") via bidirectional
        # containment, guarded to non-trivial names — so the alias key
        # 'Harbor Builders Inc.' (harborbuildersinc) still matches the matrix
        # raw 'Mc Bride Builders' (mcbridebuilders). Mirrors apply_exclusions.
        for key_norm, val in norm_map.items():
            if min(len(block_norm), len(key_norm)) < 4:
                continue
            if block_norm in key_norm or key_norm in block_norm:
                return val
        return None

    for b in parsed.blocks:
        # try raw name, current display name (each normalized)
        new_name = (_lookup(normalize_name(b.raw_name))
                    or _lookup(normalize_name(b.name)))
        if new_name and new_name != b.name:
            line = (f"DISPLAY ALIAS: '{b.raw_name}' (col {b.start_col_letter}) "
                    f"shown as '{new_name}' (raw name retained for audit; "
                    f"Marvin §1.5).")
            b.name = new_name
            parsed.log.append(line)
            log.append(line)
    return log


def apply_exclusions(parsed: "ParsedMatrix", exclude_names,
                     reason: str = "human ruling (Falke set-aside)") -> List[str]:
    """Apply a human bidder-exclusion ruling to an already-parsed matrix.

    Default skill behavior is include-all-and-flag (never auto-drop; Marvin
    §1.4). This is the SEPARATE, explicit channel for APPLYING Falke's curation
    decision after the flagging — e.g. setting aside Harbor / Borealis per the
    §1.4 ruling so the curated gold field reproduces.

    `exclude_names`: iterable of bidder names (display or raw); matched on the
    NORMALIZED name (normalize_name). Matching is tolerant of common firm-suffix
    drift via bidirectional containment, so the ruling 'Harbor Builders Inc.'
    (norm 'harborbuildersinc') matches the block raw-name 'Mc Bride Builders'
    (norm 'mcbridebuilders'), and 'Harbor' matches it too. A short guard
    (>=4 chars) prevents accidental over-matching. Each matched, currently-
    included block is set included=False with an audit drop_reason, and every
    exclusion is logged and returned. A ruling name that matches nothing is
    logged as a no-op (so a typo is visible, not silently ignored). Returns the
    list of log lines.
    """
    log: List[str] = []
    norm_targets = {normalize_name(n): n
                    for n in (exclude_names or []) if n and str(n).strip()}
    if not norm_targets:
        return log

    def _match(block_norm: str, target_norm: str) -> bool:
        if not block_norm or not target_norm:
            return False
        if block_norm == target_norm:
            return True
        # tolerate suffix/qualifier drift ("...Inc.", "...Builders") via
        # bidirectional containment, guarded to non-trivial names.
        if min(len(block_norm), len(target_norm)) < 4:
            return False
        return block_norm in target_norm or target_norm in block_norm

    matched: set = set()
    for b in parsed.blocks:
        if not b.included:
            continue
        for nt, raw in norm_targets.items():
            if _match(b.norm, nt):
                b.included = False
                b.drop_reason = (
                    f"EXCLUDED by ruling: '{raw}' — {reason} "
                    f"(Marvin §1.4: human applies the set-aside; not auto-dropped)."
                )
                matched.add(nt)
                line = (f"EXCLUSION (ruling): '{b.name}' col {b.start_col_letter} "
                        f"(total {b.grand_total}) removed from scored field — "
                        f"matched ruling '{raw}'; {reason}.")
                parsed.log.append(line)
                log.append(line)
                break
    for nt, raw in norm_targets.items():
        if nt not in matched:
            line = (f"EXCLUSION no-op: ruling named '{raw}' but no included block "
                    f"matched normalized '{nt}' — check the name spelling.")
            parsed.log.append(line)
            log.append(line)
    return log


# ----------------------------------------------------------------------------
# parser
# ----------------------------------------------------------------------------
class MatrixParser:
    def __init__(self, matrix_cfg: Dict):
        self.cfg = matrix_cfg or {}
        # NOTE: block_subheaders is retained in config as documentation of the
        # expected quartet, but matching is done by classify_subheader (bucket
        # by normalized prefix), NOT by literal text — the real matrix carries
        # source typos ("$/SXFX") and double-spaces that defeat literal matches.
        self.gt_label = _norm_label(self.cfg.get("grand_total_label",
                                                 "GRAND TOTAL CONSTRUCTION COST"))
        self.cc_subtotal_label = _norm_label(self.cfg.get(
            "construction_subtotal_label", "CONSTRUCTION COST SUBTOTAL"))
        self.gsf_label = _norm_label(self.cfg.get("gsf_label", "TOTAL GSF"))
        self.keep = self.cfg.get("duplicate_keep", "first")

    # -- public API ----------------------------------------------------------
    def detect_sf(self, xlsx_path: str) -> Tuple[Optional[int], Optional[float]]:
        """Read ONLY the matrix's own Row-10 'TOTAL GSF' (label + value), without
        the full block/grand-total detection.

        Backs the CLI's SF suggest-and-confirm gate: the matrix GSF is offered as
        a SUGGESTED default the user confirms (--sf-confirmed) or overrides
        (--sf-basis). Reuses the SAME ``_locate_gsf`` detector the full parse
        uses (no second GSF heuristic), so the suggested value and the value the
        audit later sees are identical. Returns (gsf_row, gsf_value); either may
        be None when no labeled GSF / no numeric value is present.

        Sheet note: the SF details line lives on the producer's Bid_Form mirror
        (row 2), not the leveled view, so this metadata probe reads the mirror
        when present (else the first sheet) regardless of which sheet the run
        consumes. The SF basis is a SUGGESTION the user must confirm — it is
        not a compared quantity, so Marvin's one-card-one-sheet rule (which
        binds parsed dollar quantities) does not apply to it."""
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet = self.cfg.get("sheet_name")
        if sheet is None:
            sheet = (MIRROR_SHEET if MIRROR_SHEET in wb.sheetnames
                     else wb.sheetnames[0])
        if sheet not in wb.sheetnames:
            raise MatrixStructureError(
                f"Sheet {sheet!r} not in workbook {wb.sheetnames}.")
        grid = Grid(wb[sheet])
        return self._locate_gsf(grid, grid.max_row)

    def parse(self, xlsx_path: str, peer_fraction: float = 0.5) -> ParsedMatrix:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        # ---- producer-version contract (Floyd verdict f): read + enforce the
        # workbook stamp BEFORE any structural parsing. Unstamped = pre-stamp/
        # legacy, logged and allowed; out-of-range = hard stop (exit 2).
        stamp = read_producer_stamp(wb)
        stamp_log = check_producer_stamp(stamp)
        # ---- sheet selection is an explicit, ruled decision (Marvin P0-7) —
        # never "first sheet in the workbook".
        sheet, sheet_mode = resolve_sheet(
            list(wb.sheetnames), self.cfg.get("sheet_name"))
        ws = wb[sheet]
        # PERFORMANCE: read the whole used range into memory exactly ONCE, then
        # every detector reads from this in-memory grid instead of doing repeated
        # openpyxl random cell access (which is O(n) per cell on a read-only
        # workbook and was the round-3 ~46-minute regression). Dimensions come
        # from the materialized grid so they stay consistent with what we read.
        grid = Grid(ws)
        max_row = grid.max_row
        max_col = grid.max_col

        log: List[str] = [
            f"sheet={sheet} (mode={sheet_mode}; explicit selection — Marvin "
            f"P0-7) dims={max_row}x{max_col}"]
        if stamp_log:
            log.append(stamp_log)
        else:
            log.append("no producer stamp on workbook — pre-stamp/legacy or "
                       "non-producer file; provenance unconfirmed.")

        # ---- Stage-6b quarantine banner (Marvin P0-7 hard rule 5): a workbook
        # that failed the producer's own self-check must never be consumed
        # silently. Flag here (LOUD in the log); audit C17 blocks on it.
        quarantine = False
        for qr in range(1, min(QUARANTINE_SCAN_ROWS, max_row) + 1):
            for qc in (1, 2, 3):
                v = grid.cell(qr, qc)
                if isinstance(v, str) and QUARANTINE_SIGNATURE in v.lower():
                    quarantine = True
                    break
            if quarantine:
                break
        if quarantine:
            log.append(
                "QUARANTINE BANNER DETECTED: this workbook carries the "
                "producer's Stage-6b POST_WRITE_TIEOUT_FAILURE banner "
                "(\"AUTOMATED CHECK FAILED\"). Figures failed the matrix "
                "tool's own self-check — the audit will BLOCK this run "
                "(C17); do not deliver a scorecard from this workbook."
            )

        header_row = self._detect_header_row(grid, max_row, max_col)
        log.append(f"detected bidder-name row = {header_row}")

        blocks, width, stride = self._detect_blocks(
            grid, header_row, max_col, max_row)
        log.append(
            f"detected {len(blocks)} bidder blocks; "
            f"measured block_width={width}, stride={stride}; "
            f"starts={[b.start_col_letter for b in blocks]}"
        )

        gt_row, gt_label = self._locate_grand_total(grid, max_row, blocks)
        log.append(f"grand-total row = {gt_row} (label {gt_label!r})")

        gsf_row, gsf_val = self._locate_gsf(grid, max_row)
        if gsf_row:
            log.append(
                f"matrix GSF detected at row {gsf_row} = {gsf_val} — REPORTED "
                f"ONLY; NEVER used for $/SF (Marvin §3)."
            )

        division_rows = self._detect_division_rows(grid, max_row, gt_row)
        log.append(f"detected {len(division_rows)} CSI division subtotal rows")

        # read grand totals + completeness per block
        for b in blocks:
            b.grand_total = self._read_block_value(grid, gt_row, b)
            b.populated_divisions = self._count_populated_divisions(
                grid, division_rows, b, max_col=max_col)

        self._resolve_duplicates(blocks, log)
        self._flag_completeness(blocks, division_rows, peer_fraction, log)

        return ParsedMatrix(
            sheet_name=sheet,
            header_row=header_row,
            block_width=width,
            block_stride=stride,
            grand_total_row=gt_row,
            grand_total_label=gt_label,
            gsf_value=gsf_val,
            gsf_row=gsf_row,
            blocks=blocks,
            division_rows=division_rows,
            log=log,
            construction_subtotal_row=getattr(self, "_cc_subtotal_row", None),
            sheet_mode=sheet_mode,
            producer_stamp=stamp,
            quarantine_flag=quarantine,
            grid=grid,
        )

    # -- detection internals -------------------------------------------------
    # how far below the name row the sub-header quartet may sit. In the real
    # validation matrix the quartet is 4 rows below the names (names row 8,
    # intervening project/GSF rows 9-11, quartet row 12), so the quartet is NOT
    # immediately beneath the names. We scan a window rather than only +1/+2.
    QUARTET_SCAN_WINDOW = 12

    def _detect_header_row(self, grid, max_row: int, max_col: int) -> int:
        """Find the bidder-name row.

        If matrix.header_row_hint is set it FORCES the name row (only validated,
        never overridden). Otherwise: the row whose cells name firms AND which
        has the COST/$/SF sub-header quartet somewhere below it (allowing
        intervening rows — the quartet need not be immediately beneath)."""
        hint = self.cfg.get("header_row_hint")
        if hint:
            # Forced: trust the hint as the name row. Validate softly (warn via
            # exception only if it truly has no names) but do NOT silently fall
            # back to auto-detect — an explicit hint must win (Defect 1.3).
            name_cells = sum(
                1 for c in range(1, max_col + 1)
                if self._looks_like_name(grid.cell(hint, c))
            )
            if name_cells < 1:
                raise MatrixStructureError(
                    f"header_row_hint={hint} forced but that row has no "
                    f"firm-name cells. Fix the hint or unset it to auto-detect."
                )
            return int(hint)

        best = None
        for r in range(1, min(30, max_row) + 1):
            # how many of the four quartet buckets appear on ANY row in the
            # scan window below this candidate (allowing intervening rows)?
            quartet_row = self._find_quartet_row(grid, r, max_row, max_col)
            quartet_hits = quartet_row[1] if quartet_row else 0
            # how many non-empty text cells in this row look like names?
            name_cells = sum(
                1 for c in range(1, max_col + 1)
                if self._looks_like_name(grid.cell(r, c))
            )
            score = quartet_hits * 10 + name_cells
            if quartet_hits >= 2 and name_cells >= 2:
                if best is None or score > best[1]:
                    best = (r, score)
        if best is None:
            # FALLBACK (create-matrix / falke-bid-tools house format): the
            # plugin engine writes the sub-header row ABOVE the bidder names
            # (headers row 4, names row 5) — inverted vs the reference layout
            # this detector was built against. Only when NO reference-orientation
            # candidate exists, look for a name row just BELOW the sub-header
            # row. Kept as a fallback so the original orientation always wins.
            best = self._detect_names_below_subheaders(grid, max_row, max_col)
        if best is None:
            raise MatrixStructureError(
                "Could not detect the bidder-name row: no row had >=2 firm-name "
                "cells with the COST/$/SF sub-header quartet below it (nor a "
                "name row just below a sub-header row — the create-matrix "
                "orientation). Set matrix.header_row_hint in config to override."
            )
        return best[0]

    # In the create-matrix (falke-bid-tools) house format the bidder names sit
    # 1 row below the sub-header row (headers row 4, names row 5). Scan a small
    # window so an extra inserted row doesn't break detection.
    NAMES_BELOW_SUBHEADER_WINDOW = 3

    def _detect_names_below_subheaders(self, grid, max_row: int, max_col: int):
        """Fallback name-row detector for sub-headers-ABOVE-names layouts.

        Locates the row carrying the most distinct sub-header buckets (>=2 —
        the create-matrix Bid_Form mirror carries only COST SUBTOTALS + $/SF,
        so the full quartet is not required), then picks the row within a small
        window BELOW it that carries the most DISTINCT firm-name cells.
        Distinct-ness matters: the row under the names repeats the project name
        per bidder (identical values) and must not win over the name row.
        Returns (row, score) like the primary detector, or None."""
        best_q = None
        for r in range(1, min(30, max_row) + 1):
            buckets = set()
            for c in range(1, max_col + 1):
                bk = classify_subheader(grid.cell(r, c))
                if bk is not None:
                    buckets.add(bk)
            if len(buckets) >= 2 and (best_q is None or len(buckets) > best_q[1]):
                best_q = (r, len(buckets))
        if best_q is None:
            return None
        subheader_row = best_q[0]
        best = None
        upper = min(subheader_row + self.NAMES_BELOW_SUBHEADER_WINDOW, max_row)
        for r in range(subheader_row + 1, upper + 1):
            distinct = {
                normalize_name(str(grid.cell(r, c)).strip())
                for c in range(1, max_col + 1)
                if self._looks_like_name(grid.cell(r, c))
            }
            if len(distinct) >= 2 and (best is None or len(distinct) > best[1]):
                best = (r, len(distinct))
        return best

    def _find_quartet_row(self, grid, header_row: int, max_row: int, max_col: int):
        """Find the row carrying the sub-header quartet for `header_row`.

        Scans rows (header_row+1 .. header_row+QUARTET_SCAN_WINDOW), returning
        (row, distinct_bucket_count) for the row with the most distinct quartet
        buckets, or None if no row has >=2 buckets. Buckets are matched by
        classify_subheader (tolerant of typos/spacing), NOT literal text."""
        best = None
        upper = min(header_row + self.QUARTET_SCAN_WINDOW, max_row)
        for rr in range(header_row + 1, upper + 1):
            buckets = set()
            for c in range(1, max_col + 1):
                bk = classify_subheader(grid.cell(rr, c))
                if bk is not None:
                    buckets.add(bk)
            if len(buckets) >= 2 and (best is None or len(buckets) > best[1]):
                best = (rr, len(buckets))
        return best

    def _find_quartet_row_above(self, grid, header_row: int, max_col: int):
        """Find the sub-header row ABOVE `header_row` (create-matrix layout).

        Scans rows (header_row-NAMES_BELOW_SUBHEADER_WINDOW .. header_row-1),
        returning (row, distinct_bucket_count) for the row with the most
        distinct sub-header buckets, or None if no row has >=2. Same bucket
        matching (classify_subheader) as the below-scan."""
        best = None
        lower = max(1, header_row - self.NAMES_BELOW_SUBHEADER_WINDOW)
        for rr in range(lower, header_row):
            buckets = set()
            for c in range(1, max_col + 1):
                bk = classify_subheader(grid.cell(rr, c))
                if bk is not None:
                    buckets.add(bk)
            if len(buckets) >= 2 and (best is None or len(buckets) > best[1]):
                best = (rr, len(buckets))
        return best

    @staticmethod
    def _looks_like_name(v) -> bool:
        if v is None:
            return False
        s = str(v).strip()
        if not s or s.isdigit():
            return False
        # at least one letter, not a pure label like 'COST'
        if not re.search(r"[A-Za-z]", s):
            return False
        # exclude the sub-header labels (incl. source typos like '$/SXFX') and
        # the GSF row label — these are structure, not firm names.
        if classify_subheader(s) is not None:
            return False
        return _norm_label(s) != "total gsf"

    def _detect_blocks(self, grid, header_row: int, max_col: int, max_row: int
                       ) -> Tuple[List[BidderBlock], int, int]:
        """Build ordered (name, start_col) by scanning the header row, then
        MEASURE block width (the quartet span) and stride (gap to next name).

        The sub-header quartet is located ONCE (it lives on a single row that
        may be several rows below the names — decoupled from the name row), then
        each bidder block maps its quartet columns from that row."""
        named_cols: List[Tuple[int, str]] = []
        for c in range(1, max_col + 1):
            v = grid.cell(header_row, c)
            if self._looks_like_name(v):
                named_cols.append((c, str(v).strip()))
        if len(named_cols) < 1:
            raise MatrixStructureError(
                f"No bidder names found on detected header row {header_row}.")

        # MEASURE stride from spacing between consecutive named columns.
        strides = [named_cols[i + 1][0] - named_cols[i][0]
                   for i in range(len(named_cols) - 1)]
        stride = self._mode(strides) if strides else 5
        # remember the measured stride: the completeness counter uses it as the
        # block column-span for its resilient fallback (see _count_populated_divisions).
        self._stride = stride

        # Locate the quartet row below the names (allowing intervening rows).
        quartet = self._find_quartet_row(grid, header_row, max_row, max_col)
        if quartet is None:
            # create-matrix orientation: the sub-header row sits ABOVE the
            # names (headers row 4, names row 5). Scan the small window above
            # the name row before giving up — mirrors _detect_names_below_
            # subheaders, so both detectors agree on the orientation.
            quartet = self._find_quartet_row_above(grid, header_row, max_col)
        if quartet is None:
            raise MatrixStructureError(
                f"No COST/$/SF sub-header quartet row found below name row "
                f"{header_row} (scanned {self.QUARTET_SCAN_WINDOW} rows) nor "
                f"just above it (scanned {self.NAMES_BELOW_SUBHEADER_WINDOW} "
                f"rows — the create-matrix orientation).")
        self._quartet_row = quartet[0]

        blocks: List[BidderBlock] = []
        widths: List[int] = []
        for (start_col, raw) in named_cols:
            cols = self._map_subheaders(
                grid, self._quartet_row, start_col, stride, max_col)
            if not cols:
                # a named cell with no quartet beneath -> not a bidder block; skip
                continue
            width = max(cols.values()) - start_col + 1
            widths.append(width)
            blocks.append(BidderBlock(
                raw_name=raw,
                name=display_name(raw),
                norm=normalize_name(raw),
                start_col=start_col,
                cols=cols,
            ))
        if not blocks:
            raise MatrixStructureError(
                "Header row had names but none had the COST/$/SF sub-header "
                "quartet below; cannot form bidder blocks.")
        width = self._mode(widths)
        return blocks, width, stride

    # how many rows around the located quartet row to also scan when resolving a
    # block's sub-headers. The real validation matrix wraps the two-word labels
    # ("COST \nSUBTOTALS", "$/SXFX \nSUBTOTALS") and some blocks render the
    # "SUBTOTALS" half one row below the "COST"/"$/SF" half, so a single fixed
    # quartet row misses the *_SUBTOTAL buckets for those blocks (the Defect:
    # false 0/20 root cause — Crest col F / Dorne col K never resolved into cols).
    QUARTET_ROW_PROBE = 2

    def _map_subheaders(self, grid, quartet_row: int, start_col: int,
                        stride: int, max_col: int) -> Dict[str, int]:
        """Map each quartet BUCKET to its column for the block starting at
        start_col. Keyed by the canonical bucket constants (SUB_COST, ...) —
        tolerant of source typos/spacing/newlines because the match goes through
        classify_subheader, not literal text.

        Resilience (Defect: false 0/20). Two source quirks broke the *_SUBTOTAL
        buckets for specific blocks (Crest @ E–H, Dorne @ J–M):
          1. the labels carry in-cell newlines ("COST \\nSUBTOTALS"); and
          2. on some blocks the wrapped "SUBTOTALS" half sits one row off the
             single globally-located quartet row, so reading only that one row
             resolved COST but not COST_SUBTOTALS.
        We therefore (a) scan the block's FULL span [start_col, start_col+stride)
        across a small row window around the quartet row, and (b) if the
        COST_SUBTOTAL/PSF buckets are still unresolved while COST is present,
        derive them POSITIONALLY from the canonical contiguous quartet layout
        (COST, COST_SUBTOTAL, $/SF, $/SF_SUBTOTAL). This guarantees the
        completeness counter can read the correct COST_SUBTOTALS column per block
        and never manufactures a false zero (Marvin §1.4)."""
        out: Dict[str, int] = {}
        upper = min(start_col + max(stride, 4), max_col + 1)
        # scan a small row window around the located quartet row to absorb
        # blocks whose wrapped "SUBTOTALS" half is offset by a row.
        row_lo = max(1, quartet_row - self.QUARTET_ROW_PROBE)
        row_hi = quartet_row + self.QUARTET_ROW_PROBE
        for c in range(start_col, upper):
            if any(k for k, v in out.items() if v == c):
                continue
            for rr in range(row_lo, row_hi + 1):
                bk = classify_subheader(grid.cell(rr, c))
                if bk is not None and bk not in out:
                    out[bk] = c
                    break
        # require a cost-bearing column to consider this a block. The reference
        # layout always carries the per-line COST column; the create-matrix
        # Bid_Form mirror writes 2-column groups (COST SUBTOTALS | $/SF) with
        # NO plain COST column, so COST_SUBTOTAL alone also qualifies —
        # _read_block_value already probes COST first, then COST_SUBTOTAL.
        if SUB_COST not in out:
            if SUB_COST_SUBTOTAL not in out:
                return {}
            # no COST anchor -> the canonical-contiguous positional backfill
            # below cannot run (it is anchored on the COST column); the
            # classified buckets are the whole block.
            return out
        # positional backfill: the source quartet is contiguous starting at the
        # COST column. If a *_SUBTOTAL / $/SF bucket failed to classify for THIS
        # block (newline/offset quirk), place it at its canonical offset so the
        # division-subtotal column is never lost. Only fills columns inside the
        # block span and never overwrites a classified bucket.
        cost_c = out[SUB_COST]
        canonical = [SUB_COST, SUB_COST_SUBTOTAL, SUB_PSF, SUB_PSF_SUBTOTAL]
        used_cols = set(out.values())
        for offset, bucket in enumerate(canonical):
            target = cost_c + offset
            if bucket in out:
                continue
            if target >= upper or target in used_cols:
                continue
            out[bucket] = target
            used_cols.add(target)
        return out

    # Exact-label set for the compared total. The producer's Bid_Form mirror
    # writes the display label 'GRAND TOTAL' (col B) beside the col-A machine
    # key; the leveled sheet and the legacy reference carry the full label
    # (matched via the configured gt_label). Exact matches only — NEVER a
    # substring scan against prose (the v0.4.0 legend break, P0-3).
    GT_EXACT_LABELS = ("grand total", "grand total construction cost")
    # col-A machine key the producer writes on the Bid_Form footer.
    GT_MACHINE_KEY = "grand_total"

    def _locate_grand_total(self, grid, max_row: int,
                            blocks: Optional[List[BidderBlock]] = None
                            ) -> Tuple[int, str]:
        """Find the compared-total row. NEVER the CONSTRUCTION COST SUBTOTAL
        row (Marvin §1.2).

        Detection order (P0-3, Floyd verdict b — surgical machine-key/fenced
        fallback, not a format-adapter layer):
          1. the CONFIGURED exact label (case/space-insensitive);
          2. the producer's col-A MACHINE KEY 'GRAND_TOTAL' (house format);
          3. a GT_EXACT_LABELS exact match (e.g. the mirror's bare
             'GRAND TOTAL');
          4. FENCED fallback: the lowest row strictly below the construction
             cost subtotal whose label contains the WORD 'total' (word-
             boundary — 'subtotal' inside legend prose can never match) AND
             which carries a numeric value in at least one bidder column
             (structural fence: a legend/prose row has no bidder-column
             values). The v0.4.0 unified-legend row fails BOTH fences.

        Side effect: records the detected CONSTRUCTION COST SUBTOTAL row on
        ``self._cc_subtotal_row`` (None if absent) so the QA fingerprint test
        and the division-row scan can use that pre-markup boundary. It is
        recorded, never used as the compared total."""
        cc_subtotal_row = None
        candidate_total_rows: List[Tuple[int, str]] = []
        gt_configured: Optional[Tuple[int, str]] = None
        gt_machine_key: Optional[Tuple[int, str]] = None
        gt_exact: Optional[Tuple[int, str]] = None
        word_total_re = re.compile(r"(?<![a-z0-9])total(?![a-z0-9])")
        for r in range(1, max_row + 1):
            # machine key lives in col A even when the display label row
            # resolves via col B — check it independently of the label scan.
            key_a = grid.cell(r, 1)
            if (key_a is not None and gt_machine_key is None
                    and _norm_label(key_a).replace(" ", "_") ==
                    self.GT_MACHINE_KEY):
                # report the col-B display label when present (the machine key
                # itself is the anchor, not the board-facing label)
                disp = grid.cell(r, 2)
                gt_machine_key = (
                    r, str(disp).strip() if isinstance(disp, str) and
                    disp.strip() else str(key_a).strip())
            for col in (2, 3, 1):  # B, C, A label columns
                label = grid.cell(r, col)
                if label is None:
                    continue
                nl = _norm_label(label)
                if nl == self.gt_label and gt_configured is None:
                    gt_configured = (r, str(label).strip())
                if nl in self.GT_EXACT_LABELS and gt_exact is None:
                    gt_exact = (r, str(label).strip())
                if nl == self.cc_subtotal_label:
                    cc_subtotal_row = r
                if (word_total_re.search(nl)
                        and nl != self.cc_subtotal_label):
                    candidate_total_rows.append((r, str(label).strip()))
                break  # first non-empty label column wins for this row
        self._cc_subtotal_row = cc_subtotal_row
        for found in (gt_configured, gt_machine_key, gt_exact):
            if found is not None:
                return found
        # FENCED fallback: lowest word-'total' row strictly below the
        # construction cost subtotal that carries a numeric in a bidder column.
        if cc_subtotal_row is not None:
            below = [(r, lab) for (r, lab) in candidate_total_rows
                     if r > cc_subtotal_row
                     and self._row_has_bidder_value(grid, r, blocks)]
            if below:
                r, lab = max(below, key=lambda t: t[0])
                return r, lab
        raise GrandTotalNotFoundError(
            "Could not locate 'GRAND TOTAL CONSTRUCTION COST' (nor the "
            "producer's col-A GRAND_TOTAL machine key, nor an exact "
            "'GRAND TOTAL' label) and no safe fallback total row exists below "
            "the markup adders. Refusing to use 'CONSTRUCTION COST SUBTOTAL' "
            "(pre-markup, not apples-to-apples). Check the matrix or set "
            "matrix.grand_total_label in config."
        )

    @staticmethod
    def _row_has_bidder_value(grid, row: int,
                              blocks: Optional[List[BidderBlock]]) -> bool:
        """Structural prose fence: True when the row carries a numeric in at
        least one detected bidder column. Legend/prose rows carry none. With no
        blocks supplied (defensive), the fence is inert (returns True) — the
        word-boundary fence above still holds."""
        if not blocks:
            return True
        for b in blocks:
            for c in b.cols.values():
                if isinstance(grid.cell(row, c), (int, float)):
                    return True
        return False

    def _locate_gsf(self, grid, max_row: int) -> Tuple[Optional[int], Optional[float]]:
        for r in range(1, max_row + 1):
            for col in (2, 3, 1):
                label = grid.cell(r, col)
                if label is None:
                    continue
                if _norm_label(label) == self.gsf_label:
                    # value usually sits to the right of the label
                    for c in range(col + 1, col + 8):
                        v = grid.cell(r, c)
                        if isinstance(v, (int, float)):
                            return r, float(v)
                    return r, None
                break
        # FALLBACK (create-matrix / falke-bid-tools format): no labeled
        # TOTAL-GSF row exists; the SF basis appears in the row-2 details line
        # as a '|'-separated segment like '12,000 GSF' (label configurable, so
        # any SF-suffixed token qualifies). Scan the top title rows' col A.
        # As with the labeled read, the value is SUGGESTED/REPORTED only —
        # NEVER used for $/SF (Marvin §3); the CLI gate still confirms it.
        return self._gsf_from_details_text(grid, max_row)

    # how many top rows to scan for the details-line GSF fallback (the
    # create-matrix writer puts the details on row 2; scan a few for safety).
    GSF_DETAILS_SCAN_ROWS = 6
    # a details segment carrying the SF basis: '<number> <label ending in SF>'
    # e.g. '12,000 GSF' or '100,000 balcony SF'.
    _GSF_SEGMENT_RE = re.compile(
        r"^([\d,]+(?:\.\d+)?)\s+[^|]*SF$", re.IGNORECASE)

    def _gsf_from_details_text(self, grid, max_row: int
                               ) -> Tuple[Optional[int], Optional[float]]:
        for r in range(1, min(self.GSF_DETAILS_SCAN_ROWS, max_row) + 1):
            v = grid.cell(r, 1)
            if not isinstance(v, str) or "|" not in v:
                continue
            for seg in v.split("|"):
                m = self._GSF_SEGMENT_RE.match(seg.strip())
                if m:
                    try:
                        return r, float(m.group(1).replace(",", ""))
                    except ValueError:  # pragma: no cover - regex guards this
                        continue
        return None, None

    def _detect_division_rows(self, grid, max_row: int, gt_row: int
                              ) -> List[Tuple[int, str]]:
        """Per-division SUBTOTAL rows above the grand-total row.

        The per-division dollar VALUES live on the SUBTOTAL rows (e.g.
        'GENERAL CONDITIONS SUBTOTAL', 'WOOD & PLASTICS SUBTOTAL'), which sit a
        few rows BELOW each 'DIV 0x 00 00' header. The header rows themselves
        are blank in the cost-subtotal column for most bidders, so detecting
        them (the old 2-digit/'Division' regex) made the completeness counter
        read 0/20 and starved the fingerprint of Crest's row-49 subtotal.

        Predicate: a row whose label (cols 3 then 2) CONTAINS 'subtotal'
        (case/space/newline-insensitive), EXCLUDING the compared grand total
        ('GRAND TOTAL CONSTRUCTION COST'), the overall pre-markup
        'CONSTRUCTION COST SUBTOTAL', and the bare 'SUBTOTAL' (the OH&P/markup
        running total, whose normalized label is EXACTLY 'subtotal' with no
        division-name prefix) — none of these is a per-division subtotal. Genuine
        CSI division subtotals always carry a division-name prefix before the
        word (e.g. 'GENERAL CONDITIONS SUBTOTAL', 'MEHANICAL - HVAC SUBTOTAL').

        SCAN BOUND (P0-3 ride-along, C-R4): CSI division subtotals live ABOVE
        the construction cost subtotal; the footer fee block between it and the
        grand total carries 'Fees Subtotal', which is a MARKUP row, not a 21st
        CSI division. The scan therefore stops at the construction-subtotal row
        when one was detected (else at the grand-total row, the old bound)."""
        rows: List[Tuple[int, str]] = []
        cc_row = getattr(self, "_cc_subtotal_row", None)
        upper = cc_row if cc_row is not None else gt_row
        for r in range(1, upper):
            for col in (3, 2, 1):
                label = grid.cell(r, col)
                if label is None:
                    continue
                nl = _norm_label(label)
                if ("subtotal" in nl
                        and nl != self.gt_label
                        and nl != self.cc_subtotal_label
                        and nl != "subtotal"):
                    rows.append((r, str(label).strip()))
                break
        return rows

    # -- value reads ---------------------------------------------------------
    def _read_block_value(self, grid, row: int, b: BidderBlock) -> Optional[float]:
        """Read a numeric from a block's COST (then COST SUBTOTALS) column."""
        for key in (SUB_COST, SUB_COST_SUBTOTAL):
            c = b.cols.get(key)
            if c is None:
                continue
            v = grid.cell(row, c)
            if isinstance(v, (int, float)):
                return float(v)
        return None

    def _count_populated_divisions(self, grid, division_rows, b: BidderBlock,
                                   max_col: Optional[int] = None) -> int:
        """Count division-subtotal rows this block populates.

        At a CSI division-SUBTOTAL row the value lives in the COST_SUBTOTALS
        (running-subtotal) column; the per-line COST column is blank there.
        We therefore probe COST_SUBTOTAL FIRST, then COST (same proven ordering
        as fingerprint_test), so a populated subtotal is never missed.

        Resilience (Defect: false 0/20): if the COST_SUBTOTAL bucket failed to
        resolve into b.cols for this block (a per-block classify miss — the
        source headers carry typos/newlines like 'COST \\nSUBTOTALS'), reading
        only the per-line COST column would wrongly report 0 because that column
        is empty at every subtotal row. To avoid that false zero we fall back to
        scanning ALL columns in the block span [start_col, start_col+stride) at
        the division row. A division is 'populated' if ANY numeric appears in the
        block's own columns on that row — never auto-drop on a counter artifact
        (Marvin §1.4)."""
        cost_col = b.cols.get(SUB_COST)
        sub_col = b.cols.get(SUB_COST_SUBTOTAL)
        # block column span used as the resilient fallback when a bucket is
        # missing. Derive the span from the block's OWN resolved quartet columns
        # (max mapped col) so a mis-measured stride can never shrink the window
        # below the COST_SUBTOTALS column — then widen by the measured stride.
        # Defaults keep the window >= the canonical 4-wide quartet.
        stride = getattr(self, "_stride", None) or 5
        resolved_hi = max(b.cols.values()) + 1 if b.cols else b.start_col + 4
        span_hi = max(b.start_col + stride, resolved_hi, b.start_col + 4)
        if max_col is not None:
            span_hi = min(span_hi, max_col + 1)
        n = 0
        for (r, _label) in division_rows:
            populated = False
            # 1) bucket columns, COST_SUBTOTAL first (where subtotals actually sit)
            for c in (sub_col, cost_col):
                if c is None:
                    continue
                v = grid.cell(r, c)
                if isinstance(v, (int, float)) and v != 0:
                    populated = True
                    break
            # 2) fallback: ONLY when the COST_SUBTOTAL bucket failed to resolve
            #    for this block (the original false-zero scenario). Scanning the
            #    whole span when sub_col IS resolved would wrongly count rows that
            #    carry only a $/SF value where the cost subtotal is blank, so the
            #    fallback is gated on a missing cost-subtotal column. With the
            #    positional backfill above this path is rarely taken, but it keeps
            #    a misclassified block from ever manufacturing a false zero.
            if not populated and sub_col is None:
                for c in range(b.start_col, span_hi):
                    v = grid.cell(r, c)
                    if isinstance(v, (int, float)) and v != 0:
                        populated = True
                        break
            if populated:
                n += 1
        return n

    # -- duplicate + completeness rules --------------------------------------
    def _resolve_duplicates(self, blocks: List[BidderBlock], log: List[str]) -> None:
        seen: Dict[str, BidderBlock] = {}
        for b in blocks:
            if b.norm in seen:
                kept = seen[b.norm]
                # default: keep left-most (earlier). Drop this later one.
                if self.keep == "first":
                    b.included = False
                    b.drop_reason = (
                        f"duplicate of '{kept.name}' (col {kept.start_col_letter}); "
                        f"kept left-most per duplicate_keep='first' (Marvin §1.3)"
                    )
                    log.append(
                        f"DUPLICATE: '{b.name}' col {b.start_col_letter} "
                        f"(total {b.grand_total}) DROPPED; kept col "
                        f"{kept.start_col_letter} (total {kept.grand_total}). "
                        f"OPEN QUESTION: totals differ by "
                        f"{abs((b.grand_total or 0) - (kept.grand_total or 0)):,.0f} "
                        f"— supersession vs re-quote not resolvable from matrix "
                        f"(Marvin §11)."
                    )
                else:  # keep last
                    kept.included = False
                    kept.drop_reason = f"duplicate; kept later block per keep='last'"
                    seen[b.norm] = b
            else:
                seen[b.norm] = b

    def _flag_completeness(self, blocks, division_rows, peer_fraction, log) -> None:
        """Flag (NOT drop) blocks missing divisions that >= peer_fraction of
        peers populate (Marvin §1.4). Final include/exclude is Falke's call."""
        if not division_rows:
            return
        included = [b for b in blocks if b.included]
        if len(included) < 2:
            return
        median_pop = sorted(b.populated_divisions for b in included)[len(included) // 2]
        threshold = peer_fraction * median_pop
        for b in included:
            if b.populated_divisions < threshold:
                msg = (
                    f"COMPLETENESS OUTLIER: '{b.name}' populates "
                    f"{b.populated_divisions}/{len(division_rows)} CSI division "
                    f"subtotals (peer median {median_pop}). Flagged for Falke "
                    f"ruling — NOT auto-dropped (Marvin §1.4)."
                )
                b.flags.append(msg)
                log.append(msg)

    @staticmethod
    def _mode(values: List[int]) -> int:
        if not values:
            return 0
        counts: Dict[int, int] = {}
        for v in values:
            counts[v] = counts.get(v, 0) + 1
        return max(counts.items(), key=lambda kv: (kv[1], -kv[0]))[0]
