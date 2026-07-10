"""MECHANICAL layer — deterministic facts from matrix + the two PARAMETERS.

Covers Marvin §3 ($/SF), §4 (Section B tiers), §5 (Section C Bid + Variance
columns), §9 (ranking), and §2.2 (QA fingerprint test).

NOTHING here is judgment. Every output is reproducible from (matrix cells +
sf_basis + baseline band). The tier cuts come from config so Falke can tune.
"""
from __future__ import annotations

import math
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from .config import Config
from .matrix import (SUB_COST, SUB_COST_SUBTOTAL, BidderBlock, ParsedMatrix)


# tier identifiers (stable keys; map to Anna's CSS classes in render layer)
TIER_TOP = "TOP"
TIER_MID = "MID"
TIER_DEFENSIVE = "DEFENSIVE"
TIER_PREMIUM = "PREMIUM"
TIER_RISK = "RISK"

TIER_LABELS = {
    TIER_TOP: "TOP TIER (aligned)",
    TIER_MID: "MID (aggressive)",
    TIER_DEFENSIVE: "DEFENSIVE (above baseline)",
    TIER_PREMIUM: "PREMIUM (far above)",
    TIER_RISK: "HIGH RISK (under baseline)",
}
TIER_CSS = {
    TIER_TOP: "tier-top",
    TIER_MID: "tier-mid",
    TIER_DEFENSIVE: "tier-defensive",
    TIER_PREMIUM: "tier-premium",
    TIER_RISK: "tier-risk",
}


@dataclass
class BidderMechanical:
    name: str
    start_col_letter: str
    total: float                 # Row 164 grand total ($)
    per_sf: int                  # round(total / sf_basis)
    tier: str
    bid_m: float                 # total / 1e6, 2dp
    variance_m: float            # bid_m - variance_mid (2dp)
    variance_frac: float         # signed fractional variance from variance_mid
    flags: List[str] = field(default_factory=list)


def compute_per_sf(total: float, sf_basis: float) -> int:
    """$/SF = round(total / sf_basis). PARAMETER=sf_basis (Marvin §3)."""
    return int(round(total / sf_basis))


def assign_tier(per_sf: float, cfg: Config) -> str:
    """Section B tier from $/SF vs band (Marvin §4.1). MECHANICAL given band.

    $/SF is reported as a displayed INTEGER (round(total/sf_basis)), but the raw
    band bounds are fractional (e.g. 3.35e6/16000 = 209.4). Comparing the
    integer $/SF against the fractional band would push a bidder printed at the
    band edge (printed at the floored band floor) one tier too low. We
    therefore compare against the band bounds FLOORED to the displayed integer
    so the boundary aligns with what the card prints (Marvin §4.1). The MID and
    PREMIUM floors keep using the unrounded band (they are interior cuts, not
    edges a bidder lands exactly on).
    """
    run = cfg.run
    tcfg = cfg.block("tiers")
    band_low = run.band_low_per_sf
    band_high = run.band_high_per_sf
    band_low_int = math.floor(band_low)
    band_high_int = math.floor(band_high)
    mid_floor = tcfg["mid_floor_frac_of_band_low"] * band_low
    premium_floor = tcfg["premium_floor_frac_of_band_high"] * band_high

    if per_sf < mid_floor:
        return TIER_RISK
    if per_sf < band_low_int:
        return TIER_MID
    if per_sf <= band_high_int:
        return TIER_TOP
    if per_sf <= premium_floor:
        return TIER_DEFENSIVE
    return TIER_PREMIUM


def build_mechanical(parsed: ParsedMatrix, cfg: Config) -> List[BidderMechanical]:
    """One BidderMechanical per INCLUDED bidder, in matrix order."""
    run = cfg.run
    out: List[BidderMechanical] = []
    for b in parsed.included_blocks:
        if b.grand_total is None:
            # explicit, not silent: a block with no readable total is surfaced
            bm = BidderMechanical(
                name=b.name, start_col_letter=b.start_col_letter,
                total=float("nan"), per_sf=0, tier=TIER_RISK,
                bid_m=float("nan"), variance_m=float("nan"),
                variance_frac=float("nan"),
                flags=["NO GRAND-TOTAL VALUE READ — verify matrix block."],
            )
            out.append(bm)
            continue
        per_sf = compute_per_sf(b.grand_total, run.sf_basis)
        tier = assign_tier(per_sf, cfg)
        bid_m = round(b.grand_total / 1e6, 2)
        variance_m = round(bid_m - run.variance_mid, 2)
        variance_frac = (b.grand_total / 1e6 - run.variance_mid) / run.variance_mid
        out.append(BidderMechanical(
            name=b.name,
            start_col_letter=b.start_col_letter,
            total=b.grand_total,
            per_sf=per_sf,
            tier=tier,
            bid_m=bid_m,
            variance_m=variance_m,
            variance_frac=variance_frac,
            flags=list(b.flags),
        ))
    return out


# ----------------------------------------------------------------------------
# QA fingerprint test (Marvin §2.2 / §11) — bid-anchoring detector
# ----------------------------------------------------------------------------
@dataclass
class FingerprintHit:
    baseline_label: str
    baseline_value: float
    bidder_name: str
    bidder_subtotal_label: str
    bidder_value: float
    pct_delta: float


def fingerprint_test(
    baseline_lines: List[Dict],
    parsed: ParsedMatrix,
    xlsx_path: str,
    cfg: Config,
) -> List[FingerprintHit]:
    """Flag any supplied baseline line within tolerance% of ANY bidder subtotal.

    Reproduces the construction-cost-subtotal ($3,000,000 ~ a bidder's pre-markup
    subtotal) and finishes-division ($590,000 ~ a bidder's division subtotal)
    fingerprints. A hit means the 'independent' baseline may be
    bid-anchored — a board disclosure item, NOT an auto-fail (Marvin §2.2).

    baseline_lines: [{"label": str, "value": float}, ...] (the nine trade lines
    + subtotals supplied as PARAMETER input).
    """
    tol = cfg.block("qa").get("fingerprint_tolerance_pct", 0.2) / 100.0

    # Read from the SAME in-memory Grid the parser built (absolute 1-based, fast)
    # rather than reopening the workbook and doing read_only ``ws.cell()`` random
    # access. In read_only mode ``ws.cell(row, col)`` is unsupported/empty, so the
    # old path read None for EVERY subtotal cell and returned 0 hits — the
    # fingerprint-test root cause (same subtotal-read family as the completeness
    # false 0/20). If a grid is not attached (older callers) we rebuild one.
    grid = parsed.grid
    if grid is None:
        import openpyxl
        from .matrix import Grid
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        grid = Grid(wb[parsed.sheet_name])

    # gather every bidder subtotal-ish numeric per block: the per-division
    # subtotal rows + the CONSTRUCTION COST SUBTOTAL row (a pre-markup
    # construction-cost subtotal lives HERE, not on a division row) + the grand
    # total. Dedupe rows.
    candidate_rows: List[Tuple[int, str]] = list(parsed.division_rows)
    if parsed.construction_subtotal_row is not None:
        candidate_rows.append(
            (parsed.construction_subtotal_row, "CONSTRUCTION COST SUBTOTAL"))
    candidate_rows.append((parsed.grand_total_row, parsed.grand_total_label))
    seen_rows = set()
    deduped_rows = []
    for (r, rlabel) in candidate_rows:
        if r in seen_rows:
            continue
        seen_rows.add(r)
        deduped_rows.append((r, rlabel))

    hits: List[FingerprintHit] = []
    for line in baseline_lines:
        bval = line.get("value")
        if not isinstance(bval, (int, float)) or bval == 0:
            continue
        for b in parsed.blocks:  # all blocks, incl. dropped duplicates
            # COST_SUBTOTAL first (where running/division subtotals actually sit),
            # then the per-line COST column.
            for label_key in (SUB_COST_SUBTOTAL, SUB_COST):
                col = b.cols.get(label_key)
                if col is None:
                    continue
                for (r, rlabel) in deduped_rows:
                    v = grid.cell(r, col)
                    if not isinstance(v, (int, float)) or v == 0:
                        continue
                    if abs(v - bval) / bval <= tol:
                        hits.append(FingerprintHit(
                            baseline_label=line.get("label", "?"),
                            baseline_value=float(bval),
                            bidder_name=b.name,
                            bidder_subtotal_label=str(rlabel),
                            bidder_value=float(v),
                            pct_delta=abs(v - bval) / bval * 100.0,
                        ))
    return hits


# ----------------------------------------------------------------------------
# ranking (Marvin §9) — descending by Overall /100; tiebreak by Pricing then $/SF
# ----------------------------------------------------------------------------
def rank_bidders(
    rows: List[Dict],
) -> List[Dict]:
    """rows: [{name, overall, pricing_score, per_sf, ...}]. Returns sorted desc.

    Tiebreak: higher pricing_score, then lower per_sf-distance-above-band
    (best-value preference). Caller supplies per_sf_over_band if available.
    Bidders whose overall is None sort last (provisional / not fully scored).
    """
    def key(r):
        ov = r.get("overall")
        ov_key = ov if ov is not None else -1
        return (
            -ov_key,
            -(r.get("pricing_score") or 0),
            r.get("per_sf_over_band", 0),
        )
    return sorted(rows, key=key)
