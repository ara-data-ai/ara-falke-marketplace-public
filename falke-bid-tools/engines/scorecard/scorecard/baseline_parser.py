"""Parse a Falke baseline sheet into the scorecard's internal baseline
representation — the band plus the trade lines.

Two callers, ONE parser (P1-4):
  * ``parse_baseline_xlsx(path)`` — the individual ``--baseline`` file, in the
    shipped baseline-template.xlsx format (sheet ``Baseline``).
  * ``parse_baseline_sheet(ws, path, ...)`` — the run pack's Baseline tab, which
    carries the SAME band + trade lines PLUS the P1-6 provenance block.

Returns (band_low, band_high, band_mid, baseline_lines):
  - band_*: float ($M) — read from the Band labels
  - baseline_lines: list[dict] the engine already expects
    (keys: scope, basis, cost, value, kind[optional])

Raises ValueError with a user-friendly message if the sheet is missing, band
values are absent/non-numeric, or no trade lines are found. The CLI prints it as
``[STOP] ...`` and exits 2.

WHY THIS PARSER IS LABEL-ADDRESSED (Marvin R2 — a hard rule, not a style
preference)
--------------------------------------------------------------------------
It used to hard-code rows 3–5 for the band and row 8 for the first trade line.
That works right up until a spec adds a field. P1-6 adds a six-field provenance
block to exactly this header area, and under a row-index parser that shifts
every downstream read — the same shape of bug as the P0-3 grand-total break,
where a detector that located a row by guessing instead of by anchoring found
the legend prose and reported None totals on a board document.

So: every scalar is found by its label in column A (case/space-insensitive),
the trade table by a header-row scan, and every block terminates on a fully-blank
row. The shipped baseline-template.xlsx parses IDENTICALLY under this rule — its
labels already match — so nothing about the legacy channel changes except that
adding a row to it is now free instead of fatal.
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

from . import pack_schema as ps

BASELINE_SHEET = "Baseline"


def _label_row(ws, label: str) -> Optional[int]:
    """Row whose column A carries ``label`` (case/space-insensitive), or None."""
    target = ps.norm_label(label)
    for row in range(1, ws.max_row + 1):
        if ps.norm_label(ws.cell(row=row, column=1).value) == target:
            return row
    return None


def _band_value(ws, path: str, label: str) -> float:
    row = _label_row(ws, label)
    if row is None:
        raise ValueError(
            f"'{path}': the Baseline sheet has no '{label}' label in column A. "
            f"Fill out the Band section (do not delete or rename its labels)."
        )
    val = ws.cell(row=row, column=2).value
    if val is None or str(val).strip() == "":
        raise ValueError(
            f"Band value '{label}' is missing (row {row}, column B) in "
            f"'{path}'. Fill in the Band section of the baseline."
        )
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(
            f"Band value '{label}' (row {row}, column B) must be numeric; "
            f"got {val!r} in '{path}'."
        )


def _trade_header_row(ws, path: str) -> int:
    """Locate the trade-line table by scanning for its header row (R2)."""
    headers = [ps.norm_label(h) for h in ps.BASELINE_TRADE_HEADERS]
    for row in range(1, ws.max_row + 1):
        if ps.norm_label(ws.cell(row=row, column=1).value) != headers[0]:
            continue
        got = [ps.norm_label(ws.cell(row=row, column=c).value)
               for c in range(1, len(headers) + 1)]
        if got == headers:
            return row
    raise ValueError(
        f"'{path}': could not find the trade-line header row "
        f"({' | '.join(ps.BASELINE_TRADE_HEADERS)}) on the Baseline sheet. "
        f"Do not reshape the template."
    )


def parse_baseline_sheet(ws, path: str) -> Tuple[float, float, float,
                                                 List[Dict[str, Any]]]:
    """Parse the band + trade lines from an OPEN Baseline worksheet.

    Shared by the individual --baseline file and the run pack's Baseline tab.
    Provenance fields are NOT read here — they are the pack's own surface and
    are parsed by run_pack.py, because the legacy template has no such block and
    requiring one of it would break the escape hatch that legacy workbooks
    depend on (Marvin §9.1: legitimate INDEFINITELY).
    """
    band_low = _band_value(ws, path, ps.B_BAND_LOW)
    band_high = _band_value(ws, path, ps.B_BAND_HIGH)
    band_mid = _band_value(ws, path, ps.B_BAND_MID)

    header_row = _trade_header_row(ws, path)

    # Columns: A=scope, B=basis, C=cost_str, D=value, E=kind
    lines: List[Dict[str, Any]] = []
    row = header_row + 1
    while True:
        scope_val = ws.cell(row=row, column=1).value
        basis_val = ws.cell(row=row, column=2).value
        cost_val = ws.cell(row=row, column=3).value
        value_val = ws.cell(row=row, column=4).value
        kind_val = ws.cell(row=row, column=5).value

        # Stop at the first fully-blank row (all five columns empty/None)
        if all(v is None or str(v).strip() == ""
               for v in (scope_val, basis_val, cost_val, value_val, kind_val)):
            break

        # Skip rows where scope is blank (e.g. trailing partial rows)
        if scope_val is None or str(scope_val).strip() == "":
            row += 1
            continue

        scope_str = str(scope_val).strip()
        basis_str = str(basis_val).strip() if basis_val is not None else ""

        # Value must be numeric
        if value_val is None:
            raise ValueError(
                f"Trade line row {row}: 'Value' column (D) is blank for "
                f"scope '{scope_str}' in '{path}'. Each row needs a numeric value."
            )
        try:
            value_num = float(value_val)
        except (TypeError, ValueError):
            raise ValueError(
                f"Trade line row {row}: 'Value' column (D) must be numeric; "
                f"got {value_val!r} for scope '{scope_str}' in '{path}'."
            )
        value_int = int(round(value_num))

        # cost_str: use as-is if present, otherwise format from value
        if cost_val is not None and str(cost_val).strip():
            cost_str = str(cost_val).strip()
        else:
            cost_str = f"${value_int:,.0f}"

        line: Dict[str, Any] = {
            "scope": scope_str,
            "basis": basis_str,
            "cost": cost_str,
            "value": value_int,
        }
        if kind_val is not None and str(kind_val).strip():
            line["kind"] = str(kind_val).strip()

        lines.append(line)
        row += 1

    if not lines:
        raise ValueError(
            f"No trade lines found in '{path}' (expected data starting at row "
            f"{header_row + 1} of the '{ws.title}' sheet). Fill in the "
            f"trade-line section."
        )

    return band_low, band_high, band_mid, lines


def parse_baseline_xlsx(
    path: str,
) -> Tuple[float, float, float, List[Dict[str, Any]]]:
    """Parse a standalone baseline xlsx file (the --baseline escape hatch).

    Returns (band_low, band_high, band_mid, baseline_lines).
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError(
            "openpyxl is required to read xlsx baseline files "
            "(pip install openpyxl)."
        ) from exc

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open xlsx file '{path}': {exc}") from exc

    if BASELINE_SHEET not in wb.sheetnames:
        available = ", ".join(wb.sheetnames) if wb.sheetnames else "(none)"
        raise ValueError(
            f"Expected a sheet named '{BASELINE_SHEET}' in '{path}'; "
            f"found: {available}."
        )

    return parse_baseline_sheet(wb[BASELINE_SHEET], path)
