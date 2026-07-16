"""Parse the two REQUIRED per-run scoring xlsx inputs.

Every real render needs TWO Falke-filled files (NO fallback — the CLI
hard-stops with exit 2 without them; see cli.py):

  * the SCORING FRAMEWORK (scoring-framework-template.xlsx format) — the
    categories, weights, and descriptions that drive Section D and the
    Overall /100 weighting. Sheet ``Scoring_Framework``:
      Row 1: title
      Row 2: headers — Category | Short Label | Weight (%) | What it captures
      Row 3+: one category per row (stop at first fully-blank row)

  * the DETAILED CATEGORY SCORES (category-scores-template.xlsx format) — the
    per-bidder 1–10 scores that drive Section E. Sheet ``Category_Scores``:
      Row 1: title
      Row 2: headers — Firm | one column per framework Short Label
      Row 3+: one SCORED bidder per row (stop at first fully-blank row)
    The Overall /100 is COMPUTED by the engine and never supplied here.

These two files are the SINGLE SOURCE OF TRUTH for category weights and 1–10
scores: they supersede the config ``weights`` block and the old ``--overrides``
qual-scores JSON for any run that supplies them. Categories/weights may differ
per run — Sections D/E render dynamically from the framework.

All failures raise ValueError with a user-friendly, actionable message (the
CLI prints it as ``[STOP] ...`` and exits 2).

SHARED WITH THE RUN PACK (P1-4 — binding, Marvin §3.4)
------------------------------------------------------
``parse_framework_table`` and ``parse_scores_table`` take an already-open
worksheet plus the header row, so the run pack's Framework/Scores tabs
(run_pack.py, where the header row is FOUND BY SCAN per R2) and the individual
``--scoring-framework`` / ``--category-scores`` files (where it is row 2, per
the shipped templates) are read by the SAME code with the SAME semantics. This
is not incidental reuse — Marvin made it binding: "the pack parser and the
individual-flag parser must be the same code reading the same semantics. No
divergence, ever."

That is what made the blank-cell question answer itself. P1-2 has since landed:
a blank score means NOT YET SCORED, and both channels inherited the new semantic
in the same commit, because there is only one place it lives. That is the reuse
paying out exactly as ruled.

BLANK-CELL SEMANTICS (P1-2 — Marvin's provisional-pathway ruling)
-----------------------------------------------------------------
A blank score cell means the category has not yet been scored for that bidder.
Nothing else — never zero, never a middle value, never "doesn't apply". It is
carried as None and NEVER omitted (an omitted key would make "blank" and
"column absent" indistinguishable, so coverage would be computed off a shape
rather than off a fact).

Blanks are self-penalizing: the engine does not rescale, so an unscored category
costs the bidder its full weight. That property is why there is no ``n/a``
sentinel — one would RAISE a score by removing evidence.

There is exactly ONE hard stop, and it is degenerate rather than a judgment
about sufficiency: a grid with zero scored cells across the entire field is the
blank template, not a partial evaluation record. The tool refuses to render
nothing; it never refuses to render little.
"""
from __future__ import annotations

import re
from typing import Any, Dict, List, Optional

from .matrix import normalize_name
from .scoring import BidderScores, CategoryScore

FRAMEWORK_SHEET = "Scoring_Framework"
SCORES_SHEET = "Category_Scores"

FRAMEWORK_HEADERS = ("Category", "Short Label", "Weight (%)", "What it captures")
SCORES_FIRM_HEADER = "Firm"

WEIGHT_SUM_TOL = 0.01
SCORE_MIN, SCORE_MAX = 1.0, 10.0


def _norm_header(s: Any) -> str:
    """Case/space-insensitive header key ('Short  Label' -> 'short label')."""
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _load_sheet(path: str, sheet: str, kind: str):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ValueError(
            "openpyxl is required to read xlsx scoring inputs "
            "(pip install openpyxl)."
        ) from exc
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open {kind} xlsx '{path}': {exc}") from exc
    if sheet not in wb.sheetnames:
        available = ", ".join(wb.sheetnames) if wb.sheetnames else "(none)"
        raise ValueError(
            f"Expected a sheet named '{sheet}' in '{path}'; found: {available}. "
            f"Use the shipped template so the sheet name matches."
        )
    return wb[sheet]


def parse_scoring_framework(path: str) -> List[Dict[str, Any]]:
    """Parse a scoring-framework xlsx.

    Returns a list of dicts, one per category, in sheet order:
      {"category", "short_label", "weight" (float, percent), "description",
       "key" (normalized short-label slug used internally)}

    Validates: >=1 category row; non-empty Category + Short Label; numeric
    weights summing to 100 (+/-0.01); no duplicate short labels.
    """
    ws = _load_sheet(path, FRAMEWORK_SHEET, "scoring-framework")

    # header row 2 sanity check (catch a file built off the wrong template)
    got = [_norm_header(ws.cell(row=2, column=c).value) for c in (1, 2, 3)]
    want = [_norm_header(h) for h in FRAMEWORK_HEADERS[:3]]
    if got != want:
        raise ValueError(
            f"'{path}': row 2 must carry the template headers "
            f"{' | '.join(FRAMEWORK_HEADERS)}; found "
            f"{[ws.cell(row=2, column=c).value for c in range(1, 5)]}. "
            f"Fill out scoring-framework-template.xlsx (do not reshape it)."
        )
    return parse_framework_table(ws, path, header_row=2)


def parse_framework_table(ws, path: str, *, header_row: int) -> List[Dict[str, Any]]:
    """Parse the framework table on an OPEN worksheet, starting one row below
    ``header_row`` and stopping at the first fully-blank row.

    Shared verbatim by the individual --scoring-framework file (header_row=2,
    fixed by the shipped template) and the run pack's Framework tab (header_row
    found by scan, per R2). Validation is identical for both: >=1 category row;
    non-empty Category + Short Label; numeric weights summing to 100 (+/-0.01);
    no duplicate short labels.
    """
    rows: List[Dict[str, Any]] = []
    r = header_row + 1
    while True:
        vals = [ws.cell(row=r, column=c).value for c in (1, 2, 3, 4)]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        category, short_label, weight, description = vals
        if category is None or str(category).strip() == "":
            raise ValueError(
                f"'{path}' row {r}: 'Category' (column A) is blank. Every "
                f"framework row needs a category name."
            )
        if short_label is None or str(short_label).strip() == "":
            raise ValueError(
                f"'{path}' row {r}: 'Short Label' (column B) is blank for "
                f"category '{str(category).strip()}'. The short label is what "
                f"the Category_Scores column headers must match."
            )
        try:
            weight_f = float(weight)
        except (TypeError, ValueError):
            raise ValueError(
                f"'{path}' row {r}: 'Weight (%)' (column C) must be numeric; "
                f"got {weight!r} for category '{str(category).strip()}'."
            )
        rows.append({
            "category": str(category).strip(),
            "short_label": str(short_label).strip(),
            "weight": weight_f,
            "description": str(description).strip() if description is not None else "",
            "key": normalize_name(str(short_label)),
        })
        r += 1

    if not rows:
        raise ValueError(
            f"No framework rows found in '{path}' (expected data starting at "
            f"row {header_row + 1} of the '{ws.title}' sheet). Every run needs "
            f"at least one scoring category."
        )

    # duplicate short labels (matched on the normalized slug)
    seen: Dict[str, str] = {}
    for row in rows:
        if row["key"] in seen:
            raise ValueError(
                f"'{path}': duplicate Short Label — '{row['short_label']}' "
                f"collides with '{seen[row['key']]}'. Short labels must be "
                f"unique (they become the Category_Scores column headers)."
            )
        if not row["key"]:
            raise ValueError(
                f"'{path}': Short Label '{row['short_label']}' contains no "
                f"letters/digits — give it a usable label."
            )
        seen[row["key"]] = row["short_label"]

    total = sum(row["weight"] for row in rows)
    if abs(total - 100.0) > WEIGHT_SUM_TOL:
        hint = ""
        if abs(total - 1.0) <= 0.001:
            hint = (" (The weights sum to 1.0 — if the Weight column is "
                    "Excel-percent-formatted, the cells store fractions; enter "
                    "plain numbers like 25 instead.)")
        raise ValueError(
            f"'{path}': framework weights must sum to 100 (+/-{WEIGHT_SUM_TOL}); "
            f"they sum to {total:g}. Adjust the Weight (%) column.{hint}"
        )
    return rows


def parse_category_scores(
    path: str,
    framework: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Optional[float]]]:
    """Parse a category-scores xlsx against an already-parsed framework.

    Returns {firm (as written): {framework short_label: score-or-None}} — every
    framework category gets an entry for every firm, and a BLANK cell carries
    None rather than being omitted (P1-2 §1.1).

    Validates: score columns exactly match the framework's short labels
    (order-insensitive; the error names missing/extra columns); every score that
    IS present is numeric and within 1–10; no duplicate firms (matched on the
    same normalized-name rule the matrix parser uses).

    A blank is NOT an error — it means NOT YET SCORED, and it costs the bidder
    the category's full weight (the engine never rescales). The single hard stop
    is degenerate: a grid with zero scored cells anywhere is the blank template,
    not a partial evaluation record (§1.2).
    """
    ws = _load_sheet(path, SCORES_SHEET, "category-scores")

    # ---- header row 2: Firm | <short labels...> ----
    if _norm_header(ws.cell(row=2, column=1).value) != _norm_header(SCORES_FIRM_HEADER):
        raise ValueError(
            f"'{path}': row 2 column A must be the '{SCORES_FIRM_HEADER}' "
            f"header; found {ws.cell(row=2, column=1).value!r}. Fill out "
            f"category-scores-template.xlsx (do not reshape it)."
        )
    return parse_scores_table(ws, path, framework, header_row=2)


def parse_scores_table(
    ws,
    path: str,
    framework: List[Dict[str, Any]],
    *,
    header_row: int,
    skip_firms: Optional[List[str]] = None,
) -> Dict[str, Dict[str, Optional[float]]]:
    """Parse the scores table on an OPEN worksheet against a parsed framework.

    Shared verbatim by the individual --category-scores file (header_row=2) and
    the run pack's Scores tab (header_row found by scan, per R2). Blank-cell
    semantics are therefore identical by construction, which is exactly the
    point (Marvin §3.4): the pack CARRIES the semantic, it never redefines it.

    ``skip_firms`` drops named rows BEFORE they are validated. Only the pack
    passes it, and only with the firms its Settings tab excludes — because the
    pack's Firm column is producer-filled at matrix time, days before the
    operator rules a bidder out, so an excluded bidder necessarily still has a
    row. Marvin §3.4 is explicit that deleting the row is NOT the mechanism
    (it breaks the roster reconcile); Settings is, with a reason. Skipping
    before validation is what lets the operator exclude a bidder without first
    scoring the bidder they just excluded.

    This does not diverge the two channels: the rows that ARE parsed are parsed
    identically. The individual-flag path passes nothing here, keeping its own
    long-standing contract that an excluded bidder simply must not appear.
    """
    file_labels: List[str] = []
    c = 2
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == "":
            break
        file_labels.append(str(v).strip())
        c += 1

    fw_by_key = {row["key"]: row["short_label"] for row in framework}
    file_by_key: Dict[str, str] = {}
    for lab in file_labels:
        k = normalize_name(lab)
        if k in file_by_key:
            raise ValueError(
                f"'{path}': duplicate score column '{lab}' (also present as "
                f"'{file_by_key[k]}'). One column per framework Short Label."
            )
        file_by_key[k] = lab

    missing = [fw_by_key[k] for k in fw_by_key if k not in file_by_key]
    extra = [file_by_key[k] for k in file_by_key if k not in fw_by_key]
    if missing or extra:
        parts = [f"'{path}': score columns must exactly match the scoring "
                 f"framework's Short Labels."]
        if missing:
            parts.append("Missing column(s): " + ", ".join(missing) + ".")
        if extra:
            parts.append("Unexpected column(s): " + ", ".join(extra) + ".")
        parts.append("Framework short labels: "
                     + ", ".join(row["short_label"] for row in framework) + ".")
        raise ValueError(" ".join(parts))

    # column index per framework label (order-insensitive match)
    col_for_label: Dict[str, int] = {}
    for idx, lab in enumerate(file_labels, start=2):
        col_for_label[normalize_name(lab)] = idx

    # ---- data rows (one below the header, to the first fully-blank row) ----
    scores: Dict[str, Dict[str, Optional[float]]] = {}
    seen_firms: Dict[str, str] = {}
    skip_keys = {normalize_name(f) for f in (skip_firms or [])}
    r = header_row + 1
    n_cols = 1 + len(file_labels)
    while True:
        vals = [ws.cell(row=r, column=cc).value for cc in range(1, n_cols + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        firm_val = vals[0]
        if firm_val is None or str(firm_val).strip() == "":
            raise ValueError(
                f"'{path}' row {r}: scores present but 'Firm' (column A) is "
                f"blank. Every scores row needs the bidder's name."
            )
        firm = str(firm_val).strip()
        fk = normalize_name(firm)
        if fk in skip_keys:
            # excluded by an explicit, reasoned ruling — not scored, not
            # validated, and not silently dropped (the ruling and its reason are
            # on the Settings tab and in the run json).
            r += 1
            continue
        if fk in seen_firms:
            raise ValueError(
                f"'{path}': duplicate firm '{firm}' (also listed as "
                f"'{seen_firms[fk]}'). One row per scored bidder."
            )
        seen_firms[fk] = firm

        firm_scores: Dict[str, Optional[float]] = {}
        for row in framework:
            col = col_for_label[row["key"]]
            raw = ws.cell(row=r, column=col).value
            if raw is None or str(raw).strip() == "":
                # BLANK = NOT YET SCORED. Always. Never zero, never a middle
                # value, never "doesn't apply" (Marvin P1-2 §1.1).
                #
                # Set to None, NEVER omitted — this is his binding build rule and
                # the one he'd most regret leaving implicit. An omitted key makes
                # "blank" and "column absent" indistinguishable, so coverage would
                # be computed off a SHAPE rather than off a FACT. Same instinct as
                # R2: never let an absence be inferred from a shape.
                #
                # Blanks are safe because the engine does not rescale: an unscored
                # category costs the bidder its full weight, so a blank is
                # self-penalizing and can never be used to advantage a favourite.
                # That property is exactly why there is no `n/a` sentinel (§1.3) —
                # a sentinel that drops a category from one bidder's denominator
                # would RAISE their number by removing evidence, which is a
                # per-cell rank-manipulation surface on the one tab the operator
                # types into. We killed the curve for less.
                firm_scores[row["short_label"]] = None
                continue
            try:
                score = float(raw)
            except (TypeError, ValueError):
                raise ValueError(
                    f"'{path}' row {r}: score for '{firm}' / "
                    f"'{row['short_label']}' must be numeric 1–10; got {raw!r}."
                )
            if not (SCORE_MIN <= score <= SCORE_MAX):
                raise ValueError(
                    f"'{path}' row {r}: score for '{firm}' / "
                    f"'{row['short_label']}' must be within 1–10; got {score:g}."
                )
            firm_scores[row["short_label"]] = (
                int(score) if float(score).is_integer() else score)
        scores[firm] = firm_scores
        r += 1

    if not scores:
        raise ValueError(
            f"No bidder rows found in '{path}' (expected data starting at "
            f"row {header_row + 1} of the '{ws.title}' sheet). One row per "
            f"SCORED bidder."
        )

    # THE ONE HARD STOP (Marvin P1-2 §1.2), and it is DEGENERATE — not a
    # judgment about whether the evaluation is far enough along.
    #
    #   The tool does not rule on whether the evaluation is far enough along.
    #   It rules on whether the document tells the truth about how far along
    #   it is.
    #
    # Sufficiency is the precon lead's judgment and it is the judgment we hired
    # him for; a tool that second-guesses it invents a threshold it cannot
    # defend and trains the operator to route around it. So a wholly-unscored
    # bidder renders, a never-evaluated category renders, 1-of-64 renders. The
    # tool refuses to render NOTHING. It never refuses to render LITTLE.
    #
    # And this stop is not paternalism, because THE MATRIX ALREADY IS THE
    # ZERO-COVERAGE ARTIFACT: a PM who wants price, $/SF, tiers and variance
    # before scoring anything already has that document, in more detail. The
    # scorecard's reason to exist is the COMBINATION. With nothing to combine it
    # is the matrix at lower resolution wearing a board document's clothes.
    if not any(v is not None
               for cells in scores.values() for v in cells.values()):
        raise ValueError(
            f"No category scores were supplied in '{path}' — every cell in the "
            f"grid is blank. A scorecard combines the price picture with the "
            f"qualitative evaluation; with no scores entered there is nothing "
            f"for it to combine, and the leveled matrix already carries every "
            f"price fact on its own. Score what you know and re-run."
        )
    return scores


def build_scores_from_inputs(
    name: str,
    framework: List[Dict[str, Any]],
    firm_scores: Dict[str, Optional[float]],
    *,
    run_id: Optional[str] = None,
) -> BidderScores:
    """Assemble a BidderScores from the xlsx inputs (single source of truth).

    Coverage is whatever the operator actually entered — NOT always 100%
    (P1-2 restored the provisional pathway; the old docstring's claim that
    "coverage is always 100% by construction" was true only because the parser
    rejected the blanks that would have reached it, which is precisely the dead
    pathway F3 identified).

    A blank arrives here as None and is recorded as genuinely unscored:
    `is_scored` keys on `effective_score is not None`, so the category drops out
    of the weighted sum and out of coverage on its own. Provenance is recorded
    per category, and an unscored one says so rather than claiming evidence it
    does not have.

    `firm_scores` is indexed directly, not `.get`-ed, on purpose: the parser
    guarantees an entry for every framework category (blanks carry None), so a
    missing key is a genuine bug and should be a loud KeyError — never a silent
    "unscored".
    """
    cats: Dict[str, CategoryScore] = {}
    for row in framework:
        score = firm_scores[row["short_label"]]
        scored = score is not None
        cats[row["key"]] = CategoryScore(
            category=row["key"],
            score=score,
            confidence="high" if scored else "low",
            evidence_status="sufficient" if scored else "absent",
            source="category_scores_xlsx",
            rationale=(f"Supplied via Category Scores xlsx "
                       f"('{row['short_label']}')." if scored else
                       f"Not yet scored ('{row['short_label']}' left blank)."),
            run_id=run_id,
        )
    return BidderScores(name=name, categories=cats)
