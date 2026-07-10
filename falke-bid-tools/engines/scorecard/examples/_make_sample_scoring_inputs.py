"""Generate the SYNTHETIC sample scoring-input example files.

Run once to (re)create them:
    python3 examples/_make_sample_scoring_inputs.py

Outputs:
  examples/sample_scoring_framework.xlsx — the 8-category Falke framework
    (identical to the shipped scoring-framework-template.xlsx rows: that IS
    Falke's current framework and the sample card's Section D).
  examples/sample_category_scores.xlsx — the synthetic sample card's Section-E
    1-10 scores for the 7 SCORED bidders, sourced from
    examples/sample_gold_overrides.json (the same values the sample tests use).

All firms and figures are fictional — no client data. These are the synthetic
validation EXAMPLE inputs for --scoring-framework / --category-scores (see
examples/sample_run.yaml), not skill defaults.
"""
from __future__ import annotations

import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(HERE), "templates"))

from _make_scoring_framework_template import (FRAMEWORK_ROWS,  # noqa: E402
                                              make_template as make_framework)

FRAMEWORK_OUT = os.path.join(HERE, "sample_scoring_framework.xlsx")
SCORES_OUT = os.path.join(HERE, "sample_category_scores.xlsx")
GOLD_OVERRIDES = os.path.join(HERE, "sample_gold_overrides.json")

# framework short label -> gold-overrides category key
LABEL_TO_KEY = {
    "Pricing": "pricing", "Scope": "scope", "Condo Exp": "condo_exp",
    "CO Risk": "co_risk", "Reputation": "reputation",
    "Financial": "financial", "Controls": "controls", "Docs": "docs",
}
# sample card firm order (best -> worst on the published card)
FIRM_ORDER = ["Acme", "Borealis", "Cascade", "Dorne", "Crest",
              "Fjord", "Granite"]


def make_scores(out_path: str = SCORES_OUT) -> None:
    with open(GOLD_OVERRIDES, "r", encoding="utf-8") as fh:
        gold = {k: v for k, v in json.load(fh).items()
                if not k.startswith("_")}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category_Scores"
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    labels = [r[1] for r in FRAMEWORK_ROWS]
    n_cols = 1 + len(labels)
    ws.cell(row=1, column=1,
            value="DETAILED CATEGORY SCORES (1-10) — Sample Condominium · "
                  "Lobby Renovation (sample-card Section E)").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    for col_idx, header in enumerate(["Firm"] + labels, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx > 1:
            cell.alignment = Alignment(horizontal="center")

    for i, firm in enumerate(FIRM_ORDER):
        ws.cell(row=3 + i, column=1, value=firm)
        for j, label in enumerate(labels, start=2):
            ws.cell(row=3 + i, column=j,
                    value=gold[firm][LABEL_TO_KEY[label]]["score"])

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)].width = 12
    ws.freeze_panes = "A3"
    wb.save(out_path)
    print(f"Example written: {out_path}")


if __name__ == "__main__":
    make_framework(FRAMEWORK_OUT)
    make_scores()
