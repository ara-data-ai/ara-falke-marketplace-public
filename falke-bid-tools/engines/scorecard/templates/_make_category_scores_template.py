"""Generate the category-scores-template.xlsx file.

Run once to (re)create the template:
    python3 templates/_make_category_scores_template.py

Output: templates/category-scores-template.xlsx
Pre-populated with generic placeholder bidder rows ('Bidder 1'…'Bidder 4') and
BLANK score cells — real bidder names differ every run, so Falke replaces the
placeholders with the actual scored bidders and fills every score cell (1–10).
The score column headers must match the Short Labels in the scoring-framework
file; the Overall /100 is COMPUTED by the engine and never supplied here.
"""
from __future__ import annotations

import os
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "category-scores-template.xlsx")

# one column per Short Label in scoring-framework-template.xlsx (same order)
SHORT_LABELS = ["Pricing", "Scope", "Condo Exp", "CO Risk",
                "Reputation", "Financial", "Controls", "Docs"]
PLACEHOLDER_FIRMS = ["Bidder 1", "Bidder 2", "Bidder 3", "Bidder 4"]

HEADER_NOTE = (
    "Scores are 1–10 (10 = best). Column headers must match the Short Labels "
    "in the scoring-framework file. One row per SCORED bidder in the matrix — "
    "replace the placeholder names with the real bidders (add/remove rows as "
    "needed; excluded bidders must NOT appear). The Overall /100 is computed "
    "by the engine — never add it here."
)


def make_template(out_path: str = OUT_PATH) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category_Scores"

    # ---- styles ----
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    n_cols = 1 + len(SHORT_LABELS)

    # ---- Row 1: title ----
    ws.cell(row=1, column=1,
            value="DETAILED CATEGORY SCORES (1–10) — Falke Bid-Comparison "
                  "Scorecard").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # ---- Row 2: column headers (Firm | one per Short Label) ----
    firm_header = ws.cell(row=2, column=1, value="Firm")
    firm_header.font = header_font
    firm_header.fill = header_fill
    firm_header.comment = Comment(HEADER_NOTE, "Falke Scorecard")
    for col_idx, label in enumerate(SHORT_LABELS, start=2):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ---- Rows 3+: placeholder bidder rows, score cells left BLANK ----
    for i, firm in enumerate(PLACEHOLDER_FIRMS):
        ws.cell(row=3 + i, column=1, value=firm)

    # ---- column widths ----
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)].width = 12

    # ---- freeze row 2 (header row) ----
    ws.freeze_panes = "A3"

    wb.save(out_path)
    print(f"Template written: {out_path}")


if __name__ == "__main__":
    make_template()
