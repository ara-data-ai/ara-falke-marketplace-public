"""Generate the baseline-template.xlsx file.

Run once to (re)create the template:
    python3 templates/_make_baseline_template.py

Output: templates/baseline-template.xlsx
Pre-populated with generic synthetic example values so Falke can see the
expected format and edit in place.
"""
from __future__ import annotations

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "baseline-template.xlsx")

# ---- example data (synthetic; illustrative round figures, not a real project) ----
BAND_LOW = 3.35
BAND_HIGH = 3.55
BAND_MID = 3.45

TRADE_LINES = [
    # (scope, basis, cost_str, value, kind)
    ("Interior demolition",               "Modeled allowance",    "$200,000",   200_000,  None),
    ("Flooring + slab prep",              "Modeled allowance",    "$590,000",   590_000,  None),
    ("Ceilings (framing/drywall/feature)","Modeled allowance",    "$320,000",   320_000,  None),
    ("Electrical + lighting",             "Modeled allowance",    "$400,000",   400_000,  None),
    ("Fire alarm relocation",             "Modeled allowance",    "$220,000",   220_000,  None),
    ("Plumbing / ADA restrooms",          "Modeled allowance",    "$240,000",   240_000,  None),
    ("Millwork / reception / veneer",     "Modeled allowance",    "$450,000",   450_000,  None),
    ("Glass / specialty elements",        "Modeled allowance",    "$180,000",   180_000,  None),
    ("General conditions (occupied high-rise)", "Modeled allowance", "$400,000", 400_000, None),
    ("Direct trades subtotal",            "Sum of trade lines",   "$3,000,000", 3_000_000, "subtotal"),
    ("GC OH&P (modeled) 12%",             "12% x direct subtotal","$360,000",   360_000,  "subtotal"),
]


def make_template(out_path: str = OUT_PATH) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline"

    # ---- styles ----
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    label_font = Font(bold=True)
    subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ---- Row 1: title ----
    ws.cell(row=1, column=1, value="COST BASELINE — Falke Bid-Comparison Scorecard").font = title_font
    ws.merge_cells("A1:E1")

    # ---- Row 2: blank ----

    # ---- Rows 3–5: band values ----
    band_labels = [
        (3, "Band Low ($M)",  BAND_LOW),
        (4, "Band High ($M)", BAND_HIGH),
        (5, "Band Mid ($M)",  BAND_MID),
    ]
    for row_num, label, value in band_labels:
        cell_label = ws.cell(row=row_num, column=1, value=label)
        cell_label.font = label_font
        cell_val = ws.cell(row=row_num, column=2, value=value)
        cell_val.number_format = "0.00"

    # ---- Row 6: blank separator ----

    # ---- Row 7: column headers ----
    col_headers = ["Scope", "Basis", "Cost ($)", "Value", "Kind"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ---- Rows 8+: trade line data ----
    for i, (scope, basis, cost_str, value, kind) in enumerate(TRADE_LINES):
        row_num = 8 + i
        ws.cell(row=row_num, column=1, value=scope)
        ws.cell(row=row_num, column=2, value=basis)
        ws.cell(row=row_num, column=3, value=cost_str)
        ws.cell(row=row_num, column=4, value=value)
        if kind:
            ws.cell(row=row_num, column=5, value=kind)
        # light fill for subtotal rows
        if kind == "subtotal":
            for col_idx in range(1, 6):
                ws.cell(row=row_num, column=col_idx).fill = subtotal_fill

    # ---- column widths ----
    col_widths = {"A": 40, "B": 30, "C": 14, "D": 12, "E": 10}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ---- freeze row 7 (header row) ----
    ws.freeze_panes = "A8"

    wb.save(out_path)
    print(f"Template written: {out_path}")


if __name__ == "__main__":
    make_template()
