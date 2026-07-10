"""Generate the scoring-framework-template.xlsx file.

Run once to (re)create the template:
    python3 templates/_make_scoring_framework_template.py

Output: templates/scoring-framework-template.xlsx
Pre-populated with Falke's CURRENT 8-category framework (the board scorecard's
Section D) as the sensible starting point — Falke edits categories/weights per
run. Weights must sum to 100; the Short Label column is what the
category-scores file's column headers must match.
"""
from __future__ import annotations

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "scoring-framework-template.xlsx")

# ---- Falke's current framework (board scorecard Section D) ----
# (category, short_label, weight_pct, what it captures)
FRAMEWORK_ROWS = [
    ("Market-aligned pricing", "Pricing", 25,
     "Closeness to takeoff baseline and South Florida $/SF realism (reduces "
     "CO drift probability)."),
    ("Scope completeness / clarity", "Scope", 15,
     "Quality of inclusions/exclusions/allowances; fewer silent omissions "
     "score higher."),
    ("Condo-specific execution experience", "Condo Exp", 15,
     "Occupied high-rise/common-area execution (phasing, protection, "
     "logistics, resident sensitivity)."),
    ("Change order exposure risk", "CO Risk", 15,
     "Likelihood of drift based on under-baseline pricing, allowance "
     "structure, and coordination triggers."),
    ("Reputation & longevity", "Reputation", 10,
     "Public footprint and credibility signals (track record, references)."),
    ("Financial strength / stability", "Financial", 10,
     "Implied resilience (cash flow, ability to carry subs) given price "
     "posture and scale."),
    ("Project controls & infrastructure", "Controls", 5,
     "RFI/submittal/change control discipline; tools and reporting cadence."),
    ("Documentation quality / professionalism", "Docs", 5,
     "Completeness/accuracy of forms; clarity and consistency of submission."),
]


def make_template(out_path: str = OUT_PATH) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoring_Framework"

    # ---- styles ----
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    # ---- Row 1: title ----
    ws.cell(row=1, column=1,
            value="SCORING FRAMEWORK — Falke Bid-Comparison Scorecard "
                  "(weights must sum to 100)").font = title_font
    ws.merge_cells("A1:D1")

    # ---- Row 2: column headers ----
    for col_idx, header in enumerate(
            ["Category", "Short Label", "Weight (%)", "What it captures"],
            start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ---- Rows 3+: framework rows ----
    for i, (category, short_label, weight, captures) in enumerate(FRAMEWORK_ROWS):
        row_num = 3 + i
        ws.cell(row=row_num, column=1, value=category)
        ws.cell(row=row_num, column=2, value=short_label)
        ws.cell(row=row_num, column=3, value=weight).number_format = "0.##"
        cap = ws.cell(row=row_num, column=4, value=captures)
        cap.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- column widths ----
    for col_letter, width in {"A": 38, "B": 14, "C": 12, "D": 72}.items():
        ws.column_dimensions[col_letter].width = width

    # ---- freeze row 2 (header row) ----
    ws.freeze_panes = "A3"

    wb.save(out_path)
    print(f"Template written: {out_path}")


if __name__ == "__main__":
    make_template()
