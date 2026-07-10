"""Baseline-confirmation gate — CLI process-level contract.

The modeled cost baseline is the yardstick the whole scorecard hangs off, and it
can show signs of being bid-derived (the baseline-anchoring fingerprint). So the
skill REQUIRES the baseline be PREVIEWED and explicitly CONFIRMED before any
scorecard is built — mirroring how --sf-basis is required. This file locks that
contract end-to-end through cli.main():

  * --preview-baseline  -> echoes the baseline (band in $M AND $/SF) + runs the
    bid-anchoring fingerprint check, EXITS 0, renders NOTHING.
  * a render run WITHOUT --baseline-confirmed -> [STOP], exit 2, NO artifacts.
  * a render run WITH  --baseline-confirmed   -> renders normally, exit 0.
  * the preview fingerprint section surfaces a known bid-anchoring HIT when the
    supplied baseline collides with a bidder subtotal.

The client xlsx is gitignored / absent in builder worktrees, so these
tests BUILD a small synthetic matrix (same structural format as the real one) and
plant a baseline line that collides with a bidder subtotal to force the hit —
self-contained, no client data required.
"""
from __future__ import annotations

import json
import os

import openpyxl

from scorecard.cli import main

from .conftest import SIMPLE_FRAMEWORK_ROWS, write_framework_xlsx, write_scores_xlsx

PROJECT = "PREVIEW TEST · Synthetic"
SF_BASIS = 10000          # band_low_per_sf = 2.00e6/1e4 = 200, high = 220
BAND_LOW = 2.00
BAND_HIGH = 2.20
MID = 2.10

# One bidder whose FINISHES division subtotal (612,300) collides with the
# planted baseline "Flooring" line (612,000) inside the 0.2% tolerance ->
# reproduces a bid-anchoring fingerprint signature (synthetic firms/figures).
SYNTH_BIDDERS = [
    # (name, grand_total, [div subtotals: GenConds, Concrete, Finishes])
    ("Crestline Builders LLC", 2_100_000, [300_000, 250_000, 612_300]),
    ("Borealis Group",       1_900_000, [240_000, 210_000, 150_000]),
]


def _build_synthetic_matrix(path: str) -> None:
    """Minimal 2-bidder matrix in the real structural format (mirrors
    tests/test_generic_matrix.py): names every 5 cols, the COST quartet a few
    rows below, three division SUBTOTAL rows in the COST SUBTOTALS column, a
    pre-markup construction subtotal, a grand total, and a planted GSF."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"

    NAME_ROW, QUARTET_ROW = 4, 8
    DIV_ROWS = [11, 12, 13]
    CC_SUBTOTAL_ROW, GT_ROW, GSF_ROW = 15, 18, 20
    FIRST_COL, STRIDE = 4, 5

    ws.cell(row=1, column=1, value="SYNTHETIC — Baseline-gate test matrix")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUBTOTAL_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    for r, lab in zip(DIV_ROWS, ["GENERAL CONDITIONS SUBTOTAL",
                                 "CONCRETE SUBTOTAL", "FINISHES SUBTOTAL"]):
        ws.cell(row=r, column=2, value=lab)

    for i, (name, total, divs) in enumerate(SYNTH_BIDDERS):
        c_cost = FIRST_COL + i * STRIDE
        ws.cell(row=NAME_ROW, column=c_cost, value=name)
        ws.cell(row=QUARTET_ROW, column=c_cost, value="COST")
        ws.cell(row=QUARTET_ROW, column=c_cost + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c_cost + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c_cost + 3, value="$/SF SUBTOTALS")
        ws.cell(row=GT_ROW, column=c_cost, value=total)
        for r, dv in zip(DIV_ROWS, divs):
            ws.cell(row=r, column=c_cost + 1, value=dv)
        ws.cell(row=CC_SUBTOTAL_ROW, column=c_cost, value=total * 0.85)
        ws.cell(row=GSF_ROW, column=c_cost, value=8500)

    wb.save(path)


def _baseline_json(path: str) -> None:
    """Baseline with a 'Flooring' line at 612,000 — collides with Crestline's
    612,300 Finishes subtotal (0.05%) -> a fingerprint HIT."""
    lines = [
        {"scope": "Interior demolition", "basis": "Modeled allowance",
         "cost": "$200,000", "value": 200_000},
        {"scope": "Flooring", "basis": "Modeled allowance",
         "cost": "$612,000", "value": 612_000},
        {"scope": "Electrical + lighting", "basis": "Modeled allowance",
         "cost": "$400,000", "value": 400_000},
        {"scope": "Direct trades subtotal", "basis": "Sum of trade lines",
         "cost": "$1,194,000", "value": 1_194_000, "kind": "subtotal"},
        {"scope": "GC OH&P (modeled) 12%", "basis": "12% x direct subtotal",
         "cost": "$143,280", "value": 143_280, "kind": "subtotal"},
    ]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(lines, fh)


def _common_argv(matrix, baseline, tmp_path):
    # includes the two REQUIRED per-run scoring xlsx inputs (harmlessly ignored
    # by --preview-baseline, which is exempt from the scoring-inputs gate)
    fw = write_framework_xlsx(str(tmp_path / "framework.xlsx"),
                              SIMPLE_FRAMEWORK_ROWS)
    labels = [r[1] for r in SIMPLE_FRAMEWORK_ROWS]
    # distinct pricing scores: an exact Overall tie with the pricier bidder
    # listed first would trip audit C5 (ranking tiebreak) at full coverage
    cs = write_scores_xlsx(str(tmp_path / "scores.xlsx"), labels,
                           [(n, [8 - i, 7, 7])
                            for i, (n, _, _) in enumerate(SYNTH_BIDDERS)])
    return [
        "--matrix", matrix, "--project-name", PROJECT,
        "--sf-basis", str(SF_BASIS), "--band-low", str(BAND_LOW),
        "--band-high", str(BAND_HIGH), "--mid", str(MID),
        "--baseline", baseline,
        "--scoring-framework", fw, "--category-scores", cs,
    ]


def _setup(tmp_path):
    matrix = str(tmp_path / "synthetic_matrix.xlsx")
    baseline = str(tmp_path / "baseline.json")
    _build_synthetic_matrix(matrix)
    _baseline_json(baseline)
    return matrix, baseline


def _no_artifacts(out_dir):
    for fn in ("scorecard.pdf", "scorecard.html", "scorecard_run.json"):
        assert not os.path.exists(os.path.join(out_dir, fn)), fn


# ---------------------------------------------------------------------------
def test_preview_baseline_exits_zero_echoes_band_and_renders_nothing(
        tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_common_argv(matrix, baseline, tmp_path) + ["--preview-baseline",
                                                "--out-dir", out])
    assert rc == 0
    s = capsys.readouterr().out
    # band echoed in BOTH $M and $/SF
    assert "$2.00M–$2.20M" in s
    assert "mid $2.10M" in s
    assert "$200–$220/SF" in s        # 2.00e6/1e4 .. 2.20e6/1e4
    assert "mid $210/SF" in s
    # trade-scope lines + the fingerprint SECTION are present
    assert "Trade-scope lines:" in s
    assert "Flooring" in s
    assert "Bid-anchoring fingerprint check:" in s
    # NOTHING rendered
    assert "No scorecard rendered" in s
    _no_artifacts(out)


def test_render_without_baseline_confirmed_stops_exit2_no_artifacts(
        tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_common_argv(matrix, baseline, tmp_path) + ["--out-dir", out])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[STOP] Baseline not confirmed" in err
    assert "--preview-baseline" in err
    assert "--baseline-confirmed" in err
    _no_artifacts(out)


def test_render_with_baseline_confirmed_renders_exit0(tmp_path):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_common_argv(matrix, baseline, tmp_path) + [
        "--baseline-confirmed", "--html-only", "--no-audit", "--out-dir", out])
    assert rc == 0
    assert os.path.exists(os.path.join(out, "scorecard.html"))
    assert os.path.exists(os.path.join(out, "scorecard_run.json"))


def test_preview_surfaces_bid_anchoring_fingerprint_hit(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    rc = main(_common_argv(matrix, baseline, tmp_path) + ["--preview-baseline"])
    assert rc == 0
    s = capsys.readouterr().out
    # the planted Flooring 612,000 ~ Crestline 612,300 hit must surface, with the
    # "may be bid-derived" warning and the bidder named.
    assert "⚠ Baseline line 'Flooring'" in s
    assert "Crestline" in s
    assert "may be bid-derived" in s
    assert "No bid-anchoring fingerprints detected." not in s
