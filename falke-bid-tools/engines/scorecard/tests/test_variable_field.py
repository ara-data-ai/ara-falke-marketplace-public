"""Flexibility / genericity harness — VARIABLE bidders AND variable line items.

The matrix always varies in (a) the number of bidders and (b) the number of CSI
division subtotal rows. test_generic_matrix.py already proves variable bidders
(4-bidder HARBORVIEW); this file HARDENS the claim across the full range the
brief calls out, with synthetic matrices that differ from the validation set in BOTH axes:

  * 2 bidders (minimum scored field);
  * 11 and 12 bidders (MORE than the validation set's 10/7);
  * MORE CSI divisions than the validation set (25);
  * FEWER / sparse divisions (3; and a block that populates only some of them);
  * division LABELS not in the validation set (custom restoration-scope names);
  * a bidder block with EXTRA / MISSING line items vs its peers.

For each it asserts block detection, grand-total detection, $/SF (on the SUPPLIED
basis), tiering vs the SUPPLIED band, contiguous ranking over ALL bidders, and
that the deterministic self-audit RUNS to a verdict — all WITHOUT any hardcoded
bidder count or division set. The engine measures count/width/stride and detects
divisions by the '...SUBTOTAL' predicate, so nothing here is single-project-shaped.

NOTE on C8's DIV_BAND (16-20): that is a WARN-only PLAUSIBILITY heuristic and the
documented C8 contract (see test_audit.py) — it is NOT a bidder-count or
division-SET assumption and never changes any number, tier, or rank. On a matrix
with far fewer/more divisions C8 legitimately raises a WARN disclosure; that is
correct behavior, so these tests assert the audit reaches a VERDICT (no BLOCKER
from genericity), not that it is WARN-free.
"""
from __future__ import annotations

from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl

from scorecard import audit as A
from scorecard.config import load_config
from scorecard.matrix import MatrixParser
from scorecard.pipeline import run_scorecard
from scorecard.render import build_context

SF_BASIS = 10_000          # band_low_per_sf = 200, band_high_per_sf = 220
BAND_LOW, BAND_HIGH, MID = 2.00, 2.20, 2.10
FIRST_COL, STRIDE = 4, 5


def _build(path: str, bidders: Sequence[Tuple[str, float]],
           div_labels: Sequence[str],
           div_values: Optional[Dict[str, List[Optional[float]]]] = None,
           gsf: float = 8_500) -> None:
    """Write a matrix with an ARBITRARY number of bidders and division rows.

    bidders     : [(name, grand_total), ...] — any length >= 1.
    div_labels  : the CSI division-SUBTOTAL row labels — any length, any names.
    div_values  : optional {name: [v0, v1, ...]} per-division values; None in a
                  slot leaves that division blank for that block (missing line
                  item). Defaults to every block populating every division.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"
    NAME_ROW, QUARTET_ROW = 4, 8
    DIV_ROWS = list(range(11, 11 + len(div_labels)))
    CC_SUB_ROW = DIV_ROWS[-1] + 2 if DIV_ROWS else 13
    GT_ROW = CC_SUB_ROW + 2
    GSF_ROW = GT_ROW + 2

    ws.cell(row=1, column=1, value="SYNTHETIC VARIABLE-FIELD matrix")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUB_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    ws.cell(row=GSF_ROW, column=3, value=gsf)
    for r, lab in zip(DIV_ROWS, div_labels):
        ws.cell(row=r, column=2, value=lab)

    default_vals = [10_000 + j * 100 for j in range(len(div_labels))]
    for i, (name, total) in enumerate(bidders):
        c = FIRST_COL + i * STRIDE
        ws.cell(row=NAME_ROW, column=c, value=name)
        ws.cell(row=QUARTET_ROW, column=c, value="COST")
        ws.cell(row=QUARTET_ROW, column=c + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c + 3, value="$/SF SUBTOTALS")
        ws.cell(row=GT_ROW, column=c, value=total)
        vals = (div_values or {}).get(name, default_vals)
        for r, v in zip(DIV_ROWS, vals):
            if v is not None:
                ws.cell(row=r, column=c + 1, value=v)
        ws.cell(row=CC_SUB_ROW, column=c, value=total * 0.85)
    wb.save(path)


def _cfg():
    return load_config(overrides={
        "sf_basis": SF_BASIS, "band_low": BAND_LOW, "band_high": BAND_HIGH,
        "modeled_mid_takeoff": MID})


def _eighteen_labels(n: int) -> List[str]:
    """n generic CSI-style division-subtotal labels (names not from the validation set)."""
    return [f"DIVISION {i + 1:02d} SUBTOTAL" for i in range(n)]


def _assert_field_correct(xlsx: str, bidders, *, n_divs: int):
    """Shared assertions: detection, $/SF, tiers, ranking, audit-runs — all
    data-driven, NO hardcoded counts."""
    cfg = _cfg()
    parser = MatrixParser(cfg.block("matrix"))
    parsed = parser.parse(xlsx)

    # block detection counts the field it MEASURES (no 10/7 assumption)
    assert len(parsed.included_blocks) == len(bidders), \
        [b.raw_name for b in parsed.blocks]
    # grand-total row located by label, each total read from the COST column
    assert parsed.grand_total_label.upper().startswith("GRAND TOTAL")
    by_block = {b.name: b for b in parsed.included_blocks}
    for name, total in bidders:
        assert by_block[name].grand_total == total, name
    # division detection counts the rows present (no fixed division SET)
    assert len(parsed.division_rows) == n_divs

    result = run_scorecard(xlsx, cfg, project_name="VARIABLE FIELD · Test")
    out = {b["name"]: b for b in result["bidders"]}
    assert set(out) == {n for n, _ in bidders}
    # $/SF on the SUPPLIED basis (not the planted GSF)
    for name, total in bidders:
        assert out[name]["per_sf"] == round(total / SF_BASIS), name
    # every bidder gets a tier from the band rule
    valid = {"TOP", "MID", "DEFENSIVE", "PREMIUM", "RISK"}
    assert all(out[n]["tier"] in valid for n, _ in bidders)
    # ranking is contiguous 1..N over the WHOLE field
    ranks = sorted(b["rank"] for b in result["bidders"])
    assert ranks == list(range(1, len(bidders) + 1))
    assert len(result["ranking"]) == len(bidders)
    # render context threads the WHOLE field through (one Section-B row/bidder)
    ctx = build_context(result, cfg)
    assert len(ctx["bid_rows"]) == len(bidders)
    # the deterministic audit RUNS to a verdict with no genericity BLOCKER
    ar = A.audit(parsed, cfg, result)
    assert ar.verdict in (A.V_PASS, A.V_WARN, A.V_FAIL)
    # C1/C2/C5 (totals, $/SF, ranking) must not blocker-fail on a clean field
    for nm in ("C1", "C2", "C5", "C10"):
        c = next(ck for ck in ar.checks if ck.name == nm)
        assert c.status == A.PASS, c.verdict_line
    return result


# ---- 2 bidders (minimum) -----------------------------------------------------
def test_two_bidders(tmp_path):
    xlsx = str(tmp_path / "two.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_900_000)]
    _build(xlsx, bidders, _eighteen_labels(18))
    _assert_field_correct(xlsx, bidders, n_divs=18)


# ---- 11 and 12 bidders (more than the validation set's 10) -------------------
def test_eleven_bidders(tmp_path):
    xlsx = str(tmp_path / "eleven.xlsx")
    bidders = [(f"Bidder {i:02d} Co", 1_700_000 + i * 60_000) for i in range(11)]
    _build(xlsx, bidders, _eighteen_labels(18))
    _assert_field_correct(xlsx, bidders, n_divs=18)


def test_twelve_bidders(tmp_path):
    xlsx = str(tmp_path / "twelve.xlsx")
    bidders = [(f"Firm {chr(65 + i)} LLC", 1_650_000 + i * 55_000) for i in range(12)]
    _build(xlsx, bidders, _eighteen_labels(18))
    _assert_field_correct(xlsx, bidders, n_divs=18)


# ---- MORE divisions than the validation set (25) -----------------------------
def test_more_divisions_than_validation_set(tmp_path):
    xlsx = str(tmp_path / "manydiv.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_900_000),
               ("Cascade Co", 2_500_000)]
    _build(xlsx, bidders, _eighteen_labels(25))
    _assert_field_correct(xlsx, bidders, n_divs=25)


# ---- FEWER / sparse divisions (3) --------------------------------------------
def test_fewer_divisions(tmp_path):
    xlsx = str(tmp_path / "fewdiv.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_900_000)]
    _build(xlsx, bidders, _eighteen_labels(3))
    result = _assert_field_correct(xlsx, bidders, n_divs=3)
    # a low-division field legitimately raises the C8 WARN disclosure (correct,
    # not a genericity failure): it must be a WARN, never a BLOCKER.
    cfg = _cfg()
    ar = A.audit(result["parsed"], cfg, result)
    c8 = next(c for c in ar.checks if c.name == "C8")
    assert c8.severity == A.WARN              # WARN-only, can't blocker the run


def test_sparse_divisions_some_blocks_missing(tmp_path):
    """A block populates only SOME division rows (sparse) — still detected,
    scored, and ranked; the false-zero BLOCKER must NOT fire (it populates >0)."""
    xlsx = str(tmp_path / "sparse.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_900_000),
               ("Cascade Co", 2_500_000)]
    labels = _eighteen_labels(10)
    # Cascade leaves the last 6 divisions blank (sparse); peers populate all 10.
    div_values = {
        "Cascade Co": [10_000 + j * 100 if j < 4 else None for j in range(10)],
    }
    _build(xlsx, bidders, labels, div_values=div_values)
    cfg = _cfg()
    parsed = MatrixParser(cfg.block("matrix")).parse(xlsx)
    by = {b.name: b for b in parsed.included_blocks}
    assert by["Cascade Co"].populated_divisions == 4
    assert by["Apex Builders"].populated_divisions == 10
    result = run_scorecard(xlsx, cfg, project_name="SPARSE · Test")
    assert {b["name"] for b in result["bidders"]} == {n for n, _ in bidders}
    ar = A.audit(parsed, cfg, result)
    c8 = next(c for c in ar.checks if c.name == "C8")
    # sparse Cascade (4 divisions, peer median 10) is flagged, never auto-dropped
    assert c8.severity == A.WARN
    assert all(b.included for b in parsed.included_blocks)


# ---- division LABELS not in the validation set -------------------------------
def test_non_validation_set_division_labels(tmp_path):
    xlsx = str(tmp_path / "customlabels.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_900_000)]
    # restoration-scope labels that never appear in the validation set's CSI set
    labels = [
        "BALCONY SPALL REPAIR SUBTOTAL",
        "POST-TENSION CABLE REMEDIATION SUBTOTAL",
        "WINDOW & DOOR REPLACEMENT SUBTOTAL",
        "WATERPROOFING & COATINGS SUBTOTAL",
        "STRUCTURAL STEEL REINFORCEMENT SUBTOTAL",
    ]
    _build(xlsx, bidders, labels)
    cfg = _cfg()
    parsed = MatrixParser(cfg.block("matrix")).parse(xlsx)
    # divisions detected purely by the '...SUBTOTAL' predicate, NOT a name list
    assert len(parsed.division_rows) == 5
    detected = {lab.upper() for _, lab in parsed.division_rows}
    assert "BALCONY SPALL REPAIR SUBTOTAL" in detected
    _assert_field_correct(xlsx, bidders, n_divs=5)


# ---- a bidder block with EXTRA / MISSING line items vs peers ------------------
def test_extra_and_missing_line_items(tmp_path):
    """One block populates an extra division a peer leaves blank, and another
    block misses a division its peers populate — the per-block populated count is
    measured independently; nothing assumes a uniform line-item count."""
    xlsx = str(tmp_path / "ragged.xlsx")
    bidders = [("Apex Builders", 2_100_000), ("Borealis Group", 1_950_000),
               ("Cascade Co", 2_050_000)]
    labels = _eighteen_labels(12)
    div_values = {
        # Apex: all 12 (the 'extra' relative to Borealis)
        "Apex Builders": [10_000 + j * 100 for j in range(12)],
        # Borealis: misses divisions 10-11 (only 10 populated)
        "Borealis Group": [10_000 + j * 100 if j < 10 else None for j in range(12)],
        # Cascade: all 12
        "Cascade Co": [10_000 + j * 100 for j in range(12)],
    }
    _build(xlsx, bidders, labels, div_values=div_values)
    cfg = _cfg()
    parsed = MatrixParser(cfg.block("matrix")).parse(xlsx)
    by = {b.name: b for b in parsed.included_blocks}
    assert by["Apex Builders"].populated_divisions == 12
    assert by["Borealis Group"].populated_divisions == 10
    assert by["Cascade Co"].populated_divisions == 12
    # all three still scored + ranked (ragged line items never drop a bidder)
    _assert_field_correct(xlsx, bidders, n_divs=12)
