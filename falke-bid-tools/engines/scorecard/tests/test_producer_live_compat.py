"""LIVE cross-engine compat gate (P0-2 — Floyd consolidated ruling, C-R1 ≡
Boris E3, merged): generate a workbook with the IN-TREE matrix engine (same
plugin, same version), parse it with THIS scorecard, and tie the numbers out
cell-vs-parsed against the workbook's own machine-key footer.

This is the test that would have failed the v0.4.0 release before it shipped:
the committed producer fixtures froze a producer that no longer existed, so
the suite stayed green while production broke ("fixture-based contract testing
without a regeneration trigger is a contract with a dead counterparty").
Because it is pytest-shaped, release.sh gates [2/8] and [6/8] run it for free.

Fully synthetic bids (fictional firms/dollars — Floyd fixture rule); the
workbooks are generated into a tmp dir per session, never committed. The
committed tests/fixtures/create_matrix_{2,4}bidders.xlsx remain as v0.3-era
BACK-COMPAT pins (test_create_matrix_compat.py); this module covers the
producer that exists NOW.
"""
from __future__ import annotations

import importlib.util
import json
import os
import shutil
import sys

import openpyxl
import pytest

from scorecard.cli import main as cli_main
from scorecard.config import load_config
from scorecard.errors import MatrixStructureError, ProducerVersionError
from scorecard.matrix import (LEVELED_SHEET, MIRROR_SHEET, MatrixParser,
                              read_producer_stamp)
from scorecard.pipeline import audit_run, run_scorecard

HERE = os.path.dirname(os.path.abspath(__file__))
GENERATOR = os.path.join(HERE, "fixtures", "_make_create_matrix_fixtures.py")
MATRIX_ENGINE = os.path.abspath(os.path.join(HERE, "..", "..", "matrix"))

pytestmark = pytest.mark.skipif(
    not os.path.isdir(os.path.join(MATRIX_ENGINE, "src")),
    reason="in-tree matrix engine not present (engines/matrix)")

# synthetic ground truth (mirrors the generator's seeds):
# 4 priced CSI divisions, base dollars scaled by (1 + 0.08*i) per bidder.
_DIV_BASES = {
    "DIV 01 00 00": 250_000.0,
    "DIV 03 00 00": 400_000.0,
    "DIV 07 00 00": 300_000.0,
    "DIV 09 00 00": 150_000.0,
}
_FIRMS = ["Alpine Restoration Group", "Bayside Builders LLC",
          "Cypress Construction Co.", "Driftwood Contractors"]


def _expected_totals(n):
    return {
        _FIRMS[i]: round(sum(_DIV_BASES.values()) * (1 + 0.08 * i))
        for i in range(n)
    }


@pytest.fixture(scope="session")
def generated(tmp_path_factory):
    """Generate 2- and 4-bidder workbooks with the CURRENT producer, once."""
    out = tmp_path_factory.mktemp("live_producer")
    spec = importlib.util.spec_from_file_location("_mk_live", GENERATOR)
    mk = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mk)
    return {n: str(mk.build(n, out)) for n in (2, 4)}


def _cfg():
    return load_config(overrides={
        "sf_basis": 10_000, "band_low": 1.00, "band_high": 1.20,
        "modeled_mid_takeoff": 1.10})


def _machine_key_footer(path, sheet=MIRROR_SHEET):
    """INDEPENDENT read of the mirror's machine-key GRAND_TOTAL footer row:
    row located by col-A key, per-bidder values by the bidder name columns
    (names row 5, COST-SUBTOTALS column of each 2-col group)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    gt_row = next(i for i, r in enumerate(rows)
                  if r and str(r[0] or "").strip() == "GRAND_TOTAL")
    names_row = next(r for r in rows if r and any(f in r for f in _FIRMS))
    out = {}
    for c, v in enumerate(names_row):
        if v in _FIRMS:
            out[v] = rows[gt_row][c]
    return out


# ---------------------------------------------------------------------------
# THE gate: current producer output -> default parse -> machine-key tie-out
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("n", [2, 4])
def test_default_path_grand_totals_tie_out_to_machine_key_footer(generated, n):
    parsed = MatrixParser(_cfg().block("matrix")).parse(generated[n])

    # the DEFAULT sheet is the ruled leveled view — never an ordering accident
    assert parsed.sheet_name == LEVELED_SHEET
    assert parsed.sheet_mode == "leveled"

    expected = _expected_totals(n)
    got = {b.raw_name: b.grand_total for b in parsed.included_blocks}
    assert got == expected, got

    # cell-vs-parsed: the mirror's machine-key footer must carry the same
    # values (independent openpyxl read — no parser code involved)
    footer = _machine_key_footer(generated[n])
    assert footer == expected, footer


@pytest.mark.parametrize("n", [2, 4])
def test_division_subtotals_tie_out_cell_vs_parsed(generated, n):
    """Division-level tie-out on the DEFAULT (leveled) sheet: exactly the 20
    CSI division-subtotal rows are detected (the footer 'Fees Subtotal' is a
    markup row, never a 21st division — P0-3 ride-along), each bidder
    populates exactly the 4 priced divisions, and the priced cells re-read
    from the workbook equal the synthetic seeds."""
    parsed = MatrixParser(_cfg().block("matrix")).parse(generated[n])
    assert len(parsed.division_rows) == 20, \
        [lab for _r, lab in parsed.division_rows]
    assert not any("fees" in lab.lower() for _r, lab in parsed.division_rows)

    priced_short = {"general requirements", "concrete",
                    "thermal & moisture protection", "finishes"}
    for i, b in enumerate(parsed.included_blocks):
        assert b.populated_divisions == 4, (b.raw_name, b.populated_divisions)
        factor = 1 + 0.08 * i
        sub_col = b.cols.get("cost_subtotal") or b.cols.get("cost")
        seen = {}
        for (r, lab) in parsed.division_rows:
            v = parsed.grid.cell(r, sub_col)
            if isinstance(v, (int, float)) and v != 0:
                seen[lab.lower().replace(" subtotal", "")] = float(v)
        expected_cells = {
            name.lower(): round(base * factor)
            for code, base in _DIV_BASES.items()
            for name in [{"DIV 01 00 00": "General Requirements",
                          "DIV 03 00 00": "Concrete",
                          "DIV 07 00 00": "Thermal & Moisture Protection",
                          "DIV 09 00 00": "Finishes"}[code]]
        }
        assert set(seen) == set(expected_cells), (b.raw_name, seen)
        for k, v in expected_cells.items():
            assert abs(seen[k] - v) < 1.0, (b.raw_name, k, seen[k], v)


def test_mirror_path_reads_grand_totals_via_machine_key(generated):
    """P0-3 regression: the as-submitted mirror (explicit --sheet) reads its
    grand totals via the col-A GRAND_TOTAL machine key — the v0.4.0 unified
    legend can never be selected again (its prose row carries no bidder-column
    numerics and 'subtotal' has no word-boundary 'total')."""
    mc = dict(_cfg().block("matrix"))
    mc["sheet_name"] = MIRROR_SHEET
    parsed = MatrixParser(mc).parse(generated[4])
    assert parsed.sheet_mode == "mirror"
    assert parsed.grand_total_label.strip().upper() == "GRAND TOTAL"
    got = {b.raw_name: b.grand_total for b in parsed.included_blocks}
    assert got == _expected_totals(4)
    # the chosen row must be ABOVE the legend block (which contains the
    # 'subtotal rows is house formatting' prose line that broke v0.4.0)
    legend_rows = [r for r in range(1, parsed.grid.max_row + 1)
                   if isinstance(parsed.grid.cell(r, 3), str)
                   and "house formatting" in parsed.grid.cell(r, 3).lower()]
    assert legend_rows, "expected the v0.4.0 unified legend on the mirror"
    assert parsed.grand_total_row < min(legend_rows)


def test_producer_stamp_written_and_in_range(generated):
    wb = openpyxl.load_workbook(generated[4], read_only=True)
    stamp = read_producer_stamp(wb)
    assert stamp["producer"] == "falke-bid-tools/matrix"
    assert stamp["format_version"] == "0.4.0"
    parsed = MatrixParser(_cfg().block("matrix")).parse(generated[4])
    assert parsed.producer_stamp == stamp
    assert any("inside supported range" in l for l in parsed.log)


def test_producer_stamps_run_and_project_identity(generated):
    """P1-4 §10.1: the stamp gained a run identity (the pack's binding rested on
    a field that did not exist) and the project identity the pack's I3 check
    re-derives against. Both are additive doc properties — no visible geometry,
    and an older consumer that ignores them still parses the workbook, which is
    why they do not move PRODUCER_FORMAT_VERSION."""
    stamp = read_producer_stamp(
        openpyxl.load_workbook(generated[4], read_only=True))
    assert stamp["run_id"] and len(stamp["run_id"]) == 12
    assert stamp["project_name"] == "Harborview Synthetic Test Tower"
    assert stamp["project_address"] == "1 Test Quay, Test City FL 00000"
    assert stamp["sf_basis_label"] == "GSF"

    # opaque + collision-free: two runs never share one
    other = read_producer_stamp(
        openpyxl.load_workbook(generated[2], read_only=True))
    assert other["run_id"] != stamp["run_id"]


def test_out_of_range_producer_stamp_hard_stops(generated, tmp_path):
    """A NEWER-format stamp than SUPPORTED_PRODUCER refuses to parse."""
    path = str(tmp_path / "future.xlsx")
    shutil.copy(generated[2], path)
    wb = openpyxl.load_workbook(path)
    for p in wb.custom_doc_props.props:
        if p.name == "falke_bid_tools.format_version":
            p.value = "9.9"
    wb.save(path)
    with pytest.raises(ProducerVersionError):
        MatrixParser(_cfg().block("matrix")).parse(path)


def test_missing_leveled_sheet_on_producer_workbook_hard_stops(generated,
                                                               tmp_path):
    """Marvin P0-7 hard rule 2: a producer workbook without the default
    leveled sheet exits with a hard stop NAMING the expected sheet — never a
    silent first-sheet fallback."""
    path = str(tmp_path / "no_leveled.xlsx")
    shutil.copy(generated[2], path)
    wb = openpyxl.load_workbook(path)
    del wb[LEVELED_SHEET]
    wb.save(path)
    with pytest.raises(MatrixStructureError) as ei:
        MatrixParser(_cfg().block("matrix")).parse(path)
    assert LEVELED_SHEET in str(ei.value)


def test_quarantine_banner_blocks_via_audit(generated, tmp_path):
    """Marvin P0-7 hard rule 5: a workbook carrying the producer's Stage-6b
    RED quarantine banner is flagged at parse and BLOCKED by audit C17."""
    path = str(tmp_path / "quarantined.xlsx")
    shutil.copy(generated[2], path)
    wb = openpyxl.load_workbook(path)
    for sheet in (MIRROR_SHEET, LEVELED_SHEET):
        wb[sheet].cell(row=1, column=1).value = (
            "⚠ AUTOMATED CHECK FAILED — DO NOT RELY ON THE FLAGGED FIGURES "
            "FOR AN AWARD DECISION.")
    wb.save(path)

    cfg = _cfg()
    result = run_scorecard(path, cfg, project_name="Quarantine Probe")
    assert result["parsed"].quarantine_flag is True
    ar, _paths = audit_run(result, cfg, str(tmp_path / "out"))
    c17 = next(c for c in ar.checks if c.name == "C17")
    assert c17.status == "fail" and c17.severity == "BLOCKER"
    assert ar.verdict == "FAIL"


def test_cross_sheet_gt_mismatch_is_audit_blocker(generated, tmp_path):
    """Marvin P0-7 hard rule 4: mirror-GT == leveled-GT is a producer
    invariant; the audit's independent cross-sheet re-read (C18) BLOCKS when
    it is broken."""
    path = str(tmp_path / "tampered.xlsx")
    shutil.copy(generated[2], path)
    wb = openpyxl.load_workbook(path)
    ws = wb[MIRROR_SHEET]
    gt_row = next(r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(row=r, column=1).value or "").strip()
                  == "GRAND_TOTAL")
    # bump the first bidder's mirror GT (col C = first group's COST SUBTOTALS)
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=gt_row, column=c).value
        if isinstance(v, (int, float)):
            ws.cell(row=gt_row, column=c).value = v + 12_345
            break
    wb.save(path)

    cfg = _cfg()
    result = run_scorecard(path, cfg, project_name="Tamper Probe")
    ar, _paths = audit_run(result, cfg, str(tmp_path / "out"))
    c18 = next(c for c in ar.checks if c.name == "C18")
    assert c18.status == "fail" and c18.severity == "BLOCKER"
    assert ar.verdict == "FAIL"


def test_full_pipeline_and_audit_clean_on_live_producer_output(generated,
                                                               tmp_path):
    """End-to-end on the CURRENT producer's output, default path: correct
    totals/$/SF, ranking, the leveled disclosure recorded, and a self-audit
    with no blockers (C1/C2/C14/C17/C18 all pass live)."""
    cfg = _cfg()
    result = run_scorecard(generated[4], cfg,
                           project_name="Harborview Synthetic Test Tower")
    expected = _expected_totals(4)
    by_name = {b["name"]: b for b in result["bidders"]}
    for name, total in expected.items():
        assert by_name[name]["total"] == total
        assert by_name[name]["per_sf"] == round(total / 10_000)
    # This run supplies no scoring xlsx, so coverage is partial and — under
    # Marvin's P1-2 ruling — the field is listed alphabetically and carries NO
    # rank. The mechanical facts asserted above are unaffected.
    assert result["full_coverage"] is False
    assert all("rank" not in b for b in result["bidders"])

    assert result["sheet"]["name"] == LEVELED_SHEET
    assert result["sheet"]["mode"] == "leveled"
    assert "Leveled/Normalized view" in result["sheet"]["disclosure"]

    ar, _paths = audit_run(result, cfg, str(tmp_path / "out"))
    assert ar.counts["blocker"] == 0, [
        c.verdict_line for c in ar.checks if c.status == "fail"]
    c18 = next(c for c in ar.checks if c.name == "C18")
    assert c18.status == "pass"


# ===========================================================================
# THE RUN-PACK SEAM (P1-4) — executed live, producer -> consumer
# ===========================================================================
# The pack is a producer->consumer schema contract, and the P0-2 lesson is that
# a contract only exists if it is EXECUTED across the seam in the suite
# release.sh gates. Fixture-based contract testing without a regeneration
# trigger is a contract with a dead counterparty — so these tests emit the pack
# with the IN-TREE matrix engine and consume it through THIS scorecard, every
# run. Nothing here is a committed fixture.


def _emit_pack(tmp_path, matrix_path, *, run_id=None, exclusions=(),
               standing=None, firms=None):
    """Emit a run pack with the in-tree matrix engine (the live producer)."""
    import openpyxl as _x
    sys.path.insert(0, MATRIX_ENGINE)
    from src.normalize import normalize_bid
    from src.run_config import RunInputs
    from src.scorecard_pack import emit_scorecard_pack, read_standing_framework

    spec = importlib.util.spec_from_file_location("_mk_pack", GENERATOR)
    mk = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mk)

    if run_id is None:
        props = {p.name: p.value
                 for p in _x.load_workbook(matrix_path).custom_doc_props.props}
        run_id = props["falke_bid_tools.run_id"]

    n = len(firms) if firms is not None else 4
    bids = [normalize_bid(mk._doc(i)) for i in range(n)]
    run = RunInputs(project_name=PROJECT, project_address=ADDRESS,
                    gross_sf=12_000.0, sf_basis_label="GSF",
                    sf_source="explicit")
    out = tmp_path / f"packdir_{abs(hash((run_id, n, tuple(exclusions))))}"
    return str(emit_scorecard_pack(
        out_dir=out, matrix_path=matrix_path, run=run, run_id=run_id,
        bids=bids, matrix_exclusions=list(exclusions),
        standing=standing or read_standing_framework(None)))


def _label_row(ws, label):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower() == str(label).strip().lower():
            return r
    raise AssertionError(f"label {label!r} not found on {ws.title}")


def _fill_pack(path, out, *, basis=None, lock_date=None, note=None,
               weights=None, bid_open="2026-06-01", scored="2026-06-15",
               settings_extra=None):
    """Fill a pack the way a Falke operator would."""
    wb = openpyxl.load_workbook(path)

    ws = wb["Settings"]
    ws.cell(row=_label_row(ws, "Bid Opening Date"), column=2).value = bid_open
    if settings_extra:
        row = ws.max_row + 1
        for k, v in settings_extra:
            ws.cell(row=row, column=1).value = k
            ws.cell(row=row, column=2).value = v
            row += 1

    ws = wb["Baseline"]
    for label, value in (("Band Low ($M)", 1.00), ("Band High ($M)", 1.20),
                         ("Band Mid ($M)", 1.10),
                         ("Provenance", "independent"),
                         ("Estimator of Record", "J. Rivera, Falke Atlantic")):
        ws.cell(row=_label_row(ws, label), column=2).value = value
    hdr = _label_row(ws, "Scope")
    for i, line in enumerate((
            ("Structural repairs", "Modeled allowance", "$600,000", 600_000, None),
            ("Waterproofing", "Modeled allowance", "$400,000", 400_000, None))):
        for c, v in enumerate(line, start=1):
            ws.cell(row=hdr + 1 + i, column=c).value = v

    ws = wb["Framework"]
    if basis is not None:
        ws.cell(row=_label_row(ws, "Framework Basis"), column=2).value = basis
    if lock_date is not None:
        ws.cell(row=_label_row(ws, "Framework Lock Date"), column=2).value = lock_date
    if note is not None:
        ws.cell(row=_label_row(ws, "Ruling Note"), column=2).value = note
    if weights is not None:
        hdr = _label_row(ws, "Category")
        for i, w in enumerate(weights):
            ws.cell(row=hdr + 1 + i, column=3).value = w

    ws = wb["Scores"]
    ws.cell(row=_label_row(ws, "Scoring Completed Date"), column=2).value = scored
    hdr = _label_row(ws, "Firm")
    r = hdr + 1
    while ws.cell(row=r, column=1).value:
        for c in range(2, 20):
            if ws.cell(row=hdr, column=c).value:
                ws.cell(row=r, column=c).value = 7
        r += 1

    wb.save(out)
    return str(out)


def _run_cli(matrix, tmp_path, *extra):
    return cli_main([
        "--matrix", matrix, "--project-name", PROJECT,
        "--out-dir", str(tmp_path / "out"), "--html-only", *extra])


PROJECT = "Harborview Synthetic Test Tower"
ADDRESS = "1 Test Quay, Test City FL 00000"


@pytest.fixture
def filled_pack(generated, tmp_path):
    return _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "filled.xlsx")


# ---------------------------------------------------------------------------
# THE gate: one upload, live producer -> live consumer, full render
# ---------------------------------------------------------------------------
def test_one_upload_path_renders_end_to_end(generated, filled_pack, tmp_path):
    """The whole point of P1-4: the operator brings ONE workbook and the card
    renders — with the firm names never re-typed."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0

    out = tmp_path / "out"
    assert (out / "scorecard.html").exists()
    run = json.loads((out / "scorecard_run.json").read_text())
    assert run["input_channel"] == "pack"
    assert run["pack"]["binding"]["tier"] == "I4"
    assert run["pack"]["framework_basis"] == "standing"
    # the P1-6 seam: parsed and RECORDED now, enforced later
    assert run["pack"]["baseline_provenance"]["provenance"] == "independent"
    assert run["pack"]["baseline_provenance"]["estimator_of_record"]
    # every bidder in the matrix is scored, from producer-filled names
    assert len(run["bidders"]) == 4


def test_pack_firm_column_is_the_matrix_roster_in_matrix_order(generated,
                                                               tmp_path):
    """The re-keying killer (§3.4): the Firm column IS the pipeline's roster."""
    wb = openpyxl.load_workbook(_emit_pack(tmp_path, generated[4]))
    ws = wb["Scores"]
    hdr = _label_row(ws, "Firm")
    firms, r = [], hdr + 1
    while ws.cell(row=r, column=1).value:
        firms.append(ws.cell(row=r, column=1).value)
        r += 1
    parsed = MatrixParser(_cfg().block("matrix")).parse(generated[4])
    assert firms == [b.raw_name for b in parsed.blocks]


# ---------------------------------------------------------------------------
# PRE-FILLED IS NOT PRE-CONFIRMED — Floyd's protected list, live (§5)
# ---------------------------------------------------------------------------
def test_pack_does_not_satisfy_the_sf_gate(generated, filled_pack, tmp_path,
                                           capsys):
    """The pack pre-fills the SF echo. It must NOT satisfy the SF gate: a
    suggestion does not become a confirmation by traveling through a
    spreadsheet."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--baseline-confirmed")
    assert rc == 2
    assert "SF basis not confirmed" in capsys.readouterr().err


def test_pack_does_not_satisfy_the_baseline_gate(generated, filled_pack,
                                                 tmp_path, capsys):
    """The pack carries the band. It must NOT satisfy the baseline gate."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed")
    assert rc == 2
    assert "Baseline not confirmed" in capsys.readouterr().err


def test_hand_added_confirmation_key_is_rejected(generated, tmp_path, capsys):
    """R1 + R5, the load-bearing pair: no confirmation field exists in the
    schema, and an unknown Settings key hard-stops naming itself. A helpful
    operator (or a maintainer chasing convenience) who adds `sf_confirmed: yes`
    gets a stop, not a silence."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "smuggled.xlsx",
                      settings_extra=[("sf_confirmed", "yes")])
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack,
                  "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert "unrecognized field" in err and "sf_confirmed" in err


def test_inputs_and_individual_flags_are_mutually_exclusive(generated,
                                                           filled_pack,
                                                           tmp_path, capsys):
    """§9.2: one channel per run. No merge semantics, no precedence rules."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--baseline", "whatever.xlsx", "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "--baseline" in capsys.readouterr().err


@pytest.mark.parametrize("flag,value", [
    ("--band-low", "0.50"),
    ("--band-high", "9.99"),
    ("--mid", "0.75"),
    # floats: `--band-low 0` is falsy, so a truthiness check would let a zero
    # band through the very gate meant to stop it.
    ("--band-low", "0"),
])
def test_band_flags_cannot_override_the_packs_baseline_tab(generated,
                                                           filled_pack,
                                                           tmp_path, capsys,
                                                           flag, value):
    """§9.2 + R6 — regression pin (Boris, 2026-07-16).

    The band flags used to silently win over the pack's Baseline tab: --inputs
    with --band-low 0.50 rendered a $0.50M card off a pack that said $1.00M, at
    exit 0. That is two channels for one fact with an implicit "the flag wins"
    precedence rule — forbidden — and it defeats the baseline-confirmation gate,
    because the owner confirms the pack's band at preview and a flag then
    substitutes a different one into the render.
    """
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  flag, value, "--sf-confirmed", "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert flag in err
    assert "one channel or the other" in err


def test_band_flags_still_override_on_the_escape_hatch_path(generated,
                                                            tmp_path):
    """The other half of the fix: the band flags remain the DOCUMENTED override
    on the individual --baseline path (§9.1, legitimate indefinitely). Only the
    pack combination is barred — barring them outright would break the legacy
    and archival-re-render paths that have no pack to edit."""
    fw = os.path.join(HERE, "..", "templates", "scoring-framework-template.xlsx")
    scores, baseline = _individual_inputs(tmp_path)
    rc = _run_cli(generated[4], tmp_path, "--baseline", str(baseline),
                  "--scoring-framework", fw, "--category-scores", str(scores),
                  "--band-low", "0.90", "--band-high", "1.30", "--mid", "1.05",
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    assert (tmp_path / "out" / "scorecard.html").exists()


# ---------------------------------------------------------------------------
# The declaration-keyed drift model (§4.4)
# ---------------------------------------------------------------------------
def _standing(tmp_path, rows):
    """Build Falke's standing-framework.xlsx — the artifact that does not exist
    yet (§10.2). Proves the seam accepts one without redesign."""
    sys.path.insert(0, MATRIX_ENGINE)
    from src.scorecard_pack import read_standing_framework
    path = tmp_path / "standing-framework.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standing_Framework"
    ws.cell(row=1, column=1, value="Version")
    ws.cell(row=1, column=2, value="2.0")
    ws.cell(row=2, column=1, value="Effective Date")
    ws.cell(row=2, column=2, value="2026-01-01")
    for c, h in enumerate(("Category", "Short Label", "Weight (%)",
                           "What it captures"), start=1):
        ws.cell(row=4, column=c, value=h)
    for i, (cat, label, weight, desc) in enumerate(rows):
        for c, v in enumerate((cat, label, weight, desc), start=1):
            ws.cell(row=5 + i, column=c, value=v)
    wb.save(path)
    return read_standing_framework(str(path))


@pytest.fixture
def standing_rows():
    sys.path.insert(0, os.path.join(HERE, "..", "scorecard"))
    from scorecard.pack_schema import DEFAULT_FRAMEWORK_ROWS
    return list(DEFAULT_FRAMEWORK_ROWS)


def _audit_check(out_dir, name):
    data = json.loads((out_dir / "audit.json").read_text())
    return next(c for c in data["checks"] if c["name"] == name)


def test_undeclared_drift_is_a_blocker(generated, tmp_path, standing_rows):
    """W1/W2 — the steering vector, and the only one. The pack declares
    'standing' but the weights are not the standing weights: the declaration
    contradicts its own content."""
    standing = _standing(tmp_path, standing_rows)
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], standing=standing),
                      tmp_path / "drifted.xlsx",
                      basis="standing",
                      weights=[40, 10, 10, 10, 10, 10, 5, 5])
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    c19 = _audit_check(tmp_path / "out", "C19")
    assert c19["severity"] == "BLOCKER" and c19["status"] == "fail"
    # exit-contract v2 (P1-1): an audit BLOCKER is exit 3 — "delivered WITH a
    # blocker", artifacts on disk, lead with the disclosure. It used to be the
    # overloaded exit 1 ("environment / nothing written"), which said the
    # opposite of what happened.
    assert rc == 3


def test_declared_drift_is_a_warn_plus_disclosure(generated, tmp_path,
                                                  standing_rows):
    """W5 — legitimate, declared, disclosed. Exactly as designed: the control is
    not prevention, it is declaration + detection + disclosure."""
    standing = _standing(tmp_path, standing_rows)
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], standing=standing),
                      tmp_path / "declared.xlsx",
                      basis="revised-post-opening",
                      note="Board approved revised weights on 2026-06-10.",
                      weights=[40, 10, 10, 10, 10, 10, 5, 5])
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    c19 = _audit_check(tmp_path / "out", "C19")
    assert c19["severity"] == "WARN" and c19["status"] == "fail"

    # the disclosure is MANDATORY and UNCONDITIONAL on the card
    html = (tmp_path / "out" / "scorecard.html").read_text()
    assert "revised after bids" in html
    assert "Board approved revised weights on 2026-06-10." in html
    assert "so the board can weigh them accordingly" in html


def test_standing_declaration_matching_standing_is_clean(generated, tmp_path,
                                                         standing_rows):
    """The honest path is the lazy path: declaring 'standing' with unmodified
    weights costs the operator nothing and passes clean."""
    standing = _standing(tmp_path, standing_rows)
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], standing=standing),
                      tmp_path / "clean.xlsx")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    assert _audit_check(tmp_path / "out", "C19")["status"] == "pass"
    html = (tmp_path / "out" / "scorecard.html").read_text()
    assert "standing evaluation framework (version 2.0" in html


def test_semantic_hash_ignores_prose(generated, tmp_path, standing_rows):
    """§4.3 — the hash is SEMANTIC, not byte-level. Falke rewording a 'What it
    captures' cell must not fire a fiduciary alarm; an alarm that cries wolf is
    worse than no alarm."""
    reworded = [(c, l, w, "totally different prose here")
                for c, l, w, _d in standing_rows]
    standing = _standing(tmp_path, standing_rows)
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], standing=standing),
                      tmp_path / "reworded.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Framework"]
    hdr = _label_row(ws, "Category")
    for i, row in enumerate(reworded):
        ws.cell(row=hdr + 1 + i, column=4).value = row[3]
    wb.save(pack)

    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    assert _audit_check(tmp_path / "out", "C19")["status"] == "pass"


def test_no_standing_framework_degrades_to_warn_and_claims_nothing(
        generated, filled_pack, tmp_path):
    """W8 (§10.2) — TODAY'S REALITY. Falke has no standing framework, so the
    check WARNs always and the card must NOT claim a standing framework that
    does not exist."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    c19 = _audit_check(tmp_path / "out", "C19")
    assert c19["severity"] == "WARN" and c19["status"] == "fail"
    assert "no standing evaluation framework was on file" in c19["verdict_line"]

    html = (tmp_path / "out" / "scorecard.html").read_text()
    assert "No standing evaluation framework was on file for this run" in html
    assert "standing evaluation framework (version" not in html


def test_w8_plus_revised_post_opening_puts_both_sentences_on_the_card(
        generated, tmp_path):
    """Floyd C-1 regression pin — THE combination that describes every run
    until Falke adopts a standing framework.

    W8 used to return early and swallow the operator's declaration with it, so
    an operator who honestly declared a post-opening re-weighting produced a
    board packet where that fact was invisible in every human-readable artifact
    (only the run JSON had it). That is the F1 failure class, and it punished
    exactly the operator the declaration model exists to reward.

    The standing-CLAIM is conditional; the declaration is not. Both get said.
    """
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "w8_revised.xlsx",
                      basis="revised-post-opening",
                      note="Board approved revised weights on 2026-06-10.",
                      weights=[40, 10, 10, 10, 10, 10, 5, 5])
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0

    for artifact in ("scorecard.html", "scorecard_summary.html"):
        text = (tmp_path / "out" / artifact).read_text()
        # (1) the conditional standing-framework claim — still honest
        assert "No standing evaluation framework was on file for this run" in text, artifact
        # (2) the operator's declaration — MANDATORY, UNCONDITIONAL (§4.5)
        assert "revised after bids were opened" in text, artifact
        assert "Board approved revised weights on 2026-06-10." in text, artifact
        assert "so the board can weigh them accordingly" in text, artifact
        # and it must still not claim a standing framework it does not have
        assert "standing evaluation framework (version" not in text, artifact


def test_w8_ruling_note_without_a_period_still_reads_as_two_sentences(
        generated, tmp_path):
    """Floyd F-4 — the note is interpolated mid-paragraph on the card's most
    sensitive line. An operator who omits the final period must not produce
    '…on 2026-06-10 Weights set after bid opening…'."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "noperiod.xlsx",
                      basis="revised-post-opening",
                      note="Board approved revised weights on 2026-06-10")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    html = (tmp_path / "out" / "scorecard.html").read_text()
    assert "on 2026-06-10. Weights set after bid opening" in html
    assert "2026-06-10 Weights" not in html


def test_w8_plus_project_specific_composes_the_declared_lock_date(
        generated, tmp_path):
    """Floyd C-3 pin — W8 + `project-specific` describes real runs until Falke
    adopts a standing framework.

    Suppression is CLAUSE level, not branch level: suppress only clauses that
    assert a fact the tool does not have; never suppress a clause that carries
    a fact the operator declared. So the lock date and the pre-opening claim
    RENDER (the W8 sentence is basis-independent and speaks to the provenance
    of the content, never to WHEN — and W3 is a BLOCKER keyed on that very
    date), the "differ from Falke's standing framework" comparison does NOT
    (there is nothing on file to differ from), and the award-file pointer may.
    """
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "w8_ps.xlsx",
                      basis="project-specific",
                      bid_open="2026-06-01", lock_date="2026-05-20",
                      note="Roofing-only package.")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0

    for artifact in ("scorecard.html", "scorecard_summary.html"):
        text = (tmp_path / "out" / artifact).read_text()
        # the claim the tool cannot make
        assert "No standing evaluation framework was on file for this run" in text, artifact
        assert "differ from Falke&#39;s standing framework" not in text, artifact
        assert "differ from Falke's standing framework" not in text, artifact
        # the facts the operator DID declare
        assert "locked on 2026-05-20, before bids were opened" in text, artifact
        assert "recorded in the award file" in text, artifact
        # and no stutter from the naive composition
        assert "set for this project" not in text, artifact


def test_w8_plus_standing_declaration_is_suppressed_whole(generated,
                                                          filled_pack,
                                                          tmp_path):
    """The other half of C-3: `standing` IS suppressed whole under W8, and that
    stays correct — the entire sentence is a standing-framework claim
    ("applied unmodified" relative to what?). Nothing in it is an
    operator-declared fact the W8 sentence does not already carry."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    html = (tmp_path / "out" / "scorecard.html").read_text()
    assert "No standing evaluation framework was on file for this run" in html
    assert "applied unmodified" not in html
    assert "locked on" not in html


def test_standing_framework_present_declaration_reaches_the_summary_too(
        generated, tmp_path, standing_rows):
    """C-1's summary half: the declaration is a board-facing disclosure, so it
    travels on the summary as well as the card — the same text, imported from
    one function. Two human-readable documents in one packet must not disagree
    about what the board has to weigh."""
    standing = _standing(tmp_path, standing_rows)
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], standing=standing),
                      tmp_path / "sum_decl.xlsx",
                      basis="revised-post-opening",
                      note="Board approved revised weights on 2026-06-10.",
                      weights=[40, 10, 10, 10, 10, 10, 5, 5])
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    summary = (tmp_path / "out" / "scorecard_summary.html").read_text()
    assert "revised after bids were opened" in summary
    assert "Board approved revised weights on 2026-06-10." in summary
    # no standing-framework claim is suppressed here — a reference exists
    assert "No standing evaluation framework was on file" not in summary


# ---------------------------------------------------------------------------
# The two-clock coherence audit (§4.4 W3/W4)
# ---------------------------------------------------------------------------
def test_project_specific_locked_after_bid_opening_blocks(generated, tmp_path):
    """W3 — the declaration claims a pre-opening lock and states a post-opening
    date. A self-contradiction, never a disagreement with the tool."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "w3.xlsx",
                      basis="project-specific",
                      lock_date="2026-06-05", bid_open="2026-06-01",
                      note="Roofing-only package.")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    c20 = _audit_check(tmp_path / "out", "C20")
    assert c20["severity"] == "BLOCKER" and c20["status"] == "fail"
    assert rc == 3   # exit-contract v2: delivered WITH an audit blocker


def test_framework_locked_after_scoring_completed_blocks(generated, tmp_path):
    """W4 — the plan was locked after the record was made."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "w4.xlsx",
                      basis="project-specific",
                      bid_open="2026-06-01", lock_date="2026-05-20",
                      scored="2026-05-01", note="Roofing-only package.")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    c20 = _audit_check(tmp_path / "out", "C20")
    assert c20["severity"] == "BLOCKER" and c20["status"] == "fail"
    assert rc == 3   # exit-contract v2: delivered WITH an audit blocker


def test_missing_ruling_note_is_an_input_gate_stop(generated, tmp_path, capsys):
    """W7 — exit 2, not an audit finding: a required cell is empty."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "w7.xlsx",
                      basis="revised-post-opening", note="")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "Ruling Note" in capsys.readouterr().err


# ---------------------------------------------------------------------------
# The binding table (§8.3)
# ---------------------------------------------------------------------------
def test_cross_project_pack_hard_stops(generated, tmp_path, capsys):
    """I3 — exit 2, ALWAYS, no warning tier. Cross-project contamination puts
    Building A's bidders on Building B's card; there is no legitimate case."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "wrongbldg.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Settings"]
    ws.cell(row=_label_row(ws, "Project Name"), column=2).value = "Other Tower"
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "project name mismatch" in capsys.readouterr().err


def test_run_id_mismatch_with_reconciling_roster_warns_not_blocks(
        generated, tmp_path):
    """I5 (§8.4) — the corrected-matrix re-run is real, common and legitimate.
    Hard-refusing would force the operator to re-key eight firms, recreating the
    exact failure class the pack exists to kill. run_id is EVIDENCE; the roster
    is the gate."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4], run_id="deadbeef0001"),
                      tmp_path / "i5.xlsx")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    assert run["pack"]["binding"]["tier"] == "I5"
    c22 = _audit_check(tmp_path / "out", "C22")
    assert c22["severity"] == "WARN" and c22["status"] == "fail"


def test_identity_less_matrix_names_that_the_i3_check_could_not_run(
        generated, filled_pack, tmp_path, capsys):
    """Floyd F-3 — a matrix stamped as pipeline-produced but carrying no
    identity (a pre-P1-4 v0.4.0 workbook; Falke has these on disk today) makes
    I3 silently unable to run, and I3 is the one tier Marvin ruled "exit 2,
    ALWAYS. No warning tier." The I7 WARN must NAME that gap rather than leave
    it inferable — same principle as C-1: say the thing you know."""
    legacy = str(tmp_path / "legacy_matrix.xlsx")
    shutil.copy(generated[4], legacy)
    wb = openpyxl.load_workbook(legacy)
    keep = [p for p in list(wb.custom_doc_props.props)
            if p.name in ("falke_bid_tools.producer",
                          "falke_bid_tools.format_version")]
    while len(wb.custom_doc_props.props):
        wb.custom_doc_props.props.pop()
    for p in keep:
        wb.custom_doc_props.append(p)
    wb.save(legacy)

    rc = cli_main(["--matrix", legacy, "--inputs", filled_pack,
                   "--project-name", PROJECT, "--out-dir", str(tmp_path / "out"),
                   "--html-only", "--sf-confirmed", "--baseline-confirmed"])
    assert rc == 0
    assert "cross-project check (I3) COULD NOT RUN" in capsys.readouterr().err

    c22 = _audit_check(tmp_path / "out", "C22")
    assert c22["severity"] == "WARN" and c22["status"] == "fail"
    assert "COULD NOT RUN" in c22["verdict_line"]
    assert c22["evidence"]["cross_project_check_ran"] is False

    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    assert run["pack"]["binding"]["cross_project_check_ran"] is False


def test_roster_mismatch_hard_stops_naming_both_directions(generated, tmp_path,
                                                           capsys):
    """I6 — the real failure class. The message must name the difference in BOTH
    directions: "which firm is missing" and "which firm is unexpected" are
    different repairs."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "i6.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Scores"]
    ws.cell(row=_label_row(ws, "Firm") + 1, column=1).value = "Ghost Builders"
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert "In the pack but not in the matrix: Ghost Builders" in err
    assert "In the matrix but not in the pack: Alpine Restoration Group" in err


def test_edited_producer_field_hard_stops(generated, tmp_path, capsys):
    """I8 / R3 — the sheet lock is advisory UI and trivially removed, so the
    lock proves nothing. The parser's re-derivation is the integrity."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]), tmp_path / "i8.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Settings"]
    ws.cell(row=_label_row(ws, "Matrix Format Version"), column=2).value = "9.9"
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "Matrix Format Version" in capsys.readouterr().err


def test_matrix_exclusion_block_cannot_be_edited(generated, tmp_path, capsys):
    """§3.1 — an operator may ADD exclusions; an operator may not quietly delete
    the record of one the matrix ruled. Re-derived from the matrix's own AUDIT
    sheet, never trusted from the pack."""
    pack = _fill_pack(
        _emit_pack(tmp_path, generated[4],
                   exclusions=[("stale.json", "JSON parse error")]),
        tmp_path / "excl.xlsx")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "Matrix Exclusions" in capsys.readouterr().err


def test_additional_exclusion_needs_a_reason(generated, tmp_path, capsys):
    """No silent drops, ever (§1.5). An exclusion is a ruling; the award file
    has to say why it was made."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "noreason.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Settings"]
    row = _label_row(ws, "Additional Exclusions") + 2
    ws.cell(row=row, column=1).value = "Bayside Builders LLC"
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "no reason given" in capsys.readouterr().err


def test_additional_exclusion_applies_and_aliases_kill_the_json(generated,
                                                                tmp_path):
    """The Settings tab replaces the operator-facing JSON (Boris §B.4 / P1-5):
    aliases and exclusions are typed in the workbook they already have."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "settings.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Settings"]
    row = _label_row(ws, "Additional Exclusions") + 2
    ws.cell(row=row, column=1).value = "Driftwood Contractors"
    ws.cell(row=row, column=2).value = "Board set-aside: no condo references."
    row = _label_row(ws, "Display Aliases") + 2
    ws.cell(row=row, column=1).value = "Alpine Restoration Group"
    ws.cell(row=row, column=2).value = "Alpine"
    wb.save(pack)

    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 0
    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    names = [b["name"] for b in run["bidders"]]
    assert "Alpine" in names
    assert "Driftwood Contractors" not in names


def test_exclusion_naming_a_non_bidder_hard_stops(generated, tmp_path, capsys):
    """Defense in depth for R1: a key smuggled INTO a table block (where it
    reads as table data, not as an unknown scalar) is still rejected, by name,
    because an exclusion must name a real bidder. Correct on its own merits too
    — an exclusion naming nobody is a typo the operator wants to know about."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "ghostexcl.xlsx")
    wb = openpyxl.load_workbook(pack)
    ws = wb["Settings"]
    row = _label_row(ws, "Additional Exclusions") + 2
    ws.cell(row=row, column=1).value = "sf_confirmed"
    ws.cell(row=row, column=2).value = "yes"
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert "sf_confirmed" in err
    assert ("is not a bidder in this matrix" in err
            or "does not match this matrix" in err)


# ---------------------------------------------------------------------------
# The escape hatch (§9) stays open
# ---------------------------------------------------------------------------
def _individual_inputs(tmp_path):
    """Build the hand-built --category-scores + --baseline pair (the escape
    hatch's inputs). Returns (scores_path, baseline_path)."""
    scores = tmp_path / "scores.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category_Scores"
    ws.cell(row=2, column=1, value="Firm")
    labels = ["Pricing", "Scope", "Condo Exp", "CO Risk", "Reputation",
              "Financial", "Controls", "Docs"]
    for c, lab in enumerate(labels, start=2):
        ws.cell(row=2, column=c, value=lab)
    for r, firm in enumerate(_FIRMS, start=3):
        ws.cell(row=r, column=1, value=firm)
        for c in range(2, len(labels) + 2):
            ws.cell(row=r, column=c, value=7)
    wb.save(scores)

    baseline = tmp_path / "baseline.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline"
    for r, (lab, v) in enumerate((("Band Low ($M)", 1.0),
                                  ("Band High ($M)", 1.2),
                                  ("Band Mid ($M)", 1.1)), start=3):
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=2, value=v)
    for c, h in enumerate(("Scope", "Basis", "Cost ($)", "Value", "Kind"),
                          start=1):
        ws.cell(row=7, column=c, value=h)
    ws.cell(row=8, column=1, value="Structural repairs")
    ws.cell(row=8, column=2, value="Modeled allowance")
    ws.cell(row=8, column=4, value=1_000_000)
    wb.save(baseline)
    return scores, baseline


def test_individual_flags_against_a_stamped_matrix_warn(generated, tmp_path):
    """§9.3 — using hand-built inputs when a pack SHOULD exist means the firm
    names were not pipeline-originated. A WARN in the award file, never a block:
    the operator may have legitimately lost the file, and blocking on a guess
    would be the tool overreaching."""
    fw = os.path.join(HERE, "..", "templates", "scoring-framework-template.xlsx")
    scores, baseline = _individual_inputs(tmp_path)

    rc = _run_cli(generated[4], tmp_path, "--baseline", str(baseline),
                  "--scoring-framework", fw, "--category-scores", str(scores),
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    assert run["input_channel"] == "individual"
    assert run["pack"] is None
    c21 = _audit_check(tmp_path / "out", "C21")
    assert c21["severity"] == "WARN" and c21["status"] == "fail"
    # C19/C20/C22 must not invent findings out of an absent pack
    for name in ("C19", "C20", "C22"):
        assert _audit_check(tmp_path / "out", name)["severity"] == "INFO"


# ---------------------------------------------------------------------------
# C-2 — pack CONTENT errors stop cleanly; they never traceback (Floyd C-2)
# ---------------------------------------------------------------------------
# The three parsers extracted so the pack and the individual flags run the SAME
# code raise plain ValueError. The pack guard caught only PackError, so every
# one of these escaped as a traceback at exit 1 — with a good message nobody
# saw, and an exit code that collides with "a card exists and it has audit
# blockers on it". 216 tests were green over a 7/7 crash because not one of
# these cases had a pin. They do now.
#
# These are not exotic: Marvin left the Framework table unlocked deliberately
# (§3.3) and the pack hands the operator a pre-filled empty score grid, so
# weights-not-100 and a forgotten cell are the two likeliest slips in the
# product.

def _break_pack(src, dst, sheet, mutate):
    wb = openpyxl.load_workbook(src)
    mutate(wb[sheet])
    wb.save(dst)
    return str(dst)


def _row_below(ws, label):
    return _label_row(ws, label) + 1


def _clear(ws, row, col):
    """Blank a cell. NOT ws.cell(row, col, value=None) — openpyxl treats a None
    value as "don't set" and silently no-ops, which made the first draft of the
    blank-score pin pass against the unfixed code."""
    ws.cell(row=row, column=col).value = None


@pytest.mark.parametrize("name,sheet,mutate,expect", [
    ("weights_sum_110", "Framework",
     lambda ws: ws.cell(row=_row_below(ws, "Category"), column=3, value=35),
     "sum to 100"),
    ("weight_is_text", "Framework",
     lambda ws: ws.cell(row=_row_below(ws, "Category"), column=3, value="lots"),
     "must be numeric"),
    ("duplicate_short_labels", "Framework",
     lambda ws: ws.cell(row=_row_below(ws, "Category") + 1, column=2,
                        value="Pricing"),
     "duplicate Short Label"),
    ("score_out_of_range", "Scores",
     lambda ws: ws.cell(row=_row_below(ws, "Firm"), column=2, value=99),
     "within 1"),
    ("score_is_text", "Scores",
     lambda ws: ws.cell(row=_row_below(ws, "Firm"), column=2, value="great"),
     "must be numeric"),
    ("baseline_band_is_text", "Baseline",
     lambda ws: ws.cell(row=_label_row(ws, "Band Low ($M)"), column=2,
                        value="about a million"),
     "must be numeric"),
])
def test_pack_content_errors_stop_cleanly_without_a_traceback(
        generated, filled_pack, tmp_path, capsys, name, sheet, mutate, expect):
    pack = _break_pack(filled_pack, tmp_path / f"{name}.xlsx", sheet, mutate)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    err = capsys.readouterr().err
    assert rc == 2, f"{name}: expected the input-gate stop, got exit {rc}"
    assert err.startswith("[STOP]"), f"{name}: not an operator-facing stop"
    assert "Traceback" not in err, f"{name}: crashed instead of stopping"
    assert expect in err, f"{name}: message did not name the problem: {err}"


def test_pack_error_is_a_valueerror_so_the_wider_catch_loses_nothing():
    """The one-word fix is only safe because of this relationship — pin it, so a
    future refactor that decouples PackError from ValueError fails here rather
    than silently reopening C-2."""
    from scorecard.run_pack import PackError
    assert issubclass(PackError, ValueError)


# ===========================================================================
# EXIT-CONTRACT v2 + the PRELIMINARY watermark, live (P1-1)
# ===========================================================================

def _visible(html: str) -> str:
    """Strip HTML/CSS comments — what a reader can actually see. The templates
    document the watermark mechanism in prose that contains the word
    PRELIMINARY, and a test that cannot tell a comment from a banner would
    either fail on a clean card or pass on a marked one."""
    import re
    html = re.sub(r"<!--.*?-->", "", html, flags=re.S)
    return re.sub(r"/\*.*?\*/", "", html, flags=re.S)

def test_clean_run_carries_no_watermark_anywhere(generated, filled_pack,
                                                 tmp_path):
    """The control that must never misfire. A clean run's artifacts must have no
    watermark node, no watermark CSS, and no PRELIMINARY anywhere — a mark on
    every card is a mark on no card.

    Note this run is PASS-WITH-WARNINGS (W8 fires on every run until Falke
    adopts a standing framework), which is exit 0 and therefore unmarked. If
    that ever regresses to "warnings are preliminary", every honest card in
    production sprouts a watermark and the control dies the same week."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    assert run["audit_verdict"] == "PASS-WITH-WARNINGS"
    assert run["exit_code"] == 0
    assert run["watermark"] == []
    for artifact in ("scorecard.html", "scorecard_summary.html"):
        text = (tmp_path / "out" / artifact).read_text()
        # the rendered signals, not the source: the template's explanatory
        # comments and the unused .prelim-* class definitions legitimately ship
        # on every card. What must be absent is anything a READER can see.
        assert '<div class="prelim-banner">' not in text, artifact
        assert "data:image/svg+xml" not in text, artifact   # the tiling mark
        assert "PRELIMINARY" not in _visible(text), artifact


def test_audit_blocker_exits_3_and_stamps_every_artifact(generated, tmp_path):
    """Exit 3 = delivered WITH an audit blocker: the artifacts EXIST (that is
    the contract) and therefore must disclose themselves on their face."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "blocked.xlsx",
                      basis="project-specific",
                      lock_date="2026-06-05", bid_open="2026-06-01",
                      note="Roofing-only package.")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 3

    out = tmp_path / "out"
    # DELIVERED, not withheld — that is exit 3's contract — but NOT under the
    # deliverable's name (Marvin P1-2 §2.3: the PDF is what travels).
    assert (out / "scorecard-PRELIMINARY.html").exists()
    assert not (out / "scorecard.html").exists()
    run = json.loads((out / "scorecard_run.json").read_text())
    assert run["exit_code"] == 3
    assert run["audit_verdict"] == "FAIL"
    assert run["watermark"] == ["audit blocker"]

    for artifact in ("scorecard-PRELIMINARY.html",
                     "scorecard_summary-PRELIMINARY.html"):
        text = (out / artifact).read_text()
        assert "PRELIMINARY — audit blocker" in text, artifact
        # the tiling root-background mark — the layer that reaches every page
        assert "background-image" in text and "data:image/svg+xml" in text, artifact


def test_audit_runs_before_the_artifacts_are_written(generated, tmp_path):
    """C-R12 / verdict (e): the audit used to run AFTER the render, so a FAIL
    left a clean-looking board PDF on disk. Ordering is now load-bearing — the
    render cannot stamp a verdict that does not exist yet — so pin it by mtime
    rather than by reading the code."""
    pack = _fill_pack(_emit_pack(tmp_path, generated[4]),
                      tmp_path / "order.xlsx",
                      basis="project-specific",
                      lock_date="2026-06-05", bid_open="2026-06-01",
                      note="Roofing-only package.")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 3
    out = tmp_path / "out"
    assert (out / "audit.json").stat().st_mtime_ns <= \
        (out / "scorecard-PRELIMINARY.html").stat().st_mtime_ns


def test_no_audit_cannot_produce_an_artifact_that_looks_audited(generated,
                                                                filled_pack,
                                                                tmp_path):
    """--no-audit survives for debugging (Floyd verdict (e) keeps the flag), and
    the skill prohibits it for board runs. The engine's job is narrower and
    absolute: an unaudited artifact must never be mistakable for a checked one.
    This is also where Floyd's "audit pending" phrase is actually TRUE."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed", "--no-audit")
    assert rc == 0
    out = tmp_path / "out"
    assert not (out / "audit.json").exists()
    run = json.loads((out / "scorecard_run.json").read_text())
    assert run["audit_verdict"] is None
    assert run["watermark"] == ["not audited"]
    for artifact in ("scorecard-PRELIMINARY.html",
                     "scorecard_summary-PRELIMINARY.html"):
        text = (out / artifact).read_text()
        assert "PRELIMINARY — not audited" in text, artifact


def test_unreadable_matrix_is_exit_1_not_a_traceback(generated, filled_pack,
                                                     tmp_path, capsys):
    """Exit 1 = environment / nothing to do, nothing written (Boris §D, Floyd
    verdict (e)). Before P1-1 a bad --matrix path raised a raw
    FileNotFoundError out of openpyxl as an uncaught traceback — the right code
    by accident, with a stack trace where the operator message should be. It is
    the single most likely operator slip in the product."""
    rc = cli_main(["--matrix", str(tmp_path / "nope.xlsx"),
                   "--inputs", filled_pack, "--project-name", PROJECT,
                   "--out-dir", str(tmp_path / "out"), "--html-only",
                   "--sf-confirmed", "--baseline-confirmed"])
    assert rc == 1
    err = capsys.readouterr().err
    assert err.startswith("[STOP]")
    assert "Traceback" not in err
    assert "not found" in err
    assert not (tmp_path / "out").exists()   # nothing written


# ===========================================================================
# P1-2 — the restored provisional pathway (Marvin's ruling, his named pins)
# ===========================================================================

def _blank_scores(pack_path, out, *, leave_scored=0, scoring_date=None):
    """Blank the score grid, optionally leaving `leave_scored` cells filled.

    The completion date is CLEARED unless one is asked for — that is the honest
    mid-evaluation state (§5.3), and leaving the fixture's date behind would
    make every partial pin a self-contradiction stop instead."""
    wb = openpyxl.load_workbook(pack_path)
    ws = wb["Scores"]
    date_row = _label_row(ws, "Scoring Completed Date")
    if scoring_date is None:
        _clear(ws, date_row, 2)
    else:
        ws.cell(row=date_row, column=2).value = scoring_date
    hdr = _label_row(ws, "Firm")
    left = leave_scored
    r = hdr + 1
    while ws.cell(row=r, column=1).value:
        for c in range(2, 20):
            if ws.cell(row=hdr, column=c).value:
                if left > 0:
                    ws.cell(row=r, column=c).value = 7
                    left -= 1
                else:
                    _clear(ws, r, c)
        r += 1
    wb.save(out)
    return str(out)


def test_fully_blank_pack_hard_stops(generated, filled_pack, tmp_path, capsys):
    """MARVIN'S NAMED PIN — "the one that must never render again" (§2.1).

    A freshly emitted pack is 100% blank, and it is now the product's DEFAULT
    starting state. Had P1-2 simply stopped rejecting blanks, this input would
    have rendered a board card that RANKED the field and NAMED a recommended
    best value on zero evidence: rank_bidders' keys collapse (`-(None or 0)` →
    0 for everyone), leaving only per_sf_over_band, and summary.py named
    ranking[0] with no coverage guard. The close-call hedge needs two Overall
    numbers to fire, so at zero coverage it could not — and the summary took its
    most CONFIDENT branch. Certainty inversely proportional to evidence.
    """
    pack = _blank_scores(filled_pack, tmp_path / "allblank.xlsx")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert "No category scores were supplied" in err
    assert "nothing for it to combine" in err
    assert not (tmp_path / "out").exists()      # nothing rendered, ever


def test_coherent_partial_record_exits_4_and_withholds_every_claim(
        generated, filled_pack, tmp_path):
    """§3.3 + §4.4 — a coherent partial record renders, provisionally: no rank,
    no leader, no Overall; alphabetical; watermarked; exit 4."""
    pack = _blank_scores(filled_pack, tmp_path / "partial.xlsx", leave_scored=5)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 4

    out = tmp_path / "out"
    run = json.loads((out / "scorecard_run.json").read_text())
    assert run["exit_code"] == 4
    assert run["full_coverage"] is False
    assert run["watermark"] == ["evaluation incomplete"]
    # no bidder carries a rank or an Overall
    assert all(b["rank"] is None for b in run["bidders"])
    assert all(b["overall"]["numeric"] is None for b in run["bidders"])

    # _visible: the templates document the mechanism in comments that name the
    # very things being withheld, and a test that cannot tell a comment from a
    # section would pass on a card that still ranks.
    card = _visible((out / "scorecard-PRELIMINARY.html").read_text())
    assert "Listed alphabetically — not ranked" in card
    assert "Final Hierarchy" not in card           # Section G absent entirely
    assert "Rank #" not in card                    # Section F label gone
    assert "PRELIMINARY — evaluation incomplete" in card
    assert "Pending — " in card                    # the withheld Overall
    assert "Scoring progress" in card              # the worklist

    summary = _visible((out / "scorecard_summary-PRELIMINARY.html").read_text())
    assert "EVALUATION IN PROGRESS — no recommendation yet" in summary
    assert "Recommendation" not in summary
    for firm in _FIRMS[:4]:
        assert f'class="winner-name">{firm}' not in summary


def test_partial_plus_a_declared_completion_date_hard_stops(generated,
                                                            filled_pack,
                                                            tmp_path, capsys):
    """§5.3 — the cells say provisional, the date says final. exit 2, NOT a
    BLOCKER: the contradiction is about WHICH DOCUMENT TO BUILD, and you cannot
    render-then-flag your way out of not knowing what to render (§5.4)."""
    pack = _blank_scores(filled_pack, tmp_path / "contradict.xlsx",
                         leave_scored=5, scoring_date="2026-06-15")
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    err = capsys.readouterr().err
    assert "declares scoring completed on 2026-06-15" in err
    assert "clear the date and render provisionally" in err


def test_full_coverage_without_a_completion_date_hard_stops(generated,
                                                            filled_pack,
                                                            tmp_path, capsys):
    """§5.3 — every category scored means the record IS complete; declare when."""
    wb = openpyxl.load_workbook(filled_pack)
    ws = wb["Scores"]
    _clear(ws, _label_row(ws, "Scoring Completed Date"), 2)
    pack = str(tmp_path / "nodate.xlsx")
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 2
    assert "declare when it was completed" in capsys.readouterr().err


def test_partial_without_a_date_is_the_lazy_path_and_carries_no_friction(
        generated, filled_pack, tmp_path):
    """§5.3 — THE HONEST PATH IS THE LAZY PATH. Mid-evaluation the operator does
    nothing: leaves the date alone, leaves cells blank, gets the card."""
    pack = _blank_scores(filled_pack, tmp_path / "lazy.xlsx", leave_scored=9)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 4          # delivered, provisional, zero friction


def test_a_wholly_unscored_bidder_renders(generated, filled_pack, tmp_path):
    """§1.2 — scoring the top three after interviews and the rest later is the
    actual workflow. Blocking it forecloses the thing this item restores."""
    wb = openpyxl.load_workbook(filled_pack)
    ws = wb["Scores"]
    hdr = _label_row(ws, "Firm")
    _clear(ws, _label_row(ws, "Scoring Completed Date"), 2)
    for c in range(2, 20):                       # blank ONE bidder entirely
        if ws.cell(row=hdr, column=c).value:
            _clear(ws, hdr + 1, c)
    pack = str(tmp_path / "onebidder.xlsx")
    wb.save(pack)
    rc = _run_cli(generated[4], tmp_path, "--inputs", pack, "--sf-confirmed",
                  "--baseline-confirmed")
    assert rc == 4
    run = json.loads((tmp_path / "out" / "scorecard_run.json").read_text())
    assert len(run["bidders"]) == 4              # still on the card, unassessed


def test_full_coverage_run_is_unchanged(generated, filled_pack, tmp_path):
    """§3.3.4 / his named regression pin: THIS RULING CHANGES NOTHING ABOUT A
    COMPLETE RUN. Ranked, Overall printed, winner named, no watermark, exit 0,
    and the deliverable's own filename."""
    rc = _run_cli(generated[4], tmp_path, "--inputs", filled_pack,
                  "--sf-confirmed", "--baseline-confirmed")
    assert rc == 0
    out = tmp_path / "out"
    assert (out / "scorecard.html").exists()
    assert not (out / "scorecard-PRELIMINARY.html").exists()
    run = json.loads((out / "scorecard_run.json").read_text())
    assert run["full_coverage"] is True
    assert run["watermark"] == []
    assert sorted(b["rank"] for b in run["bidders"]) == [1, 2, 3, 4]
    assert all(b["overall"] is not None for b in run["bidders"])
    card = _visible((out / "scorecard.html").read_text())
    assert "Final Hierarchy" in card
    assert "Listed alphabetically" not in card
    assert "PRELIMINARY" not in card
