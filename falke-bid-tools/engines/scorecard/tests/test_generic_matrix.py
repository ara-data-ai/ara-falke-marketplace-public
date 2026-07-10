"""Genericity proof — the skill is a project-agnostic TEMPLATE, not a single-project
one-off.

This test programmatically BUILDS a small synthetic bid-comparison matrix as an
.xlsx (openpyxl, written to pytest's tmp_path) in the SAME structural FORMAT as
the validation matrix but DELIBERATELY DIFFERENT in every project-specific
way:

  * a different project ("HARBORVIEW TOWER · Amenity Deck"),
  * a DIFFERENT vendor count — FOUR bidders, not the validation set's 7/10,
  * different firm names (Apex / Borealis / Cascade / Delta), and
  * different dollar numbers, SF basis, and band.

It then runs the SAME parser + pipeline (the validation gold path uses) and asserts
the skill detects exactly 4 bidder blocks, reads each grand total, computes $/SF
on the SUPPLIED SF basis, assigns the right tier vs a SUPPLIED band, ranks all 4,
and threads the SUPPLIED project name into the output context.

The block detector measures count/width/stride from the sheet, so nothing here
is single-project-shaped: it is the "not a one-off" proof. The validation gold suite
(tests/test_integration_sample.py) stays intact and is the validation example.
"""
from __future__ import annotations

import openpyxl
import pytest

from scorecard.config import load_config
from scorecard.matrix import MatrixParser
from scorecard.mechanical import (TIER_DEFENSIVE, TIER_MID, TIER_RISK, TIER_TOP)
from scorecard.pipeline import run_scorecard
from scorecard.render import build_context

# ---- SYNTHETIC PROJECT — intentionally unlike the validation set in every input ----
SYNTH_PROJECT = "HARBORVIEW TOWER · Amenity Deck"
SYNTH_SF_BASIS = 10000          # not 16000
SYNTH_BAND_LOW = 2.00           # band_low_per_sf  = 2.00e6 / 1e4 = 200
SYNTH_BAND_HIGH = 2.20          # band_high_per_sf = 2.20e6 / 1e4 = 220
SYNTH_MID = 2.10
# mid_floor = 0.90 * 200 = 180 ; premium_floor = 1.20 * 220 = 264
# Four bidders, each chosen to land in a DISTINCT tier vs the band above.
SYNTH_BIDDERS = [
    ("Apex Builders LLC", 2_100_000, TIER_TOP),         # 210 $/SF -> TOP (in band)
    ("Borealis Group", 1_700_000, TIER_RISK),           # 170 $/SF -> RISK (< mid_floor)
    ("Cascade Construction Co.", 2_500_000, TIER_DEFENSIVE),  # 250 -> DEFENSIVE
    ("Delta Contractors", 1_900_000, TIER_MID),         # 190 $/SF -> MID
]
# per-division subtotals each block populates (mirrors the real layout: values
# live in the COST SUBTOTALS column at division-SUBTOTAL rows).
SYNTH_DIVISION_VALUES = {
    "Apex Builders LLC":        [300000, 250000, 180000],
    "Borealis Group":           [240000, 210000, 150000],
    "Cascade Construction Co.": [360000, 300000, 220000],
    "Delta Contractors":        [275000, 230000, 165000],
}


def _build_synthetic_matrix(path: str) -> None:
    """Write a 4-bidder synthetic matrix in the real structural format.

    Layout mirrors the real matrix GENERICALLY (no hard-coded project constants):
      row 1: project title (free text, ignored by the parser)
      row 4: bidder-name row — names placed every 5 columns (stride 5, width 4)
      row 8: the COST / COST SUBTOTALS / $/SF / $/SF SUBTOTALS quartet, sitting a
             few rows BELOW the names (the real matrix has intervening rows too)
      rows 11-13: three CSI division SUBTOTAL rows ("<DIV> SUBTOTAL") whose dollar
             values live in each block's COST SUBTOTALS column
      row 15: CONSTRUCTION COST SUBTOTAL (pre-markup; never the compared total)
      row 18: GRAND TOTAL CONSTRUCTION COST — the compared total, in COST column
      row 20: TOTAL GSF (detected, reported, NEVER used for $/SF)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"

    NAME_ROW = 4
    QUARTET_ROW = 8
    DIV_ROWS = [11, 12, 13]
    CC_SUBTOTAL_ROW = 15
    GT_ROW = 18
    GSF_ROW = 20
    FIRST_COL = 4          # blocks start at column D (label cols A-C on the left)
    STRIDE = 5             # bidder columns 5 apart -> measured stride 5, width 4

    ws.cell(row=1, column=1, value="HARBORVIEW TOWER — Amenity Deck Bid Comparison")

    # left-hand label column (col B) for the structural rows
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUBTOTAL_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    div_labels = ["GENERAL CONDITIONS SUBTOTAL", "CONCRETE SUBTOTAL",
                  "FINISHES SUBTOTAL"]
    for r, lab in zip(DIV_ROWS, div_labels):
        ws.cell(row=r, column=2, value=lab)

    # per-bidder blocks: name on NAME_ROW, quartet sub-headers on QUARTET_ROW,
    # grand total in the COST column, division subtotals in COST SUBTOTALS.
    for i, (name, total, _tier) in enumerate(SYNTH_BIDDERS):
        c_cost = FIRST_COL + i * STRIDE
        c_cost_sub = c_cost + 1
        c_psf = c_cost + 2
        c_psf_sub = c_cost + 3
        ws.cell(row=NAME_ROW, column=c_cost, value=name)
        ws.cell(row=QUARTET_ROW, column=c_cost, value="COST")
        ws.cell(row=QUARTET_ROW, column=c_cost_sub, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c_psf, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c_psf_sub, value="$/SF SUBTOTALS")
        # grand total -> COST column (where _read_block_value reads first)
        ws.cell(row=GT_ROW, column=c_cost, value=total)
        # division subtotals -> COST SUBTOTALS column (where they actually sit)
        for r, dv in zip(DIV_ROWS, SYNTH_DIVISION_VALUES[name]):
            ws.cell(row=r, column=c_cost_sub, value=dv)
        # pre-markup construction subtotal (must NOT be picked as the total)
        ws.cell(row=CC_SUBTOTAL_ROW, column=c_cost, value=total * 0.85)
        # a deliberately DIFFERENT GSF than the SF basis, to prove it's ignored
        ws.cell(row=GSF_ROW, column=c_cost, value=8500)

    wb.save(path)


def _cfg():
    """Config driven ONLY by the synthetic project's own supplied inputs."""
    return load_config(overrides={
        "sf_basis": SYNTH_SF_BASIS,
        "band_low": SYNTH_BAND_LOW,
        "band_high": SYNTH_BAND_HIGH,
        "modeled_mid_takeoff": SYNTH_MID,
    })


def test_parser_detects_four_blocks_not_single_project_shape(tmp_path):
    """The detector measures the field — it must find EXACTLY 4 blocks (not
    7/10), measure stride 5 / width 4, and read each bidder's grand total."""
    xlsx = str(tmp_path / "harborview_matrix.xlsx")
    _build_synthetic_matrix(xlsx)

    parser = MatrixParser(_cfg().block("matrix"))
    parsed = parser.parse(xlsx)

    assert parsed.sheet_name == "Bid_Comparison"
    assert len(parsed.blocks) == 4, [b.raw_name for b in parsed.blocks]
    assert len(parsed.included_blocks) == 4
    assert parsed.block_stride == 5
    assert parsed.block_width == 4
    # grand total located by label, and NOT the pre-markup construction subtotal
    assert parsed.grand_total_label.upper().startswith("GRAND TOTAL")
    # detected GSF is the 8500 we planted — reported only, never used for $/SF
    assert parsed.gsf_value == 8500
    # each bidder's grand total read from the COST column (not the 0.85 subtotal)
    by_name = {b.name: b for b in parsed.included_blocks}
    for name, total, _tier in SYNTH_BIDDERS:
        assert by_name[name].grand_total == total, name
    # three CSI division subtotal rows detected, each block populating all three
    assert len(parsed.division_rows) == 3
    for b in parsed.included_blocks:
        assert b.populated_divisions == 3, (b.name, b.populated_divisions)


def test_pipeline_ranks_four_bidders_with_supplied_basis_band_and_name(tmp_path):
    """End to end on the synthetic project: $/SF on the SUPPLIED basis, tiers vs
    the SUPPLIED band, all 4 ranked, and the SUPPLIED project name in context."""
    xlsx = str(tmp_path / "harborview_matrix.xlsx")
    _build_synthetic_matrix(xlsx)
    cfg = _cfg()

    result = run_scorecard(xlsx, cfg, project_name=SYNTH_PROJECT)
    by_name = {b["name"]: b for b in result["bidders"]}

    # all 4 present, none dropped
    assert set(by_name) == {n for n, _t, _x in SYNTH_BIDDERS}

    # $/SF computed on the SUPPLIED SF basis (10000), NOT the planted GSF (8500)
    for name, total, _tier in SYNTH_BIDDERS:
        assert by_name[name]["per_sf"] == round(total / SYNTH_SF_BASIS), name
    # sanity: using the GSF would give a different number — prove it isn't
    assert by_name["Apex Builders LLC"]["per_sf"] == 210
    assert round(2_100_000 / 8500) != 210

    # tiers assigned vs the SUPPLIED band (one bidder in each of 4 tiers)
    for name, _total, tier in SYNTH_BIDDERS:
        assert by_name[name]["tier"] == tier, (name, by_name[name]["tier"])

    # ranking covers all 4, contiguous ranks 1..4
    ranks = sorted(b["rank"] for b in result["bidders"])
    assert ranks == [1, 2, 3, 4]
    assert len(result["ranking"]) == 4

    # the SUPPLIED project name threads into the meta + render context (no
    # hard-coded project name leaks in)
    assert result["meta"]["project_name"] == SYNTH_PROJECT
    ctx = build_context(result, cfg)
    assert ctx["project_name"] == SYNTH_PROJECT
    assert "SAMPLE CONDOMINIUM" not in ctx["project_name"].upper()
    # the footer carries the SUPPLIED SF basis, not the validation set's 16000
    assert "10,000" in result["meta"]["footer_note"]
    assert "16,000" not in result["meta"]["footer_note"]
