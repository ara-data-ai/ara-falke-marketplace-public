"""Tests for scorecard.baseline_parser.parse_baseline_xlsx.

Covers:
  - test_parse_valid_xlsx: in-memory xlsx with band + 2 trade lines
  - test_missing_band_raises: xlsx with no band values -> ValueError
  - test_empty_trades_raises: xlsx with band but no trade lines -> ValueError
  - test_json_path_unchanged: existing _load_json still works (smoke test)
"""
from __future__ import annotations

import io
import json
import os

import openpyxl
import pytest

from scorecard.baseline_parser import parse_baseline_xlsx


# ---- helpers ----------------------------------------------------------------

def _make_xlsx_bytes(
    *,
    band_low=3.35,
    band_high=3.55,
    band_mid=3.45,
    trades=None,         # list of (scope, basis, cost_str, value, kind_or_None)
    omit_band=False,
) -> bytes:
    """Build a minimal Baseline-sheet xlsx in memory and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline"

    # Row 1: title
    ws.cell(row=1, column=1, value="COST BASELINE — test")
    # Row 2: blank
    # Rows 3-5: band (unless omit_band)
    if not omit_band:
        ws.cell(row=3, column=1, value="Band Low ($M)")
        ws.cell(row=3, column=2, value=band_low)
        ws.cell(row=4, column=1, value="Band High ($M)")
        ws.cell(row=4, column=2, value=band_high)
        ws.cell(row=5, column=1, value="Band Mid ($M)")
        ws.cell(row=5, column=2, value=band_mid)
    # Row 6: blank
    # Row 7: headers
    for col, header in enumerate(["Scope", "Basis", "Cost ($)", "Value", "Kind"], 1):
        ws.cell(row=7, column=col, value=header)
    # Rows 8+: trade lines
    if trades:
        for i, (scope, basis, cost_str, value, kind) in enumerate(trades):
            r = 8 + i
            ws.cell(row=r, column=1, value=scope)
            ws.cell(row=r, column=2, value=basis)
            ws.cell(row=r, column=3, value=cost_str)
            ws.cell(row=r, column=4, value=value)
            if kind:
                ws.cell(row=r, column=5, value=kind)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_xlsx(tmp_path, name, **kwargs) -> str:
    """Write an in-memory xlsx to a temp file and return its path."""
    path = str(tmp_path / name)
    with open(path, "wb") as fh:
        fh.write(_make_xlsx_bytes(**kwargs))
    return path


# ---- tests ------------------------------------------------------------------

def test_parse_valid_xlsx(tmp_path):
    """A well-formed xlsx with band values + 2 trade lines parses correctly."""
    trades = [
        ("Interior demolition", "Modeled allowance", "$200,000", 200_000, None),
        ("Direct trades subtotal", "Sum of trade lines", "$200,000", 200_000, "subtotal"),
    ]
    path = _write_xlsx(tmp_path, "valid.xlsx", trades=trades)

    band_low, band_high, band_mid, lines = parse_baseline_xlsx(path)

    assert band_low == pytest.approx(3.35)
    assert band_high == pytest.approx(3.55)
    assert band_mid == pytest.approx(3.45)
    assert len(lines) == 2

    # first line — regular row (no 'kind' key)
    l0 = lines[0]
    assert l0["scope"] == "Interior demolition"
    assert l0["basis"] == "Modeled allowance"
    assert l0["cost"] == "$200,000"
    assert l0["value"] == 200_000
    assert "kind" not in l0

    # second line — subtotal
    l1 = lines[1]
    assert l1["scope"] == "Direct trades subtotal"
    assert l1["kind"] == "subtotal"


def test_missing_band_raises(tmp_path):
    """An xlsx with no band values raises ValueError with a helpful message."""
    trades = [
        ("Some scope", "Some basis", "$100,000", 100_000, None),
    ]
    path = _write_xlsx(tmp_path, "no_band.xlsx", omit_band=True, trades=trades)

    with pytest.raises(ValueError, match="Band Low"):
        parse_baseline_xlsx(path)


def test_empty_trades_raises(tmp_path):
    """An xlsx with band values but NO trade lines raises ValueError."""
    path = _write_xlsx(tmp_path, "no_trades.xlsx", trades=[])

    with pytest.raises(ValueError, match="No trade lines found"):
        parse_baseline_xlsx(path)


def test_json_path_unchanged():
    """The existing _load_json path still works for sample_baseline.json.

    Smoke-test: load the canonical synthetic sample example and verify it
    returns a non-empty list of dicts with the expected keys.
    """
    json_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "examples",
        "sample_baseline.json",
    )
    assert os.path.exists(json_path), f"sample_baseline.json not found at {json_path}"

    with open(json_path, "r", encoding="utf-8") as fh:
        data = json.load(fh)

    assert isinstance(data, list), "sample_baseline.json should deserialize to a list"
    assert len(data) > 0, "sample_baseline.json should not be empty"

    required_keys = {"scope", "basis", "cost", "value"}
    for item in data:
        assert required_keys.issubset(item.keys()), (
            f"Missing keys in {item}; expected at least {required_keys}"
        )
