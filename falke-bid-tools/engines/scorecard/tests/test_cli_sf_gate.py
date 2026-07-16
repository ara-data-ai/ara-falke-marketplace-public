"""SF-basis SUGGEST-AND-CONFIRM gate — CLI process-level contract.

The skill no longer HARD-REFUSES to look at the matrix's own SF. It now READS
the matrix Row-10 'TOTAL GSF' and offers it as a SUGGESTED default — but a render
still REQUIRES the user to either supply an explicit --sf-basis (override) OR
pass --sf-confirmed to ACCEPT the matrix GSF. A render with neither hard-stops
(exit 2) with a message that NAMES the matrix SF. This file locks that contract
end-to-end through cli.main():

  (a) --sf-basis explicit        -> renders, $/SF uses THAT value;
  (b) --sf-confirmed (no basis)  -> renders, $/SF uses the matrix Row-10 GSF;
  (c) neither (a render)         -> [STOP] exit 2, message suggests the matrix SF;
  (d) the matrix Row-10 GSF is correctly read by the gate (matches the planted
      value, drives $/SF under --sf-confirmed);
  + --preview-baseline surfaces the matrix-suggested SF and renders nothing.

Self-contained: builds a small synthetic matrix in the real structural format
(no client data; the client xlsx is gitignored in builder worktrees). The
planted GSF (12,000) is DISTINCT from the explicit override (10,000) so the two
$/SF outcomes are provably different.
"""
from __future__ import annotations

import json
import os

import openpyxl

from scorecard.cli import main

from .conftest import (SIMPLE_FRAMEWORK_ROWS, write_framework_xlsx,
                       write_scores_xlsx)

PROJECT = "SF-GATE TEST · Synthetic"
EXPLICIT_SF = 10_000        # explicit override; total/10000 = 210 for the 2.1M bidder
MATRIX_GSF = 12_000         # planted Row-10 GSF; total/12000 = 175 -> different $/SF
BAND_LOW, BAND_HIGH, MID = 2.00, 2.20, 2.10

# A bidder at $2.10M: $/SF = 210 on the explicit 10,000 basis, 175 on the 12,000
# matrix GSF — the two confirmed paths must produce DIFFERENT, provable numbers.
SYNTH_BIDDERS = [
    ("Apex Builders LLC", 2_100_000, [300_000, 250_000, 180_000]),
    ("Borealis Group",    1_900_000, [240_000, 210_000, 150_000]),
]


def _build_matrix(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"
    NAME_ROW, QUARTET_ROW = 4, 8
    DIV_ROWS = [11, 12, 13]
    CC_SUBTOTAL_ROW, GT_ROW, GSF_ROW = 15, 18, 20
    FIRST_COL, STRIDE = 4, 5

    ws.cell(row=1, column=1, value="SYNTHETIC — SF-gate test matrix")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUBTOTAL_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    # the matrix's own TOTAL GSF value sits to the RIGHT of the label (col C)
    ws.cell(row=GSF_ROW, column=3, value=MATRIX_GSF)
    for r, lab in zip(DIV_ROWS, ["GENERAL CONDITIONS SUBTOTAL",
                                 "CONCRETE SUBTOTAL", "FINISHES SUBTOTAL"]):
        ws.cell(row=r, column=2, value=lab)

    for i, (name, total, divs) in enumerate(SYNTH_BIDDERS):
        c = FIRST_COL + i * STRIDE
        ws.cell(row=NAME_ROW, column=c, value=name)
        ws.cell(row=QUARTET_ROW, column=c, value="COST")
        ws.cell(row=QUARTET_ROW, column=c + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c + 3, value="$/SF SUBTOTALS")
        ws.cell(row=GT_ROW, column=c, value=total)
        for r, dv in zip(DIV_ROWS, divs):
            ws.cell(row=r, column=c + 1, value=dv)
        ws.cell(row=CC_SUBTOTAL_ROW, column=c, value=total * 0.85)
    wb.save(path)


def _baseline_json(path: str) -> None:
    lines = [
        {"scope": "Interior demolition", "basis": "Modeled allowance",
         "cost": "$200,000", "value": 200_000},
        {"scope": "Direct trades subtotal", "basis": "Sum", "cost": "$200,000",
         "value": 200_000, "kind": "subtotal"},
    ]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(lines, fh)


def _setup(tmp_path):
    matrix = str(tmp_path / "sf_gate_matrix.xlsx")
    baseline = str(tmp_path / "baseline.json")
    _build_matrix(matrix)
    _baseline_json(baseline)
    return matrix, baseline


def _scoring_argv(tmp_path, firm_names):
    """The two REQUIRED per-run scoring xlsx inputs for a synthetic field
    (exempt for --preview-baseline; the SF stop fires after the scoring gate)."""
    fw = write_framework_xlsx(str(tmp_path / "framework.xlsx"),
                              SIMPLE_FRAMEWORK_ROWS)
    labels = [r[1] for r in SIMPLE_FRAMEWORK_ROWS]
    # DISTINCT descending pricing scores so no two bidders tie on Overall
    # (an exact tie with the pricier bidder listed first trips audit C5).
    cs = write_scores_xlsx(str(tmp_path / "scores.xlsx"), labels,
                           [(n, [max(1, 8 - i), 7, 7])
                            for i, n in enumerate(firm_names)])
    return ["--scoring-framework", fw, "--category-scores", cs]


def _base_argv(matrix, baseline, tmp_path):
    return (["--matrix", matrix, "--project-name", PROJECT,
             "--band-low", str(BAND_LOW), "--band-high", str(BAND_HIGH),
             "--mid", str(MID), "--baseline", baseline]
            + _scoring_argv(tmp_path, [n for n, _, _ in SYNTH_BIDDERS]))


def _run_json(out_dir):
    with open(os.path.join(out_dir, "scorecard_run.json"), encoding="utf-8") as fh:
        return json.load(fh)


def _no_artifacts(out_dir):
    """A gate stop writes NOTHING — under either name. P1-2 §2.3 gives a
    non-deliverable artifact a distinct filename, so an absence assertion that
    only knew the deliverable's name would pass while a PRELIMINARY card sat on
    disk beside it."""
    for fn in ("scorecard.pdf", "scorecard.html", "scorecard_run.json",
               "scorecard-PRELIMINARY.pdf", "scorecard-PRELIMINARY.html"):
        assert not os.path.exists(os.path.join(out_dir, fn)), fn


# (d) the gate reads the matrix Row-10 GSF -------------------------------------
def test_gate_reads_matrix_row10_gsf():
    from scorecard.config import load_config
    from scorecard.matrix import MatrixParser
    import tempfile
    with tempfile.TemporaryDirectory() as d:
        matrix = os.path.join(d, "m.xlsx")
        _build_matrix(matrix)
        cfg = load_config(overrides={"sf_basis": 1, "band_low": BAND_LOW,
                                     "band_high": BAND_HIGH, "modeled_mid_takeoff": MID})
        gsf_row, gsf_val = MatrixParser(cfg.block("matrix")).detect_sf(matrix)
        assert gsf_val == MATRIX_GSF
        assert gsf_row == 20


# (a) explicit --sf-basis renders and uses THAT value --------------------------
def test_explicit_sf_basis_renders_and_uses_it(tmp_path):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_base_argv(matrix, baseline, tmp_path) + [
        "--sf-basis", str(EXPLICIT_SF), "--baseline-confirmed",
        "--html-only", "--no-audit", "--out-dir", out])
    assert rc == 0
    rj = _run_json(out)
    apex = next(b for b in rj["bidders"] if "Apex" in b["name"])
    assert apex["per_sf"] == round(2_100_000 / EXPLICIT_SF)   # 210, not 175


# (b) --sf-confirmed (no --sf-basis) renders using the matrix GSF --------------
def test_sf_confirmed_renders_using_matrix_gsf(tmp_path):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_base_argv(matrix, baseline, tmp_path) + [
        "--sf-confirmed", "--baseline-confirmed",
        "--html-only", "--no-audit", "--out-dir", out])
    assert rc == 0
    rj = _run_json(out)
    apex = next(b for b in rj["bidders"] if "Apex" in b["name"])
    assert apex["per_sf"] == round(2_100_000 / MATRIX_GSF)    # 175, the matrix GSF
    assert apex["per_sf"] != round(2_100_000 / EXPLICIT_SF)


# (b') --sf-confirmed render passes the audit (C3 honors explicit confirmation) -
def test_sf_confirmed_render_passes_audit_c3(tmp_path):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_base_argv(matrix, baseline, tmp_path) + [
        "--sf-confirmed", "--baseline-confirmed", "--html-only", "--out-dir", out])
    # audit ON: a confirmed-matrix SF must NOT trip C3 (would return 1 on FAIL)
    assert rc == 0
    with open(os.path.join(out, "audit.json"), encoding="utf-8") as fh:
        audit = json.load(fh)
    c3 = next(c for c in audit["checks"] if c["name"] == "C3")
    assert c3["status"] == "pass"
    assert c3["evidence"]["sf_source"] == "matrix-confirmed"


# (c) neither -> hard-stop exit 2 naming the matrix SF -------------------------
def test_render_without_sf_basis_or_confirm_stops_exit2(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_base_argv(matrix, baseline, tmp_path) + ["--baseline-confirmed",
                                              "--out-dir", out])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[STOP] SF basis not confirmed" in err
    assert f"{MATRIX_GSF:,.0f}" in err          # the matrix SF is named: "12,000"
    assert "--sf-basis" in err and "--sf-confirmed" in err
    _no_artifacts(out)


# + preview surfaces the matrix-suggested SF and renders nothing ----------------
def test_preview_surfaces_matrix_suggested_sf(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_base_argv(matrix, baseline, tmp_path) + ["--preview-baseline",
                                              "--out-dir", out])
    assert rc == 0
    s = capsys.readouterr().out
    assert "SUGGESTED from matrix Row-10" in s
    assert f"{MATRIX_GSF:,.0f}" in s
    assert "--sf-confirmed" in s and "--sf-basis" in s
    _no_artifacts(out)


# end-to-end gate on a VARIABLE field (11 bidders x 25 divisions) through main():
# preview -> STOP -> --sf-confirmed render -> explicit-override render. Stands in
# for the manual --preview-baseline / render smoke run on a synthetic edge matrix
# (the build sandbox only runs `python -m pytest`), exercising the gate end to end
# on a non-single-project-shaped matrix.
def _build_edge_matrix(path: str, gsf: int) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"
    NAME_ROW, QUARTET_ROW = 4, 8
    n_div = 25
    DIV_ROWS = list(range(11, 11 + n_div))
    CC_SUB_ROW = DIV_ROWS[-1] + 2
    GT_ROW = CC_SUB_ROW + 2
    GSF_ROW = GT_ROW + 2
    ws.cell(row=1, column=1, value="EDGE 11x25 smoke matrix")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUB_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    ws.cell(row=GSF_ROW, column=3, value=gsf)
    for j, r in enumerate(DIV_ROWS):
        ws.cell(row=r, column=2, value=f"DIVISION {j + 1:02d} SUBTOTAL")
    for i in range(11):
        c = 4 + i * 5
        ws.cell(row=NAME_ROW, column=c, value=f"Restorer {i:02d}")
        ws.cell(row=QUARTET_ROW, column=c, value="COST")
        ws.cell(row=QUARTET_ROW, column=c + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c + 3, value="$/SF SUBTOTALS")
        total = 3_000_000 + i * 120_000
        ws.cell(row=GT_ROW, column=c, value=total)
        for j, r in enumerate(DIV_ROWS):
            ws.cell(row=r, column=c + 1, value=10_000 + j * 100)
        ws.cell(row=CC_SUB_ROW, column=c, value=total * 0.85)
    wb.save(path)


def test_edge_matrix_gate_end_to_end(tmp_path, capsys):
    matrix = str(tmp_path / "edge.xlsx")
    baseline = str(tmp_path / "baseline.json")
    _build_edge_matrix(matrix, gsf=15_000)
    _baseline_json(baseline)
    common = (["--matrix", matrix, "--project-name", "EDGE 11x25 . Smoke",
               "--band-low", "3.2", "--band-high", "3.6", "--mid", "3.4",
               "--baseline", baseline]
              + _scoring_argv(tmp_path, [f"Restorer {i:02d}" for i in range(11)]))

    # preview: surfaces the suggested matrix SF, renders nothing
    rc = main(common + ["--preview-baseline"])
    assert rc == 0
    s = capsys.readouterr().out
    assert "15,000" in s and "SUGGESTED from matrix Row-10" in s

    # render with neither -> STOP exit 2 naming the matrix SF
    out1 = str(tmp_path / "o1")
    rc = main(common + ["--baseline-confirmed", "--html-only", "--no-audit",
                        "--out-dir", out1])
    assert rc == 2
    assert "15,000" in capsys.readouterr().err
    _no_artifacts(out1)

    # --sf-confirmed render uses the matrix SF for all 11 bidders
    out2 = str(tmp_path / "o2")
    rc = main(common + ["--sf-confirmed", "--baseline-confirmed", "--html-only",
                        "--out-dir", out2])
    assert rc == 0
    rj = _run_json(out2)
    assert len(rj["bidders"]) == 11
    r07 = next(b for b in rj["bidders"] if "Restorer 07" in b["name"])
    assert r07["per_sf"] == round((3_000_000 + 7 * 120_000) / 15_000)

    # explicit-override render uses the supplied basis instead
    out3 = str(tmp_path / "o3")
    rc = main(common + ["--sf-basis", "12000", "--baseline-confirmed",
                        "--html-only", "--out-dir", out3])
    assert rc == 0
    rj3 = _run_json(out3)
    r07b = next(b for b in rj3["bidders"] if "Restorer 07" in b["name"])
    assert r07b["per_sf"] == round((3_000_000 + 7 * 120_000) / 12_000)
