"""Shared fixtures + SYNTHETIC sample-card ground truth.

All firms and figures here are fictional. The end-to-end validation matrix is a
client binary that is gitignored and absent from the shipped bundle, so the
integration tests that need it skip; the constants below describe the synthetic
sample card the suite validates against."""
import json
import os

import pytest

HERE = os.path.dirname(os.path.abspath(__file__))            # .../tests
SKILL_ROOT = os.path.dirname(HERE)                            # .../scorecard
# The client Inputs/ folder lives under the engine project root, which is the
# PARENT of the engine package — NOT two levels up (that over-climbed one level
# higher, so the matrix xlsx never resolved and every gold test silently
# SKIPPED — false green). Resolve absolutely against the project root.
PROJECT_ROOT = os.path.dirname(SKILL_ROOT)                    # the engine project root

SAMPLE_XLSX = os.path.join(
    PROJECT_ROOT, "Inputs",
    "Sample Condominium - Bid Comparison Matrix.xlsx")
BASELINE_JSON = os.path.join(SKILL_ROOT, "examples", "sample_baseline.json")
GOLD_OVERRIDES_JSON = os.path.join(
    SKILL_ROOT, "examples", "sample_gold_overrides.json")
# the two REQUIRED per-run scoring xlsx inputs — synthetic sample validation
# examples (framework = the 8 Falke rows; scores = the sample card's Section-E
# values). Fully fictional firms/figures; no client data.
SAMPLE_FRAMEWORK_XLSX = os.path.join(
    SKILL_ROOT, "examples", "sample_scoring_framework.xlsx")
SAMPLE_SCORES_XLSX = os.path.join(
    SKILL_ROOT, "examples", "sample_category_scores.xlsx")

# bidders excluded by a §1.4 set-aside ruling (applied via --exclude)
GOLD_EXCLUSIONS = ["Harbor Builders Inc.", "Borealis Builders Solutions"]

# display-name aliases (Marvin §1.5): the matrix carries full legal firm names
# but the sample board card uses short names. Keys are matched on the normalized
# raw/display name; the raw matrix name is retained in the run log for audit.
# (A dotted acronym is already handled by display_name normalization, so it is
# NOT listed here.)
GOLD_ALIASES = {
    "Acme Restoration": "Acme",
    "Granite Remodel Group": "Granite",
    "Harbor Builders Inc.": "Harbor",
    "Borealis Builders Solutions": "Borealis",
}

# HISTORICAL sample-card Overall (curve ON, 100% coverage; Cascade via +5
# bonus). The curve was RETIRED (P0-6) — these no longer reproduce from a live
# run, and that is correct; kept only as the archived-card record.
GOLD_OVERALL = {
    "Acme": 84, "Borealis": 82, "Cascade": 75, "Dorne": 69,
    "Crest": 65, "Fjord": 56, "Granite": 51,
}

# THE CURRENT CONTRACT (post-P0-6): Overall = the honest weighted average.
# Published synthetic wavg per firm (the modeling anchor wavg column).
GOLD_WAVG = {
    "Acme": 82.0, "Borealis": 80.0, "Crest": 72.0, "Cascade": 70.0,
    "Dorne": 69.0, "Fjord": 47.0, "Granite": 39.0,
}
# ranking under the raw weighted average — note Crest (#3) sits ABOVE Cascade
# and Dorne here; the retired curve demoted Crest to #5 via the $/SF penalty.
GOLD_RANK_ORDER_RAW = ["Acme", "Borealis", "Crest", "Cascade", "Dorne",
                       "Fjord", "Granite"]

# ---- sample-card run parameters ----
SF_BASIS = 16000
BAND_LOW = 3.35
BAND_HIGH = 3.55
MID = 3.40
VARIANCE_MID = 3.45

# ---- sample-card grand-total row totals for the 7 KEPT bidders ----
# (drops Harbor, Borealis, duplicate Dorne; keeps Dorne J)
GOLD_TOTALS = {
    "Crest": 4400000,
    "Dorne": 3680000,
    "Fjord": 2080000,
    "Granite": 1950000,
    "Acme": 3360000,
    "Borealis": 3370000,
    "Cascade": 3050000,
}

# ---- sample-card $/SF (Marvin §3) ----
GOLD_PER_SF = {
    "Crest": 275,
    "Dorne": 230,
    "Borealis": 211,
    "Acme": 210,
    "Cascade": 191,
    "Fjord": 130,
    "Granite": 122,
}

# ---- sample-card tiers (Marvin §4.1) ----
GOLD_TIERS = {
    "Borealis": "TOP",
    "Acme": "TOP",
    "Cascade": "MID",
    "Fjord": "RISK",
    "Granite": "RISK",
    "Dorne": "DEFENSIVE",
    "Crest": "PREMIUM",
}

# ---- dropped bidders (must NOT appear in included field) ----
DROPPED = {"Harbor", "Mc Bride Builders", "Borealis Builders Solutions"}

# ---- HISTORICAL sample card ranking order (Marvin §9, curve ON — retired) ----
GOLD_RANK_ORDER = ["Acme", "Borealis", "Cascade", "Dorne", "Crest", "Fjord", "Granite"]


@pytest.fixture(scope="session")
def sample_xlsx_available():
    return os.path.exists(SAMPLE_XLSX)


@pytest.fixture
def baseline_lines():
    with open(BASELINE_JSON, "r", encoding="utf-8") as fh:
        return json.load(fh)


@pytest.fixture
def run_overrides():
    """sf_basis + band overrides for the synthetic sample gold run."""
    return {
        "sf_basis": SF_BASIS,
        "band_low": BAND_LOW,
        "band_high": BAND_HIGH,
        "modeled_mid_takeoff": MID,
        "variance_mid": VARIANCE_MID,
    }


# ---------------------------------------------------------------------------
# xlsx builders for the two REQUIRED scoring inputs (shared by the CLI tests —
# self-contained synthetic files, no client data)
# ---------------------------------------------------------------------------
# a minimal valid 3-category framework for synthetic CLI runs (sums to 100)
SIMPLE_FRAMEWORK_ROWS = [
    ("Market-aligned pricing", "Pricing", 50, "Closeness to baseline."),
    ("Scope completeness", "Scope", 30, "Inclusions/exclusions quality."),
    ("Documentation quality", "Docs", 20, "Form completeness."),
]


def write_framework_xlsx(path, rows=None):
    """Write a Scoring_Framework xlsx in the template layout.

    rows: [(category, short_label, weight_pct, captures), ...]
    """
    import openpyxl
    rows = SIMPLE_FRAMEWORK_ROWS if rows is None else rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoring_Framework"
    ws.cell(row=1, column=1, value="SCORING FRAMEWORK — test fixture")
    for c, h in enumerate(
            ["Category", "Short Label", "Weight (%)", "What it captures"],
            start=1):
        ws.cell(row=2, column=c, value=h)
    for i, (cat, lab, w, cap) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=lab)
        ws.cell(row=r, column=3, value=w)
        ws.cell(row=r, column=4, value=cap)
    wb.save(path)
    return path


def write_scores_xlsx(path, labels, firm_rows):
    """Write a Category_Scores xlsx in the template layout.

    labels: score column headers (the framework short labels).
    firm_rows: [(firm_name, [score, ...]), ...] aligned with labels.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category_Scores"
    ws.cell(row=1, column=1, value="DETAILED CATEGORY SCORES — test fixture")
    for c, h in enumerate(["Firm"] + list(labels), start=1):
        ws.cell(row=2, column=c, value=h)
    for i, (firm, scores) in enumerate(firm_rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=firm)
        for j, s in enumerate(scores, start=2):
            ws.cell(row=r, column=j, value=s)
    wb.save(path)
    return path
