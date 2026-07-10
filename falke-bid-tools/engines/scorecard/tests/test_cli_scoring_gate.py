"""Scoring-inputs gate — CLI process-level contract (REQUIRED, NO fallback).

Every real render needs TWO Falke-filled xlsx inputs: the scoring framework
(--scoring-framework) and the detailed category scores (--category-scores).
There is NO default and nothing is reused from a previous run — every run has
different bidders and may carry a different framework. This file locks the
contract end-to-end through cli.main():

  * a render WITHOUT --scoring-framework -> [STOP] exit 2, no artifacts;
  * a render WITHOUT --category-scores   -> [STOP] exit 2, no artifacts;
  * --overrides (the superseded qual-scores JSON) -> [STOP] exit 2;
  * --preview-baseline does NOT require the two files (baseline/SF preview);
  * scores firms must match the SCORED bidder field (mismatches listed);
  * a MODIFIED framework (6 categories, different weights) renders Sections
    D/E dynamically — the card carries the run's categories/weights, not a
    hardcoded 8.

Self-contained: builds a small synthetic matrix in the real structural format
(no client data; the client xlsx is gitignored in builder worktrees).
"""
from __future__ import annotations

import json
import os

import openpyxl

from scorecard.cli import main
from .conftest import write_framework_xlsx, write_scores_xlsx

PROJECT = "SCORING-GATE TEST · Synthetic"
SF_BASIS = 10_000
BAND_LOW, BAND_HIGH, MID = 2.00, 2.20, 2.10

SYNTH_BIDDERS = [
    ("Apex Builders LLC", 2_100_000, [300_000, 250_000, 180_000]),
    ("Borealis Group",    1_900_000, [240_000, 210_000, 150_000]),
]

# a 6-category framework with DIFFERENT weights than the Falke 8 — proves the
# render is data-driven, not hardcoded.
SIX_CAT_ROWS = [
    ("Market-aligned pricing", "Pricing", 30, "Closeness to baseline."),
    ("Scope completeness", "Scope", 20, "Inclusions/exclusions quality."),
    ("Schedule realism", "Schedule", 15, "Duration and phasing credibility."),
    ("Safety program", "Safety", 15, "EMR history and site protection."),
    ("Warranty strength", "Warranty", 10, "Coverage beyond statutory."),
    ("Documentation quality", "Docs", 10, "Form completeness."),
]
SIX_CAT_LABELS = [r[1] for r in SIX_CAT_ROWS]


def _build_matrix(path: str) -> None:
    """Minimal 2-bidder matrix in the real structural format (mirrors
    tests/test_cli_baseline_gate.py)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"
    NAME_ROW, QUARTET_ROW = 4, 8
    DIV_ROWS = [11, 12, 13]
    CC_SUBTOTAL_ROW, GT_ROW, GSF_ROW = 15, 18, 20
    FIRST_COL, STRIDE = 4, 5

    ws.cell(row=1, column=1, value="SYNTHETIC — scoring-gate test matrix")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUBTOTAL_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    for r, lab in zip(DIV_ROWS, ["GENERAL CONDITIONS SUBTOTAL",
                                 "CONCRETE SUBTOTAL", "FINISHES SUBTOTAL"]):
        ws.cell(row=r, column=2, value=lab)

    for i, (name, total, divs) in enumerate(SYNTH_BIDDERS):
        c = FIRST_COL + i * STRIDE
        ws.cell(row=NAME_ROW, column=c, value=name)
        ws.cell(row=QUARTET_ROW, column=c, value="COST")
        ws.cell(row=QUARTET_ROW, column=c + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c + 3, value="$/SF SUBTOTALS")
        ws.cell(row=GT_ROW, column=c, value=total)
        for r, dv in zip(DIV_ROWS, divs):
            ws.cell(row=r, column=c + 1, value=dv)
        ws.cell(row=CC_SUBTOTAL_ROW, column=c, value=total * 0.85)
    wb.save(path)


def _baseline_json(path: str) -> None:
    lines = [
        {"scope": "Interior demolition", "basis": "Modeled allowance",
         "cost": "$200,000", "value": 200_000},
        {"scope": "Direct trades subtotal", "basis": "Sum", "cost": "$200,000",
         "value": 200_000, "kind": "subtotal"},
    ]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(lines, fh)


def _setup(tmp_path):
    matrix = str(tmp_path / "matrix.xlsx")
    baseline = str(tmp_path / "baseline.json")
    _build_matrix(matrix)
    _baseline_json(baseline)
    return matrix, baseline


def _scoring_files(tmp_path, rows=SIX_CAT_ROWS, firms=None):
    labels = [r[1] for r in rows]
    fw = write_framework_xlsx(str(tmp_path / "framework.xlsx"), rows)
    firms = firms if firms is not None else [
        ("Apex Builders LLC", [9, 8, 7, 8, 7, 8][:len(labels)]),
        ("Borealis Group",    [6, 7, 6, 5, 6, 7][:len(labels)]),
    ]
    cs = write_scores_xlsx(str(tmp_path / "scores.xlsx"), labels, firms)
    return fw, cs


def _render_argv(matrix, baseline, out):
    return ["--matrix", matrix, "--project-name", PROJECT,
            "--sf-basis", str(SF_BASIS), "--band-low", str(BAND_LOW),
            "--band-high", str(BAND_HIGH), "--mid", str(MID),
            "--baseline", baseline, "--baseline-confirmed",
            "--html-only", "--no-audit", "--out-dir", out]


def _no_artifacts(out_dir):
    for fn in ("scorecard.pdf", "scorecard.html", "scorecard_run.json"):
        assert not os.path.exists(os.path.join(out_dir, fn)), fn


# ---------------------------------------------------------------------------
# hard-stops (exit 2, no-fallback messaging, no artifacts)
# ---------------------------------------------------------------------------
def test_render_without_scoring_framework_stops_exit2(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    _, cs = _scoring_files(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + ["--category-scores", cs])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[STOP] Scoring Framework not provided" in err
    assert "CANNOT be produced without it" in err
    assert "There is no default" in err
    assert "scoring-framework-template.xlsx" in err
    assert "--scoring-framework" in err
    _no_artifacts(out)


def test_render_without_category_scores_stops_exit2(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    fw, _ = _scoring_files(tmp_path)
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + ["--scoring-framework", fw])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[STOP] Detailed Category Scores not provided" in err
    assert "CANNOT be produced without them" in err
    assert "There is no default" in err
    assert "category-scores-template.xlsx" in err
    assert "--category-scores" in err
    _no_artifacts(out)


def test_overrides_superseded_stops_exit2(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    fw, cs = _scoring_files(tmp_path)
    ov = str(tmp_path / "overrides.json")
    with open(ov, "w", encoding="utf-8") as fh:
        json.dump({"Apex Builders LLC": {"pricing": {"score": 9}}}, fh)
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + [
        "--scoring-framework", fw, "--category-scores", cs, "--overrides", ov])
    assert rc == 2
    err = capsys.readouterr().err
    assert "--overrides is superseded" in err
    _no_artifacts(out)


def test_preview_baseline_does_not_require_scoring_files(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    out = str(tmp_path / "out")
    rc = main(["--matrix", matrix, "--project-name", PROJECT,
               "--sf-basis", str(SF_BASIS), "--band-low", str(BAND_LOW),
               "--band-high", str(BAND_HIGH), "--mid", str(MID),
               "--baseline", baseline, "--preview-baseline",
               "--out-dir", out])
    assert rc == 0
    s = capsys.readouterr().out
    assert "COST BASELINE PREVIEW" in s
    assert "No scorecard rendered" in s
    _no_artifacts(out)


def test_invalid_framework_stops_exit2_via_cli(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    bad_rows = [("A cat", "A", 60, ""), ("B cat", "B", 30, "")]  # sums to 90
    fw = write_framework_xlsx(str(tmp_path / "bad_fw.xlsx"), bad_rows)
    cs = write_scores_xlsx(str(tmp_path / "cs.xlsx"), ["A", "B"],
                           [("Apex Builders LLC", [9, 8]),
                            ("Borealis Group", [6, 7])])
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + [
        "--scoring-framework", fw, "--category-scores", cs])
    assert rc == 2
    assert "sum to 100" in capsys.readouterr().err
    _no_artifacts(out)


# ---------------------------------------------------------------------------
# firm cross-check vs the SCORED bidder field
# ---------------------------------------------------------------------------
def test_firm_mismatch_vs_matrix_stops_and_lists_names(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    # scores name a firm that is NOT in the matrix and OMIT Borealis
    fw, cs = _scoring_files(tmp_path, firms=[
        ("Apex Builders LLC", [9, 8, 7, 8, 7, 8]),
        ("Ghost Contracting", [6, 7, 6, 5, 6, 7]),
    ])
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + [
        "--scoring-framework", fw, "--category-scores", cs])
    assert rc == 2
    err = capsys.readouterr().err
    assert "do not match the scored bidder field" in err
    assert "Borealis Group" in err          # scored bidder without a row
    assert "Ghost Contracting" in err       # unknown firm with a row
    _no_artifacts(out)


def test_scores_row_for_excluded_bidder_is_an_error(tmp_path, capsys):
    matrix, baseline = _setup(tmp_path)
    fw, cs = _scoring_files(tmp_path)   # rows for BOTH synthetic bidders
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + [
        "--scoring-framework", fw, "--category-scores", cs,
        "--exclude", "Borealis Group"])
    assert rc == 2
    err = capsys.readouterr().err
    assert "unknown/excluded bidder" in err
    assert "Borealis Group" in err
    _no_artifacts(out)


# ---------------------------------------------------------------------------
# end-to-end dynamic render: 6 categories, different weights
# ---------------------------------------------------------------------------
def test_e2e_render_with_modified_framework_is_dynamic(tmp_path):
    matrix, baseline = _setup(tmp_path)
    fw, cs = _scoring_files(tmp_path)   # the 6-category framework
    out = str(tmp_path / "out")
    rc = main(_render_argv(matrix, baseline, out) + [
        "--scoring-framework", fw, "--category-scores", cs])
    assert rc == 0
    html_path = os.path.join(out, "scorecard.html")
    assert os.path.exists(html_path)
    with open(html_path, encoding="utf-8") as fh:
        html = fh.read()

    # Section D: every framework category + weight, dynamically rendered
    for cat, lab, w, cap in SIX_CAT_ROWS:
        assert cat in html, cat
        assert f"{w}%" in html, (lab, w)
    # Section E: every short label appears as a score column header
    for lab in SIX_CAT_LABELS:
        assert lab in html, lab
    # categories NOT in this run's framework must NOT appear
    for absent in ("Condo Exp", "CO Risk", "Reputation & longevity"):
        assert absent not in html, absent

    # the run JSON carries the xlsx-derived Overall (single source of truth),
    # computed from the xlsx weights: Apex = 9*.30 + 8*.20 + 7*.15 + 8*.15 +
    # 7*.10 + 8*.10 = 8.05 -> 80.5 on the /100 scale, at 100% coverage
    with open(os.path.join(out, "scorecard_run.json"), encoding="utf-8") as fh:
        rj = json.load(fh)
    apex = next(b for b in rj["bidders"] if "Apex" in b["name"])
    assert apex["overall"]["weighted_average"] == 80.5
    assert apex["overall"]["coverage"] == 1.0
