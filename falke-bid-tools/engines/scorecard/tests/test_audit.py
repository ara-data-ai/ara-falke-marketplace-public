"""Self-audit tests (Marvin's rubric C1..C16 + the P0-7 checks C17/C18).

Every check has >=1 PASS case (on a clean synthetic run) and >=1 FAIL/WARN case
(on a deliberately broken input). Fixtures are SYNTHETIC (a 4-bidder matrix
written to tmp_path) — the client xlsx is gitignored and absent from the
worktree, so nothing here requires it. The synthetic matrix mirrors the real
structural format (mirrors tests/test_generic_matrix.py).
"""
from __future__ import annotations

import copy

import openpyxl
import pytest

from scorecard import audit as A
from scorecard.config import load_config
from scorecard.mechanical import (TIER_DEFENSIVE, TIER_MID, TIER_RISK, TIER_TOP)
from scorecard.pipeline import audit_run, run_scorecard

# ---- SYNTHETIC PROJECT (intentionally unlike any client project) ----
SF_BASIS = 10000
BAND_LOW = 2.00       # band_low_per_sf  = 200
BAND_HIGH = 2.20      # band_high_per_sf = 220
MID = 2.10
PROJECT = "HARBORVIEW TOWER · Amenity Deck"
GSF = 8500            # planted matrix GSF (distinct from SF_BASIS)

# four bidders, one per tier vs the band above
BIDDERS = [
    ("Apex Builders LLC", 2_100_000, TIER_TOP),          # 210 -> TOP
    ("Borealis Group", 1_700_000, TIER_RISK),            # 170 -> RISK (<180 mid floor)
    ("Cascade Construction Co.", 2_500_000, TIER_DEFENSIVE),  # 250 -> DEFENSIVE
    ("Delta Contractors", 1_900_000, TIER_MID),          # 190 -> MID
]
# 18 division subtotals/block so C8's [16,20] band passes for a clean run
DIV_COUNT = 18


def _build_matrix(path, *, gsf=GSF, totals=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid_Comparison"
    NAME_ROW, QUARTET_ROW = 4, 8
    DIV_ROWS = list(range(11, 11 + DIV_COUNT))
    CC_SUB_ROW = DIV_ROWS[-1] + 2
    GT_ROW = CC_SUB_ROW + 2
    GSF_ROW = GT_ROW + 2
    FIRST_COL, STRIDE = 4, 5

    ws.cell(row=1, column=1, value="HARBORVIEW TOWER — Amenity Deck Bid Comparison")
    ws.cell(row=GT_ROW, column=2, value="GRAND TOTAL CONSTRUCTION COST")
    ws.cell(row=CC_SUB_ROW, column=2, value="CONSTRUCTION COST SUBTOTAL")
    ws.cell(row=GSF_ROW, column=2, value="TOTAL GSF")
    for i, r in enumerate(DIV_ROWS):
        ws.cell(row=r, column=2, value=f"DIVISION {i+1:02d} SUBTOTAL")

    totals = totals or {n: t for n, t, _ in BIDDERS}
    for i, (name, _t, _tier) in enumerate(BIDDERS):
        total = totals[name]
        c_cost = FIRST_COL + i * STRIDE
        ws.cell(row=NAME_ROW, column=c_cost, value=name)
        ws.cell(row=QUARTET_ROW, column=c_cost, value="COST")
        ws.cell(row=QUARTET_ROW, column=c_cost + 1, value="COST SUBTOTALS")
        ws.cell(row=QUARTET_ROW, column=c_cost + 2, value="$/SF")
        ws.cell(row=QUARTET_ROW, column=c_cost + 3, value="$/SF SUBTOTALS")
        ws.cell(row=GT_ROW, column=c_cost, value=total)
        for j, r in enumerate(DIV_ROWS):
            ws.cell(row=r, column=c_cost + 1, value=10000 + j * 100)
        ws.cell(row=CC_SUB_ROW, column=c_cost, value=total * 0.85)
        ws.cell(row=GSF_ROW, column=c_cost, value=gsf)
    wb.save(path)


def _cfg(**ov):
    base = {"sf_basis": SF_BASIS, "band_low": BAND_LOW, "band_high": BAND_HIGH,
            "modeled_mid_takeoff": MID}
    base.update(ov)
    return load_config(overrides=base)


def _run(tmp_path, *, gsf=GSF, totals=None, cfg=None):
    xlsx = str(tmp_path / "matrix.xlsx")
    _build_matrix(xlsx, gsf=gsf, totals=totals)
    cfg = cfg or _cfg()
    result = run_scorecard(xlsx, cfg, project_name=PROJECT)
    return result, cfg


# ============================================================================
# whole-audit PASS on a clean run + artifacts
# ============================================================================
def test_clean_run_audit_passes(tmp_path):
    result, cfg = _run(tmp_path)
    ar = A.audit(result["parsed"], cfg, result)
    blockers = [c for c in ar.checks if c.severity == A.BLOCKER and c.status == A.FAIL]
    assert not blockers, [c.verdict_line for c in blockers]
    assert ar.verdict in (A.V_PASS, A.V_WARN), ar.verdict
    assert ar.overall_status == A.PASS


def test_artifacts_written(tmp_path):
    import json
    import os
    result, cfg = _run(tmp_path)
    ar, paths = audit_run(result, cfg, str(tmp_path))
    assert os.path.exists(paths["report_md"])
    assert os.path.exists(paths["audit_json"])
    data = json.load(open(paths["audit_json"]))
    assert data["verdict"] in (A.V_PASS, A.V_WARN, A.V_FAIL)
    # C1..C16 + C17/C18 (P0-7) + C19..C22 (P1-4 run pack)
    assert len(data["checks"]) == 22
    md = open(paths["report_md"]).read()
    assert "Overall verdict:" in md
    assert "## Checks" in md


# ============================================================================
# C1 — Totals reconcile to Row 164
# ============================================================================
def test_c1_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c1(result["parsed"], cfg, result).status == A.PASS


def test_c1_fail_tampered_total(tmp_path):
    """Tamper a bidder's emitted total so it no longer matches Row 164."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]          # keep real parse (re-derivation source)
    r["bidders"][0]["total"] += 50_000      # tamper headline total
    c = A.check_c1(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C2 — $/SF math
# ============================================================================
def test_c2_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c2(result["parsed"], cfg, result).status == A.PASS


def test_c2_fail_bad_per_sf(tmp_path):
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"][0]["per_sf"] += 7          # drift the displayed $/SF
    c = A.check_c2(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C3 — SF basis is NOT the matrix GSF (must-never)
# ============================================================================
def test_c3_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c3(result["parsed"], cfg, result).status == A.PASS


def test_c3_fail_sf_basis_equals_gsf(tmp_path):
    """SF basis == matrix GSF is the documented trap -> BLOCKER."""
    cfg = _cfg(sf_basis=GSF, band_low=1.6, band_high=1.8, modeled_mid_takeoff=1.7)
    result, cfg = _run(tmp_path, cfg=cfg)
    c = A.check_c3(result["parsed"], cfg, result)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C4 — Tier matches band rule
# ============================================================================
def test_c4_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c4(result["parsed"], cfg, result).status == A.PASS


def test_c4_fail_wrong_tier(tmp_path):
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"][0]["tier"] = "RISK"        # flip a TOP bidder to RISK
    c = A.check_c4(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C5 — Ranking integrity
# ============================================================================
def test_c5_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c5(result["parsed"], cfg, result).status == A.PASS


def test_c5_fail_duplicate_rank(tmp_path):
    """Ranks exist only at FULL coverage now, so the fixture declares it."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["full_coverage"] = True
    for i, b in enumerate(r["bidders"], start=1):
        b["rank"] = i
    r["bidders"][0]["rank"] = r["bidders"][1]["rank"]   # duplicate rank
    c = A.check_c5(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ---- C5 regression: partial-coverage equal-Overall ties must NOT FAIL -------
def _partial_result(*, ranked: bool):
    """A partial-coverage result — with or without ranks."""
    def bd(name, total, rank):
        b = {"name": name, "total": total, "per_sf": int(round(total / SF_BASIS)),
             "tier": "MID", "bid_m": total / 1e6,
             "overall": {"numeric": None, "applied": False, "coverage": 0.6,
                         "weighted_average": None,
                         "display": "Pending — 3 of 8 categories outstanding"},
             "section_c": {}, "flags": []}
        if ranked:
            b["rank"] = rank
        return b
    bidders = [bd("Alpha", 1_800_000, 1), bd("Bravo", 1_900_000, 2),
               bd("Charlie", 2_000_000, 3), bd("Delta", 2_100_000, 4)]
    return {"meta": {"run_id": "t"}, "bidders": bidders, "log": [],
            "fingerprints": [], "full_coverage": False}


def test_c5_partial_coverage_no_rank_passes():
    """Marvin P1-2 §4.3 — C5's partial-coverage SPECIAL CASE IS DELETED. It used
    to accept equal-Overall ties "at partial coverage"; there is nothing to tie
    now, because a provisional run carries no Overall and no rank. C5's contract
    at partial coverage is simply: no bidder carries a rank, and THE ABSENCE IS
    THE CHECK. This ruling removes a special case rather than adding one."""
    r = _partial_result(ranked=False)
    c = A.check_c5(None, _cfg(), r)
    assert c.status == A.PASS, c.verdict_line
    assert c.evidence["full_coverage"] is False


def test_c5_partial_coverage_with_a_rank_blocks():
    """A rank computed on ragged coverage orders the evaluator's calendar, not
    the bidders — the engine does not rescale, so an unscored category costs its
    full weight and whoever got scored first floats up."""
    r = _partial_result(ranked=True)
    c = A.check_c5(None, _cfg(), r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "no bidder may carry a rank" in c.verdict_line


def _full_coverage_result():
    """A ranked, fully-scored result: equal Overall (63), tiebreak orders by
    ascending total -> ranks 1..4."""
    r = _partial_result(ranked=True)
    r["full_coverage"] = True
    for b in r["bidders"]:
        b["overall"].update({"coverage": 1.0, "numeric": 63,
                             "weighted_average": 63, "display": "63"})
    return r


def test_c5_full_coverage_tiebreak_still_blocks():
    """TRUE POSITIVE kept: at FULL coverage, equal Overall with a HIGHER total
    ranked above a lower total is a real tiebreak violation -> BLOCKER."""
    r = _full_coverage_result()
    # invert the tiebreak: put the HIGHEST total at rank 1
    r["bidders"][0]["rank"], r["bidders"][3]["rank"] = 4, 1
    c = A.check_c5(None, _cfg(), r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER, c.verdict_line


def test_c5_inversion_blocks_at_full_coverage():
    """TRUE POSITIVE kept: a genuine Overall inversion (higher rank, strictly
    lower Overall) is a BLOCKER.

    It used to be asserted at PARTIAL coverage, which is no longer a reachable
    state: a provisional run carries no Overall to invert and no rank to invert
    it against. The true positive itself is unchanged and still guarded here."""
    r = _full_coverage_result()
    r["bidders"][0]["overall"]["numeric"] = 40   # rank 1 now BELOW rank 2's 63
    c = A.check_c5(None, _cfg(), r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER, c.verdict_line


# ============================================================================
# C6 — Variance signs / magnitudes
# ============================================================================
def test_c6_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c6(result["parsed"], cfg, result).status == A.PASS


def test_c6_fail_sign_flip(tmp_path):
    """An over-band bidder (per_sf high) forced to a bid_m under mid -> the
    re-derived variance sign disagrees with the $/SF-vs-mid sign."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    cas = next(b for b in r["bidders"] if b["per_sf"] >= 220)
    cas["bid_m"] = MID - 0.30               # bid under mid while per_sf stays high
    c = A.check_c6(r["parsed"], cfg, r)
    assert c.status == A.FAIL


# ============================================================================
# C7 — Duplicate handling logged
# ============================================================================
def test_c7_pass_no_dups(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c7(result["parsed"], cfg, result).status == A.PASS


def _dup_parsed(*, drop_one=True, keep="first"):
    """Two blocks share a normalized name. When drop_one is True the later block
    is marked included=False with a duplicate drop_reason (mirrors the matrix
    dedup), modelling the recurring-firm duplicate situation: one survivor, one
    dropped col-AX block. When drop_one is False BOTH stay included (genuine failure)."""
    from scorecard.matrix import (SUB_COST, BidderBlock, ParsedMatrix,
                                  normalize_name)

    def blk(raw, col, total, included=True, drop=None):
        return BidderBlock(raw_name=raw, name=raw, norm=normalize_name(raw),
                           start_col=col, cols={SUB_COST: col}, grand_total=total,
                           included=included, drop_reason=drop)
    kept = blk("Dorne", 10, 2_100_000)
    dropped = blk("Dorne", 50, 2_050_000,
                  included=not drop_one,
                  drop=("duplicate of 'Dorne' (col J); kept left-most" if drop_one else None))
    return ParsedMatrix(
        sheet_name="X", header_row=4, block_width=4, block_stride=5,
        grand_total_row=18, grand_total_label="GRAND TOTAL", gsf_value=8500,
        gsf_row=20, blocks=[kept, dropped], division_rows=[(11, "D1")])


def _dup_firm_result(*, log_drop=True, second_included=False):
    bidders = [{"name": "Dorne", "total": 2_100_000, "per_sf": 210,
                "tier": "TOP", "bid_m": 2.10, "rank": 1,
                "overall": {"numeric": 80, "applied": False, "coverage": 1.0,
                            "weighted_average": 80, "display": "80"},
                "section_c": {}, "flags": []}]
    if second_included:   # genuine failure: a SECOND scored 'Dorne' survives
        b2 = dict(bidders[0]); b2 = copy.deepcopy(bidders[0]); b2["rank"] = 2
        bidders.append(b2)
    log = []
    if log_drop:
        log = ["DUPLICATE: 'Dorne' col AX (total 2,050,000) DROPPED; kept col J "
               "(total 2,100,000). OPEN QUESTION: totals differ by 50,000."]
    return {"meta": {"run_id": "t"}, "bidders": bidders, "log": log,
            "fingerprints": [], "full_coverage": True}


def test_c7_recurring_firm_dup_one_survivor_logged_passes():
    """REGRESSION (false BLOCKER): the dropped duplicate (included=False) must NOT
    count as a survivor. Exactly one 'Dorne' survives AND the drop is logged ->
    C7 PASS (this is the live condition that wrongly BLOCKED)."""
    parsed = _dup_parsed(drop_one=True)
    result = _dup_firm_result(log_drop=True, second_included=False)
    c = A.check_c7(parsed, _cfg(), result)
    assert c.status == A.PASS, c.verdict_line
    assert c.evidence["duplicates"][0]["kept_in_output"] == 1
    assert c.evidence["duplicates"][0]["logged"] is True


def test_c7_fail_two_survivors_both_included():
    """TRUE POSITIVE: two same-named bidders BOTH included/scored -> 2 survivors
    -> BLOCKER (even with a drop line in the log)."""
    parsed = _dup_parsed(drop_one=False)          # both included
    result = _dup_firm_result(log_drop=True, second_included=True)
    c = A.check_c7(parsed, _cfg(), result)
    assert c.status == A.FAIL and c.severity == A.BLOCKER, c.verdict_line


def test_c7_fail_silent_dup(tmp_path):
    """Two blocks share a normalized name but the run log carries no DUPLICATE
    drop line -> silent dedup BLOCKER."""
    from scorecard.matrix import (SUB_COST, BidderBlock, ParsedMatrix,
                                  normalize_name)

    def blk(raw, col, total):
        return BidderBlock(raw_name=raw, name=raw, norm=normalize_name(raw),
                           start_col=col, cols={SUB_COST: col}, grand_total=total)
    parsed = ParsedMatrix(
        sheet_name="X", header_row=4, block_width=4, block_stride=5,
        grand_total_row=18, grand_total_label="GRAND TOTAL", gsf_value=8500,
        gsf_row=20, blocks=[blk("Apex", 4, 2_100_000), blk("Apex", 9, 2_100_000)],
        division_rows=[(11, "D1")])
    result = {
        "meta": {"run_id": "t"},
        "bidders": [{"name": "Apex", "total": 2_100_000, "per_sf": 210,
                     "tier": "TOP", "bid_m": 2.10, "rank": 1,
                     "overall": {"numeric": 80, "applied": False, "coverage": 1.0,
                                 "weighted_average": 80, "display": "80"},
                     "section_c": {}, "flags": []}],
        "log": [],   # NO duplicate log line -> silent drop
        "fingerprints": [],
    }
    c = A.check_c7(parsed, _cfg(), result)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C8 — Completeness flags (false-zero trap + out-of-band WARN)
# ============================================================================
def test_c8_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c8(result["parsed"], cfg, result).status == A.PASS


def test_c8_blocker_false_zero(tmp_path):
    result, cfg = _run(tmp_path)
    result["parsed"].included_blocks[0].populated_divisions = 0   # the false zero
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


def test_c8_warn_out_of_band(tmp_path):
    result, cfg = _run(tmp_path)
    result["parsed"].included_blocks[0].populated_divisions = 5   # below [16,20]
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.FAIL and c.severity == A.WARN


def test_c8_inclusive_floor_16_passes(tmp_path):
    """REGRESSION (off-by-one WARN): 16 is the INCLUSIVE band floor; 16<=pd<=20
    must PASS. The prior bound treated 16 as out_of_band."""
    result, cfg = _run(tmp_path)
    for b in result["parsed"].included_blocks:
        b.populated_divisions = 16
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.PASS, c.verdict_line


def test_c8_inclusive_ceiling_20_passes(tmp_path):
    """REGRESSION: 20 is the INCLUSIVE band ceiling -> PASS."""
    result, cfg = _run(tmp_path)
    for b in result["parsed"].included_blocks:
        b.populated_divisions = 20
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.PASS, c.verdict_line


def test_c8_below_floor_still_warns(tmp_path):
    """TRUE POSITIVE kept: 15 (just under the inclusive floor) still WARNs —
    the completeness disclosure working, as on a genuinely-low bidder
    (13/14)."""
    result, cfg = _run(tmp_path)
    result["parsed"].included_blocks[0].populated_divisions = 15
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.FAIL and c.severity == A.WARN, c.verdict_line


def test_c8_above_ceiling_still_warns(tmp_path):
    """TRUE POSITIVE kept: 21 (just over the inclusive ceiling) still WARNs."""
    result, cfg = _run(tmp_path)
    result["parsed"].included_blocks[0].populated_divisions = 21
    c = A.check_c8(result["parsed"], cfg, result)
    assert c.status == A.FAIL and c.severity == A.WARN, c.verdict_line


# ============================================================================
# C9 — Fingerprint disclosure
# ============================================================================
def test_c9_pass_no_fingerprints(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c9(result["parsed"], cfg, result).status == A.PASS


def test_c9_fail_undisclosed(tmp_path):
    from scorecard.mechanical import FingerprintHit
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["fingerprints"] = [FingerprintHit("Demo", 100.0, "Apex Builders LLC",
                                        "D1", 100.2, 0.2)]
    r["log"] = ["run start"]   # no FINGERPRINT: line
    c = A.check_c9(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.WARN


# ============================================================================
# C10 — No silently auto-dropped bidder
# ============================================================================
def test_c10_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c10(result["parsed"], cfg, result).status == A.PASS


def test_c10_fail_silent_drop(tmp_path):
    """A matrix bidder absent from output with no ruling/duplicate log -> BLOCKER."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"] = r["bidders"][1:]   # drop one bidder from output, no log reason
    r["log"] = ["run start"]
    c = A.check_c10(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C11 — Overall = honest weighted average (curve retired, P0-6)
# ============================================================================
def test_c11_pass_overall_is_raw_wavg(tmp_path):
    """Clean run: every emitted Overall re-derives as the raw weighted average
    of its category scores."""
    result, cfg = _run(tmp_path)
    c = A.check_c11(result["parsed"], cfg, result)
    assert c.status == A.PASS, c.verdict_line


def test_c11_fail_adjusted_overall_blocks(tmp_path):
    """Anything that nudges the ranked number away from the recomputed raw
    weighted average (a resurrected curve, a bonus, a manual edit) is a
    BLOCKER — P0-6, Floyd verdict d."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    b = r["bidders"][0]
    b["overall"]["numeric"] = (b["overall"]["numeric"] or 0) + 5.0  # "adjusted"
    c = A.check_c11(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C12 — Qualitative coverage gating
# ============================================================================
def test_c12_pass(tmp_path):
    """Default partial-coverage run: must be flagged provisional, not curved."""
    result, cfg = _run(tmp_path)
    assert A.check_c12(result["parsed"], cfg, result).status == A.PASS


def _partial(result):
    """A partial-coverage copy of a run, with every emitted coverage figure
    RE-DERIVED from the scores actually present.

    The base fixture is the legacy scaffold path, which is ALREADY partial — so
    an earlier draft of this helper that assumed 1.0 and wrote `1.0 - weight`
    produced a coverage figure that did not match its own scores, and C12(e)
    caught it. That is the check doing exactly its job on a lying fixture; the
    honest thing is to derive the number rather than assert it."""
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["full_coverage"] = False
    r["watermark"] = [{"token": "evaluation incomplete", "detail": "x"}]
    key = r["categories"][0]["key"]
    r["bidders"][0]["scores"][key] = None
    weights = {c["key"]: c["weight_pct"] / 100.0 for c in r["categories"]}
    for b in r["bidders"]:
        b["overall"]["coverage"] = round(
            sum(w for k, w in weights.items()
                if (b.get("scores") or {}).get(k) is not None), 6)
        b["overall"]["numeric"] = None
        b.pop("rank", None)
    return r


def test_c12_partial_pass_when_no_claim_is_made(tmp_path):
    """C12's new contract (Marvin P1-2 §4.2): the document makes no claim the
    run is not entitled to. It stops policing a label and starts re-deriving an
    ABSENCE — absences are cheap to check and impossible to fudge."""
    result, cfg = _run(tmp_path)
    r = _partial(result)
    c = A.check_c12(r["parsed"], cfg, r, summary_context={"winner_name": ""})
    assert c.status == A.PASS, c.verdict_line


def test_c12_blocks_a_rank_at_partial_coverage(tmp_path):
    result, cfg = _run(tmp_path)
    r = _partial(result)
    r["bidders"][0]["rank"] = 1              # a rank the run is not entitled to
    c = A.check_c12(r["parsed"], cfg, r, summary_context={"winner_name": ""})
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "carry a rank" in c.verdict_line


def test_c12_blocks_an_overall_at_partial_coverage(tmp_path):
    result, cfg = _run(tmp_path)
    r = _partial(result)
    r["bidders"][0]["overall"]["numeric"] = 82.0
    c = A.check_c12(r["parsed"], cfg, r, summary_context={"winner_name": ""})
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "carry an Overall" in c.verdict_line


def test_c12_blocks_a_named_leader_at_partial_coverage(tmp_path):
    """A front-runner is a ranking claim — §3.4."""
    result, cfg = _run(tmp_path)
    r = _partial(result)
    c = A.check_c12(r["parsed"], cfg, r,
                    summary_context={"winner_name": "Acme Restoration"})
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "names a leader" in c.verdict_line


def test_c12_blocks_a_missing_watermark_at_partial_coverage(tmp_path):
    result, cfg = _run(tmp_path)
    r = _partial(result)
    r["watermark"] = []                      # never marked PRELIMINARY
    c = A.check_c12(r["parsed"], cfg, r, summary_context={"winner_name": ""})
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "watermark" in c.verdict_line


def test_c12_blocks_a_coverage_figure_that_does_not_re_derive(tmp_path):
    """(e) — coverage is RE-DERIVED from the scored cells against the framework
    weights, never echoed."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"][0]["overall"]["coverage"] = 0.42     # a lie
    c = A.check_c12(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER
    assert "does not re-derive" in c.verdict_line


# ============================================================================
# C13 — PII / scope discipline
# ============================================================================
def test_c13_pass_clean_report(tmp_path):
    result, cfg = _run(tmp_path)
    ar = A.audit(result["parsed"], cfg, result)
    report = A.render_report_md(ar, result["parsed"], cfg, result)
    assert A.check_c13(result["parsed"], cfg, result, report_text=report).status == A.PASS


def test_c13_fail_ssn_leak(tmp_path):
    result, cfg = _run(tmp_path)
    leaked = "Some narrative with an SSN 123-45-6789 in it."
    c = A.check_c13(result["parsed"], cfg, result, report_text=leaked)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


def test_c13_fail_email_leak(tmp_path):
    result, cfg = _run(tmp_path)
    c = A.check_c13(result["parsed"], cfg, result,
                    report_text="contact bidder@example.com for terms")
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C14 — Bucket separation
# ============================================================================
def test_c14_pass(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c14(result["parsed"], cfg, result).status == A.PASS


def test_c14_fail_bucket_leak(tmp_path):
    """Total inflated above Row 164 (as if alternates were folded in) -> BLOCKER."""
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"][0]["total"] += 120_000     # alternate/allowance fold-in shape
    c = A.check_c14(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.BLOCKER


# ============================================================================
# C15 — Alias round-trip
# ============================================================================
def test_c15_pass_no_alias(tmp_path):
    result, cfg = _run(tmp_path)
    assert A.check_c15(result["parsed"], cfg, result).status == A.PASS


def test_c15_pass_with_alias(tmp_path):
    xlsx = str(tmp_path / "m.xlsx")
    _build_matrix(xlsx)
    cfg = _cfg()
    aliases = {"Apex Builders LLC": "Apex"}
    result = run_scorecard(xlsx, cfg, project_name=PROJECT, aliases=aliases)
    c = A.check_c15(result["parsed"], cfg, result, aliases=aliases)
    assert c.status == A.PASS


def test_c15_fail_untraceable(tmp_path):
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    r["bidders"][0]["name"] = "Phantom Co"   # not in matrix, no alias
    c = A.check_c15(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.WARN


# ============================================================================
# C16 — Exclusion register coverage
# ============================================================================
def test_c16_info_when_no_narrative(tmp_path):
    result, cfg = _run(tmp_path)
    c = A.check_c16(result["parsed"], cfg, result)
    assert c.severity == A.INFO and c.status == A.PASS


def test_c16_fail_missing_register(tmp_path):
    result, cfg = _run(tmp_path)
    r = copy.deepcopy(result)
    r["parsed"] = result["parsed"]
    # a bidder carries exclusion narrative but the register has NO matching entry
    r["exclusions_text"] = {"Apex Builders LLC": "Excludes hazmat abatement."}
    r["exclusion_register"] = {}
    c = A.check_c16(r["parsed"], cfg, r)
    assert c.status == A.FAIL and c.severity == A.WARN
